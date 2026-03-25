import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print('Missing dependency: openpyxl')
    print('Install with: pip3 install openpyxl')
    sys.exit(1)


ROOT_DIR = Path(__file__).resolve().parent
EXCEL_DIR = ROOT_DIR / 'rate' / 'excel'
WEBP_DIR = ROOT_DIR / 'rate' / 'webp'
DEFAULT_IMAGE = 'linshi.webp'
LOCAL_IMAGE_PREFIX = '/rate/webp/'

TARGET_FILES = [
    {
        'file_name': 'reel.xlsx',
        'mode': 'id',
    },
    {
        'file_name': 'rod.xlsx',
        'mode': 'existing',
    },
    {
        'file_name': 'lure.xlsx',
        'mode': 'existing',
    },
]


def normalize_numeric_string(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def resolve_target_image_name(row, mode):
    if mode == 'id':
        raw_id = row[0].value
        if raw_id is None:
            return DEFAULT_IMAGE

        file_name = f'{normalize_numeric_string(raw_id)}.webp'
        return file_name if (WEBP_DIR / file_name).exists() else DEFAULT_IMAGE

    current_value = str(row[7].value or '').strip()
    if not current_value:
        return DEFAULT_IMAGE

    current_name = Path(current_value).name
    if not current_name:
        return DEFAULT_IMAGE

    return current_name if (WEBP_DIR / current_name).exists() else DEFAULT_IMAGE


def normalize_excel_file(file_name, mode):
    excel_path = EXCEL_DIR / file_name
    if not excel_path.exists():
        print(f'Skip: {excel_path} not found.')
        return

    print(f'Loading workbook: {excel_path}')
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    max_row = worksheet.max_row
    print(f'Processing {file_name}, rows={max_row - 1}')

    for row_index in range(2, max_row + 1):
        row = worksheet[row_index]
        target_image_name = resolve_target_image_name(row, mode)
        target_value = f'{LOCAL_IMAGE_PREFIX}{target_image_name}'
        worksheet.cell(row=row_index, column=8).value = target_value

    workbook.save(excel_path)
    print(f'Saved workbook: {excel_path}')

    csv_path = excel_path.with_suffix('.csv')
    with csv_path.open('w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(list(row))
    print(f'Updated CSV: {csv_path}')


def main():
    print('Start normalizing gear Excel image paths...')
    print(f'Excel directory: {EXCEL_DIR}')
    print(f'Image directory: {WEBP_DIR}')
    print(f'Path prefix: {LOCAL_IMAGE_PREFIX}')

    for target in TARGET_FILES:
        normalize_excel_file(target['file_name'], target['mode'])

    print('Done.')


if __name__ == '__main__':
    main()
