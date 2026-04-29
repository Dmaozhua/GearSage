from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")
TARGET_ID = "DRD10144"
TARGET_FIELD = "guide_use_hint"
NEW_HINT = "SC 全能底操：Neko、No Sinker、Down Shot、Texas 和 Free Rig 為主，實心竿尖也能兼顧 Spinnerbait、Vibration、Crankbait 的鬆線回收與觸感。"


def n(value):
    return str(value or "").strip()


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    for field in ["id", TARGET_FIELD]:
        if field not in col:
            raise RuntimeError(f"missing rod_detail column: {field}")

    changed = []
    for row in range(2, ws.max_row + 1):
        if n(ws.cell(row, col["id"]).value) != TARGET_ID:
            continue
        cell = ws.cell(row, col[TARGET_FIELD])
        old = n(cell.value)
        if old != NEW_HINT:
            cell.value = NEW_HINT
            changed.append({"id": TARGET_ID, "field": TARGET_FIELD, "old": old, "new": NEW_HINT})
        break
    else:
        raise RuntimeError(f"target detail row not found: {TARGET_ID}")

    wb.save(XLSX_PATH)
    print({"file": str(XLSX_PATH), "changed": changed})


if __name__ == "__main__":
    main()
