import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = ROOT / "GearSage-client/pkgGear/data_raw/megabass_rod_import.xlsx"
REPORT_PATH = ROOT / "GearSage-client/pkgGear/data_raw/megabass_rod_master_field_optimization_report.json"

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")


MASTER_UPDATES = {
    "ARMS SUPER LEGGERA": {
        "type_tips": "鲈鱼 / 枪柄 / 高端旗舰",
        "player_positioning": "高端鲈钓 / ARMS 定制血统 / 强力泛用",
        "player_selling_points": "高密度复合空白与定制握把 / 覆盖搜索饵、大饵和强力场景 / 偏高端玩家与收藏实钓",
    },
    "ARMS SUPER LEGGERA SPINNING": {
        "type_tips": "鲈鱼 / 直柄 / 高端精细",
        "player_positioning": "高端鲈钓 / 直柄精细 / 轻线操控",
        "player_selling_points": "ARMS 高端空白与轻量手感 / 面向 finesse、轻线和细腻控饵 / 适合重视感度与操控的进阶玩家",
    },
    "EVOLUZION": {
        "type_tips": "鲈鱼 / 枪柄 / 钛纤维复合",
        "player_positioning": "高端鲈钓 / 钛纤维高感度 / 精细到卷阻饵",
        "player_selling_points": "钛纤维复合带来清晰反馈和韧性 / 轻量控饵、moving bait 与强力型号并存 / 适合重视手感的进阶玩家",
    },
    "DESTROYER T.S": {
        "type_tips": "鲈鱼 / 枪柄 / 大饵巨物",
        "player_positioning": "大饵 / swimbait / 巨物强力",
        "player_selling_points": "中重到超重负荷取向清楚 / 面向 swimbait、粗线和高负荷搏鱼 / 适合巨鲈和重障碍场景",
    },
    "Brand New DESTROYER": {
        "type_tips": "鲈鱼 / 枪直柄 / 高端全技法",
        "player_positioning": "高端鲈钓主力 / 精细到强力全技法",
        "player_selling_points": "从 Ultra Light 到 Extra Heavy 覆盖完整 / 轻线精细、硬饵、底钓和强障碍分工清楚 / 适合作为核心竿组搭建",
    },
    "OROCHI X10": {
        "type_tips": "鲈鱼 / 枪柄为主 / 高感专项",
        "player_positioning": "高感度专项 / 轻量控饵到大饵强力",
        "player_selling_points": "X10 系列强调高感度和专项分工 / 覆盖轻量控饵、moving bait 与 swimbait / 适合按技法精细搭配",
    },
    "LEVANTE": {
        "type_tips": "鲈鱼 / 枪直柄 / 主力泛用",
        "player_positioning": "主力鲈钓 / 常用技法覆盖 / 组件配置完整",
        "player_selling_points": "价位更亲近且型号覆盖主流技法 / Fuji Alconite 导环和 hook keeper 配置更完整 / 适合建立日常主力竿组",
    },
    "VALKYRIE World Expedition": {
        "type_tips": "鲈鱼 / 多节旅行 / 远征强力",
        "player_positioning": "旅行远征 / 多节便携 / 中重饵和大鱼",
        "player_selling_points": "多节结构便于远行携带 / 中重到超重负荷覆盖明确 / 适合 swimbait、粗线和跨场景大鱼应对",
    },
    "TRIZA": {
        "type_tips": "鲈鱼 / 三节旅行 / 技法分工",
        "player_positioning": "便携鲈钓 / 三节结构 / 精细到强力",
        "player_selling_points": "三节结构兼顾移动钓行和实钓分工 / 从 Ultra Light 到 Extra Heavy 覆盖完整 / 适合旅行时保留完整技法选择",
    },
    "Pagani TRAD": {
        "type_tips": "鲈鱼 / 枪柄 / 复古轻量",
        "player_positioning": "复古硬饵 / 小型饵操控 / 收藏实钓",
        "player_selling_points": "传统调性和小型硬饵操控取向清楚 / 适合慢节奏控饵、浅场搜索和复古风格玩家 / 收藏属性与实钓价值并存",
    },
    "GREATHUNTING HUNTSMAN": {
        "type_tips": "鳟鱼 / 溪流 / 高端便携",
        "player_positioning": "山岳溪流鳟 / 便携精密抛投",
        "player_selling_points": "面向山岳溪流和移动钓行 / 短尺便携与精准落点更重要 / 适合小型米诺和溪流鳟控线",
    },
    "GreatHunting X-GLASS COMPOSITE": {
        "type_tips": "鳟鱼 / 玻璃复合 / 溪流湖泊",
        "player_positioning": "鳟鱼硬饵 / 玻璃调性 / 咬口包容",
        "player_selling_points": "玻璃复合调性提升咬口包容度 / 适合小型硬饵、spoon 和近中距离控饵 / 对溪流与湖泊鳟鱼更友好",
    },
    "GREATHUNTING MOUNTAIN STREAM EDITION": {
        "type_tips": "鳟鱼 / 山溪 / 多节便携",
        "player_positioning": "山溪鳟鱼 / 短尺多节 / 轻量米诺",
        "player_selling_points": "短尺多节适合山溪穿行和背包携带 / 小型米诺、spoon 操作更直接 / 面向溪流落点和近距离控线",
    },
    "GREATHUNTING RIVER&LAKE EDITION": {
        "type_tips": "鳟鱼 / 本流湖泊 / 远投控线",
        "player_positioning": "本流湖泊鳟 / 远投搜索 / 大型目标",
        "player_selling_points": "长度和规格更偏远投与控线 / 覆盖本流、湖泊和大型鳟鱼 / 适合需要线弧管理和搜索效率的场景",
    },
    "GREATHUNTING TRACKING BUDDY": {
        "type_tips": "鳟鱼 / 湖泊拖钓 / Lead core",
        "player_positioning": "湖泊拖钓 / Lead core / Downrigger",
        "player_selling_points": "面向 Lead core 与 Downrigger 拖钓场景 / 适合湖泊大型鳟鱼和持续搜索 / 不按常规 casting/spinning 竿理解",
    },
}


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod"]
    col = {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
    required = ["id", "model", "type_tips", "player_positioning", "player_selling_points"]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing rod columns: {missing}")

    changes = []
    by_field = defaultdict(int)
    unmatched = []

    for row_idx in range(2, ws.max_row + 1):
        model = n(ws.cell(row_idx, col["model"]).value)
        update = MASTER_UPDATES.get(model)
        if not update:
            unmatched.append(model)
            continue

        source_id = n(ws.cell(row_idx, col["id"]).value)
        for field in ["type_tips", "player_positioning", "player_selling_points"]:
            cell = ws.cell(row_idx, col[field])
            old_value = n(cell.value)
            new_value = update[field]
            if old_value == new_value:
                continue
            cell.value = new_value
            cell.fill = YELLOW_FILL
            by_field[field] += 1
            changes.append({
                "row": row_idx,
                "id": source_id,
                "model": model,
                "field": field,
                "old_value": old_value,
                "new_value": new_value,
            })

    wb.save(XLSX_PATH)

    report = {
        "xlsx_file": str(XLSX_PATH),
        "changed_cells": len(changes),
        "changed_rows": len({item["id"] for item in changes}),
        "yellow_fill": "FFFFFF00",
        "by_field": dict(sorted(by_field.items())),
        "unmatched_models": unmatched,
        "changes": changes,
        "source_boundary": "Official Megabass series descriptions are primary. Whitelist evidence is used only to refine player-facing series positioning for matched bass series.",
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "changed_cells": report["changed_cells"],
        "changed_rows": report["changed_rows"],
        "by_field": report["by_field"],
        "unmatched_models": report["unmatched_models"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
