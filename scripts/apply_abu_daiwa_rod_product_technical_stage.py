#!/usr/bin/env python3
import copy
import json
import re
import subprocess
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from bs4 import BeautifulSoup
from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
FIELD = "product_technical"
DELIMITER = " / "
REPORT_PATH = DATA_DIR / "abu_daiwa_rod_product_technical_stage_report.json"
EVIDENCE_PATH = DATA_DIR / "abu_daiwa_rod_product_technical_official_evidence.json"


def norm(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def normalize_key(value):
    text = norm(value).lower()
    text = re.sub(r"【[^】]*】", " ", text)
    text = re.sub(r"\([^)]*\)|（[^）]*）", " ", text)
    text = re.sub(r"^(?:20\d{2}|2[3-6])\s+", " ", text)
    text = re.sub(r"\b(?:k|tw|rt|st)\b", " ", text)
    text = re.sub(r"[^a-z0-9+]+", " ", text)
    return norm(text)


def merge_terms(terms):
    merged = []
    for term in terms:
        value = norm(term)
        if value and value not in merged:
            merged.append(value)
    return DELIMITER.join(merged)


def headers(ws):
    return [cell.value for cell in ws[1]]


def col_map(ws):
    return {name: idx + 1 for idx, name in enumerate(headers(ws))}


def sheet_values_without_field(ws):
    h = headers(ws)
    keep_cols = [idx + 1 for idx, name in enumerate(h) if name != FIELD]
    return [
        tuple(ws.cell(row=row, column=col).value for col in keep_cols)
        for row in range(1, ws.max_row + 1)
    ]


def require_detail_product_column(ws):
    h = headers(ws)
    if FIELD not in h:
        raise RuntimeError(f"{ws.title} missing {FIELD}")
    idx = h.index(FIELD)
    if idx == 0 or h[idx - 1] != "Description":
        raise RuntimeError(f"{ws.title}.{FIELD} must be after Description")
    if idx + 1 >= len(h) or h[idx + 1] != "Extra Spec 1":
        raise RuntimeError(f"{ws.title}.{FIELD} must be before Extra Spec 1")
    return idx + 1


def validate_value(value, brand, row_id):
    text = norm(value)
    if not text:
        return []
    errors = []
    if "/" in text and DELIMITER not in text:
        errors.append({"brand": brand, "row": row_id, "value": text, "reason": "bad slash delimiter"})
    for bad in ["白名单", "玩家", "钓组", "饵型", "guide_use_hint", "recommended", "来源"]:
        if bad in text:
            errors.append({"brand": brand, "row": row_id, "value": text, "reason": f"forbidden fragment {bad}"})
    if len(text) > 260:
        errors.append({"brand": brand, "row": row_id, "value": text, "reason": "value looks too long for technical terms"})
    return errors


def abu_terms_from_features(features):
    text = "\n".join(norm(item) for item in features or [])
    terms = []
    patterns = [
        ("Powerlux 1000", r"Powerlux(?:®|™)?\s*1000(?!\d)"),
        ("Powerlux 500", r"Powerlux(?:®|™)?\s*500(?!\d)"),
        ("Powerlux 200", r"Powerlux(?:®|™)?\s*200(?!\d)"),
        ("Powerlux 100", r"Powerlux(?:®|™)?\s*100(?!\d)"),
        ("ROCS", r"ROCS(?:™)?|Robotically Optimized Casting System"),
        ("IntraCarbon", r"IntraCarbon(?:™)?"),
        ("CCRS", r"CCRS(?:™)?|Carbon Constructed Reel Seat"),
        ("Carbon cross wrapped butt section", r"Carbon cross wrapped butt section"),
        ("Carbon split grip", r"Carbon split grip"),
        ("30-ton graphite", r"30[- ]?ton graphite"),
        ("36-ton graphite", r"36[- ]?ton graphite"),
        ("24-ton graphite", r"24[- ]?ton graphite|24[- ]?Ton graphite"),
        ("Solid carbon blank", r"Solid carbon blanks?"),
        ("Composite blend construction", r"Composite blend construction"),
        ("PVD coated one-piece stainless steel guides", r"PVD coated one-piece stainless steel guides"),
        ("Titanium alloy guides with zirconium inserts", r"Titanium alloy guides with .*?zirconi(?:um|a) inserts"),
        ("Titanium guides with zirconium inserts", r"Titanium guides with .*?zirconi(?:um|a) inserts"),
        ("Stainless steel guides with zirconium inserts", r"Stainless [Ss]teel guides with .*?Zirconium inserts|Stainless steel guides with zirconium inserts"),
        ("Stainless steel guides with aluminum oxide inserts", r"Stainless steel guides with aluminum oxide inserts"),
        ("Stainless steel guides with stainless steel inserts", r"Stainless Steel guides with stainless steel inserts"),
    ]
    for term, pattern in patterns:
        if re.search(pattern, text, flags=re.I):
            terms.append(term)
    return merge_terms(terms)


def apply_abu():
    normalized_path = DATA_DIR / "abu_rods_normalized.json"
    xlsx_path = DATA_DIR / "abu_rod_import.xlsx"
    data = json.loads(normalized_path.read_text(encoding="utf-8"))
    values_by_key = {}
    evidence = []
    changed_variants = []

    for index, item in enumerate(data):
        rod_id = f"ABR{1000 + index}"
        value = abu_terms_from_features(item.get("features") or [])
        evidence.append({
            "brand": "Abu Garcia",
            "rod_id": rod_id,
            "model": item.get("model"),
            "source_url": item.get("source_url"),
            "source_section": "official product page Features accordion",
            "features": item.get("features") or [],
            FIELD: value,
        })
        for variant in item.get("variants") or []:
            sku = norm(variant.get("sku"))
            before = norm(variant.get(FIELD))
            variant[FIELD] = value
            values_by_key[(rod_id, sku)] = value
            if before != value:
                changed_variants.append({"rod_id": rod_id, "sku": sku, "old": before, "new": value})

    normalized_path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    result = apply_workbook_values(
        xlsx_path=xlsx_path,
        brand="abu",
        values_by_key=values_by_key,
        shade_script=ROOT / "scripts/shade_abu_rod_detail_groups.py",
    )
    result["normalized_changed_variants"] = len(changed_variants)
    return result, evidence


def daiwa_terms_from_tech_block(section):
    lines = [norm(line) for line in section.get_text("\n", strip=True).splitlines()]
    terms = []
    skip = {"DAIWA 技術", "DAIWA技术", "TECHNOLOGY", "技術"}
    for line in lines:
        if not line or line in skip:
            continue
        if len(line) > 60:
            continue
        if any(fragment in line for fragment in ["：", "。", "，", "透過", "為了", "由於", "此外", "※"]):
            continue
        if re.search(r"(AGS|CWS|SVF|HVF|X45|X4|3DX|MEGA TOP|ZERO|AIR SENSOR|V-JOINT|ESS|SMT|BRAIDING)", line, re.I):
            terms.append(line)
    return merge_terms(terms)


def load_daiwa_cache_terms():
    cache_dir = DATA_DIR / "daiwa_tw_product_technical_cache"
    records = []
    for path in sorted(cache_dir.glob("*.html")):
        if path.name.startswith("list_"):
            continue
        soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
        title = norm(soup.find("title").get_text(" ", strip=True) if soup.find("title") else "")
        product_name = title.split("(釣竿)", 1)[0].split("｜", 1)[0].strip()
        section = soup.select_one("[class*=tech]")
        value = daiwa_terms_from_tech_block(section) if section else ""
        records.append({
            "file": path.name,
            "product_name": product_name,
            "key": normalize_key(product_name),
            FIELD: value,
        })
    return records


def daiwa_match_value(rod_id, model, model_cn, records):
    model_text = f"{model} {model_cn}"
    key = normalize_key(model_text)

    explicit = {
        "DR1001": "銀狼SX",
        "DR1007": "STEEZ（紡車型）",
        "DR1008": "STEEZ（小烏龜款式）",
        "DR1015": "STEEZ（紡車型）",
        "DR1016": "STEEZ（小烏龜款式）",
        "DR1022": "EMERALDAS STOIST RT",
        "DR1038": "月下美人 岩魚",
        "DR1045": "MORETHAN WISEMEN",
        "DR1048": "MORETHAN",
    }
    if rod_id in explicit:
        wanted = normalize_key(explicit[rod_id])
        for record in records:
            if record["key"] == wanted:
                return record

    candidates = []
    for record in records:
        rec_key = record["key"]
        if not rec_key:
            continue
        if key == rec_key:
            candidates.append((3, record))
        elif len(rec_key) >= 5 and rec_key in key:
            candidates.append((2, record))
        elif len(key) >= 5 and key in rec_key:
            candidates.append((1, record))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], len(item[1]["key"])), reverse=True)
    return candidates[0][1]


def apply_daiwa():
    xlsx_path = DATA_DIR / "daiwa_rod_import.xlsx"
    records = load_daiwa_cache_terms()

    wb = load_workbook(xlsx_path, data_only=False)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rh = col_map(rod_ws)
    dh = col_map(detail_ws)
    old_master_col = rh.get(FIELD)
    existing_by_rod = {}
    if FIELD in dh:
        for row_idx in range(2, detail_ws.max_row + 1):
            rod_id = norm(detail_ws.cell(row=row_idx, column=dh["rod_id"]).value)
            value = norm(detail_ws.cell(row=row_idx, column=dh[FIELD]).value)
            if value and rod_id not in existing_by_rod:
                existing_by_rod[rod_id] = value
    values_by_rod = {}
    evidence = []
    for row_idx in range(2, rod_ws.max_row + 1):
        rod_id = norm(rod_ws.cell(row=row_idx, column=rh["id"]).value)
        model = norm(rod_ws.cell(row=row_idx, column=rh["model"]).value)
        model_cn = norm(rod_ws.cell(row=row_idx, column=rh["model_cn"]).value)
        old_master_value = norm(rod_ws.cell(row=row_idx, column=old_master_col).value) if old_master_col else ""
        match = daiwa_match_value(rod_id, model, model_cn, records)
        match_value = norm(match.get(FIELD)) if match else ""
        value = old_master_value or existing_by_rod.get(rod_id, "") or match_value
        values_by_rod[rod_id] = value
        evidence.append({
            "brand": "Daiwa",
            "rod_id": rod_id,
            "model": model,
            "model_cn": model_cn,
            "matched_official_product": match.get("product_name") if match else "",
            "cache_file": match.get("file") if match else "",
            "source_section": "official DAIWA 技術 module",
            "migrated_from_old_master_field": bool(old_master_value),
            "preserved_from_existing_detail_field": bool(not old_master_value and existing_by_rod.get(rod_id, "")),
            FIELD: value,
        })

    values_by_key = {}
    for row_idx in range(2, detail_ws.max_row + 1):
        rod_id = norm(detail_ws.cell(row=row_idx, column=dh["rod_id"]).value)
        sku = norm(detail_ws.cell(row=row_idx, column=dh["SKU"]).value)
        values_by_key[(rod_id, sku)] = values_by_rod.get(rod_id, "")

    result = apply_workbook_values(
        xlsx_path=xlsx_path,
        brand="daiwa",
        values_by_key=values_by_key,
        shade_script=ROOT / "scripts/shade_daiwa_rod_detail_groups_stage12.py",
    )
    result["official_cache_products"] = len(records)
    result["matched_rod_products"] = sum(1 for item in evidence if item[FIELD])
    return result, evidence


def apply_workbook_values(xlsx_path, brand, values_by_key, shade_script):
    wb = load_workbook(xlsx_path, data_only=False)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]

    rod_before = sheet_values_without_field(rod_ws)
    detail_before = sheet_values_without_field(detail_ws)
    rod_master_removed = False
    h_rod = headers(rod_ws)
    if FIELD in h_rod:
        rod_ws.delete_cols(h_rod.index(FIELD) + 1)
        rod_master_removed = True
    field_col = require_detail_product_column(detail_ws)
    c = col_map(detail_ws)
    changes = []
    errors = []
    for row_idx in range(2, detail_ws.max_row + 1):
        rod_id = norm(detail_ws.cell(row=row_idx, column=c["rod_id"]).value)
        sku = norm(detail_ws.cell(row=row_idx, column=c["SKU"]).value)
        key = (rod_id, sku)
        value = values_by_key.get(key, "")
        old = norm(detail_ws.cell(row=row_idx, column=field_col).value)
        if old != value:
            detail_ws.cell(row=row_idx, column=field_col).value = value
            changes.append({"row": row_idx, "rod_id": rod_id, "sku": sku, "old": old, "new": value})
        errors.extend(validate_value(value, brand, f"{rod_id}:{sku}"))

    if sheet_values_without_field(rod_ws) != rod_before:
        raise RuntimeError(f"{brand}: rod sheet changed outside {FIELD}")
    if sheet_values_without_field(detail_ws) != detail_before:
        raise RuntimeError(f"{brand}: rod_detail changed outside {FIELD}")
    if errors:
        raise RuntimeError(f"{brand}: validation failed: {errors[:5]}")

    wb.save(xlsx_path)
    if shade_script.exists():
        subprocess.run(["python3", str(shade_script)], check=True)

    wb_check = load_workbook(xlsx_path, data_only=True)
    detail_check = wb_check["rod_detail"]
    c2 = col_map(detail_check)
    values = [
        norm(detail_check.cell(row=row, column=c2[FIELD]).value)
        for row in range(2, detail_check.max_row + 1)
    ]
    return {
        "file": xlsx_path.name,
        "detail_rows": detail_check.max_row - 1,
        "changed_cells": len(changes),
        "nonempty": sum(1 for value in values if value),
        "unique_values": len(set(value for value in values if value)),
        "rod_master_product_technical_removed": rod_master_removed,
        "changes_sample": changes[:20],
    }


def main():
    report = {
        "field": FIELD,
        "scope": "rod_detail.product_technical only",
        "source_policy": "official product page technology/features modules only",
        "delimiter": DELIMITER,
        "updated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "brands": {},
    }
    evidence = {"items": []}

    abu_result, abu_evidence = apply_abu()
    report["brands"]["abu"] = abu_result
    evidence["items"].extend(abu_evidence)

    daiwa_result, daiwa_evidence = apply_daiwa()
    report["brands"]["daiwa"] = daiwa_result
    evidence["items"].extend(daiwa_evidence)

    report["summary"] = {
        "brands_processed": list(report["brands"].keys()),
        "total_changed_cells": sum(item["changed_cells"] for item in report["brands"].values()),
        "nonempty_by_brand": {brand: item["nonempty"] for brand, item in report["brands"].items()},
        "unique_by_brand": {brand: item["unique_values"] for brand, item in report["brands"].items()},
    }
    report["top_values"] = {}
    for brand in report["brands"]:
        path = DATA_DIR / f"{brand}_rod_import.xlsx"
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["rod_detail"]
        c = col_map(ws)
        counter = Counter(
            norm(row[c[FIELD] - 1])
            for row in ws.iter_rows(min_row=2, values_only=True)
            if norm(row[c[FIELD] - 1])
        )
        report["top_values"][brand] = counter.most_common(20)

    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    EVIDENCE_PATH.write_text(json.dumps(evidence, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
