#!/usr/bin/env python3
import copy
import json
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
REPORT_PATH = DATA_DIR / "rod_product_technical_field_report.json"

FIELD = "product_technical"
DETAIL_ANCHOR = "Description"


def norm(value):
    return str(value or "").strip()


def headers(ws):
    return [cell.value for cell in ws[1]]


def header_index(ws, field):
    current = headers(ws)
    return current.index(field) + 1 if field in current else None


def copy_style(source, target):
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)


def cell_snapshots(ws, col):
    values = []
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        values.append(
            {
                "value": cell.value,
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
                "alignment": copy.copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
            }
        )
    return values


def apply_snapshot(snapshot, cell):
    cell.value = snapshot["value"]
    cell.font = copy.copy(snapshot["font"])
    cell.fill = copy.copy(snapshot["fill"])
    cell.border = copy.copy(snapshot["border"])
    cell.alignment = copy.copy(snapshot["alignment"])
    cell.number_format = snapshot["number_format"]
    cell.protection = copy.copy(snapshot["protection"])


def delete_rod_product_column(ws):
    col = header_index(ws, FIELD)
    if not col:
        return False, {}

    id_col = header_index(ws, "id")
    values_by_id = {}
    if id_col:
        for row in range(2, ws.max_row + 1):
            rod_id = ws.cell(row=row, column=id_col).value
            value = ws.cell(row=row, column=col).value
            if norm(value):
                values_by_id[rod_id] = value

    ws.delete_cols(col)
    return True, values_by_id


def ensure_detail_column(ws):
    current_col = header_index(ws, FIELD)
    anchor_col = header_index(ws, DETAIL_ANCHOR)
    if not anchor_col:
        raise RuntimeError(f"{ws.title} missing {DETAIL_ANCHOR}")

    snapshots = cell_snapshots(ws, current_col) if current_col else None
    if current_col:
        desired_col = anchor_col + 1
        if current_col == desired_col:
            return current_col, False
        ws.delete_cols(current_col)
        anchor_col = header_index(ws, DETAIL_ANCHOR)

    insert_col = anchor_col + 1
    ws.insert_cols(insert_col)
    source_col = anchor_col
    ws.column_dimensions[ws.cell(row=1, column=insert_col).column_letter].width = (
        ws.column_dimensions[ws.cell(row=1, column=source_col).column_letter].width
    )

    if snapshots:
        for row, snapshot in enumerate(snapshots, start=1):
            apply_snapshot(snapshot, ws.cell(row=row, column=insert_col))
    else:
        for row in range(1, ws.max_row + 1):
            source = ws.cell(row=row, column=source_col)
            target = ws.cell(row=row, column=insert_col)
            copy_style(source, target)
            target.value = FIELD if row == 1 else ""

    ws.cell(row=1, column=insert_col).value = FIELD
    return insert_col, True


def validate_value(value, file_name, row):
    text = norm(value)
    if not text:
        return []
    errors = []
    if "/" in text and " / " not in text:
        errors.append(f"{file_name}:rod_detail:{row} uses slash without ` / `")
    for fragment in ["白名单", "玩家", "钓组", "饵型", "来源"]:
        if fragment in text:
            errors.append(f"{file_name}:rod_detail:{row} contains forbidden fragment `{fragment}`")
    return errors


def migrate_normalized_json(path):
    data = json.loads(path.read_text(encoding="utf-8"))
    items = data if isinstance(data, list) else data.get("items", []) if isinstance(data, dict) else []
    if not isinstance(items, list):
        return None

    item_removed = 0
    variant_written = 0
    variant_nonempty = 0
    for item in items:
        if not isinstance(item, dict):
            continue
        item_value = item.pop(FIELD, None)
        if item_value is not None:
            item_removed += 1
        variants = item.get("variants") or []
        if item_value:
            for variant in variants:
                if isinstance(variant, dict) and not norm(variant.get(FIELD)):
                    variant[FIELD] = item_value
        for variant in variants:
            if isinstance(variant, dict) and FIELD in variant:
                variant_written += 1
                if norm(variant.get(FIELD)):
                    variant_nonempty += 1

    if item_removed or variant_written:
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {
        "file": path.name,
        "item_product_technical_removed": item_removed,
        "variant_product_technical_count": variant_written,
        "variant_product_technical_nonempty": variant_nonempty,
    }


def main():
    report = {
        "field": FIELD,
        "scope": "rod_detail SKU level only; rod master must not contain product_technical",
        "detail_position": "after Description, before Extra Spec 1",
        "workbooks": [],
        "normalized": [],
        "errors": [],
    }

    for path in sorted(DATA_DIR.glob("*rod*normalized*.json")):
        result = migrate_normalized_json(path)
        if result and (
            result["item_product_technical_removed"]
            or result["variant_product_technical_count"]
        ):
            report["normalized"].append(result)

    for path in sorted(DATA_DIR.glob("*_rod_import.xlsx")):
        wb = load_workbook(path)
        if "rod" not in wb.sheetnames or "rod_detail" not in wb.sheetnames:
            continue

        rod_ws = wb["rod"]
        detail_ws = wb["rod_detail"]

        rod_removed, master_values = delete_rod_product_column(rod_ws)
        detail_col, detail_inserted_or_moved = ensure_detail_column(detail_ws)
        rod_id_col = header_index(detail_ws, "rod_id")

        migrated_count = 0
        if master_values and rod_id_col:
            for row in range(2, detail_ws.max_row + 1):
                current = detail_ws.cell(row=row, column=detail_col).value
                rod_id = detail_ws.cell(row=row, column=rod_id_col).value
                if not norm(current) and rod_id in master_values:
                    detail_ws.cell(row=row, column=detail_col).value = master_values[rod_id]
                    migrated_count += 1

        nonempty = 0
        for row in range(2, detail_ws.max_row + 1):
            value = detail_ws.cell(row=row, column=detail_col).value
            if norm(value):
                nonempty += 1
            report["errors"].extend(validate_value(value, path.name, row))

        wb.save(path)
        current_headers = headers(detail_ws)
        idx = current_headers.index(FIELD)
        report["workbooks"].append(
            {
                "file": path.name,
                "rod_product_technical_removed": rod_removed,
                "detail_product_technical_inserted_or_moved": detail_inserted_or_moved,
                "detail_product_technical_nonempty": nonempty,
                "migrated_from_master_rows": migrated_count,
                "previous_header": current_headers[idx - 1] if idx > 0 else None,
                "next_header": current_headers[idx + 1] if idx + 1 < len(current_headers) else None,
            }
        )

    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if report["errors"]:
        raise RuntimeError("product_technical validation failed; see report")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
