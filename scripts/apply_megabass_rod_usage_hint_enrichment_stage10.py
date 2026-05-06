import json
import re
import subprocess
from collections import Counter
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook

import apply_megabass_rod_recommended_rig_pairing_stage9 as stage9


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('megabass_rod_import.xlsx')
REPORT_PATH = resolve_data_raw('megabass_rod_usage_hint_enrichment_report.json')
SHADE_SCRIPT = ROOT / "scripts/shade_megabass_rod_detail_groups_stage2.py"

MAX_ITEMS = 8


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def add(matches, label, pos, source="description"):
    if not label:
        return
    for item in matches:
        if item["label"] == label:
            item["pos"] = min(item["pos"], pos)
            return
    matches.append({"label": label, "pos": pos, "source": source})


def add_if(matches, text, label, patterns):
    best = None
    for pattern in patterns:
        found = re.search(pattern, text, re.I)
        if found:
            best = found.start() if best is None else min(best, found.start())
    if best is not None:
        add(matches, label, best)


def add_group_if(matches, text, labels, patterns):
    best = None
    for pattern in patterns:
        found = re.search(pattern, text, re.I)
        if found:
            best = found.start() if best is None else min(best, found.start())
    if best is None:
        return
    for offset, label in enumerate(labels):
        add(matches, label, best + offset)


def labels_from_description(row):
    text = lower(f"{row.get('SKU')} {row.get('Code Name')} {row.get('Description')}")
    matches = []

    add_if(matches, text, "Bug Lure", [r"\bmushi\b", r"\bsiglett\b", r"虫"])
    add_if(matches, text, "Popper", [r"\bpopx\b", r"\bpopmax\b", r"\bpopper\b", r"popping", r"ポップ"])
    add_if(matches, text, "Pencil Bait", [r"\bdog-?x\b", r"\bgatta-?x\b", r"\bpencil\b", r"dog[- ]?walk", r"ドッグウォーク"])
    add_if(matches, text, "Topwater Plug", [r"topwater", r"surface game", r"surface lures?", r"surface plugs?", r"floating minnows?", r"waterside topwater"])
    add_if(matches, text, "I-shaped Plug", [r"i-shaped", r"\bi[- ]?slide\b"])
    add_if(matches, text, "Jerkbait", [r"jerkbait", r"vision oneten", r"oneten", r"jerking", r"twitch", r"twitching", r"whip twitch"])
    add_if(matches, text, "Minnow", [r"\bminnows?\b", r"小型ミノー", r"ミノー", r"米诺"])
    add_if(matches, text, "Shad", [r"\bshad\b", r"shading-x"])
    add_if(matches, text, "Deep Crankbait", [r"deep crank", r"super[- ]?deep crank", r"large crank", r"big crank", r"granade"])
    add_if(matches, text, "Crankbait", [r"crankbait", r"cranking", r"\bcranks?\b", r"small crank", r"mini crank", r"light cranking", r"shallow running crankbait"])
    add_if(matches, text, "Vibration", [r"vibration", r"\bvib\b", r"blade bait"])
    add_if(matches, text, "Spinnerbait", [r"spinnerbait", r"spinner bait"])
    add_if(matches, text, "Chatterbait", [r"chatterbait", r"bladed jigs?", r"bladed swim jig", r"wild header", r"刀片铅头钩", r"刀片式铅头钩"])
    add_if(matches, text, "Swim Jig", [r"swim jig", r"泳饵铅头钩"])
    add_if(matches, text, "Small Rubber Jig", [r"small rubber jigs?", r"スモールラバージグ", r"スモラバ"])
    add_if(matches, text, "Football Jig", [r"football jig"])
    add_if(matches, text, "Rubber Jig", [r"rubber jigs?", r"\bjigs?\b", r"橡胶铅头钩", r"橡膠鉛頭鉤"])
    add_if(matches, text, "Jighead Rig", [r"jighead", r"jig head", r"mid[- ]?strolling", r"shakyhead", r"shaky head", r"ジグヘッド"])
    add_if(matches, text, "Light Texas Rig", [r"light texas"])
    add_if(matches, text, "Heavy Texas Rig", [r"heavy texas"])
    add_if(matches, text, "Texas Rig", [r"\btexas\b", r"テキサス"])
    add_if(matches, text, "Free Rig", [r"free rig"])
    add_if(matches, text, "Neko Rig", [r"\bneko\b", r"ネコリグ"])
    add_if(matches, text, "Down Shot", [r"down[- ]?shot", r"drop[- ]?shot", r"dropshot", r"ダウンショット"])
    add_if(matches, text, "No Sinker", [r"no[- ]?sinker", r"weightless", r"無鉛", r"ノーシンカー"])
    add_if(matches, text, "Wacky Rig", [r"wacky", r"ワッキー"])
    add_if(matches, text, "Nose Hook Rig", [r"nose-hook", r"nose hooking"])
    add_if(matches, text, "Frog", [r"\bfrog\b", r"frogging", r"heavy mat", r"フロッグ"])
    add_if(matches, text, "Punching", [r"punching", r"punch\b"])
    add_if(matches, text, "Dark Sleeper", [r"dark sleeper"])
    add_if(matches, text, "Sleeper Craw", [r"sleeper craw"])
    add_if(matches, text, "Craw", [r"\bcraw\b"])
    add_if(matches, text, "Swimbait", [r"swimbait", r"swim bait", r"magdraft", r"vatalion", r"bottom-tracing swimbait"])
    add_if(matches, text, "Big Bait", [r"bigbait", r"big bait", r"giant bait", r"giant,?\s*40 cm", r"40 cm\+ bait", r"30-35 cm class", r"20-30 cm", r"magnum[- ]size lures?", r"magnum-sized lures?", r"100\+?g class", r"100g class", r"ビッグベイト", r"マグナムサイズルアー"])
    add_if(matches, text, "Magnum Spoon", [r"magnum[- ]sized spoons?"])
    add_if(matches, text, "Metal Jig", [r"metal jig", r"metal type lures?", r"shore jigging"])
    add_if(matches, text, "Shore Plug", [r"shore-based", r"wading", r"surf games", r"shore plug", r"shore plugging"])
    add_if(matches, text, "Small Plug", [r"small plugs?", r"small to medium plugs?", r"tiny plugs?", r"light plugs?", r"lightweight plugs?", r"ライトプラッギング", r"タイニープラグ", r"プラグ"])
    add_if(matches, text, "Spoon", [r"\bspoon\b", r"spooning", r"スプーン"])
    add_if(matches, text, "Trout Minnow", [r"yamame", r"iwana", r"amago", r"satsuki", r"masu", r"brown trout", r"rainbow", r"trout", r"渓流", r"本流"])
    add_if(matches, text, "Lead Core Trolling", [r"lead core"])
    add_if(matches, text, "Downrigger Trolling", [r"downrigger"])

    add_group_if(matches, text, ["Down Shot", "Neko Rig", "Jighead Rig"], [r"light finesse rigs?", r"finesse rigs?", r"light rigs?", r"ライトリグ", r"軽量リグ"])
    add_group_if(matches, text, ["BFS Jighead Rig", "Small Rubber Jig", "No Sinker"], [r"bait finesse", r"ベイトフィネス", r"finesse application"])
    add_group_if(matches, text, ["Neko Rig", "Jighead Rig", "Down Shot"], [r"一点シェイク", r"one[- ]?point shake"])

    labels = [item["label"] for item in sorted(matches, key=lambda item: (item["pos"], item["label"]))]
    return stage9.normalize_labels(labels)[:MAX_ITEMS]


ZH = {
    "BFS Jighead Rig": "BFS轻量铅头/小软饵",
    "Small Rubber Jig": "小型橡胶铅头钩",
    "Rubber Jig": "橡胶铅头钩",
    "Football Jig": "Football Jig",
    "Jighead Rig": "铅头钩/中层晃饵",
    "Light Texas Rig": "轻德州",
    "Heavy Texas Rig": "重德州",
    "Texas Rig": "德州钓组",
    "Free Rig": "Free Rig",
    "Neko Rig": "Neko Rig",
    "Down Shot": "Down Shot",
    "No Sinker": "无铅软饵",
    "Wacky Rig": "Wacky",
    "Nose Hook Rig": "鼻挂软饵",
    "Frog": "Frog",
    "Punching": "Punching",
    "Jerkbait": "Jerkbait",
    "Minnow": "Minnow",
    "Trout Minnow": "鳟鱼 Minnow",
    "Small Minnow": "小型 Minnow",
    "Shad": "Shad",
    "Deep Crankbait": "Deep Crankbait",
    "Crankbait": "Crankbait",
    "Vibration": "Vibration",
    "Spinnerbait": "Spinnerbait",
    "Chatterbait": "Chatterbait",
    "Swim Jig": "Swim Jig",
    "Swimbait": "Swimbait",
    "Big Bait": "Big Bait",
    "Topwater Plug": "水面系 Plug",
    "Pencil Bait": "Pencil",
    "Popper": "Popper",
    "Bug Lure": "虫系水面饵",
    "I-shaped Plug": "I 字系 Plug",
    "Small Plug": "小型 Plug",
    "Spoon": "Spoon",
    "Magnum Spoon": "Magnum Spoon",
    "Metal Jig": "Metal Jig",
    "Shore Plug": "岸投 Plug",
    "Lead Core Trolling": "Lead Core 拖钓",
    "Downrigger Trolling": "Downrigger 拖钓",
}


def zh(label):
    return ZH.get(label, label)


def cue_text(row):
    text = lower(row.get("Description"))
    cues = []
    if re.search(r"pinpoint|accuracy|accurate|tight spots?|close quarters|pitch", text):
        cues.append("精准落点")
    if re.search(r"long[- ]distance|distance|moon-shot|遠投|ロングディスタンス|wide rivers?|lakes?", text):
        cues.append("远投控线")
    if re.search(r"cover|structure|mats?|vegetation|bush|grass|障碍|ブッシュ|水生植物", text):
        cues.append("障碍区控鱼")
    if re.search(r"fall|line tension|subtle bites?|light bites?|sensitivity|感度|喰わせ", text):
        cues.append("咬口/线张力反馈")
    if re.search(r"fast-moving|quick search|search|moving bait|巻き|cranking", text):
        cues.append("移动饵搜索")
    if re.search(r"glass|グラス|低弾性", text):
        cues.append("玻纤/低弹性追随")
    if re.search(r"solid carbon|stinger tip|ソリッド|stinger", text):
        cues.append("实心竿梢精细信号")
    return "、".join(cues[:3])


def is_trout_context(model):
    return "greathunting" in lower(model) or "great hunting" in lower(model)


def enriched_hint(row, items, model):
    if not items:
        return n(row.get("guide_use_hint"))
    primary = items[0]
    secondary = [zh(item) for item in items[1:4]]
    secondary_text = "，".join(secondary)
    cue = cue_text(row)
    cue_prefix = f"{cue}： " if cue else ""
    desc = lower(row.get("Description"))

    if set(items) & stage9.TROLLING:
        first = zh(primary)
        return f"{cue_prefix}{first}为主，导环和控线要稳定承接长时间拖曳负荷；{secondary_text or '湖鳟 spoon / plug'}作为速度和水层调整补充。"

    if is_trout_context(model):
        if re.search(r"lake|wide river|本流|湖|breakwater|long[- ]distance", desc):
            return f"{cue_prefix}{zh(primary)}为主，配合{secondary_text or 'Spoon'}做本流/湖泊远投搜索；出线顺畅和线弧稳定能帮助漂流、收线和远距离控饵。"
        return f"{cue_prefix}{zh(primary)}为主，配合{secondary_text or 'Spoon'}应对溪流精细操作；小饵低负荷出线和竿尖信号要清楚，便于连续 twitch 与短咬判断。"

    if primary in {"Frog", "Punching"}:
        return f"{cue_prefix}{zh(primary)}为主，{secondary_text or '重障碍软饵'}为辅；粗线通过、竿身受力和导环稳定性要支撑草洞、重障碍里的强制控鱼。"

    if primary in {"Big Bait", "Swimbait"}:
        return f"{cue_prefix}{zh(primary)}为主，{secondary_text or '大型移动饵'}为辅；抛投阶段需要稳定出线，回收和中鱼后要能承接大饵惯性与高负荷搏鱼。"

    if primary in {"Jerkbait", "Minnow", "Crankbait", "Deep Crankbait", "Spinnerbait", "Vibration", "Chatterbait", "Topwater Plug", "Pencil Bait", "Popper", "Shad", "Small Plug"}:
        if secondary_text:
            return f"{cue_prefix}{zh(primary)}为主，{secondary_text}为辅；抛投后要保持线弧稳定和泳层连贯，避免硬饵动作被多余线松拖慢。"
        return f"{cue_prefix}{zh(primary)}专项；导环出线和控线节奏要稳定，便于维持泳层、连续抽停幅度和长时间搜索效率。"

    if primary in {"Small Rubber Jig", "BFS Jighead Rig", "Jighead Rig", "Down Shot", "Neko Rig", "No Sinker", "Texas Rig", "Free Rig", "Light Texas Rig", "Rubber Jig", "Football Jig"}:
        if secondary_text:
            return f"{cue_prefix}{zh(primary)}为主，{secondary_text}为辅；落下段线张力、触底信号和小幅操作反馈要清楚，贴障碍时也要保持直接控线。"
        return f"{cue_prefix}{zh(primary)}专项；细线到中线出线要顺，触底、跳动和短咬信号要清楚，便于低速底操或精细诱食。"

    return f"{cue_prefix}{zh(primary)}为主，{secondary_text or '同级饵型'}为辅；出线、控线和竿尖反馈要贴合描述里的核心场景，避免把竿性用成泛用模板。"


def main():
    evidence_index = stage9.build_evidence_index()
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_col = {n(cell.value): idx + 1 for idx, cell in enumerate(rod_ws[1])}
    detail_col = {n(cell.value): idx + 1 for idx, cell in enumerate(detail_ws[1])}
    required = ["id", "rod_id", "SKU", "Code Name", "Description", "guide_use_hint", "recommended_rig_pairing"]
    missing = [field for field in required if field not in detail_col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    model_by_rod_id = {
        n(rod_ws.cell(row_idx, rod_col["id"]).value): n(rod_ws.cell(row_idx, rod_col["model"]).value)
        for row_idx in range(2, rod_ws.max_row + 1)
    }

    pairing_changes = []
    guide_changes = []
    source_counts = Counter()
    guide_unique = Counter()
    coverage = 0

    for row_idx in range(2, detail_ws.max_row + 1):
        row = {field: detail_ws.cell(row_idx, col_idx).value for field, col_idx in detail_col.items()}
        model = model_by_rod_id.get(n(row.get("rod_id")), "")
        desc_items = labels_from_description(row)
        if not is_trout_context(model):
            desc_items = [item for item in desc_items if item != "Trout Minnow"]
        if desc_items:
            items, source = desc_items, "description_explicit"
        else:
            items, source = stage9.pairing_for(row, model, evidence_index)
        pairing = " / ".join(items)
        if pairing:
            coverage += 1
        source_counts[source] += 1

        pair_cell = detail_ws.cell(row_idx, detail_col["recommended_rig_pairing"])
        old_pair = n(pair_cell.value)
        if old_pair != pairing:
            pair_cell.value = pairing
            pairing_changes.append({
                "row": row_idx,
                "id": n(row.get("id")),
                "rod_id": n(row.get("rod_id")),
                "sku": n(row.get("SKU")),
                "model": model,
                "old_value": old_pair,
                "new_value": pairing,
                "source": source,
            })

        guide = enriched_hint(row, items, model)
        guide_unique[guide] += 1
        guide_cell = detail_ws.cell(row_idx, detail_col["guide_use_hint"])
        old_guide = n(guide_cell.value)
        if old_guide != guide:
            guide_cell.value = guide
            guide_changes.append({
                "row": row_idx,
                "id": n(row.get("id")),
                "rod_id": n(row.get("rod_id")),
                "sku": n(row.get("SKU")),
                "model": model,
                "old_value": old_guide,
                "new_value": guide,
                "source": source,
            })

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "xlsx_file": str(XLSX_PATH),
        "fields": ["guide_use_hint", "recommended_rig_pairing"],
        "coverage": f"{coverage}/{detail_ws.max_row - 1}",
        "source_counts": dict(source_counts),
        "changed_recommended_rig_pairing_cells": len(pairing_changes),
        "changed_guide_use_hint_cells": len(guide_changes),
        "guide_unique_count": len(guide_unique),
        "pairing_changes": pairing_changes,
        "guide_use_hint_changes": guide_changes,
        "policy": {
            "description_explicit": "When SKU/Code Name/Description exposes concrete lure or rig terms, both fields follow that row-level description.",
            "fallback": "When Description is not concrete enough, reuse Stage 9 exact whitelist or conservative SKU/spec inference.",
        },
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "coverage": report["coverage"],
        "source_counts": dict(source_counts),
        "changed_recommended_rig_pairing_cells": len(pairing_changes),
        "changed_guide_use_hint_cells": len(guide_changes),
        "guide_unique_count": len(guide_unique),
        "report": str(REPORT_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
