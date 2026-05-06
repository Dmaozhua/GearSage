#!/usr/bin/env python3
import copy
import json
import subprocess
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "daiwa_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "daiwa_rod_product_technical_detail_migration_stage45_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_daiwa_rod_detail_groups_stage12.py"

FIELD = "product_technical"
DETAIL_ANCHOR = "Description"


def n(value):
    return str(value or "").strip()


def headers(ws):
    return [cell.value for cell in ws[1]]


def col(ws, field):
    current = headers(ws)
    return current.index(field) + 1 if field in current else None


def copy_column_style(ws, source_col, target_col):
    ws.column_dimensions[ws.cell(row=1, column=target_col).column_letter].width = (
        ws.column_dimensions[ws.cell(row=1, column=source_col).column_letter].width
    )
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row=row, column=source_col)
        target = ws.cell(row=row, column=target_col)
        target._style = copy.copy(source._style)
        if source.has_style:
            target.font = copy.copy(source.font)
            target.fill = copy.copy(source.fill)
            target.border = copy.copy(source.border)
            target.alignment = copy.copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy.copy(source.protection)


def ensure_detail_column(ws):
    current = col(ws, FIELD)
    anchor = col(ws, DETAIL_ANCHOR)
    if not anchor:
        raise RuntimeError(f"{ws.title} missing {DETAIL_ANCHOR}")
    if current:
        if current != anchor + 1:
            raise RuntimeError(f"{ws.title}.{FIELD} must be immediately after {DETAIL_ANCHOR}")
        return current, False
    insert_at = anchor + 1
    ws.insert_cols(insert_at)
    copy_column_style(ws, anchor, insert_at)
    ws.cell(row=1, column=insert_at).value = FIELD
    return insert_at, True


def values_without_field(ws):
    keep = [idx for idx, header in enumerate(headers(ws), start=1) if header != FIELD]
    return [
        tuple(ws.cell(row=row, column=column).value for column in keep)
        for row in range(1, ws.max_row + 1)
    ]


def validate_value(value, row):
    text = n(value)
    errors = []
    if text and "/" in text and " / " not in text:
        errors.append(f"row {row}: separator must be ` / `")
    for fragment in ["白名单", "玩家", "钓组", "饵型", "来源", "官网确认"]:
        if fragment in text:
            errors.append(f"row {row}: forbidden fragment {fragment}")
    return errors


def main():
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]

    rod_before = values_without_field(rod_ws)
    detail_before = values_without_field(detail_ws)

    rod_field_col = col(rod_ws, FIELD)
    detail_field_col, detail_inserted = ensure_detail_column(detail_ws)
    rod_id_col = col(rod_ws, "id")
    model_col = col(rod_ws, "model")
    detail_rod_id_col = col(detail_ws, "rod_id")
    if not rod_id_col or not detail_rod_id_col:
        raise RuntimeError("missing rod id columns")

    master_values = {}
    for row in range(2, rod_ws.max_row + 1):
        rod_id = n(rod_ws.cell(row=row, column=rod_id_col).value)
        value = n(rod_ws.cell(row=row, column=rod_field_col).value) if rod_field_col else ""
        master_values[rod_id] = {
            "model": n(rod_ws.cell(row=row, column=model_col).value) if model_col else "",
            "value": value,
        }

    changes = []
    errors = []
    for row in range(2, detail_ws.max_row + 1):
        rod_id = n(detail_ws.cell(row=row, column=detail_rod_id_col).value)
        before = n(detail_ws.cell(row=row, column=detail_field_col).value)
        value = master_values.get(rod_id, {}).get("value", before) if rod_field_col else before
        if before != value:
            detail_ws.cell(row=row, column=detail_field_col).value = value
            changes.append({"row": row, "rod_id": rod_id, "before": before, "after": value})
        errors.extend(validate_value(value, row))

    removed_rod_columns = 0
    while col(rod_ws, FIELD):
        rod_ws.delete_cols(col(rod_ws, FIELD))
        removed_rod_columns += 1

    if values_without_field(rod_ws) != rod_before:
        raise RuntimeError("Unexpected rod sheet changes outside product_technical removal")
    if values_without_field(detail_ws) != detail_before:
        raise RuntimeError("Unexpected rod_detail changes outside product_technical")
    if errors:
        raise RuntimeError(f"product_technical validation failed: {errors[:5]}")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "field": FIELD,
        "scope": "migrate Daiwa Taiwan product_technical from rod master to rod_detail SKU rows",
        "source_policy": "existing values came from daiwa.com/tw official product pages in stage44",
        "rod_product_technical_removed": removed_rod_columns,
        "rod_detail_product_technical_inserted": detail_inserted,
        "rod_rows": rod_ws.max_row - 1,
        "rod_detail_rows": detail_ws.max_row - 1,
        "master_nonempty": sum(1 for row in master_values.values() if row["value"]),
        "detail_nonempty": sum(
            1 for row in range(2, detail_ws.max_row + 1)
            if n(detail_ws.cell(row=row, column=detail_field_col).value)
        ),
        "changes": changes,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({k: report[k] for k in [
        "rod_product_technical_removed",
        "rod_detail_product_technical_inserted",
        "master_nonempty",
        "detail_nonempty",
    ]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
