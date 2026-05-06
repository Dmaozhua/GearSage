from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')

HINT_BY_ID = {
    "DRD10056": "超輕精細：AGS/CWS 減少竿尖負擔，Down Shot、Neko、No Sinker 等輕量釣組讀底、看線和微弱咬口更清楚。",
    "DRD10057": "Hover/Mid Strolling：細線出線和線弧更穩，能維持中層晃動節奏，也能兼顧 Down Shot、Neko 等輕釣組控線。",
    "DRD10058": "輕硬餌泛用：Shad、Minnow、Metal Vibration 的泳層和震動回饋更直接，連續操作與小餌切換更順。",
    "DRD10059": "BF 精細軟餌：實心竿尖配 AGS/CWS 放大中層軟餌與底部釣組變化，Hover/Mid Strolling 和 Neko 操作更細。",
    "DRD10060": "Neko/輕量底操：細線出線更穩，障礙物邊的 Neko、Down Shot、No Sinker 能更清楚讀底與控鬆線。",
    "DRD10061": "操作型硬餌：Jerkbait、Spinnerbait、Topwater 操作時線弧更穩，抽停、閃動和中層軟餌細節更容易掌握。",
    "DRD10062": "卷阻硬餌：Crankbait、Spinnerbait、Popper、Jerkbait 平收和抽停回饋更清楚，方便維持泳層與節奏。",
    "DRD10063": "重底操軟餌：Free Rig、Texas、Leaderless Down Shot 貼底時張力變化更清楚，硬底、障礙和輕咬更好判斷。",
    "DRD10064": "淺水障礙精細：No Sinker、Neko、Free Rig 在覆蓋物邊更好控線，碰障、鬆線和小口刺魚更直接。",
    "DRD10065": "高感底操：Free Rig、Light Texas、Heavy Down Shot 的底部觸感和覆蓋物回饋更明確，精細操作與快速刺魚兼顧。",
    "DRD10066": "強力底操：Texas、Free Rig、Rubber Jig 在重障礙周邊控線更穩，微咬判斷、刺魚和拉離覆蓋物更直接。",
    "DRD10067": "Wire bait/大餌：Spinnerbait、Chatterbait、Swim Jig、Swimbait 拋投和回收負荷更穩，重餌也能保持操作節奏。",
    "DRD10068": "重型多用底操：Rubber Jig、Texas、Leaderless Down Shot 與 Frog/Swimbait 切換時，粗線控線和強力刺魚更穩。",
}


def n(value):
    return " ".join(str(value or "").split())


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["guide_use_hint"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        rod_id = n(ws.cell(row, col["rod_id"]).value)
        if rod_id not in {"DR1008", "DR1009"}:
            continue
        value = HINT_BY_ID.get(detail_id)
        if not value:
            raise RuntimeError(f"missing STEEZ guide_use_hint mapping for {detail_id}")
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        if old != value:
            cell.value = value
            changed.append({"row": row, "id": detail_id, "old": old, "new": value})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-guide_use_hint changes")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed": len(changed),
        "changed_ids": [item["id"] for item in changed],
    })


if __name__ == "__main__":
    main()
