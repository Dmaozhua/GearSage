from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel
import json
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
EVIDENCE_PATH = DATA_DIR / "evergreen_rod_whitelist_player_evidence.json"
REPORT_PATH = DATA_DIR / "evergreen_rod_recommended_rig_pairing_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"

FIELD = "recommended_rig_pairing"
AFTER_FIELD = "guide_use_hint"
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")

BANNED_GENERIC = re.compile(r"\b(?:General Lure|Light Rig|Hardbait|Soft Bait|Softbait|Soft Plastic)\b", re.I)


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def unique(items):
    out = []
    for item in items:
        item = n(item)
        if item and item not in out:
            out.append(item)
    return out


def join(items, limit=14):
    return " / ".join(unique(items)[:limit])


def has(text, *patterns):
    return any(re.search(pattern, text, re.I) for pattern in patterns)


def parse_max_number(value):
    text = n(value).replace("〜", "～").replace("~", "～")
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
    return max(nums) if nums else None


def parse_lure_max_g(row):
    lure = n(row.get("LURE WEIGHT"))
    if "号" in lure and not re.search(r"\bg\b", lure, re.I):
        return None
    return parse_max_number(lure)


def first_text_index(text, patterns):
    indexes = []
    for pattern in patterns:
        m = re.search(pattern, text, re.I)
        if m:
            indexes.append(m.start())
    return min(indexes) if indexes else None


TERM_PATTERNS = [
    # Bass soft rigs and cover techniques
    ("Frog", [r"フロッグ", r"\bfrog\b"]),
    ("Punching", [r"パンチング", r"\bpunching\b"]),
    ("Heavy Texas", [r"ヘビーテキサス", r"heavy texas"]),
    ("Light Texas", [r"ライトテキサス", r"light texas"]),
    ("Texas Rig", [r"テキサスリグ", r"テキサス", r"\btexas rig\b", r"\btexas\b"]),
    ("Free Rig", [r"フリーリグ", r"\bfree rig\b"]),
    ("Leaderless Down Shot", [r"リーダーレスダウンショット", r"leaderless down shot"]),
    ("Heavy Down Shot", [r"ヘビーダウンショット", r"heavy down shot"]),
    ("Down Shot", [r"ダウンショット", r"drop shot", r"down shot"]),
    ("Neko Rig", [r"ネコリグ", r"neko ring", r"neko rig", r"\bneko\b"]),
    ("High-density No Sinker", [r"高比重ノーシンカー", r"高比重.*no sinker"]),
    ("No Sinker", [r"ノーシンカー", r"lead-free", r"no sinker", r"無鉛"]),
    ("Small Rubber Jig", [r"スモラバ", r"small rubber jig"]),
    ("Football Jig", [r"フットボールジグ", r"football jig"]),
    ("Guarded Rubber Jig", [r"ガード付きラバージグ", r"ガード付ジグ", r"guarded rubber jig"]),
    ("Rubber Jig", [r"ラバージグ", r"rubber jig", r"guarded jig", r"cover jig", r"\bjig (?:and|or) texas\b"]),
    ("Carolina Rig", [r"キャロライナ", r"carolaina", r"carolina rig"]),
    ("Heavy Carolina Rig", [r"ヘビーキャロ", r"heavy carolina"]),
    ("Jighead Rig", [r"ジグヘッドリグ", r"ジグヘッド", r"jig head", r"jighead"]),
    ("Wacky Rig", [r"ワッキー", r"wacky"]),
    ("Mid Strolling Jighead", [r"ミドスト", r"mid strolling"]),
    ("Hover Strolling", [r"ホバスト", r"hover strolling"]),
    ("Bottom Strolling", [r"ボトスト", r"bottom strolling"]),
    # Bass moving/hard baits
    ("Chatterbait", [r"ブレーデッドジグ", r"チャターベイト", r"チャター", r"chatter ?bait", r"blade bait", r"bladed jig", r"刀片铅头钩"]),
    ("Swim Jig", [r"スイムジグ", r"swimming jig", r"swim jig", r"泳饵铅头钩"]),
    ("Spinnerbait", [r"スピナーベイト", r"spinner ?bait"]),
    ("Buzzbait", [r"バズベイト", r"buzz ?bait"]),
    ("Deep Crankbait", [r"ディープクランク", r"deep crank"]),
    ("Shallow Crankbait", [r"シャロークランク", r"shallow crank"]),
    ("Crankbait", [r"クランクベイト", r"クランク", r"crank ?bait"]),
    ("Shad", [r"シャッド", r"\bshad\b"]),
    ("Jerkbait", [r"ジャークベイト", r"jerk ?bait"]),
    ("Minnow", [r"ミノー", r"minnow"]),
    ("Vibration", [r"バイブレーション", r"\bvibration\b", r"lipless"]),
    ("Metal Vibration", [r"メタルバイブレーション", r"メタルバイブ", r"metal vibration"]),
    ("Topwater Plug", [r"トップウォーター", r"top ?water", r"topwater"]),
    ("Pencil Bait", [r"ペンシル", r"pencil"]),
    ("Popper", [r"ポッパー", r"popper"]),
    ("Crawler Bait", [r"クローラーベイト", r"crawler"]),
    ("Spoon", [r"スプーン", r"\bspoon\b"]),
    ("Big Bait", [r"ビッグベイト", r"big bait"]),
    ("Swimbait", [r"スイムベイト", r"swim ?bait"]),
    ("Alabama Rig", [r"アラバマリグ", r"alabama", r"\ba-rig\b"]),
    # Salt / trout
    ("Deep Eging", [r"ディープエギング", r"deep eging"]),
    ("Tip-run Eging", [r"ティップラン", r"tip[- ]?run"]),
    ("Slack-jerk Eging", [r"スラックジャーク"]),
    ("Eging", [r"エギング", r"エギ", r"\beging\b"]),
    ("SLSJ", [r"スーパーライトショアジギング", r"\bSLSJ\b"]),
    ("Micro Jigging", [r"マイクロジギング", r"micro jigging"]),
    ("Light Jigging", [r"ライトジギング", r"light jigging"]),
    ("Slow Pitch Jig", [r"スローピッチ", r"slow pitch", r"スロージギング"]),
    ("Long Fall Jig", [r"ロングフォール", r"long fall"]),
    ("Offshore Jigging", [r"ジギング", r"jigging"]),
    ("Metal Jig", [r"メタルジグ", r"metal jig"]),
    ("Casting Tairaba", [r"キャスティングタイラバ", r"casting tairaba"]),
    ("Tairaba", [r"タイラバ", r"tairaba"]),
    ("Rockfish Rig", [r"ロックフィッシュ", r"根魚", r"rockfish"]),
    ("Ajing Jighead", [r"アジング", r"ajing"]),
    ("Mebaring Jighead", [r"メバリング", r"メバル", r"mebaring"]),
    ("Light Game Jighead", [r"ライトゲーム", r"light game"]),
    ("Small Plug", [r"小型プラグ", r"small plug"]),
    ("Seabass Minnow", [r"シーバス.*ミノー", r"中型ミノー"]),
    ("Sinking Pencil", [r"シンキングペンシル", r"シンペン", r"sinking pencil"]),
    ("Seabass Plug", [r"シーバス", r"seabass"]),
    ("Shore Plug", [r"ショアプラグ", r"shore plug", r"大型プラグ", r"ヘビープラグ", r"heavy plug"]),
    ("Trout Spoon", [r"トラウト[^。]{0,30}スプーン", r"エリア[^。]{0,30}スプーン"]),
    ("Trout Minnow", [r"トラウト[^。]{0,30}ミノー", r"渓流[^。]{0,30}ミノー"]),
    ("Trout Crankbait", [r"トラウト[^。]{0,30}クランク", r"エリア[^。]{0,30}クランク"]),
]

TERM_INDEX = {name: idx for idx, (name, _) in enumerate(TERM_PATTERNS)}


def extract_terms(text):
    text = n(text)
    found = []
    for name, patterns in TERM_PATTERNS:
        idx = first_text_index(text, patterns)
        if idx is not None:
            found.append((idx, TERM_INDEX[name], name))
    found.sort(key=lambda item: (item[0], item[1]))
    return unique(name for _, _, name in found)


def refine_terms(terms, row, context):
    terms = unique(terms)
    text = lower(context)
    env = n(row.get("player_environment"))
    sku = n(row.get("SKU")).upper()

    # More precise terms should replace broader duplicates where both are present.
    if "Deep Crankbait" in terms and "Crankbait" in terms:
        terms.remove("Crankbait")
    if "Shallow Crankbait" in terms and "Crankbait" in terms:
        terms.remove("Crankbait")
    if "Metal Vibration" in terms and "Vibration" in terms:
        terms.remove("Vibration")
    if "High-density No Sinker" in terms and "No Sinker" in terms:
        terms.remove("No Sinker")
    if "Leaderless Down Shot" in terms and "Down Shot" in terms:
        terms.remove("Down Shot")
    if "Heavy Down Shot" in terms and "Down Shot" in terms:
        terms.remove("Down Shot")

    # In saltwater jigging descriptions, plain "jig" means metal jigging, not bass rubber jig.
    if ("船" in env or "铁板" in env or "鐵板" in env) and "Rubber Jig" in terms:
        terms.remove("Rubber Jig")
    if ("木虾" in env or "木蝦" in env or "乌贼" in env) and "Rubber Jig" in terms and "Metal Jig" in terms:
        terms.remove("Rubber Jig")
    if "淡水" in env or "Bass" in env:
        salt_only = {
            "Deep Eging", "Tip-run Eging", "Slack-jerk Eging", "Eging", "SLSJ",
            "Micro Jigging", "Light Jigging", "Slow Pitch Jig", "Long Fall Jig",
            "Offshore Jigging", "Casting Tairaba", "Tairaba", "Rockfish Rig",
            "Ajing Jighead", "Mebaring Jighead", "Light Game Jighead", "Seabass Minnow",
            "Sinking Pencil", "Seabass Plug", "Shore Plug",
        }
        terms = [term for term in terms if term not in salt_only and not term.startswith("Trout ")]
    elif not ("鳟" in env or "鱒" in env):
        terms = [term for term in terms if not term.startswith("Trout ")]
    if "轻型海水" in env or "輕型海水" in env:
        light_salt_exclusions = {
            "Offshore Jigging", "Slow Pitch Jig", "Long Fall Jig", "Casting Tairaba",
            "Tairaba", "Deep Eging", "Tip-run Eging", "Slack-jerk Eging", "Eging",
        }
        terms = [term for term in terms if term not in light_salt_exclusions]
    if ("海鲈" in env or "海鱸" in env) and "Offshore Jigging" in terms:
        terms.remove("Offshore Jigging")
    if ("船" in env or "铁板" in env or "鐵板" in env) and any(
        term in terms for term in ["Slow Pitch Jig", "Long Fall Jig", "Offshore Jigging", "Light Jigging", "Micro Jigging"]
    ) and "Metal Jig" not in terms:
        terms.append("Metal Jig")

    # Do not infer Big Bait just from heavy power; keep it only when explicit.
    if "Big Bait" in terms and not has(text, "ビッグベイト", "big bait", "大型餌", "大餌"):
        terms.remove("Big Bait")

    # Product-family fallbacks. These are only allowed to seed empty rows; explicit
    # model descriptions keep their own order.
    had_explicit_terms = bool(terms)
    if not had_explicit_terms and sku.startswith(("PSLJ", "PLFJ", "PHPJ")) and "Offshore Jigging" not in terms:
        terms.insert(0, "Offshore Jigging")
    if not had_explicit_terms and sku.startswith("PSPJ"):
        if "Micro Jigging" in terms:
            pass
        elif "Light Jigging" not in terms:
            terms.insert(0, "Light Jigging")
    if not had_explicit_terms and sku.startswith(("NIM", "SSSS")) and not any(t.endswith("Eging") or t == "Eging" for t in terms):
        terms.insert(0, "Eging")
    if not had_explicit_terms and sku.startswith("ZAGS") and "Seabass Plug" not in terms:
        terms.insert(0, "Seabass Plug")
    if not had_explicit_terms and sku.startswith(("PSS", "NEOS")) and not any(t in terms for t in ["Ajing Jighead", "Mebaring Jighead", "Light Game Jighead"]):
        terms.insert(0, "Light Game Jighead")
    if not had_explicit_terms and sku.startswith(("AATS", "AMSC", "AMSS")) and not any(t.startswith("Trout") for t in terms):
        terms.insert(0, "Trout Spoon")

    return unique(terms)


def fallback_pairing(row, model):
    env = n(row.get("player_environment"))
    pos = n(row.get("player_positioning"))
    sku = n(row.get("SKU")).upper()
    power = n(row.get("POWER")).upper()
    rod_type = n(row.get("TYPE"))
    max_g = parse_lure_max_g(row)
    text = lower(f"{model} {sku} {row.get('Code Name')} {row.get('Description')} {row.get('player_positioning')} {row.get('player_selling_points')}")

    if "船" in env or "铁板" in env or "鐵板" in env:
        if sku.startswith("PLFJ"):
            return ["Long Fall Jig", "Slow Pitch Jig", "Metal Jig"]
        if sku.startswith("PSPJ"):
            return ["Micro Jigging", "Light Jigging", "Small Metal Jig"]
        if sku.startswith("PHPJ"):
            return ["Offshore Jigging", "One Pitch Jig", "Metal Jig"]
        if sku.startswith("PREJ"):
            return ["Offshore Jigging", "Power Jigging", "Metal Jig"]
        return ["Slow Pitch Jig", "Offshore Jigging", "Metal Jig"]

    if "木虾" in env or "木蝦" in env or "乌贼" in env:
        if has(text, "deep", "ディープ", "tip"):
            return ["Deep Eging", "Tip-run Eging", "Light Jigging", "Tairaba"]
        return ["Slack-jerk Eging", "Eging", "Shallow Eging"]

    if "海鲈" in env or "海鱸" in env or sku.startswith("ZAGS"):
        if max_g and max_g <= 20:
            return ["Seabass Minnow", "Sinking Pencil", "Vibration", "Jighead Rig"]
        if max_g and max_g >= 45:
            return ["Shore Plug", "Metal Jig", "Sinking Pencil", "Seabass Minnow"]
        return ["Seabass Minnow", "Vibration", "Sinking Pencil", "Metal Jig"]

    if "轻型海水" in env or "輕型海水" in env or sku.startswith(("PSS", "NEOS")):
        if has(text, "アジ", "ajing") or "AJ" in sku:
            return ["Ajing Jighead", "Light Game Jighead", "Small Metal Jig", "Small Plug"]
        if has(text, "メバル", "mebaring"):
            return ["Mebaring Jighead", "Small Plug", "Light Game Jighead"]
        if has(text, "rock", "ロック", "根魚"):
            return ["Rockfish Rig", "Jighead Rig", "Free Rig", "Small Metal Jig"]
        return ["Light Game Jighead", "Small Metal Jig", "Small Plug", "Rockfish Rig"]

    if "岩" in env:
        return ["Rockfish Rig", "Texas Rig", "Jighead Rig", "Free Rig", "Metal Jig"]

    if "鳟" in env or "鱒" in env or sku.startswith(("AATS", "AMSC", "AMSS")):
        if sku.startswith("AM"):
            return ["Trout Minnow", "Trout Spoon", "Stream Plug"]
        return ["Trout Spoon", "Trout Crankbait", "Trout Minnow"]

    if has(text, "frog", "フロッグ") or "Frog" in pos:
        return ["Frog", "Punching", "Heavy Texas", "Rubber Jig"]
    if has(text, "big bait", "ビッグベイト", "swim bait", "swimbait", "スイムベイト"):
        return ["Swimbait", "Big Bait", "Spinnerbait", "Rubber Jig", "Texas Rig"]

    if rod_type == "S":
        if power in {"XUL", "UL"} or (max_g is not None and max_g <= 5):
            return ["Jighead Rig", "Down Shot", "Neko Rig", "No Sinker", "Small Minnow"]
        if power in {"L", "ML"} or (max_g is not None and max_g <= 14):
            return ["Neko Rig", "Down Shot", "Jighead Rig", "No Sinker", "Small Rubber Jig"]
        return ["Power Finesse", "Neko Rig", "Small Rubber Jig", "Down Shot", "Small Minnow"]

    if power in {"XUL", "UL", "L"} or "BF" in sku or (max_g is not None and max_g <= 7):
        return ["Neko Rig", "Down Shot", "No Sinker", "Small Rubber Jig", "Small Minnow"]
    if "ST" in sku or "ソリッド" in text:
        return ["Neko Rig", "Down Shot", "Small Rubber Jig", "Light Texas", "No Sinker"]
    if power in {"ML"} or (max_g is not None and max_g <= 14):
        return ["Shad", "Small Crankbait", "Minnow", "Neko Rig", "No Sinker"]
    if power in {"M"} or (max_g is not None and max_g <= 21):
        return ["Texas Rig", "Spinnerbait", "Chatterbait", "Crankbait", "No Sinker"]
    if power in {"MH", "MHR", "H", "XH", "XMH", "XX", "XXXH", "XXXHR", "XXH"} or (max_g is not None and max_g >= 28):
        return ["Texas Rig", "Rubber Jig", "Spinnerbait", "Chatterbait", "Swim Jig", "Crankbait"]
    return ["Texas Rig", "Rubber Jig", "Spinnerbait", "Crankbait", "Neko Rig"]


def build_evidence_index():
    if not EVIDENCE_PATH.exists():
        return {}
    data = json.loads(EVIDENCE_PATH.read_text())
    rows = data.get("evidence", []) if isinstance(data, dict) else data
    index = {}
    for item in rows:
        if item.get("scope") != "rod_detail":
            continue
        source_text = n(item.get("source_description"))
        if source_text:
            index[n(item.get("id"))] = item
    return index


def ensure_field(ws):
    headers = [cell.value for cell in ws[1]]
    if FIELD in headers:
        return headers.index(FIELD) + 1, False
    if AFTER_FIELD not in headers:
        raise RuntimeError(f"missing {AFTER_FIELD}; cannot insert {FIELD}")
    after_col = headers.index(AFTER_FIELD) + 1
    insert_col = after_col + 1
    ws.insert_cols(insert_col)
    ws.cell(1, insert_col).value = FIELD
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, after_col)
        dst = ws.cell(row, insert_col)
        if row != 1:
            dst.value = ""
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    ws.column_dimensions[ws.cell(1, insert_col).column_letter].width = 42
    return insert_col, True


def snapshot_without_field(ws, field_col):
    headers = [cell.value for cell in ws[1]]
    result = []
    for row in range(2, ws.max_row + 1):
        result.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != field_col and header != FIELD
        })
    return result


def classify_source(official_terms, whitelist_terms):
    if official_terms:
        return "official_description"
    if whitelist_terms:
        return "whitelist_auxiliary"
    return "conservative_inference"


def consistency_flags(pairing, guide_hint):
    flags = []
    first = n(pairing).split(" / ")[0] if n(pairing) else ""
    guide = n(guide_hint)
    soft_first = first in {
        "Texas Rig", "Heavy Texas", "Light Texas", "Free Rig", "Leaderless Down Shot",
        "Heavy Down Shot", "Down Shot", "Neko Rig", "No Sinker", "High-density No Sinker",
        "Small Rubber Jig", "Football Jig", "Rubber Jig", "Jighead Rig", "Wacky Rig",
        "Mid Strolling Jighead", "Hover Strolling", "Bottom Strolling", "Rockfish Rig",
        "Ajing Jighead", "Mebaring Jighead", "Light Game Jighead",
    }
    hard_first = first in {
        "Chatterbait", "Swim Jig", "Spinnerbait", "Buzzbait", "Crankbait", "Deep Crankbait",
        "Shallow Crankbait", "Shad", "Jerkbait", "Minnow", "Vibration", "Metal Vibration",
        "Topwater Plug", "Pencil Bait", "Popper", "Crawler Bait", "Spoon", "Big Bait",
        "Swimbait", "Alabama Rig", "Seabass Minnow", "Sinking Pencil", "Seabass Plug",
        "Shore Plug", "Trout Spoon", "Trout Minnow", "Trout Crankbait",
    }
    switch_words = r"泛用|切換|切换|兼顧|兼顾|可切|後段可接|后段可接|需要擴展|需要扩展|放慢"
    if soft_first and re.search(r"硬餌|硬饵|搜索", guide) and not re.search(switch_words, guide):
        flags.append("soft_first_with_hard_search_hint")
    if hard_first and re.search(r"軟餌底操|软饵底操|底操", guide) and not re.search(switch_words, guide):
        flags.append("hard_first_with_soft_bottom_hint")
    return flags


def main():
    evidence_by_id = build_evidence_index()
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]

    rod_headers = [cell.value for cell in rod_ws[1]]
    rod_col = {name: i + 1 for i, name in enumerate(rod_headers)}
    rod_models = {
        n(rod_ws.cell(row, rod_col["id"]).value): n(rod_ws.cell(row, rod_col["model"]).value)
        for row in range(2, rod_ws.max_row + 1)
    }

    inserted_col, inserted = ensure_field(detail_ws)
    before = snapshot_without_field(detail_ws, inserted_col)

    headers = [cell.value for cell in detail_ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    required = [
        "id", "rod_id", "TYPE", "SKU", "POWER", "LURE WEIGHT", "Line Wt N F",
        "PE Line Size", "Code Name", "guide_use_hint", "player_environment",
        "player_positioning", "player_selling_points", "Description", FIELD,
    ]
    missing = [name for name in required if name not in col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    changes = []
    final_rows = []
    source_counts = {}
    consistency = []

    for row_idx in range(2, detail_ws.max_row + 1):
        row = {name: detail_ws.cell(row_idx, col[name]).value for name in required}
        detail_id = n(row["id"])
        model = rod_models.get(n(row["rod_id"]), "")
        official_text = n(row.get("Description"))
        official_terms = refine_terms(extract_terms(official_text), row, official_text)

        evidence = evidence_by_id.get(detail_id, {})
        whitelist_text = n(evidence.get("source_description"))
        whitelist_terms = refine_terms(extract_terms(whitelist_text), row, whitelist_text)

        if official_terms:
            terms = official_terms
        elif whitelist_terms:
            terms = whitelist_terms
        else:
            terms = fallback_pairing(row, model)

        value = join(terms)
        if not value or BANNED_GENERIC.search(value):
            terms = fallback_pairing(row, model)
            value = join(terms)

        source_type = classify_source(official_terms, whitelist_terms)
        source_counts[source_type] = source_counts.get(source_type, 0) + 1
        final_rows.append({
            "row": row_idx,
            "id": detail_id,
            "rod_id": n(row["rod_id"]),
            "sku": n(row["SKU"]),
            "code_name": n(row["Code Name"]),
            "recommended_rig_pairing": value,
            "source_type": source_type,
            "official_terms": official_terms,
            "whitelist_terms": whitelist_terms,
            "source_url": evidence.get("source_url", ""),
        })

        cell = detail_ws.cell(row_idx, col[FIELD])
        old = n(cell.value)
        if old != value:
            cell.value = value
            cell.fill = copy(YELLOW_FILL)
            changes.append({
                "row": row_idx,
                "id": detail_id,
                "rod_id": n(row["rod_id"]),
                "sku": n(row["SKU"]),
                "code_name": n(row["Code Name"]),
                "old": old,
                "new": value,
                "source_type": source_type,
                "official_terms": official_terms,
                "whitelist_terms": whitelist_terms,
                "source_url": evidence.get("source_url", ""),
            })

        flags = consistency_flags(value, row.get("guide_use_hint"))
        if flags:
            consistency.append({
                "id": detail_id,
                "sku": n(row["SKU"]),
                "recommended_rig_pairing": value,
                "guide_use_hint": n(row.get("guide_use_hint")),
                "flags": flags,
            })

    after = snapshot_without_field(detail_ws, col[FIELD])
    if before != after:
        raise RuntimeError("unexpected changes outside recommended_rig_pairing")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_file": str(XLSX_PATH),
        "field": FIELD,
        "inserted_column": inserted,
        "changed_rows_this_run": len(changes),
        "source_counts": source_counts,
        "coverage": {
            "filled": detail_ws.max_row - 1,
            "total": detail_ws.max_row - 1,
        },
        "banned_generic_count": 0,
        "consistency_flags": consistency,
        "final_rows": final_rows,
        "changes_this_run": changes,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "inserted_column": inserted,
        "changed_rows_this_run": len(changes),
        "source_counts": source_counts,
        "coverage": report["coverage"],
        "consistency_flags": len(consistency),
        "report_file": str(REPORT_PATH),
        "samples": changes[:10],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
