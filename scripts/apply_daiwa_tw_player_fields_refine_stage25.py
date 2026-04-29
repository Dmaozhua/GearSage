from collections import Counter
from pathlib import Path
import re

from openpyxl import load_workbook


XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")
TARGET_FIELDS = ["player_environment", "player_positioning", "player_selling_points"]


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def low(value):
    return n(value).lower()


def has(text, *patterns):
    return any(re.search(pattern, text, re.I) for pattern in patterns)


def first_terms(rig, limit=3):
    return [part.strip() for part in n(rig).split("/") if part.strip()][:limit]


def main_tech_text(rig):
    terms = first_terms(rig, 3)
    if not terms:
        return "常用釣組"
    if len(terms) == 1:
        return terms[0]
    if len(terms) == 2:
        return f"{terms[0]} 與 {terms[1]}"
    return f"{terms[0]}、{terms[1]}、{terms[2]}"


def naturalize_selling_point(value, seed=""):
    text = n(value)
    if not text:
        return text
    if "不作為釣竿技法" in text:
        return "主要用於軟絲起魚輔助，不納入釣竿技法或釣組搭配判斷。"

    text = text.replace("優先", "").replace("重點是", "重點在")
    if " / " not in text:
        return text.strip(" /。") + "。"

    tech, detail = text.split(" / ", 1)
    tech = n(tech).strip(" /，。")
    detail = n(detail).strip(" /，。")
    if detail.startswith("適合"):
        detail = detail[2:].strip()

    focus = ""
    for marker in ["，重點在", "，尤其重點在"]:
        if marker in detail:
            detail, focus = detail.split(marker, 1)
            break
    detail = detail.strip("，。 ")
    focus = focus.strip("，。 ")

    index = sum(ord(ch) for ch in seed) % 5
    if focus:
        templates = [
            f"以 {tech} 做主軸時，{detail}；{focus}會更容易掌握。",
            f"{tech} 搭配起來更順手，{detail}，尤其能放大{focus}。",
            f"偏向 {tech} 的使用節奏，{detail}時，{focus}更清楚。",
            f"用在 {tech} 這類玩法時，{detail}；主要收益是{focus}。",
            f"{tech} 是較自然的搭配方向，{detail}時能保留{focus}。",
        ]
        return templates[index]

    if "適合" in detail or "兼顧" in detail:
        templates = [
            f"以 {tech} 做主軸時，{detail}。",
            f"{tech} 搭配起來較自然，{detail}。",
            f"偏向 {tech} 的使用節奏，{detail}。",
            f"用在 {tech} 這類玩法時，{detail}。",
            f"{tech} 可以作為主要搭配方向，{detail}。",
        ]
        return templates[index]

    templates = [
        f"以 {tech} 做主軸時，{detail}會更順手。",
        f"{tech} 搭配起來較自然，{detail}時更容易發揮。",
        f"偏向 {tech} 的使用節奏，適合{detail}。",
        f"用在 {tech} 這類玩法時，{detail}會比較穩。",
        f"{tech} 可以作為主要搭配方向，{detail}時更好掌握。",
    ]
    return templates[index]


def snapshot_without_targets(ws, target_cols):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col not in target_cols
        })
    return rows


def is_bass_model(text):
    return has(
        text,
        "steez",
        "tatula",
        "black label",
        "heartland",
        "airedge",
        "wilderness",
        "vertice",
    )


def bass_profile(model, sku, text, rig):
    portable = has(model, "travel", "mobile", "wilderness", "vertice", "liberalist")
    env_suffix = "Bass 便攜" if portable else "Bass"
    primary = " / ".join(first_terms(rig, 3)).lower()
    if has(primary, "frog", "punching", "heavy texas"):
        positioning = "Frog / Punching 強障礙" if has(primary, "frog") else "Punching / Heavy Cover 打點"
        return (
            f"淡水 {env_suffix} / 重障礙",
            positioning,
            f"{main_tech_text(rig)} 優先 / 適合草洞、覆蓋物邊快速打點和中魚後拉離障礙",
        )
    if has(primary, "big bait", "swimbait", "swim bait"):
        return (
            f"淡水 {env_suffix} / 大餌",
            "Swimbait / Big Bait 強力拋投" if has(rig, "big bait") else "Swimbait 強力搜索",
            f"{main_tech_text(rig)} 優先 / 適合較重餌搜索、定點拋投和需要竿身支撐的控魚",
        )
    if has(text, "hover strolling", "mid strolling", "bottom strolling"):
        return (
            f"淡水 {env_suffix} / 中層精細",
            "Hover / Mid Strolling",
            "Strolling 系精細軟餌優先 / 適合中層控線、微小晃動和細線咬口判斷",
        )
    if has(text, "bait finesse", r"\bbf\b"):
        return (
            f"淡水 {env_suffix} / BFS",
            "BFS 精細軟餌",
            "No Sinker、Neko 和小型 Jighead 優先 / 適合低彈道輕餌拋投、障礙邊控線和短咬口判斷",
        )
    hard_count = sum(1 for p in ["crank", "shad", "minnow", "jerkbait", "topwater", "popper", "vibration"] if has(rig, p))
    wire_count = sum(1 for p in ["spinnerbait", "buzzbait", "chatterbait", "swim jig", "bladed jig"] if has(rig, p))
    soft_count = sum(1 for p in ["neko", "down shot", "no sinker", "texas", "free rig", "rubber jig", "jighead", "soft plastic"] if has(rig, p))
    if wire_count and has(primary, "spinnerbait", "buzzbait", "chatterbait", "swim jig", "bladed") and has(rig, "swimbait", "big bait"):
        return (
            f"淡水 {env_suffix} / 搜索硬餌",
            "Wire Bait / Swimbait 強力搜索",
            f"{main_tech_text(rig)} 優先 / 適合中重型移動餌與 Swimbait 類大範圍搜索，重點在回收負荷和控魚餘量",
        )
    if hard_count and wire_count and soft_count:
        return (
            f"淡水 {env_suffix} / 軟硬餌全能",
            "硬餌搜索兼精細軟餌",
            f"{main_tech_text(rig)} 優先 / 硬餌搜索和精細軟餌可在同一支竿上切換，適合不想頻繁換竿的場景",
        )
    if hard_count and (wire_count or has(text, "wire bait")):
        return (
            f"淡水 {env_suffix} / 搜索硬餌",
            "Crank / Wire Bait 搜索",
            f"{main_tech_text(rig)} 優先 / 適合平收、抽停和快速覆蓋水層，重點在節奏穩定與泳層控制",
        )
    if hard_count and soft_count:
        return (
            f"淡水 {env_suffix} / 輕量多用途",
            "小型硬餌兼輕量軟餌",
            f"{main_tech_text(rig)} 優先 / 小型硬餌與輕量軟餌都能兼顧，適合岸邊小場景和魚口不穩時切換",
        )
    if has(rig, "free rig", "texas", "leaderless", "heavy down shot", "rubber jig", "punching"):
        heavy = has(text, "heavy", "mh", "h", "重", "cover", "障礙")
        return (
            f"淡水 {env_suffix} / 底操",
            "Free Rig / Texas 底部打點" if not heavy else "中重型底操 / Cover 打點",
            f"{main_tech_text(rig)} 優先 / 適合讀底、碰障和貼近結構慢速操作，重點在張力變化和刺魚餘量",
        )
    if has(rig, "neko", "down shot", "no sinker", "jighead", "wacky", "i-shaped", "bug lure", "small rubber jig"):
        return (
            f"淡水 {env_suffix} / 精細",
            "Neko / Down Shot 精細操作",
            f"{main_tech_text(rig)} 優先 / 適合細線、輕釣組和小餌慢節奏操作，重點在控鬆線與微弱咬口判斷",
        )
    if wire_count:
        return (
            f"淡水 {env_suffix} / 移動餌",
            "Spinnerbait / Swim Jig 搜索",
            f"{main_tech_text(rig)} 優先 / 適合中速搜索、障礙邊通過和需要穩定回收負荷的移動餌",
        )
    return (
        f"淡水 {env_suffix} / 泛用",
        "軟硬餌泛用",
        f"{main_tech_text(rig)} 可作為主軸 / 適合依照水域、標點和魚口在軟餌與硬餌之間切換",
    )


def eging_profile(text, rig):
    if has(text, "mmh", "mh", "monster", "大物", "深場"):
        return (
            "海水岸拋 / 木蝦大物",
            "春季大物木蝦",
            "3.5 號以上木蝦和 Sinker Rig 優先 / 適合深場、流速較快和需要補刺支撐的大型軟絲",
        )
    if has(text, "solid", "smt", "-s", "高感", "細膩", "秋季", "run&gun"):
        return (
            "海水岸拋 / 木蝦精細",
            "高感木蝦 / 看線補刺",
            "Eging 和 Sinker Rig 優先 / 適合細膩跳蝦、看線判斷和短咬口後的快速補刺",
        )
    return (
        "海水岸拋 / 木蝦",
        "木蝦泛用",
        "Eging 和 Sinker Rig 優先 / 適合港口、堤防和岸拋木蝦，重點在連續抽竿與下沉控線",
    )


def shore_seabass_profile(model, text, rig):
    if has(text, "港灣", "運河", "urban", "pin", "精準"):
        return (
            "海水岸拋 / 港灣海鱸",
            "港灣打點硬餌",
            f"{main_tech_text(rig)} 優先 / 適合橋墩、港灣邊線和近中距離標點，重點在落點精度與控線",
        )
    if has(text, "surf", "沙灘", "河口", "遠投", "long", "青物"):
        return (
            "海水岸拋 / 河口沙灘",
            "遠投搜索",
            f"{main_tech_text(rig)} 優先 / 適合河口、沙灘和開闊岸線，重點在飛距、線弧管理與大範圍搜索",
        )
    if has(text, "lunker", "monster", "mh", "強力", "大型"):
        return (
            "海水岸拋 / 大型海鱸",
            "大型硬餌 / 強力控魚",
            f"{main_tech_text(rig)} 優先 / 適合中大型海鱸、強流和較大路亞，重點在拋投後控線與控魚餘量",
        )
    return (
        "海水岸拋 / 海鱸",
        "海鱸硬餌搜索",
        f"{main_tech_text(rig)} 優先 / 適合米諾、沉水鉛筆和震動餌搜索，重點在泳層控制和咬口傳達",
    )


def jigging_profile(model, text, rig):
    if has(text, "tachiuo", "白帶", "鏡牙"):
        return (
            "近海船釣 / 白帶魚",
            "白帶魚鐵板",
            "Tachiuo Jigging 和 Metal Jig 優先 / 適合垂直控餌、停頓誘食和輕啄咬口判斷",
        )
    if has(text, "slj", "super light", "輕鐵", "超級輕量"):
        return (
            "近海船釣 / SLJ",
            "SLJ 輕鐵板",
            "SLJ 和小型 Metal Jig 優先 / 適合輕線小鐵板、斜拋與垂直抽停切換",
        )
    if has(text, "slow", "慢鐵", "lowresponse"):
        return (
            "近海船釣 / 慢鐵",
            "慢鐵 / Slow Pitch",
            "Slow Jigging 和 Metal Jig 優先 / 適合慢節奏抽停、深場承重和下沉咬口判斷",
        )
    if has(text, "one pitch", "semi-long", "long jig"):
        return (
            "近海船釣 / 青物鐵板",
            "One Pitch / 長鐵板",
            "Metal Jig、Semi-long Jig 和 Long Jig 優先 / 適合青物節奏抽停、深水和船流變化",
        )
    return (
        "近海船釣 / 鐵板",
        "Metal Jig 船釣泛用",
        f"{main_tech_text(rig)} 優先 / 適合垂直控餌、抽停節奏和中魚後持續施壓",
    )


def offshore_casting_profile(text, rig):
    heavy = has(text, "pe6", "pe8", "pe10", "pe12", "gt", "鮪", "黃鰭", "200mm", "大型")
    return (
        "近海船拋 / 大型青物鮪魚" if heavy else "近海船拋 / 青物",
        "大型 Plug / 鮪魚 GT" if heavy else "Diving Pencil / Popper 船拋",
        f"{main_tech_text(rig)} 優先 / 適合船拋表層搜索、遠距離拋投和高負荷搏魚",
    )


def light_salt_profile(text, rig):
    if has(text, "遠投", "float", "沙灘", "大型"):
        return (
            "輕海水 / 港口遠投",
            "Float Rig / 小型 Plug 遠投",
            f"{main_tech_text(rig)} 優先 / 適合港口外側、小型磯場和需要增加搜索距離的輕海水場景",
        )
    if has(text, "rockfish", "根魚", "鮶", "texas"):
        return (
            "輕海水 / 根魚",
            "Jighead / Rockfish Rig",
            f"{main_tech_text(rig)} 優先 / 適合港口根魚、小型軟餌讀底和細線控線",
        )
    return (
        "輕海水 / 竹莢魚根魚",
        "Jighead / 小型 Plug 精細",
        f"{main_tech_text(rig)} 優先 / 適合輕量汲鉤頭、小型 Plug 和夜間細咬口判斷",
    )


def rockfish_profile(text, rig):
    if has(text, "遠投", "mh", "h", "強力", "重量級"):
        return (
            "海水岸拋 / 根魚岩礁",
            "遠投強力根魚",
            f"{main_tech_text(rig)} 優先 / 適合重釣組遠投、硬底讀底和中魚後拉離岩礁",
        )
    return (
        "海水岸拋 / 根魚岩礁",
        "Texas / Jighead 根魚底操",
        f"{main_tech_text(rig)} 優先 / 適合貼底慢搜、避開根掛和結構區細膩操作",
    )


def chining_profile(text, rig):
    if has(text, "遠投", "8ft"):
        return (
            "海水岸釣 / 黑鯛",
            "遠投 Free Rig",
            "Chining Free Rig 和 Bottom Rig 優先 / 適合遠距標點、底部拖動和細 PE 控線",
        )
    if has(text, "beast", "強悍", "power", "短尺寸", "高輸出"):
        return (
            "海水岸釣 / 黑鯛障礙",
            "強力 Chining",
            "Chining Free Rig 和 Bottom Rig 優先 / 適合障礙邊黑鯛、短距離精準拋投和中魚後壓制",
        )
    return (
        "海水岸釣 / 黑鯛",
        "Free Rig 底操",
        "Free Rig 和 Bottom Rig 優先 / 適合黑鯛底部觸感、滯線控制和輕咬口判斷",
    )


def trout_profile(text, rig):
    return (
        "淡水管理場 / 鱒魚",
        "Spoon / 小型 Plug 精細",
        f"{main_tech_text(rig)} 優先 / 適合細線慢速搜索、輕量 Spoon 操作和短咬口判斷",
    )


def wind_profile(text, rig):
    return (
        "海水岸拋 / WIND 白帶魚",
        "Jighead Dart / WIND",
        "WIND、Jighead Dart 和 Metal Jig 優先 / 適合連續抽動、下沉控線和白帶魚咬口判斷",
    )


def fallback_profile(text, rig, rod_type):
    return (
        "路亞泛用",
        "直柄多用途" if rod_type == "S" else "槍柄多用途",
        f"{main_tech_text(rig)} 可作為主軸 / 適合按標點、餌重和線組在多種路亞玩法間切換",
    )


def refine(model, sku, rod_type, power, description, rig, guide_hint):
    model_l = low(model)
    text = low(f"{model} {sku} {power} {description} {rig} {guide_hint}")
    rig_l = low(rig)

    if has(text, "landing gaff", "搭鉤"):
        return ("配件 / 起魚工具", "軟絲搭鉤", "用於軟絲起魚輔助 / 不作為釣竿技法或釣組搭配判斷")
    if has(text, "silver wolf", "chining", "黑鯛"):
        return chining_profile(text, rig)
    if has(text, "emeraldas", "eging", "木蝦", "軟絲", "烏賊"):
        return eging_profile(text, rig)
    if has(text, "hardrock", "rockfish", "根魚", "石斑", "岩魚"):
        return rockfish_profile(text, rig)
    if has(text, "月下美人", "ajing", "mebaru", "竹莢", "鮶", "無備平鮋"):
        return light_salt_profile(text, rig)
    if has(text, "wind x", "wind", "jighead dart", "白帶魚") and not has(text, "鏡牙"):
        return wind_profile(text, rig)
    if has(text, "trout", "iprimi", "鱒", "管理場", "管理池"):
        return trout_profile(text, rig)
    if is_bass_model(model_l) or has(rig_l, "texas", "neko", "down shot", "spinnerbait", "crank", "swimbait", "big bait"):
        return bass_profile(model_l, sku, text, rig)
    if has(model_l, "morethan", "lateo") or has(text, "seabass plug", "海鱸"):
        return shore_seabass_profile(model_l, text, rig)
    if has(model_l, "overthere", "dragger") or (
        not has(model_l, "saltiga c", "outrage br")
        and has(text, "surf plug", "slsj", "比目魚")
    ):
        if has(text, "slsj", "small metal jig"):
            return (
                "海水岸拋 / SLSJ",
                "小型鐵板遠投",
                f"{main_tech_text(rig)} 優先 / 適合岸拋小鐵板、米諾和沉水鉛筆搜索，重點在飛距與回收控線",
            )
        return shore_seabass_profile(model_l, text, rig)
    if (
        has(model_l, "saltiga slj", "lowresponse", "鏡牙")
        or has(text, "outrage br lj", "offshore jigging", "light jigging", "slow jig", "slow pitch", "one pitch", "slj", "鐵板")
    ):
        if not is_bass_model(model_l):
            return jigging_profile(model_l, text, rig)
    if has(model_l, "saltiga c", "outrage br") or (
        not is_bass_model(model_l)
        and has(text, "船拋", "diving pencil", "popper", "stickbait", "offshore plug", "gt", "黃鰭", "鮪")
    ):
        return offshore_casting_profile(text, rig)
    if has(text, "crossbeat"):
        if has(rig_l, "eging"):
            return eging_profile(text, rig)
        if has(rig_l, "chining"):
            return chining_profile(text, rig)
        if has(rig_l, "jighead", "small plug", "float"):
            return light_salt_profile(text, rig)
    return fallback_profile(text, rig, rod_type)


def main():
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    ws = wb["rod_detail"]

    rod_headers = [cell.value for cell in rod_ws[1]]
    detail_headers = [cell.value for cell in ws[1]]
    rod_col = {name: i + 1 for i, name in enumerate(rod_headers)}
    col = {name: i + 1 for i, name in enumerate(detail_headers)}
    for field in ["id", "rod_id", "TYPE", "SKU", "POWER", "Description", "recommended_rig_pairing", "guide_use_hint", *TARGET_FIELDS]:
        if field not in col:
            raise RuntimeError(f"missing rod_detail column: {field}")
    for field in ["id", "model"]:
        if field not in rod_col:
            raise RuntimeError(f"missing rod column: {field}")

    model_by_id = {
        n(rod_ws.cell(row, rod_col["id"]).value): n(rod_ws.cell(row, rod_col["model"]).value)
        for row in range(2, rod_ws.max_row + 1)
    }

    target_cols = {col[field] for field in TARGET_FIELDS}
    before_non_targets = snapshot_without_targets(ws, target_cols)
    before_master = [[rod_ws.cell(row, c).value for c in range(1, rod_ws.max_column + 1)] for row in range(1, rod_ws.max_row + 1)]

    changed = []
    value_counts = {field: Counter() for field in TARGET_FIELDS}
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        rod_id = n(ws.cell(row, col["rod_id"]).value)
        model = model_by_id.get(rod_id, "")
        env, pos, sell = refine(
            model=model,
            sku=n(ws.cell(row, col["SKU"]).value),
            rod_type=n(ws.cell(row, col["TYPE"]).value),
            power=n(ws.cell(row, col["POWER"]).value),
            description=n(ws.cell(row, col["Description"]).value),
            rig=n(ws.cell(row, col["recommended_rig_pairing"]).value),
            guide_hint=n(ws.cell(row, col["guide_use_hint"]).value),
        )
        sell = naturalize_selling_point(sell, detail_id)
        new_values = {
            "player_environment": env,
            "player_positioning": pos,
            "player_selling_points": sell,
        }
        row_changes = {}
        for field, new_value in new_values.items():
            cell = ws.cell(row, col[field])
            old_value = n(cell.value)
            value_counts[field][new_value] += 1
            if old_value != new_value:
                cell.value = new_value
                row_changes[field] = {"old": old_value, "new": new_value}
        if row_changes:
            changed.append({"id": detail_id, "rod_id": rod_id, "model": model, "changes": row_changes})

    after_non_targets = snapshot_without_targets(ws, target_cols)
    if before_non_targets != after_non_targets:
        raise RuntimeError("unexpected non-player-field changes in rod_detail")
    after_master = [[rod_ws.cell(row, c).value for c in range(1, rod_ws.max_column + 1)] for row in range(1, rod_ws.max_row + 1)]
    if before_master != after_master:
        raise RuntimeError("unexpected changes in rod sheet")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed_rows": len(changed),
        "changed_cells": sum(len(item["changes"]) for item in changed),
        "unique_values": {field: len(counter) for field, counter in value_counts.items()},
        "top_values": {field: counter.most_common(8) for field, counter in value_counts.items()},
    })


if __name__ == "__main__":
    main()
