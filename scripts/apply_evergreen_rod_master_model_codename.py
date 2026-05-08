#!/usr/bin/env python3
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from gear_data_paths import resolve_data_raw


XLSX_PATH = resolve_data_raw("evergreen_rod_import.xlsx")
FIELD = "model"
JP_RE = re.compile(r"[\u3040-\u30ff\u3400-\u9fff]")


def norm(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def headers(ws):
    return [cell.value for cell in ws[1]]


def col_map(ws):
    return {name: idx + 1 for idx, name in enumerate(headers(ws))}


def sheet_values_without_column(ws, excluded_col):
    cols = [col for col in range(1, ws.max_column + 1) if col != excluded_col]
    return [
        tuple(ws.cell(row=row, column=col).value for col in cols)
        for row in range(1, ws.max_row + 1)
    ]


def model_prefix(model, fallback_sku):
    model = norm(model)
    if model:
        return model.split(None, 1)[0]
    return norm(fallback_sku)


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(str(XLSX_PATH))

    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_cols = col_map(rod_ws)
    detail_cols = col_map(detail_ws)

    required_rod = {"id", FIELD}
    required_detail = {"rod_id", "SKU", "Code Name"}
    missing_rod = sorted(required_rod - set(rod_cols))
    missing_detail = sorted(required_detail - set(detail_cols))
    if missing_rod or missing_detail:
        raise RuntimeError({"missing_rod": missing_rod, "missing_detail": missing_detail})

    model_col = rod_cols[FIELD]
    before_rod_without_model = sheet_values_without_column(rod_ws, model_col)
    before_detail = [
        tuple(detail_ws.cell(row=row, column=col).value for col in range(1, detail_ws.max_column + 1))
        for row in range(1, detail_ws.max_row + 1)
    ]

    detail_by_rod_id = {}
    duplicate_detail_ids = []
    for row_idx in range(2, detail_ws.max_row + 1):
        rod_id = norm(detail_ws.cell(row=row_idx, column=detail_cols["rod_id"]).value)
        if not rod_id:
            continue
        if rod_id in detail_by_rod_id:
            duplicate_detail_ids.append(rod_id)
        detail_by_rod_id[rod_id] = row_idx
    if duplicate_detail_ids:
        raise RuntimeError({"duplicate_detail_rod_id": duplicate_detail_ids[:20]})

    changes = []
    missing_detail = []
    blank_code_name = []
    for row_idx in range(2, rod_ws.max_row + 1):
        rod_id = norm(rod_ws.cell(row=row_idx, column=rod_cols["id"]).value)
        detail_row = detail_by_rod_id.get(rod_id)
        if not detail_row:
            missing_detail.append({"row": row_idx, "rod_id": rod_id})
            continue

        code_name = norm(detail_ws.cell(row=detail_row, column=detail_cols["Code Name"]).value)
        sku = norm(detail_ws.cell(row=detail_row, column=detail_cols["SKU"]).value)
        old_model = norm(rod_ws.cell(row=row_idx, column=model_col).value)
        if not code_name:
            blank_code_name.append({"row": row_idx, "rod_id": rod_id, "model": old_model, "sku": sku})
            continue

        new_model = f"{model_prefix(old_model, sku)} {code_name}".strip()
        if old_model != new_model:
            rod_ws.cell(row=row_idx, column=model_col).value = new_model
            changes.append({
                "row": row_idx,
                "rod_id": rod_id,
                "sku": sku,
                "old": old_model,
                "new": new_model,
            })

    if missing_detail:
        raise RuntimeError({"missing_detail": missing_detail[:20]})

    if sheet_values_without_column(rod_ws, model_col) != before_rod_without_model:
        raise RuntimeError("rod sheet changed outside model column")

    after_detail = [
        tuple(detail_ws.cell(row=row, column=col).value for col in range(1, detail_ws.max_column + 1))
        for row in range(1, detail_ws.max_row + 1)
    ]
    if after_detail != before_detail:
        raise RuntimeError("rod_detail sheet changed unexpectedly")

    wb.save(XLSX_PATH)

    remaining_japanese = []
    for row_idx in range(2, rod_ws.max_row + 1):
        value = norm(rod_ws.cell(row=row_idx, column=model_col).value)
        if JP_RE.search(value):
            remaining_japanese.append({
                "row": row_idx,
                "rod_id": norm(rod_ws.cell(row=row_idx, column=rod_cols["id"]).value),
                "model": value,
            })

    report = {
        "file": str(XLSX_PATH),
        "scope": "rod.model only",
        "rod_rows": rod_ws.max_row - 1,
        "detail_rows": detail_ws.max_row - 1,
        "changed": len(changes),
        "blank_code_name": len(blank_code_name),
        "remaining_japanese_model": len(remaining_japanese),
        "changed_samples": changes[:20],
        "blank_code_name_samples": blank_code_name[:20],
        "remaining_japanese_samples": remaining_japanese[:20],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
