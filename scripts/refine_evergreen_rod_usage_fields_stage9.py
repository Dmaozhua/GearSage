from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel
import json
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "evergreen_rod_usage_fields_stage9_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"

PAIR_FIELD = "recommended_rig_pairing"
GUIDE_FIELD = "guide_use_hint"
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def has(text, *patterns):
    return any(re.search(pattern, text, re.I) for pattern in patterns)


def split_pairing(value):
    return [part.strip() for part in n(value).split(" / ") if part.strip()]


def join_terms(items, limit=3):
    return " / ".join(items[:limit])


def parse_max_number(value):
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", n(value))]
    return max(nums) if nums else None


def line_summary(row):
    parts = []
    nf = n(row.get("Line Wt N F"))
    pe = n(row.get("PE Line Size"))
    if nf:
        parts.append(nf)
    if pe:
        parts.append(pe)
    return "、".join(parts)


def spec_clause(row):
    lure = n(row.get("LURE WEIGHT"))
    line = line_summary(row)
    bits = []
    if lure:
        bits.append(f"{lure} 負荷")
    if line:
        bits.append(f"{line} 線組")
    return "，".join(bits)


SOFT_TERMS = {
    "Texas Rig", "Heavy Texas", "Light Texas", "Free Rig", "Leaderless Down Shot",
    "Heavy Down Shot", "Down Shot", "Neko Rig", "No Sinker", "High-density No Sinker",
    "Small Rubber Jig", "Football Jig", "Rubber Jig", "Guarded Rubber Jig", "Jighead Rig",
    "Wacky Rig", "Mid Strolling Jighead", "Hover Strolling", "Bottom Strolling",
    "Carolina Rig", "Heavy Carolina Rig", "Power Finesse",
}

MOVING_TERMS = {
    "Chatterbait", "Swim Jig", "Spinnerbait", "Buzzbait", "Crankbait", "Deep Crankbait",
    "Shallow Crankbait", "Shad", "Jerkbait", "Minnow", "Vibration", "Metal Vibration",
    "Topwater Plug", "Pencil Bait", "Popper", "Crawler Bait", "Spoon", "Big Bait",
    "Swimbait", "Alabama Rig", "Small Plug",
}

SALT_TERMS = {
    "Slow Pitch Jig", "Long Fall Jig", "Offshore Jigging", "Light Jigging", "Micro Jigging",
    "Metal Jig", "Deep Eging", "Tip-run Eging", "Slack-jerk Eging", "Eging", "SLSJ",
    "Casting Tairaba", "Tairaba", "Rockfish Rig", "Ajing Jighead", "Mebaring Jighead",
    "Light Game Jighead", "Seabass Minnow", "Sinking Pencil", "Seabass Plug", "Shore Plug",
}

TROUT_TERMS = {"Trout Spoon", "Trout Minnow", "Trout Crankbait"}


def has_any_term(parts, group):
    return any(part in group for part in parts)


def describe_bass(parts, row, desc):
    secondary_soft = [term for term in parts if term in SOFT_TERMS][:3]
    secondary_moving = [term for term in parts if term in MOVING_TERMS][:3]
    spec = spec_clause(row)
    cue = lower(desc)

    if parts and parts[0] in {"Big Bait", "Swimbait", "Alabama Rig"}:
        lead = join_terms(parts, 4)
        angle = "大餌/泳餌"
        detail = "重餌上竿、低彈道送入標點與回收負荷控制是主軸"
        if has(cue, "遠投", "ロングキャスト", "90m"):
            detail += "，遠投後仍要維持泳層和刺魚角度"
        return f"{angle}：{lead} 優先，{detail}；{spec or '中重型線組'}下控魚餘量更高。"

    if parts and parts[0] in {"Frog", "Punching", "Heavy Texas"}:
        lead = join_terms(parts, 4)
        return f"重障礙打點：{lead} 優先，粗線近距離打入草洞、覆蓋物或菱草邊時，落點、鬆線和中魚後拉離障礙更直接。"

    if parts and parts[0] in {"Chatterbait", "Spinnerbait", "Buzzbait", "Swim Jig"}:
        lead = join_terms([term for term in parts if term in MOVING_TERMS], 4)
        soft = f"；放慢時可切 {join_terms(secondary_soft, 2)} 做底部觸感" if secondary_soft else ""
        return f"移動餌搜索：{lead} 優先，適合連續平收、變速和碰障反應，線弧與泳層維持更穩{soft}。"

    if parts and parts[0] in {"Crankbait", "Deep Crankbait", "Shallow Crankbait", "Shad", "Minnow", "Jerkbait", "Topwater Plug", "Pencil Bait", "Popper"}:
        lead = join_terms([term for term in parts if term in MOVING_TERMS], 4)
        action = "抽停、走狗與短促 rod work" if has(cue, "トゥイッチ", "ジャーク", "ドッグウォーク", "操作") else "平收、停頓與碰障回饋"
        soft = f"；兼顧 {join_terms(secondary_soft, 2)} 時以控鬆線和落點精度為主" if secondary_soft else ""
        return f"硬餌操作/搜索：{lead} 優先，{action} 更清楚，拋投後的線弧和咬口追隨性更穩{soft}。"

    if parts and parts[0] in {"Small Rubber Jig", "Jighead Rig", "Down Shot", "Neko Rig", "No Sinker", "Wacky Rig", "Power Finesse"}:
        lead = join_terms([term for term in parts if term in SOFT_TERMS], 4)
        if has(cue, "ソリッド", "solid"):
            tip = "實心竿尖"
        elif has(cue, "細線", "フィネス", "finesse"):
            tip = "細線/輕量釣組"
        else:
            tip = "輕量釣組"
        moving = f"；需要擴展時可切 {join_terms(secondary_moving, 2)} 做反應搜索" if secondary_moving else ""
        return f"精細軟餌：{lead} 優先，{tip}下讀底、控鬆線和微弱咬口判斷更清楚{moving}。"

    if parts and parts[0] in {"Texas Rig", "Light Texas", "Free Rig", "Leaderless Down Shot", "Rubber Jig", "Football Jig", "Guarded Rubber Jig", "Carolina Rig", "Heavy Carolina Rig"}:
        lead = join_terms([term for term in parts if term in SOFT_TERMS], 4)
        cover = "草區/障礙邊" if has(cue, "カバー", "ブッシュ", "ウィード", "grass", "cover") else "硬底、台階和結構邊"
        moving = f"；後段可接 {join_terms(secondary_moving, 2)} 做搜索" if secondary_moving else ""
        return f"底操軟餌：{lead} 優先，{cover}的觸底、穿越障礙和刺魚時機更好判斷{moving}。"

    lead = join_terms(parts, 4)
    return f"Bass 技法切換：{lead} 為核心，依 {spec or '當前 lure/line 範圍'} 在軟餌、硬餌與移動餌之間切換，保持落點、控線和咬口判斷。"


def describe_jigging(parts, row, desc):
    lead = join_terms(parts, 3)
    cue = lower(desc)
    spec = spec_clause(row)
    if "Long Fall Jig" in parts:
        return f"長落鐵板：{lead} 優先，深場或重鐵板下沉段的鬆線、落速和回抽節奏更容易管理；{spec or 'PE 線組'}下避免過度立竿搏魚。"
    if "Micro Jigging" in parts:
        return f"微鐵/輕鐵：{lead} 優先，50m 內小型 Metal Jig 的輕拋、快慢收切換和小口停頓更清楚。"
    if "Light Jigging" in parts:
        return f"輕量船釣鐵板：{lead} 優先，較輕 Metal Jig 的抽停、回收變速和中近海多魚種咬口判斷更穩。"
    if "Slow Pitch Jig" in parts:
        depth = "深場" if has(cue, "深場", "deep") else "垂直水層"
        return f"慢抽鐵板：{lead} 優先，{depth}中 Metal Jig 的跳動、橫擺和下沉咬口更好控制；{spec or '對應 PE/鐵板負荷'}下長時間操作負擔更低。"
    return f"船釣鐵板：{lead} 優先，垂直控線、抽停節奏和下沉段咬口判斷更直接，適合依水深與潮流調整 Metal Jig 重量。"


def describe_eging(parts, row, desc):
    lead = join_terms(parts, 4)
    spec = spec_clause(row)
    if "Deep Eging" in parts or "Tip-run Eging" in parts:
        extra = "，也能承接 Light Jigging / Tairaba 類副用途" if any(t in parts for t in ["Light Jigging", "Tairaba"]) else ""
        return f"深場木蝦：{lead} 優先，水深、潮流和船漂移造成的負荷下，竿尖視覺咬口與抽停後持線更清楚{extra}。"
    if "Slack-jerk Eging" in parts:
        return f"Slack Jerk 木蝦：{lead} 優先，鬆線抽竿、左右 dart 和停頓抱餌的線弧控制更穩；{spec or '2.5-4 號木蝦'} 範圍內連續操作負擔較低。"
    return f"木蝦泛用：{lead} 優先，抽竿、控線、停頓和抱餌判斷更清楚，依水深與風流在標準 Eging 和加鉛節奏間切換。"


def describe_seabass(parts, row, desc):
    lead = join_terms(parts, 4)
    cue = lower(desc)
    field = "港灣、河口和小中型河川" if has(cue, "港湾", "河口", "河川", "ベイエリア") else "岸投場景"
    if parts and parts[0] in {"Jighead Rig", "Light Game Jighead"}:
        return f"海鱸輕量軟餌：{lead} 優先，{field}的小型 Jighead、輕量 Plug 和 Vibration 操作更細，短咬與近岸結構反應更容易掌握。"
    if "Shore Plug" in parts or has(cue, "サーフ", "磯", "遠投"):
        return f"海鱸岸投遠投：{lead} 優先，{field}的中長距離覆蓋、迎風出線和泳層維持更穩，Metal Jig/Plug 可依風浪切換。"
    return f"海鱸搜索：{lead} 優先，{field}的小型 Plug、Jighead 和 Vibration 操作更細，短咬與近岸結構反應更容易掌握。"


def describe_light_salt(parts, row, desc):
    lead = join_terms(parts, 4)
    cue = lower(desc)
    if "Ajing Jighead" in parts:
        return f"Ajing 精細：{lead} 優先，1g 級 Jighead 的存在感、潮流變化和吸入式小口更容易讀取，適合細 PE/氟碳前導。"
    if "Mebaring Jighead" in parts:
        return f"Mebaring/輕海水：{lead} 優先，夜間小型 Plug、Jighead 和近距離結構邊控線更穩，尺級目標也有餘量。"
    if "Rockfish Rig" in parts:
        return f"Rockfish 底操：{lead} 優先，Jighead、Texas/Free Rig 貼底穿越障礙時更容易判斷觸底和咬口。"
    vertical = "縱向操作" if has(cue, "縦", "スパイラル", "ソリッド") else "輕量小餌"
    return f"輕海水精細：{lead} 優先，{vertical}下小型 Jighead、Metal Jig 或 Small Plug 的出線、潮感和細微咬口更清楚。"


def describe_trout(parts, row, desc):
    lead = join_terms(parts, 4)
    cue = lower(desc)
    if has(cue, "渓流", "源流", "山"):
        return f"溪流鱒魚：{lead} 優先，短距離低彈道拋投、落點修正和 Minnow/Twitch 操作更精準，適合狹窄標點快速反應。"
    if any(t in parts for t in ["Topwater Plug", "Minnow", "Vibration"]):
        return f"Area Trout 操作餌：{lead} 優先，Topwater、Minnow 和 Vibration 的抽停、dart、lift 節奏更清楚，遠距也能確實掛口。"
    return f"Area Trout 精細：{lead} 優先，Spoon、Crank 和小型 Minnow 的慢速控層、細線出線和輕咬判斷更穩。"


def build_hint(row):
    parts = split_pairing(row.get(PAIR_FIELD))
    env = n(row.get("player_environment"))
    desc = n(row.get("Description"))
    if any(term in parts for term in ["Deep Eging", "Tip-run Eging", "Slack-jerk Eging", "Eging"]):
        return describe_eging(parts, row, desc)
    if any(term in parts for term in ["Slow Pitch Jig", "Long Fall Jig", "Offshore Jigging", "Light Jigging", "Micro Jigging"]):
        return describe_jigging(parts, row, desc)
    if "海鲈" in env or "海鱸" in env or any(term in parts for term in ["Seabass Minnow", "Sinking Pencil", "Seabass Plug", "Shore Plug"]):
        return describe_seabass(parts, row, desc)
    if "轻型海水" in env or "輕型海水" in env or any(term in parts for term in ["Ajing Jighead", "Mebaring Jighead", "Light Game Jighead", "Rockfish Rig"]):
        return describe_light_salt(parts, row, desc)
    if "鳟" in env or "鱒" in env or has_any_term(parts, TROUT_TERMS):
        return describe_trout(parts, row, desc)
    return describe_bass(parts, row, desc)


def has_any_term(parts, terms):
    return any(part in terms for part in parts)


def snapshot_without_targets(ws, target_cols):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col not in target_cols
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    required = [
        "id", "SKU", "Code Name", "POWER", "LURE WEIGHT", "Line Wt N F", "PE Line Size",
        "player_environment", "Description", PAIR_FIELD, GUIDE_FIELD,
    ]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    before = snapshot_without_targets(ws, {col[PAIR_FIELD], col[GUIDE_FIELD]})
    changes = []
    final_rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = {field: ws.cell(row_idx, col[field]).value for field in required}
        new_hint = build_hint(row)
        old_hint = n(row[GUIDE_FIELD])
        if old_hint != new_hint:
            cell = ws.cell(row_idx, col[GUIDE_FIELD])
            cell.value = new_hint
            cell.fill = copy(YELLOW_FILL)
            changes.append({
                "row": row_idx,
                "id": n(row["id"]),
                "sku": n(row["SKU"]),
                "recommended_rig_pairing": n(row[PAIR_FIELD]),
                "old_guide_use_hint": old_hint,
                "new_guide_use_hint": new_hint,
            })
        final_rows.append({
            "id": n(row["id"]),
            "sku": n(row["SKU"]),
            "recommended_rig_pairing": n(row[PAIR_FIELD]),
            "guide_use_hint": new_hint,
        })

    after = snapshot_without_targets(ws, {col[PAIR_FIELD], col[GUIDE_FIELD]})
    if before != after:
        raise RuntimeError("unexpected changes outside guide_use_hint/recommended_rig_pairing")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_file": str(XLSX_PATH),
        "changed_guide_use_hint_rows": len(changes),
        "changed_recommended_rig_pairing_rows": 0,
        "policy": "Guide hints are rebuilt from recommended_rig_pairing, model Description cues, usage environment, lure weight, and line rating. No Description or unrelated fields are changed.",
        "changes": changes,
        "final_rows": final_rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "changed_guide_use_hint_rows": len(changes),
        "changed_recommended_rig_pairing_rows": 0,
        "report_file": str(REPORT_PATH),
        "samples": changes[:8],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
