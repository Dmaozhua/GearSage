from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')
TARGET_ID = "DRD10120"
UPDATES = {
    "recommended_rig_pairing": "Neko Rig / No Sinker / Down Shot / Texas Rig / Leaderless Down Shot / Free Rig / Rubber Jig / Guarded Jighead / Spinnerbait / Vibration / Crankbait / Craw",
    "guide_use_hint": "FIRE WOLF 全能底操：Neko、No Sinker、Down Shot、Texas 和 Free Rig 為主，實心竿尖也能兼顧 Guarded Jighead、Spinnerbait、Vibration、Crankbait 的鬆線回收與觸感。",
}


def n(value):
    return str(value or "").strip()


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    for field in ["id", *UPDATES]:
        if field not in col:
            raise RuntimeError(f"missing rod_detail column: {field}")

    changed = []
    for row in range(2, ws.max_row + 1):
        if n(ws.cell(row, col["id"]).value) != TARGET_ID:
            continue
        for field, new_value in UPDATES.items():
            cell = ws.cell(row, col[field])
            old = n(cell.value)
            if old != new_value:
                cell.value = new_value
                changed.append({"id": TARGET_ID, "field": field, "old": old, "new": new_value})
        break
    else:
        raise RuntimeError(f"target detail row not found: {TARGET_ID}")

    wb.save(XLSX_PATH)
    print({"file": str(XLSX_PATH), "changed": changed})


if __name__ == "__main__":
    main()
