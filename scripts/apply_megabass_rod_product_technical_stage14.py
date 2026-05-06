#!/usr/bin/env python3
"""Fill Megabass rod_detail.product_technical from official Megabass technology blocks only."""

from __future__ import annotations

import copy
import hashlib
import json
import re
import subprocess
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from bs4 import BeautifulSoup
from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
RAW_PATH = DATA_DIR / "megabass_rod_raw.json"
XLSX_PATH = DATA_DIR / "megabass_rod_import.xlsx"
CACHE_DIR = DATA_DIR / "megabass_rod_product_technical_cache"
EVIDENCE_PATH = DATA_DIR / "megabass_rod_product_technical_official_evidence.json"
REPORT_PATH = DATA_DIR / "megabass_rod_product_technical_stage14_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_megabass_rod_detail_groups_stage2.py"

FIELD = "product_technical"
INSERT_AFTER = "Description"
DELIMITER = " / "

EXCLUDE_EXACT = {
    "",
    "BENDING CURVE",
    "Cloth bag",
    "専用パッケージ",
    "All models feature a three-piece design",
    "All models feature a four-piece design",
    "Natural bird’s eye peacock carved reel seat.",
    "Ergonomic carved metal short trigger.",
    "CORK GRIP",
    "Grip End",
    "Reel Sheet",
    "Reel Seat",
    "Maple Wood Grip",
    "Teak Wood Grip",
}

EXCLUDE_RE = re.compile(r"^(GH\d|GHBF\d).+ Reel Seat$|^GUIDE$|^GUIDES$", re.I)
INCLUDE_RE = re.compile(
    r"(SYSTEM|GRAPHITE|BLANK|BLANKS|SHAFT|GUIDE|GUIDES|SiC|SIC|TORZITE|TITANIUM|"
    r"FUJI|SEAT|HOOD|GRIP|JOINT|FERRULE|BALANCER|HANDLE|BRIDGE|CONCEPT|CARBON|FIBER|"
    r"EVA|YOLOY|SLX|HUNTSMAN|VALKYRIE|NANO|T3|IBCS|MBCS|IAS|IES|ITS|IFPS|T-DPS|"
    r"BACK STOP|HEAD LOCKING)",
    re.I,
)

CANONICAL_TERMS = {
    "MEGABASS BRIDGE CONSTRUCTION SEAT": "MBCS (MEGABASS BRIDGE CONSTRUCTION SEAT)(D.PAT)",
    "MBCS（PAT.）": "MBCS (MEGABASS BRIDGE CONSTRUCTION SEAT)(D.PAT)",
    "IES : ITO ERGONOMIC-CONTACT SEAT（D.PAT.）": "IES(ITO ERGONOMIC-CONTACT SEAT)（D.PAT.P）",
    "IAS : ITO ENGINEERING AIRY-FIT ERGONOMIC SEAT（D.PAT.P）": "IAS(ITO ENGINEERING AIRY-FIT ERGONOMIC SEAT)（D.PAT.P）",
    "Sic Guides": "SiC Guides",
    "Stainless SIC Guide": "Stainless SiC Guide",
}


def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def merge_terms(terms) -> str:
    merged = []
    for term in terms:
        value = norm(term)
        if value and value not in merged:
            merged.append(value)
    return DELIMITER.join(merged)


def headers(ws):
    return [cell.value for cell in ws[1]]


def header_index(ws, field):
    current = headers(ws)
    return current.index(field) + 1 if field in current else None


def copy_style(source, target):
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)


def sheet_values(ws, ignore_field=None):
    current = headers(ws)
    keep_cols = [
        idx for idx, header in enumerate(current, start=1)
        if header != ignore_field
    ]
    return [
        tuple(ws.cell(row=row, column=col).value for col in keep_cols)
        for row in range(1, ws.max_row + 1)
    ]


def ensure_detail_column(ws):
    current_col = header_index(ws, FIELD)
    anchor_col = header_index(ws, INSERT_AFTER)
    if not anchor_col:
        raise RuntimeError(f"rod_detail missing {INSERT_AFTER}")
    desired_col = anchor_col + 1
    if current_col == desired_col:
        return current_col, False
    if current_col:
        values = [ws.cell(row=row, column=current_col).value for row in range(1, ws.max_row + 1)]
        ws.delete_cols(current_col)
        anchor_col = header_index(ws, INSERT_AFTER)
        desired_col = anchor_col + 1
        ws.insert_cols(desired_col)
        for row, value in enumerate(values, start=1):
            copy_style(ws.cell(row=row, column=anchor_col), ws.cell(row=row, column=desired_col))
            ws.cell(row=row, column=desired_col).value = value
        ws.cell(row=1, column=desired_col).value = FIELD
        return desired_col, True
    ws.insert_cols(desired_col)
    for row in range(1, ws.max_row + 1):
        copy_style(ws.cell(row=row, column=anchor_col), ws.cell(row=row, column=desired_col))
        ws.cell(row=row, column=desired_col).value = FIELD if row == 1 else ""
    return desired_col, True


def cache_path(url: str) -> Path:
    return CACHE_DIR / f"{hashlib.sha1(url.encode('utf-8')).hexdigest()}.html"


def canonical_title(block) -> str:
    title = block.select_one(".page_products__technology__block__title")
    if not title:
        return ""
    en = title.select_one(".__en")
    ja = title.select_one(".__ja")
    raw = (
        norm(en.get_text(" ", strip=True) if en else "")
        or norm(ja.get_text(" ", strip=True) if ja else "")
        or norm(title.get_text(" ", strip=True))
    )
    raw = raw.replace("＆", "&")
    return CANONICAL_TERMS.get(raw, raw)


def should_include(term: str) -> bool:
    if term in EXCLUDE_EXACT:
        return False
    if EXCLUDE_RE.search(term):
        return False
    if term.startswith("All models feature"):
        return False
    return bool(INCLUDE_RE.search(term))


def content_terms(term: str, content: str):
    text = norm(content)
    terms = []
    if term in {"GUIDE", "GUIDES"}:
        if "ITO WOVEN" in text and ("TORZITE" in text or "TRZITE" in text) and "TITANIUM" in text:
            terms.extend([
                "ITO WOVEN GRAPHITE GUIDE STAGE",
                "ITO ARTIFICIAL CUSTOM THREAD WRAPPING",
                "Fuji TORZITE GUIDE RING＋TITANIUM FRAME ARMS SETTING",
            ])
        if "all double footed and double wrapped" in text.lower():
            terms.append("All double footed and double wrapped guide system")
    if not term and "OROCHI X10" in text and "guide threading" in text.lower():
        terms.append("OROCHI X10 ORIGINAL TIGHT THREADING FUJI STAINLESS FRAME SiC-S GUIDE")
    return terms


def summary_terms(summary: str):
    text = norm(summary)
    lower = text.lower()
    terms = []
    if "megabass original solid tip" in lower:
        terms.append("Megabass original solid tip")
    elif "solid tip" in lower or "solid-tip" in lower:
        terms.append("Solid Tip")
    if "stinger tip technology" in lower:
        terms.append("STINGER TIP technology")
    if "high frame k-guide stripper guide" in lower:
        terms.append("High frame K-Guide stripper guide")
    if "double-footed guides" in lower and "double-wrapped" in lower:
        terms.append("Double-footed guides and double-wrapped guide setting")
    if "#8 guides" in lower:
        terms.append("Large diameter #8 guides")
    if "#6 guides" in lower:
        terms.append("Small diameter #6 guides")
    if "high-precision spigot ferrule joint system" in lower:
        terms.append("High-precision Spigot Ferrule Joint System")
    return terms


def parse_official_terms(item):
    url = item.get("url") or ""
    path = cache_path(url)
    if not url or not path.exists():
        return {
            "url": url,
            "terms": [],
            "raw_titles": [],
            "excluded_titles": [],
            "error": "missing cached official page",
        }

    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
    terms = []
    raw_titles = []
    excluded_titles = []
    summary_text = norm(" ".join(
        node.get_text(" ", strip=True)
        for node in soup.select(".page_products__summary .__en, .page_products__summary .__ja")
    ))
    for block in soup.select(".page_products__technology__block"):
        term = canonical_title(block)
        content_node = block.select_one(".page_products__technology__block__content")
        content = norm(content_node.get_text(" ", strip=True) if content_node else "")
        raw_titles.append(term)
        included_by_title = should_include(term)
        if included_by_title:
            if term not in terms:
                terms.append(term)
        for content_term in content_terms(term, content):
            if content_term not in terms:
                terms.append(content_term)
        if not included_by_title and term:
            excluded_titles.append(term)
    description_terms = summary_terms(summary_text)
    for description_term in description_terms:
        if (
            "double-footed guides" in description_term.lower()
            and any("double footed" in term.lower() for term in terms)
        ):
            continue
        if description_term not in terms:
            terms.append(description_term)

    return {
        "url": url,
        "terms": terms,
        "description_terms": description_terms,
        "raw_titles": raw_titles,
        "excluded_titles": excluded_titles,
        "has_official_technology_section": bool(raw_titles),
    }


def is_imported_rod_item(item):
    specs = item.get("specs") or {}
    return bool(specs and (specs.get("Length") or specs.get("継数")))


def stripped_raw(data):
    clone = copy.deepcopy(data)
    for item in clone:
        if isinstance(item, dict):
            item.pop(FIELD, None)
    return clone


def update_raw(data, values_by_key):
    before = stripped_raw(data)
    nonempty = 0
    for item in data:
        if not isinstance(item, dict):
            continue
        key = (item.get("series_name"), item.get("model_name"))
        if key in values_by_key:
            value = values_by_key[key]
            item[FIELD] = value
            if value:
                nonempty += 1
        else:
            item.pop(FIELD, None)
    if stripped_raw(data) != before:
        raise RuntimeError("Unexpected raw changes outside product_technical")
    RAW_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {
        "items": len(data),
        "rod_items": len(values_by_key),
        "item_product_technical_nonempty": nonempty,
    }


def build_values(data):
    values = {}
    evidence = {}
    for item in data:
        if not is_imported_rod_item(item):
            continue
        key = (item.get("series_name"), item.get("model_name"))
        parsed = parse_official_terms(item)
        value = merge_terms(parsed["terms"])
        values[key] = value
        evidence[f"{key[0]}::{key[1]}"] = {
            "series_name": key[0],
            "sku": key[1],
            "url": parsed.get("url"),
            "product_technical": value,
            "included_terms": parsed.get("terms", []),
            "description_terms": parsed.get("description_terms", []),
            "raw_titles": parsed.get("raw_titles", []),
            "excluded_titles": parsed.get("excluded_titles", []),
            "has_official_technology_section": parsed.get("has_official_technology_section", False),
            "source_rule": "Megabass official product TECHNOLOGY blocks only",
        }
        if parsed.get("error"):
            evidence[f"{key[0]}::{key[1]}"]["error"] = parsed["error"]
    return values, evidence


def update_workbook(values_by_key):
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_before = sheet_values(rod_ws, FIELD)
    detail_before = sheet_values(detail_ws, FIELD)

    if header_index(rod_ws, FIELD):
        raise RuntimeError("rod sheet must not contain product_technical")

    detail_col, moved_or_inserted = ensure_detail_column(detail_ws)
    rod_headers = headers(rod_ws)
    detail_headers = headers(detail_ws)
    rod_id_col = rod_headers.index("id") + 1
    model_col = rod_headers.index("model") + 1
    detail_rod_id_col = detail_headers.index("rod_id") + 1
    sku_col = detail_headers.index("SKU") + 1
    rod_model_by_id = {
        rod_ws.cell(row=row, column=rod_id_col).value: rod_ws.cell(row=row, column=model_col).value
        for row in range(2, rod_ws.max_row + 1)
    }

    nonempty = 0
    changed = []
    unmatched = []
    for row in range(2, detail_ws.max_row + 1):
        model = rod_model_by_id.get(detail_ws.cell(row=row, column=detail_rod_id_col).value)
        sku = detail_ws.cell(row=row, column=sku_col).value
        key = (model, sku)
        if key not in values_by_key:
            unmatched.append({"row": row, "model": model, "sku": sku})
            value = ""
        else:
            value = values_by_key[key]
        old = norm(detail_ws.cell(row=row, column=detail_col).value)
        detail_ws.cell(row=row, column=detail_col).value = value
        if value:
            nonempty += 1
        if old != value:
            changed.append({"row": row, "model": model, "sku": sku, "old": old, "new": value})

    if sheet_values(rod_ws, FIELD) != rod_before:
        raise RuntimeError("Unexpected rod sheet changes")
    if sheet_values(detail_ws, FIELD) != detail_before:
        raise RuntimeError("Unexpected rod_detail changes outside product_technical")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)
    return {
        "rod_detail_rows": detail_ws.max_row - 1,
        "rod_detail_product_technical_nonempty": nonempty,
        "rod_detail_column_inserted_or_moved": moved_or_inserted,
        "changed_cells": len(changed),
        "changed_samples": changed[:20],
        "unmatched_rows": unmatched,
    }


def validate_values(values_by_key):
    errors = []
    for (model, sku), value in values_by_key.items():
        if "/" in value and DELIMITER not in value:
            errors.append({"model": model, "sku": sku, "error": "slash delimiter must be ` / `", "value": value})
        for fragment in ["白名单", "玩家", "钓组", "饵型", "来源"]:
            if fragment in value:
                errors.append({"model": model, "sku": sku, "error": f"forbidden fragment {fragment}", "value": value})
    return errors


def main():
    data = json.loads(RAW_PATH.read_text(encoding="utf-8"))
    values_by_key, evidence = build_values(data)
    validation_errors = validate_values(values_by_key)
    if validation_errors:
        raise RuntimeError(f"product_technical validation failed: {validation_errors[:5]}")

    raw_report = update_raw(data, values_by_key)
    xlsx_report = update_workbook(values_by_key)
    no_section = [
        key for key, item in evidence.items()
        if not item.get("has_official_technology_section")
    ]
    empty_values = [
        key for key, item in evidence.items()
        if not norm(item.get("product_technical"))
    ]
    report = {
        "field": FIELD,
        "scope": "Megabass rod_detail SKU-level product_technical only",
        "source_policy": "Official Megabass product TECHNOLOGY blocks and explicit SKU description technology phrases only; whitelist/player/spec inference is not used.",
        "delimiter": DELIMITER,
        "raw": raw_report,
        "xlsx": xlsx_report,
        "evidence": str(EVIDENCE_PATH),
        "official_pages_cached": len(list(CACHE_DIR.glob("*.html"))),
        "no_official_technology_section": no_section,
        "empty_product_technical": empty_values,
    }
    EVIDENCE_PATH.write_text(json.dumps(evidence, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if xlsx_report["unmatched_rows"]:
        raise RuntimeError("unmatched workbook rows; see report")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
