#!/usr/bin/env python3
import copy
import json
import re
import subprocess
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "ark_rod_normalized.json"
XLSX_PATH = DATA_DIR / "ark_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "ark_rod_product_technical_report.json"
EVIDENCE_PATH = DATA_DIR / "ark_rod_product_technical_official_evidence.json"
SHADE_SCRIPT = ROOT / "scripts/shade_ark_rod_detail_groups.py"

FIELD = "product_technical"
INSERT_AFTER = "Description"


TECH_PATTERNS = [
    ("ARK HPCR (High Pressure Carbon Fiber Rolling) technology", r"\b(?:ARK\s+)?HPCR\b|\bHigh[- ]Pressure Carbon[- ]Fiber Rolling Technology\b|\bARK high pressure carbon[- ]fiber rolling technology\b"),
    ("Multi-Direction Multi-Layer technology", r"\bMulti[- ]Direction Multi[- ]Layer technology\b|\bMDML Technology\b"),
    ("carbon nano tube reinforcement", r"\bcarbon nano tubes?\b|\bnano tube reinforcements?\b|\bNano tube reinforcement\b"),
    ("Fuji PTS/TVS reel seat", r"\bFuji PTS/TVS reel seat\b"),
    ("Fuji K concept Alconite guides", r"\bFuji K concept alconite guides\b"),
    ("Fuji K concept guides", r"\bFuji K[- ]concept (?:tangle[- ]free )?guides\b|\bFuji K concept guides\b"),
    ("FazLite rings", r"\bFazL?ite (?:rings|inserts)\b"),
    ("ARK Titanium Tangle-Free Guides", r"\bARK Titanium Tangle[- ]Free Guides\b"),
    ("ARK Stainless Steel Tangle-Free Guides", r"\bARK Stainless Steel Tangle[- ]Free Guides\b"),
    ("ARK Tangle-Free Guides", r"\bARK Tangle[- ]?Free Guides\b"),
    ("NanoForce Rings", r"\bNanoForce(?:™)? (?:Ring|Rings|guides)\b"),
    ("A-Ring technology", r"\bA[- ]Ring technology\b"),
    ("Black Coated Stainless Micro Guides System With Zirconium Inserts", r"\bBlack Coated Stainless Micro Guides System With Zirconium Inserts\b"),
    ("Team ARK reel seat", r"\bTeam ARK reel seat\b"),
    ("High-Visibility Strike Indicator", r"\bHigh[- ]Visibility Strike Indicator\b"),
]


def norm(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def remove_product_technical(obj):
    if isinstance(obj, dict):
        return {k: remove_product_technical(v) for k, v in obj.items() if k != FIELD}
    if isinstance(obj, list):
        return [remove_product_technical(v) for v in obj]
    return obj


def headers(ws):
    return [cell.value for cell in ws[1]]


def col_map(ws):
    return {name: idx + 1 for idx, name in enumerate(headers(ws))}


def sheet_values(ws, ignore_product_technical=False):
    h = headers(ws)
    ignored = {FIELD} if ignore_product_technical else set()
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([value for idx, value in enumerate(row) if h[idx] not in ignored])
    return rows


def ensure_detail_column(ws):
    h = headers(ws)
    if FIELD in h:
        field_col = h.index(FIELD) + 1
        desired_col = h.index(INSERT_AFTER) + 2
        if field_col == desired_col:
            return field_col, False
        values = [ws.cell(row=row_idx, column=field_col).value for row_idx in range(1, ws.max_row + 1)]
        ws.delete_cols(field_col)
        h = headers(ws)
        desired_col = h.index(INSERT_AFTER) + 2
        ws.insert_cols(desired_col)
        for row_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=desired_col).value = value
        ws.cell(row=1, column=desired_col).value = FIELD
        return desired_col, True
    anchor_col = h.index(INSERT_AFTER) + 1
    insert_col = anchor_col + 1
    ws.insert_cols(insert_col)
    ws.cell(row=1, column=insert_col).value = FIELD
    return insert_col, True


def terms_from_official_description(description):
    text = re.sub(
        r"From 2026 onward, new batches of rods will be upgraded from Fuji guides to ARK Tangle Free Guides with NanoForce™ Rings\.?",
        "",
        norm(description),
        flags=re.I,
    )
    terms = []
    evidence = []
    for term, pattern in TECH_PATTERNS:
        match = re.search(pattern, text, flags=re.I)
        if match and term not in terms:
            terms.append(term)
            start = max(match.start() - 70, 0)
            end = min(match.end() + 100, len(text))
            evidence.append(
                {
                    "term": term,
                    "matched_text": match.group(0),
                    "official_excerpt": text[start:end],
                }
            )
    return " / ".join(terms), evidence


def validate_value(value, row_idx, sku):
    text = norm(value)
    errors = []
    if not text:
        return errors
    if "/" in text and " / " not in text:
        errors.append({"row": row_idx, "sku": sku, "reason": "slash delimiter must be ` / `"})
    for bad in ["白名单", "玩家", "钓组", "饵型", "recommended", "guide_use_hint"]:
        if bad in text:
            errors.append({"row": row_idx, "sku": sku, "reason": f"forbidden fragment `{bad}`"})
    return errors


def main():
    normalized = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    normalized_before_without_field = remove_product_technical(copy.deepcopy(normalized))

    evidence_by_rod_id = {}
    value_by_rod_sku = {}
    variant_updates = []

    for index, item in enumerate(normalized):
        rod_id = f"ARKR{1000 + index}"
        value, evidence = terms_from_official_description(item.get("description", ""))
        evidence_by_rod_id[rod_id] = {
            "rod_id": rod_id,
            "title": item.get("title"),
            "source_url": item.get("source_url"),
            "product_technical": value,
            "official_evidence": evidence,
        }
        for spec in item.get("specs", []):
            sku = norm(spec.get("sku"))
            old = norm(spec.get(FIELD))
            spec[FIELD] = value
            value_by_rod_sku[(rod_id, sku)] = value
            if old != value:
                variant_updates.append({"rod_id": rod_id, "sku": sku, "old": old, "new": value})

    if remove_product_technical(normalized) != normalized_before_without_field:
        raise RuntimeError("normalized guard failed: fields outside product_technical changed")

    NORMALIZED_PATH.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = load_workbook(XLSX_PATH)
    if "rod" not in wb.sheetnames or "rod_detail" not in wb.sheetnames:
        raise RuntimeError("ark import workbook missing rod or rod_detail sheet")
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]

    if FIELD in headers(rod_ws):
        raise RuntimeError("rod master must not contain product_technical")

    rod_before = sheet_values(rod_ws)
    detail_before_without_field = sheet_values(detail_ws, ignore_product_technical=True)
    field_col, column_inserted_or_moved = ensure_detail_column(detail_ws)
    c = col_map(detail_ws)

    workbook_updates = []
    errors = []
    for row_idx in range(2, detail_ws.max_row + 1):
        rod_id = norm(detail_ws.cell(row=row_idx, column=c["rod_id"]).value)
        sku = norm(detail_ws.cell(row=row_idx, column=c["SKU"]).value)
        value = value_by_rod_sku.get((rod_id, sku), "")
        old = norm(detail_ws.cell(row=row_idx, column=field_col).value)
        if old != value:
            detail_ws.cell(row=row_idx, column=field_col).value = value
            workbook_updates.append({"row": row_idx, "rod_id": rod_id, "sku": sku, "old": old, "new": value})
        errors.extend(validate_value(value, row_idx, sku))

    if sheet_values(rod_ws) != rod_before:
        raise RuntimeError("xlsx guard failed: rod sheet changed")
    if sheet_values(detail_ws, ignore_product_technical=True) != detail_before_without_field:
        raise RuntimeError("xlsx guard failed: rod_detail fields outside product_technical changed")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    wb_check = load_workbook(XLSX_PATH, data_only=True)
    detail_check = wb_check["rod_detail"]
    hc = col_map(detail_check)
    values = []
    for row_idx in range(2, detail_check.max_row + 1):
        values.append(norm(detail_check.cell(row=row_idx, column=hc[FIELD]).value))

    evidence = {
        "brand": "Ark",
        "brand_id": 115,
        "field": FIELD,
        "source_policy": "official Ark product pages only; whitelist/player/spec inference not used",
        "delimiter": " / ",
        "series": list(evidence_by_rod_id.values()),
    }
    report = {
        "field": FIELD,
        "scope": "rod_detail SKU level only; rod master must not contain product_technical",
        "updated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source_policy": "official Ark product pages only; values are extracted from official product descriptions cached from arkrods.com",
        "only_changed_field": FIELD,
        "normalized": {
            "items": len(normalized),
            "specs": sum(len(item.get("specs", [])) for item in normalized),
            "spec_product_technical_nonempty": sum(
                1 for item in normalized for spec in item.get("specs", []) if norm(spec.get(FIELD))
            ),
            "item_product_technical_count": sum(1 for item in normalized if FIELD in item),
            "variant_updates": len(variant_updates),
        },
        "xlsx": {
            "rod_rows": rod_ws.max_row - 1,
            "rod_has_product_technical": FIELD in headers(rod_ws),
            "rod_detail_rows": len(values),
            "rod_detail_product_technical_nonempty": sum(1 for value in values if value),
            "rod_detail_product_technical_unique": len({value for value in values if value}),
            "column_inserted_or_moved": column_inserted_or_moved,
            "column_position": "after Description, before Extra Spec 1",
            "distribution": dict(Counter(values)),
            "workbook_updates": len(workbook_updates),
        },
        "errors": errors,
    }
    EVIDENCE_PATH.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if errors:
        raise RuntimeError(f"validation failed: {REPORT_PATH}")
    print(json.dumps(report["xlsx"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
