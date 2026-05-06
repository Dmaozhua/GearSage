#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import urllib.request
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from lxml import html
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "keitech_rod_normalized.json"
XLSX_PATH = DATA_DIR / "keitech_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "keitech_rod_completion_stage1_report.json"
SCHEMA_PATH = ROOT / "scripts/gear_export_schema.js"


SOURCE_URLS = [
    "https://keitech.co.jp/pages/628/",
    "https://keitech.co.jp/pages/629/",
    "https://keitech.co.jp/pages/600/",
    "https://keitech.co.jp/pages/589/",
]

MODEL_CN = {
    "NF Series 66 Casting Model": "NF 66 系列枪柄卷物精细竿",
    "NF Series 76 Flippin' Stick®": "NF 76 Flippin' Stick 翻打竿",
    "F-SPEC Series": "F-SPEC 76 精细枪柄系列",
    "76 Series": "76 系列长竿",
}

MODEL_ALIAS = {
    "NF Series 66 Casting Model": "NF 66 Casting / 巻き物フィネス",
    "NF Series 76 Flippin' Stick®": "ザ・フリッピンスティック",
    "F-SPEC Series": "フィネススペック",
    "76 Series": "ナナロク",
}

MASTER_FIELDS = {
    "NF Series 66 Casting Model": {
        "model_year": "2026",
        "series_positioning": "Fresh Water / Bass / 小型卷物与单钩移动饵精细枪柄系列",
        "main_selling_points": "North Fork SM 碳布设定，围绕小型 crank、shad、spinnerbait、swim jig 等巻物精细化使用，强调抛投精度、咬口黏着和单钩刺鱼余量。",
        "player_positioning": "Bass finesse moving bait casting 系列",
        "player_selling_points": "从 1/16oz 级小型卷物到 3/4oz 级移动饵和轻底操都有分工，适合岸钓用更细的移动饵反复打角度、控泳层和节奏。",
    },
    "NF Series 76 Flippin' Stick®": {
        "series_positioning": "Fresh Water / Bass / Heavy cover flipping 专用长竿",
        "main_selling_points": "North Fork HM 碳布，7'6'' XH moderate 设定，分为 triggerless 直线导环和 trigger spiral guide 两个版本，面向浅场 cover 翻打。",
        "player_positioning": "Heavy cover flipping / 近距离强力控鱼",
        "player_selling_points": "长竿和高弹 HM blank 带来抬线、送饵和起鱼余量，适合围绕芦苇、灌木、硬 cover 做高精度近距离翻打。",
    },
    "F-SPEC Series": {
        "series_positioning": "Fresh Water / Bass / 76 长竿 finesse cover 系列",
        "main_selling_points": "以 solid tip 与 spiral guide 把 76 长竿延伸到轻量 rig，覆盖 down shot、small rubber jig、neko、light Texas 和轻 cover jig。",
        "player_positioning": "Long bait finesse / cover finesse",
        "player_selling_points": "7'6'' 长度保留远距控线和 cover 起鱼优势，solid tip 让轻 rig 的 fall bite、细抖和食わせ更容易掌握。",
    },
    "76 Series": {
        "series_positioning": "Fresh Water / Bass / 7'6'' worm & jig 长竿系列",
        "main_selling_points": "1pc bait casting 7'6''，ML 到 XH 覆盖 worm、rubber jig、Texas、free rig 等长竿底操和 cover 攻略，提供 regular 与 spiral guide 两种布局。",
        "player_positioning": "Long rod worm / jig standard",
        "player_selling_points": "长竿便于岸钓远距控线、提线和过障，regular 与 spiral guide 可按抛投手感或高负荷抗扭需求选择。",
    },
}

RIGS = {
    "KTC662NF": "Jig Spinner / Small Shad / Tiny Crankbait / Pencil Bait / Flat Side Crankbait",
    "KTC663NF": "Jighead Swimming / Small Shad / Small Crankbait / Flat Side Crankbait",
    "KTC664NF": "Spinnerbait / Small Minnow / High Density Stickbait / Popper",
    "KTC665NF": "Spinnerbait / Swim Jig / Swimbait Jighead / Mid Crankbait",
    "KTC666NF": "Bladed Jig / Spinnerbait / Texas Rig / Free Rig / Cover Worm",
    "KTC768NFL": "Flipping / Heavy Texas Rig / Rubber Jig / Cover Worm / Punching",
    "KTC768NFL-SPG": "Flipping / Heavy Texas Rig / Rubber Jig / Cover Worm / Punching",
    "KTC762F-SPG": "Down Shot / Small Rubber Jig / Nail Sinker Rig / Small Shad",
    "KTC763F-SPG": "Down Shot / Small Rubber Jig / Neko Rig / Bladed Jig",
    "KTC764F-SPG": "Compact Cover Jig / Texas Rig / Free Rig / Leaderless Down Shot / Backslide No Sinker / Spinnerbait",
    "KTC766F-SPG": "Texas Rig / Free Rig / Football Jig / Shallow Crankbait / Spinnerbait",
    "KTC764": "Light Texas Rig / Free Rig / Rubber Jig / Neko Rig",
    "KTC765": "Texas Rig / Free Rig / Rubber Jig / Leaderless Down Shot",
    "KTC766": "Rubber Jig / Texas Rig / Free Rig / Leaderless Down Shot",
    "KTC767": "Heavy Texas Rig / Rubber Jig / Free Rig / Cover Jig",
    "KTC768": "Heavy Texas Rig / Rubber Jig / Punching / Flipping",
    "KTC764-SPG": "Light Texas Rig / Free Rig / Rubber Jig / Neko Rig",
    "KTC765-SPG": "Texas Rig / Free Rig / Rubber Jig / Leaderless Down Shot",
    "KTC766-SPG": "Rubber Jig / Texas Rig / Free Rig / Leaderless Down Shot",
    "KTC767-SPG": "Heavy Texas Rig / Rubber Jig / Free Rig / Cover Jig",
    "KTC768-SPG": "Heavy Texas Rig / Rubber Jig / Punching / Flipping",
}

DETAIL_PLAYER = {
    "KTC662NF": ("淡水 Bass / 野池・河川・湖岸 / 超轻小型卷物精细搜索", "小型卷物 finesse 入门端", "1/16-1/4oz 区间更适合小型 shad、tiny crank 和 jig spinner，低活性场景可以用反复角度和轻线控泳层逼咬。"),
    "KTC663NF": ("淡水 Bass / 开放水域与轻 cover 边 / 食わせ型小卷物", "小型 shad / jighead swimming 精细卷物", "比 662 多一点载饵和抛投余量，适合 5g 级 shad、小 crank 和轻 jighead swimming 做连续搜索。"),
    "KTC664NF": ("淡水 Bass / 岸边硬物・浅场 cover 外缘 / 单钩移动饵", "1/4oz spinnerbait 与小型硬饵泛用", "ML power 适合 baby spinnerbait、小型 minnow、popper 和高比重软饵，软硬饵切换宽容度比前两支更高。"),
    "KTC665NF": ("淡水 Bass / 岸钓中近距离 / 3/8oz 单钩移动饵", "spinnerbait / swim jig 主力型号", "3/8oz spinnerbait、swim jig 和 4-5in swimbait jighead 是核心，兼顾中型 crank，适合移动饵主线的日常岸钓。"),
    "KTC666NF": ("淡水 Bass / 岸边 cover 外缘・开阔水域 / 重一点移动饵到轻底操", "NF66 系列最泛用重端", "能覆盖 1/2oz 级 bladed jig、spinnerbait，也能切到 7-10g Texas / Free Rig，是 NF66 中最适合一支竿走动的型号。"),
    "KTC768NFL": ("淡水 Bass / 芦苇・灌木・浅场重 cover / 近距离 flipping", "triggerless flipping 专用", "没有 trigger 的握持自由度更高，适合传统 flipping 中频繁调整握位、短距离送饵和直接把鱼带离 cover。"),
    "KTC768NFL-SPG": ("淡水 Bass / 芦苇・灌木・浅场重 cover / 高负荷 flipping", "trigger + spiral guide flipping", "trigger 握持和 spiral guide 抗扭更适合重 cover 下持续压鱼，Heavy Texas、Rubber Jig 和 Punching 场景更稳定。"),
    "KTC762F-SPG": ("淡水 Bass / 消波块・石积・轻 cover / 长竿 bait finesse", "F-SPEC 最轻精细端", "适合 down shot、小 rubber jig 和 nail sinker rig 这类轻量食わせ，7'6'' 长度能维持远距控线和过障角度。"),
    "KTC763F-SPG": ("淡水 Bass / 岸边轻 cover・开阔水域 / bait finesse 泛用", "F-SPEC 旗舰泛用", "1/8-3/16oz down shot、小 rubber jig、Neko 和轻 bladed jig 都能承接，是 F-SPEC 中最像一支竿解决轻量 rig 的型号。"),
    "KTC764F-SPG": ("淡水 Bass / cover 边・硬底・浅场结构 / 轻中量底操", "compact cover jig / Texas 泛用", "3/16oz 级 Texas、Free Rig、leaderless down shot 和 compact cover jig 是核心，也能补轻 spinnerbait。"),
    "KTC766F-SPG": ("淡水 Bass / 复杂 cover・开阔硬底 / 中重型底操", "F-SPEC 强力端 cover finesse", "3/8oz sinker、football jig 和 shallow crank 都能处理，保留 solid tip 食わせ同时有更强起鱼余量。"),
    "KTC764": ("淡水 Bass / 远距岸钓・轻 cover / 7'6''轻底操", "76 regular 轻量 worm / jig", "适合 1/8-3/8oz light Texas、Free Rig 和轻 rubber jig，长竿优势在远距控线、提线和轻过障。"),
    "KTC765": ("淡水 Bass / 岸边硬物・草边 / 中量底操", "76 regular M 泛用底操", "1/4-1/2oz 区间适合 Texas、Free Rig 和 rubber jig，是 76 Series regular guide 中的日常主力。"),
    "KTC766": ("淡水 Bass / cover 外缘・开阔硬底 / 中重型 worm & jig", "76 regular MH 底操主力", "3/8-1oz 区间能覆盖 rubber jig、Texas 和 leaderless down shot，适合更重线组下远距读底和控鱼。"),
    "KTC767": ("淡水 Bass / 稠密 cover 边缘・远距重底操 / H power", "76 regular heavy cover jig", "1/2-1 1/2oz 级重 Texas、cover jig 和 free rig 更合适，长竿能提升提线和障碍物外缘控鱼效率。"),
    "KTC768": ("淡水 Bass / 重 cover・大鱼目标 / XH power", "76 regular 最强重底操", "3/4-2oz 区间用于 heavy Texas、rubber jig、punching 和 flipping，适合需要硬起鱼的浅场 cover。"),
    "KTC764-SPG": ("淡水 Bass / 远距岸钓・轻 cover / spiral guide 轻底操", "76 SPG 轻量 worm / jig", "保留 KTC764 的轻底操用途，spiral guide 更适合竿身深弯时降低扭转，控线更稳。"),
    "KTC765-SPG": ("淡水 Bass / 岸边硬物・草边 / spiral guide 中量底操", "76 SPG M 泛用底操", "Texas、Free Rig 和 rubber jig 的日常区间，spiral guide 在长竿压鱼和过障时更有稳定感。"),
    "KTC766-SPG": ("淡水 Bass / cover 外缘・开阔硬底 / spiral guide 中重底操", "76 SPG MH 底操主力", "3/8-1oz 区间适合 rubber jig、Texas 和 leaderless down shot，高负荷弯曲时抗扭表现更适合 cover 外缘控鱼。"),
    "KTC767-SPG": ("淡水 Bass / 稠密 cover 边缘・远距重底操 / spiral guide H power", "76 SPG heavy cover jig", "重 Texas、cover jig 和 free rig 是核心，spiral guide 降低长竿受力扭转，适合持续压鱼。"),
    "KTC768-SPG": ("淡水 Bass / 重 cover・大鱼目标 / spiral guide XH power", "76 SPG 最强重底操", "heavy Texas、rubber jig、punching 和 flipping 优先，spiral guide 更适合重负荷下把鱼拉离 cover。"),
}


def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def clean_sku(value) -> str:
    return norm(value).replace("【NEW】", "").replace("〖NEW〗", "").strip()


def parse_schema_array(name: str) -> list[str]:
    text = SCHEMA_PATH.read_text(encoding="utf-8")
    match = re.search(rf"{name}:\s*\[(.*?)\]", text, re.S)
    if not match:
        raise RuntimeError(f"schema array not found: {name}")
    return re.findall(r"'([^']+)'", match.group(1))


def fetch_doc(url: str):
    with urllib.request.urlopen(url, timeout=30) as response:
        return html.fromstring(response.read())


def heading_text(el) -> str:
    return norm(el.text_content())


def collect_text_until(start, stop_titles: set[str]) -> str:
    texts = []
    node = start.getnext()
    while node is not None:
        if node.tag in {"h2", "h3"} and heading_text(node) in stop_titles:
            break
        if node.tag == "h3":
            title = heading_text(node)
            if title:
                texts.append(title)
        elif node.tag not in {"table", "script", "style"}:
            text = norm(node.text_content())
            if text:
                texts.append(text)
        node = node.getnext()
    return "\n".join(texts).strip()


def main_content_lines(doc) -> list[str]:
    containers = doc.xpath('//div[contains(concat(" ", normalize-space(@class), " "), " main-area ")]')
    container = containers[0] if containers else doc
    lines = []
    for text in container.itertext():
        line = norm(text)
        if not line or line.startswith("#block"):
            continue
        lines.append(line)
    return lines


def collect_master_description(doc) -> str:
    lines = main_content_lines(doc)
    try:
        end = lines.index("Detail")
    except ValueError:
        end = next((idx for idx, line in enumerate(lines) if line == "Model List" or line == "SPEC"), len(lines))
    return "\n".join(lines[:end]).strip()


def collect_lineup_descriptions(doc) -> dict[str, str]:
    result = {}
    for h3 in doc.xpath("//h3"):
        title = heading_text(h3)
        sku_match = re.search(r"KTC[0-9A-Z-]+", title)
        if not sku_match:
            continue
        sku = clean_sku(sku_match.group(0))
        texts = []
        node = h3.getnext()
        while node is not None and node.tag != "h3":
            if node.tag not in {"script", "style", "table"}:
                text = norm(node.text_content())
                if text:
                    texts.append(text)
            node = node.getnext()
        if texts:
            result[sku] = "\n".join(texts)
    return result


def parse_spec_tables(doc) -> dict[str, dict[str, str]]:
    variants = {}
    for table in doc.xpath("//table"):
        rows = []
        for tr in table.xpath(".//tr"):
            row = [norm(cell.text_content()) for cell in tr.xpath("./th|./td")]
            if row:
                rows.append(row)
        if not rows:
            continue
        header = rows[0]
        start = next((idx for idx, value in enumerate(header) if "KTC" in value), None)
        if start is None:
            continue
        skus = [clean_sku(value) for value in header[start:] if clean_sku(value)]
        for sku in skus:
            variants.setdefault(sku, {})
        for row in rows[1:]:
            if len(row) <= start:
                continue
            label = norm(" ".join(row[:start]))
            last_value = ""
            for offset, sku in enumerate(skus):
                cell_index = start + offset
                value = row[cell_index] if cell_index < len(row) else ""
                if not value and "ガイド仕様" in label:
                    value = last_value
                if value:
                    last_value = value
                apply_spec_value(variants[sku], label, value)
    return variants


def apply_spec_value(spec: dict[str, str], label: str, value: str) -> None:
    if not value:
        return
    if "全長" in label and "グリップ" not in label and "リア" not in label:
        spec["TOTAL LENGTH"] = value
    elif "グリップ" in label and "リア" not in label:
        spec["Handle Length"] = value
    elif "リアグリップ" in label or ("リア" in label and "グリップ" in label):
        spec["Rear Grip Length"] = value
    elif "自重" in label:
        spec["WEIGHT"] = value
    elif "継数" in label:
        spec["PIECES"] = value
    elif "パワー#" in label or "パワーランク" in label:
        spec["Power Rank"] = value
    elif re.search(r"パワー$", label):
        spec["POWER_RAW"] = value
        spec["POWER"] = normalize_power(value)
    elif "テーパー" in label:
        spec["Action_RAW"] = value
        spec["Action"] = normalize_action(value)
    elif "ガイド仕様" in label:
        spec["Guide Spec"] = value
    elif "Cast Wt" in label:
        spec["LURE WEIGHT"] = value
        spec["LURE WEIGHT (oz)"] = extract_oz(value)
    elif "Line Wt" in label:
        spec["Line Wt N F"] = value
    elif "JAN" in label:
        spec["AdminCode"] = value
    elif "価格" in label:
        spec["Market Reference Price"] = value


def normalize_power(value: str) -> str:
    table = {
        "ライト": "L",
        "ライトプラス": "L+",
        "ミディアムライト": "ML",
        "ミディアム": "M",
        "ミディアムへヴィ": "MH",
        "ミディアムヘヴィ": "MH",
        "へヴィ": "H",
        "ヘヴィ": "H",
        "エクストラへヴィ": "XH",
        "エクストラヘビー": "XH",
    }
    return table.get(norm(value), norm(value))


def normalize_action(value: str) -> str:
    table = {
        "モデレートファスト": "MF",
        "モデレート": "M",
        "エクストラファスト": "XF",
        "ファスト": "F",
    }
    return table.get(norm(value), norm(value))


def extract_oz(value: str) -> str:
    text = norm(value)
    match = re.search(r"(.+?oz\.?)", text, re.I)
    return match.group(1).strip() if match else ""


def technical_terms(model: str, sku: str) -> str:
    common_torzite = "FUJIチタンフレームTORZITEトップガイド"
    common_sic = "FUJIチタンフレームSiC Kガイド"
    if model == "NF Series 66 Casting Model":
        terms = [
            "North Fork Composites SMカーボンブランク",
            "ストレートガイド",
            common_torzite,
            common_sic,
            "FUJI ECSリールシート",
            "ハードEVAグリップ",
        ]
    elif model == "NF Series 76 Flippin' Stick®":
        seat = "FUJI DPSリールシート" if sku == "KTC768NFL" else "FUJI ECSリールシート"
        guide = "スパイラルガイド" if sku.endswith("-SPG") else "ストレートガイド"
        trigger = "トリガーなし" if sku == "KTC768NFL" else "トリガー付き"
        terms = [
            "North Fork Composites HMカーボンブランク",
            guide,
            "オールダブルフットガイド",
            common_torzite,
            common_sic,
            trigger,
            seat,
            "ハードEVAグリップ",
        ]
    elif model == "F-SPEC Series":
        terms = [
            "ソリッドティップ",
            "バイアスブランク設計",
            "無塗装カーボンブランク",
            "スパイラルガイド",
            common_torzite,
            common_sic,
            "FUJI ECSリールシート",
            "ハードEVAグリップ",
        ]
    else:
        guide = "スパイラルガイド" if sku.endswith("-SPG") else "レギュラーガイド"
        terms = [
            "バイアスブランク設計",
            "無塗装カーボンブランク",
            guide,
            common_torzite,
            common_sic,
            "FUJI PTSMリールシート",
            "コルクグリップ",
        ]
    return " / ".join(terms)


def grip_type(model: str) -> str:
    if model == "76 Series":
        return "Cork Grip"
    if model == "F-SPEC Series":
        return "Hard EVA Grip（ラバーコルク補強）"
    return "Hard EVA Grip"


def guide_layout(model: str, sku: str, spec: dict[str, str]) -> str:
    guide_spec = spec.get("Guide Spec") or ("スパイラルガイド" if sku.endswith("-SPG") else "ストレートガイド")
    if model == "NF Series 66 Casting Model":
        return "ストレートガイド（9 + トップ）：FUJIチタンフレームTORZITEトップ + SiC Kガイド"
    if model == "NF Series 76 Flippin' Stick®":
        return f"{guide_spec}：FUJIチタンフレームTORZITEトップ + SiC Kガイド（オールダブルフット）"
    if model == "F-SPEC Series":
        return "スパイラルガイド（11 + トップ）：小口径で糸浮きを抑える設定"
    suffix = "スパイラルガイド" if sku.endswith("-SPG") else "レギュラーガイド"
    return f"{suffix}（11 + Tip）：FUJIチタンフレームTORZITEトップ + SiC Kガイド"


def guide_use_hint(model: str, sku: str) -> str:
    if model == "NF Series 66 Casting Model":
        return "直線 Fuji K 导环路线，适合小型卷物反复抛投与稳定收线；Torzite 顶环和 SiC K 导环偏顺滑出线，适合 finesse moving bait 控节奏。"
    if model == "NF Series 76 Flippin' Stick®":
        if sku.endswith("-SPG"):
            return "螺旋双脚导环降低重负荷弯曲时的竿身扭转，适合 heavy cover flipping 下持续压鱼和近距离起鱼。"
        return "直线双脚导环出线直接，适合传统 flipping 的短距送饵、垂直落点和高负荷控鱼。"
    if model == "F-SPEC Series":
        return "小口径螺旋导环让线始终压在导环侧，降低轻 rig 下落时的糸浮き，便于看线和读取细小 fall bite。"
    if sku.endswith("-SPG"):
        return "螺旋 K 导环降低长竿重负荷弯曲时的扭转，适合底操、cover 场景和持续压鱼。"
    return "常规 K 导环适合 7'6\" worm / jig 抛投和底操控线，出线路线直接，适合远距提线和过障。"


def fallback_description(model: str, sku: str, spec: dict[str, str]) -> str:
    if model == "NF Series 76 Flippin' Stick®":
        guide = "スパイラルガイドセッティング、トリガー付き" if sku.endswith("-SPG") else "ノーマルガイドセッティング、トリガーなし"
        return f"同じNorth Fork HMブランクスを使う7'6'' Flippin' Stick。{sku}は{guide}の仕様で、シャローカバーのフリッピングに向けたXHモデル。"
    if model == "76 Series":
        guide = "スパイラルガイド仕様" if sku.endswith("-SPG") else "レギュラーガイド仕様"
        return f"76 Seriesの{sku}。7'6'' 1ピース／ベイトキャスティング、{spec.get('POWER_RAW') or spec.get('POWER')}パワー、{guide}。ワーム、ラバージグを軸にしたロングロッド。"
    return ""


def build_official_data() -> tuple[dict[str, dict], dict[str, dict]]:
    item_data = {}
    variant_data = {}
    for url in SOURCE_URLS:
        doc = fetch_doc(url)
        model = heading_text(doc.xpath("//h2")[0])
        master_description = collect_master_description(doc)
        lineups = collect_lineup_descriptions(doc)
        specs = parse_spec_tables(doc)
        item_data[model] = {
            "url": url,
            "description": master_description,
            "specs": specs,
        }
        for sku, spec in specs.items():
            desc = lineups.get(sku) or fallback_description(model, sku, spec)
            env, pos, selling = DETAIL_PLAYER[sku]
            variant_data[sku] = {
                **spec,
                "Description": desc,
                "recommended_rig_pairing": RIGS[sku],
                "guide_layout_type": guide_layout(model, sku, spec),
                "guide_use_hint": guide_use_hint(model, sku),
                "hook_keeper_included": "1",
                "sweet_spot_lure_weight_real": "",
                "official_environment": "Fresh Water / Bass",
                "player_environment": env,
                "player_positioning": pos,
                "player_selling_points": selling,
                "product_technical": technical_terms(model, sku),
                "Grip Type": grip_type(model),
                "Extra Spec 1": f"Rear Grip Length: {spec.get('Rear Grip Length')}" if spec.get("Rear Grip Length") else "",
                "Extra Spec 2": f"Power #: {spec.get('Power Rank')}" if spec.get("Power Rank") else "",
            }
    return item_data, variant_data


def read_sheet(ws) -> list[dict[str, object]]:
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        rows.append({header: ws.cell(row=row_idx, column=idx + 1).value for idx, header in enumerate(headers)})
    return rows


def write_sheet(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.delete_rows(1, ws.max_row)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="FF365F91")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
    widths = {
        "Description": 90,
        "product_technical": 56,
        "guide_use_hint": 52,
        "recommended_rig_pairing": 42,
        "player_selling_points": 52,
        "player_environment": 36,
        "player_positioning": 30,
        "guide_layout_type": 48,
        "images": 48,
        "main_selling_points": 56,
        "series_positioning": 34,
    }
    for col_idx, header in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = widths.get(header, 18)
    ws.freeze_panes = "A2"
    fill_a = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
    fill_b = PatternFill(fill_type="solid", fgColor="FFE8F1FB")
    rod_id_header = "rod_id" if "rod_id" in headers else "id"
    group_col = headers.index(rod_id_header) + 1
    current_group = None
    group_index = -1
    for row_idx in range(2, len(rows) + 2):
        group = ws.cell(row=row_idx, column=group_col).value
        if group != current_group:
            current_group = group
            group_index += 1
        fill = fill_a if group_index % 2 == 0 else fill_b
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = copy(cell.alignment)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def merge_price(values: list[str]) -> str:
    cleaned = []
    for value in values:
        text = norm(value)
        if text and text not in cleaned:
            cleaned.append(text)
    if not cleaned:
        return ""
    if len(cleaned) == 1:
        return cleaned[0]
    return " / ".join(cleaned)


def update_normalized(item_data: dict[str, dict], variant_data: dict[str, dict]) -> dict:
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    changed_items = 0
    changed_variants = 0
    for item in data:
        model = item.get("model")
        item_info = item_data.get(model, {})
        master = MASTER_FIELDS.get(model, {})
        before_item = json.dumps(item, ensure_ascii=False, sort_keys=True)
        item["model_cn"] = MODEL_CN.get(model, item.get("model_cn") or "")
        item["alias"] = MODEL_ALIAS.get(model, "")
        item["type_tips"] = "casting"
        item["fit_style_tags"] = item.get("fit_style_tags") or "bass"
        item["Description"] = item_info.get("description") or item.get("description") or ""
        item["description"] = item["Description"]
        for key, value in master.items():
            item[key] = value
        prices = []
        for variant in item.get("variants", []):
            sku = clean_sku(variant.get("sku"))
            enrich = variant_data.get(sku)
            if not enrich:
                continue
            prices.append(enrich.get("Market Reference Price", ""))
            before_variant = json.dumps(variant, ensure_ascii=False, sort_keys=True)
            for key, value in enrich.items():
                variant[key] = value
            if json.dumps(variant, ensure_ascii=False, sort_keys=True) != before_variant:
                changed_variants += 1
        item["official_reference_price"] = merge_price(prices)
        item["market_status"] = "在售"
        if json.dumps(item, ensure_ascii=False, sort_keys=True) != before_item:
            changed_items += 1
    NORMALIZED_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {"items": len(data), "changed_items": changed_items, "changed_variants": changed_variants}


def update_workbook(item_data: dict[str, dict], variant_data: dict[str, dict]) -> dict:
    rod_headers = parse_schema_array("rodMaster")
    detail_headers = parse_schema_array("rodDetail")
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_rows = read_sheet(rod_ws)
    detail_rows = read_sheet(detail_ws)

    for row in rod_rows:
        model = row.get("model")
        master = MASTER_FIELDS.get(model, {})
        item_info = item_data.get(model, {})
        row["model_cn"] = MODEL_CN.get(model, row.get("model_cn") or "")
        row["alias"] = MODEL_ALIAS.get(model, row.get("alias") or "")
        row["type_tips"] = "casting"
        row["fit_style_tags"] = row.get("fit_style_tags") or "bass"
        row["Description"] = item_info.get("description") or row.get("Description") or ""
        for key, value in master.items():
            row[key] = value
        model_skus = [
            clean_sku(detail.get("SKU"))
            for detail in detail_rows
            if detail.get("rod_id") == row.get("id")
        ]
        row["official_reference_price"] = merge_price(
            [variant_data.get(sku, {}).get("Market Reference Price", "") for sku in model_skus]
        )
        row["market_status"] = "在售"

    changed_detail_rows = []
    for row in detail_rows:
        sku = clean_sku(row.get("SKU"))
        enrich = variant_data.get(sku)
        if not enrich:
            continue
        before = {key: row.get(key) for key in detail_headers}
        for key, value in enrich.items():
            if key in detail_headers:
                row[key] = value
        for key in [
            "Handle Length",
            "AdminCode",
            "LURE WEIGHT (oz)",
            "Grip Type",
            "guide_layout_type",
            "guide_use_hint",
            "recommended_rig_pairing",
            "hook_keeper_included",
            "sweet_spot_lure_weight_real",
            "official_environment",
            "player_environment",
            "player_positioning",
            "player_selling_points",
            "Description",
            "product_technical",
            "Extra Spec 1",
            "Extra Spec 2",
        ]:
            row[key] = enrich.get(key, row.get(key, ""))
        if before != {key: row.get(key) for key in detail_headers}:
            changed_detail_rows.append(sku)

    write_sheet(rod_ws, rod_headers, rod_rows)
    write_sheet(detail_ws, detail_headers, detail_rows)
    wb.save(XLSX_PATH)
    return {
        "rod_rows": len(rod_rows),
        "detail_rows": len(detail_rows),
        "changed_detail_rows": len(changed_detail_rows),
        "changed_detail_skus": changed_detail_rows,
        "rod_headers": rod_headers,
        "detail_headers": detail_headers,
    }


def coverage(path: Path) -> dict[str, dict[str, int]]:
    wb = load_workbook(path, read_only=True)
    result = {}
    for sheet in ["rod", "rod_detail"]:
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        counts = {}
        for idx, header in enumerate(headers, start=1):
            counts[header] = sum(
                ws.cell(row=row, column=idx).value not in (None, "")
                for row in range(2, ws.max_row + 1)
            )
        result[sheet] = counts
    return result


def main():
    item_data, variant_data = build_official_data()
    normalized_report = update_normalized(item_data, variant_data)
    workbook_report = update_workbook(item_data, variant_data)
    cov = coverage(XLSX_PATH)
    report = {
        "source_urls": SOURCE_URLS,
        "source_policy": "Keitech official product pages only for official/spec fields; player fields are conservative GearSage interpretation bounded by official descriptions and specs.",
        "normalized": normalized_report,
        "workbook": workbook_report,
        "coverage": {
            "rod": {key: cov["rod"].get(key, 0) for key in ["type_tips", "fit_style_tags", "series_positioning", "main_selling_points", "official_reference_price", "market_status", "Description", "player_positioning", "player_selling_points"]},
            "rod_detail": {key: cov["rod_detail"].get(key, 0) for key in ["Handle Length", "AdminCode", "LURE WEIGHT (oz)", "Grip Type", "guide_layout_type", "guide_use_hint", "recommended_rig_pairing", "hook_keeper_included", "sweet_spot_lure_weight_real", "official_environment", "player_environment", "player_positioning", "player_selling_points", "Description", "product_technical"]},
        },
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
