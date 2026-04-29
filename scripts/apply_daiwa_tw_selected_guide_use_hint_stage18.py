from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")
TARGET_ROD_IDS = {"DR1013", "DR1016", "DR1017", "DR1018", "DR1019", "DR1043"}


def n(value):
    return " ".join(str(value or "").split())


def lower(value):
    return n(value).lower()


def has(text, *tokens):
    return any(token.lower() in text for token in tokens)


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def steez_real_control_hint(detail_id, sku, description, pairing):
    text = lower(f"{sku} {description} {pairing}")
    if detail_id == "DRD10125":
        return "Line slack 操作：中重型軟餌與移動餌切換時，出線和鬆線控制更穩，拋投、操餌、作合到搏魚連貫性更好。"
    if detail_id == "DRD10126":
        return "UL 落下釣組：3lb 細線、1/16oz Down Shot 和小型 Plug 低速操作時，竿尖訊號與鬆線咬口更容易判斷。"
    if detail_id == "DRD10127":
        return "跳底攻略：5g Sinker、Neko/Down Shot 類底操時，硬底觸感、障礙邊張力和拉魚離障礙更直接。"
    if detail_id == "DRD10128":
        return "重型無鉛打擊：高比重 No Sinker、泳鉛鉤和打擊型釣法控線更穩，強力竿身利於刺魚和壓制。"
    if detail_id == "DRD10129":
        return "極細膩 Down Shot：超輕軟蟲與落下釣組能更清楚感知線張力變化，微弱咬口可更快反應。"
    if detail_id == "DRD10130":
        return "現代輕釣法：Soft Plastic、Down Shot、No Sinker 操作時投點精度和線張力控制更好，適合多種細膩技法切換。"
    if detail_id == "DRD10131":
        return "PE 視覺系表層：Topwater、小型 Plug、No Sinker 在表層至 1m 水層操作時，拋投精度、控線和近距離緩衝更穩。"
    if detail_id == "DRD10132":
        return "610M 全能操控：低阻力移動餌和 Neko/Down Shot 等釣組都能保持線弧穩定，拋投精度與操控感更平衡。"
    if detail_id == "DRD10133":
        return "高阻力移動餌：Spinnerbait、大型 Plug 等回收阻力高的餌更好維持線弧和泳層，遠投與追隨性更穩。"
    if detail_id == "DRD10134":
        return "PE 中量精細：表層、中層、底層的中重量軟餌控線更穩，傳遞感和掛鉤後竿身彎曲更自然。"
    if detail_id == "DRD10135":
        return "Power Finesse：輕量釣組打入隱蔽標點時，出線、控線與中魚後壓制更直接，兼顧細膩操作和竿身力量。"
    return generic_bass_hint(sku, pairing, text)


def gekkabijin_mobile_hint(sku):
    sku_text = lower(sku)
    if "ul-s" in sku_text:
        return "輕海水實心竿尖：Jighead、輕量軟餌和小型 Plug 出線更順，細線咬口與竿尖訊號更容易判斷。"
    if "l-s" in sku_text:
        return "便攜輕海水：實心竿尖利於 Jighead、Rockfish Rig 和小型 Metal Jig 讀底，港口小場景控線更清楚。"
    if "ml-t" in sku_text:
        return "便攜輕海水遠投：管狀竿尖兼顧小型 Metal Jig、Plug 和底操釣組，拋投距離與回收控線更穩。"
    return "便攜輕海水：細線小餌出線更順，港口、堤防和旅行場景切換時控線更穩。"


def generic_bass_hint(sku, pairing, text):
    sku_text = lower(sku)
    pair_text = lower(pairing)
    solid_tip = any(token in sku_text for token in ["-st", "•st", ".st", " st", "smt"])

    if has(pair_text, "frog", "punching", "heavy texas"):
        return "Frog/重障礙：粗線出線與落點控制更穩，Frog、Punching、Heavy Texas 中魚後更利於拉離覆蓋物。"
    if has(pair_text, "swimbait", "big bait", "a-rig", "wakebait"):
        return "大餌/泳餌：Swimbait、Big Bait、Swim Jig 拋投和回收負荷更穩，長距離操餌與控魚更有餘量。"
    if has(pair_text, "jighead", "down shot", "neko", "wacky", "i-shaped", "bug lure", "small rubber") and not has(
        pair_text, "free rig", "texas", "leaderless", "heavy down shot"
    ):
        if has(sku_text, "bf"):
            return "BF 精細輕餌：Neko、Down Shot、Small Rubber Jig 低彈道拋投和控鬆線更穩，障礙邊細口更好判斷。"
        if has(sku_text, "ul", "xul"):
            return "超輕精細：Jighead、No Sinker、Wacky、I-shaped Plug 出線負荷更低，細線咬口和竿尖訊號更清楚。"
        return "精細輕釣組：Jighead、Down Shot、Neko、Wacky 操作時線弧更穩，讀底、看線與微弱咬口更直接。"
    if has(pair_text, "free rig", "texas", "leaderless", "heavy down shot", "rubber jig"):
        if solid_tip or has(text, "solid", "實心", "smt"):
            return "高感底操軟餌：實心竿尖放大 Texas、Free Rig、Leaderless Down Shot 的張力變化，讀底和輕咬更清楚。"
        if has(sku_text, "mh", "h", "mh+"):
            return "強力底操軟餌：Texas、Rubber Jig、Free Rig 在障礙區控線更穩，刺魚和拉離覆蓋物更直接。"
        return "底操軟餌：Texas、Free Rig、Rubber Jig 貼底時線張力和障礙觸感更清楚，方便判斷小口與刺魚時機。"
    if has(pair_text, "crankbait", "shad", "minnow", "chatterbait", "spinnerbait", "topwater", "jerkbait"):
        if has(pair_text, "soft plastic", "down shot", "no sinker"):
            return "硬餌兼輕軟餌：Crankbait、Shad、Minnow 平收控線更穩，也能兼顧 Down Shot、No Sinker 的細線操作。"
        return "卷阻/搜索硬餌：Crankbait、Shad、Minnow、Spinnerbait 的線弧和泳層更好維持，連續平收與抽停節奏更清楚。"
    if has(pair_text, "light texas", "small crankbait", "soft plastic", "no sinker"):
        return "輕量泛用：No Sinker、Light Texas、小型 Crankbait 和 Soft Plastic 切換時，出線、控線和小餌操作更自然。"

    if has(sku_text, "mh", "h"):
        return "中重型泛用：中重餌出線和回收負荷更穩，軟餌底操、移動餌和障礙區控魚更有餘量。"
    return "Bass 泛用細化：出線順暢並兼顧軟餌、硬餌與移動餌切換，日常多技法操作更穩。"


def build_hint(rod_id, detail_id, sku, description, pairing):
    if rod_id == "DR1018":
        return steez_real_control_hint(detail_id, sku, description, pairing)
    if rod_id == "DR1043":
        return gekkabijin_mobile_hint(sku)
    return generic_bass_hint(sku, pairing, lower(f"{sku} {description} {pairing}"))


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["guide_use_hint"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        rod_id = n(ws.cell(row, col["rod_id"]).value)
        if rod_id not in TARGET_ROD_IDS:
            continue
        detail_id = n(ws.cell(row, col["id"]).value)
        sku = n(ws.cell(row, col["SKU"]).value)
        description = n(ws.cell(row, col["Description"]).value)
        pairing = n(ws.cell(row, col["recommended_rig_pairing"]).value)
        new = build_hint(rod_id, detail_id, sku, description, pairing)
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        if old != new:
            cell.value = new
            changed.append({"row": row, "id": detail_id, "rod_id": rod_id, "sku": sku, "old": old, "new": new})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-guide_use_hint changes")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed": len(changed),
        "target_rod_ids": sorted(TARGET_ROD_IDS),
        "changed_ids": [item["id"] for item in changed],
    })


if __name__ == "__main__":
    main()
