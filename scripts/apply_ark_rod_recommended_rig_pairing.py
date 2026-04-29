#!/usr/bin/env python3
import json
import re
import subprocess
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


DATA_DIR = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw")
XLSX_PATH = DATA_DIR / "ark_rod_import.xlsx"
NORMALIZED_PATH = DATA_DIR / "ark_rod_normalized.json"
REPORT_PATH = DATA_DIR / "ark_rod_recommended_rig_pairing_report.json"
SHADE_SCRIPT = Path("/Users/tommy/GearSage/scripts/shade_ark_rod_detail_groups.py")

FIELD = "recommended_rig_pairing"
GENERIC_FORBIDDEN = {"General Lure", "Light Rig", "Hardbait", "Soft Bait"}


def norm(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def key(value):
    return re.sub(r"[^A-Z0-9]+", "", norm(value).upper())


def add(items, *values):
    for value in values:
        value = norm(value)
        if value and value not in items:
            items.append(value)


def max_lure_oz(value):
    nums = []
    for raw in re.findall(r"\d+(?:/\d+)?(?:\.\d+)?", norm(value)):
        try:
            nums.append(float(raw.split("/", 1)[0]) / float(raw.split("/", 1)[1]) if "/" in raw else float(raw))
        except (ValueError, ZeroDivisionError):
            pass
    return max(nums) if nums else 0


def append_by_application(items, app):
    lower = norm(app).lower().replace("＆", "&")
    rules = [
        ("jig- n- rap", "Jigging Rap"), ("jig-n-rap", "Jigging Rap"), ("rip-n-rap", "Ripping Rap"),
        ("snap jig", "Snap Jig"), ("vertical jig", "Vertical Jig"), ("bottom bouncer", "Bottom Bouncer Rig"),
        ("telestopic trolling", "Trolling Crankbait"), ("trolling", "Trolling Crankbait"),
        ("slip-n rig", "Slip Rig"), ("panfish - perch", "Panfish Ice Jig"), ("panfish", "Panfish Jig"),
        ("perch", "Perch Spoon"), ("deadstick", "Deadstick Minnow"), ("dead stick", "Deadstick Minnow"),
        ("a-rig", "Alabama Rig"), ("umbrella", "Umbrella Rig"), ("glidebait", "Glidebait"),
        ("glide bait", "Glidebait"), ("swim bait", "Swimbait"), ("swimbait", "Swimbait"),
        ("frog", "Frog"), ("punching", "Punching Rig"), ("flipping", "Flipping Jig"),
        ("heavy cover", "Frog"), ("dock skipping", "Dock Skipping Jig"), ("skipping", "Dock Skipping Jig"),
        ("finesse jig", "Small Rubber Jig"), ("small jig", "Small Rubber Jig"),
        ("football jig", "Football Jig"), ("offshore jig", "Offshore Jig"),
        ("swimjig", "Swim Jig"), ("swim jig", "Swim Jig"),
        ("jig&worm", "Rubber Jig"), ("jig, worm", "Rubber Jig"), ("jig,worm", "Rubber Jig"),
        ("jig & worm", "Rubber Jig"), ("skipping jigs", "Dock Skipping Jig"),
        ("tube", "Tube Rig"), ("c-rig", "Carolina Rig"), ("carolinarig", "Carolina Rig"),
        ("free rig", "Free Rig"), ("weightless worm", "No Sinker"), ("weighless worm", "No Sinker"),
        ("worm", "Texas Rig"), ("texas", "Texas Rig"), ("dropshot", "Down Shot"),
        ("drop shot", "Down Shot"), ("damiki", "Damiki Rig"), ("nedrig", "Ned Rig"),
        ("ned rig", "Ned Rig"), ("ned", "Ned Rig"), ("neko rig", "Neko Rig"),
        ("nekorig", "Neko Rig"), ("neko", "Neko Rig"), ("niko", "Neko Rig"),
        ("shakyhead", "Shaky Head"), ("shaky head", "Shaky Head"),
        ("jighead minnow", "Jighead Minnow"), ("jighead", "Jighead Rig"),
        ("hairjig", "Hair Jig"), ("hair jig", "Hair Jig"), ("spybait", "Spybait"),
        ("spy bait", "Spybait"), ("mid-strolling", "Mid Strolling"), ("mid strolling", "Mid Strolling"),
        ("jerkbait", "Jerkbait"), ("mini crank", "Mini Crankbait"),
        ("small crank", "Small Crankbait"), ("squarebill", "Squarebill Crankbait"),
        ("shallow crankbait", "Shallow Crankbait"), ("shallow crank", "Shallow Crankbait"),
        ("mid-depth crank", "Mid-depth Crankbait"), ("mid-deep crankbait", "Mid-depth Crankbait"),
        ("deep diver", "Deep Crankbait"), ("deepdiver", "Deep Crankbait"),
        ("crankbait", "Crankbait"), ("crank", "Crankbait"),
        ("lipless", "Lipless Crankbait"), ("vibration", "Vibration"),
        ("small topwater", "Small Topwater Plug"), ("topwater", "Topwater Plug"),
        ("buzzbait", "Buzzbait"), ("spinnerbait", "Spinnerbait"), ("chatterbait", "Chatterbait"),
        ("chatter", "Chatterbait"), ("underspin", "Underspin"), ("blade", "Blade Bait"),
    ]
    matches = [(lower.find(token), label) for token, label in rules if lower.find(token) >= 0]
    for _, label in sorted(matches, key=lambda item: item[0]):
        add(items, label)
    if any(term in lower for term in ["jig&worm", "jig, worm", "jig,worm", "jig & worm"]):
        add(items, "Texas Rig")
    if "finesse jig" in lower or "small jig" in lower:
        add(items, "Light Texas Rig")
    if "composite" in lower and "crank" in lower:
        add(items, "Crankbait")


def infer_by_spec(row, spec):
    items = []
    title = norm(spec.get("series_title", ""))
    sku = norm(row.get("SKU", ""))
    typ = norm(row.get("TYPE", ""))
    power = norm(row.get("POWER", ""))
    action = norm(row.get("Action", ""))
    lure_max = max_lure_oz(row.get("LURE WEIGHT (oz)", ""))
    text = f"{title} {sku}".lower()

    if "ice" in text or sku.startswith(("CTI", "CIR")):
        if "QT" in sku:
            add(items, "Ice Rattle Bait", "Panfish Ice Jig")
        elif "DS" in sku:
            add(items, "Deadstick Minnow", "Walleye Jig")
        elif power in {"UL", "L"}:
            add(items, "Panfish Ice Jig", "Small Spoon")
        else:
            add(items, "Walleye Jig", "Ice Spoon")
        return items, "conservative_spec"

    if "w series" in text or sku.startswith(("IVW", "LCW", "LCWT")):
        if sku.startswith("LCWT"):
            add(items, "Trolling Crankbait", "Bottom Bouncer Rig")
        elif power in {"L", "ML"}:
            add(items, "Jigging Rap", "Snap Jig", "Slip Rig")
        else:
            add(items, "Walleye Jig", "Jigging Rap", "Bottom Bouncer Rig")
        return items, "conservative_spec"

    if "bfs" in text or sku.startswith("BF"):
        add(items, "BFS Jighead Rig", "Small Plug", "Small Rubber Jig", "No Sinker")
        return items, "official_series"

    if typ == "S":
        if power in {"UL", "L"} or lure_max <= 0.125:
            add(items, "Small Jighead Rig", "Small Minnow", "Small Spoon")
        elif power in {"ML", "M"}:
            add(items, "Down Shot", "Neko Rig", "Ned Rig", "Jighead Rig")
        else:
            add(items, "Tube Rig", "Jighead Minnow", "Free Rig", "Down Shot")
        return items, "conservative_spec"

    if power in {"H+", "XH"} or lure_max >= 2:
        add(items, "Frog", "Punching Rig", "Swimbait", "Rubber Jig")
    elif power == "H":
        add(items, "Frog", "Flipping Jig", "Texas Rig", "Rubber Jig")
    elif power in {"MH", "MH+"}:
        add(items, "Texas Rig", "Rubber Jig", "Spinnerbait", "Chatterbait")
    elif action in {"R", "MF"}:
        add(items, "Jerkbait", "Crankbait", "Topwater Plug", "Spinnerbait")
    else:
        add(items, "Texas Rig", "Small Rubber Jig", "Jerkbait", "Crankbait")
    return items, "conservative_spec"


def build_pairing(row, spec):
    app = norm(spec.get("application", ""))
    items = []
    if app and not re.fullmatch(r"all[- ]?purpose(?: casting)?|do it all(?: casting| spinning)?", app, flags=re.I):
        append_by_application(items, app)
        source = "official_application"
    else:
        source = ""
    if len(items) < 2:
        inferred, inferred_source = infer_by_spec(row, spec)
        for item in inferred:
            add(items, item)
        source = source or inferred_source
    if "Small Topwater Plug" in items and "Topwater Plug" in items:
        items.remove("Topwater Plug")
    specific_cranks = {
        "Mini Crankbait", "Small Crankbait", "Squarebill Crankbait", "Shallow Crankbait",
        "Mid-depth Crankbait", "Deep Crankbait",
    }
    if "Crankbait" in items and any(item in items for item in specific_cranks):
        items.remove("Crankbait")
    return " / ".join(items[:7]), source


def first_category(pairing):
    first = norm(pairing).split(" / ")[0] if pairing else ""
    soft = {
        "Down Shot", "Neko Rig", "Ned Rig", "Jighead Rig", "BFS Jighead Rig", "Small Jighead Rig",
        "Small Rubber Jig", "Rubber Jig", "Texas Rig", "Light Texas Rig", "Free Rig", "No Sinker",
        "Tube Rig", "Carolina Rig", "Flipping Jig", "Punching Rig", "Football Jig", "Dock Skipping Jig",
        "Damiki Rig", "Shaky Head",
    }
    hard = {
        "Jerkbait", "Crankbait", "Mini Crankbait", "Small Crankbait", "Shallow Crankbait",
        "Mid-depth Crankbait", "Deep Crankbait", "Lipless Crankbait", "Vibration", "Topwater Plug",
        "Small Topwater Plug", "Buzzbait", "Spinnerbait", "Chatterbait", "Swim Jig", "Underspin",
        "Blade Bait", "Squarebill Crankbait", "Spybait", "Small Plug", "Small Minnow", "Glidebait",
        "Swimbait", "Alabama Rig", "Umbrella Rig",
    }
    return "soft" if first in soft else "hard" if first in hard else "other"


def guide_from_pairing(pairing, row, spec, source):
    parts = [part.strip() for part in norm(pairing).split("/") if part.strip()]
    first = parts[0] if parts else ""
    app = norm(spec.get("application", ""))
    title = norm(spec.get("series_title", ""))
    sku = norm(row.get("SKU", ""))
    power = norm(row.get("POWER"))
    action = norm(row.get("Action"))
    line = norm(row.get("Line Wt N F"))
    lure = norm(row.get("LURE WEIGHT (oz)"))
    spec_note = " / ".join(x for x in [f"{power} power" if power else "", f"{action} action" if action else "", f"线号 {line}" if line else "", f"饵重 {lure} oz" if lure else ""] if x)
    basis = f"官方用途为 {app}，" if app and source == "official_application" else ""
    spec_tail = f"规格上是 {spec_note}，" if spec_note else ""
    pair_note = "、".join(parts[:4]) if parts else "推荐钓组"

    if "ice" in title.lower() or sku.startswith(("CTI", "CIR")):
        return f"{basis}{spec_tail}用于冰钓短距离控线；围绕 {pair_note}，重点是小幅操饵、竿尖信号和轻口观察，Panfish/Perch/Walleye 场景下要能分辨静置、抖动和吸口变化。"

    if first in {"Down Shot", "Neko Rig", "Ned Rig", "Damiki Rig", "Shaky Head", "Jighead Minnow", "Jighead Rig"}:
        return f"{basis}{spec_tail}适合细线精细钓组；围绕 {pair_note}，重点是轻饵抛投、线弧控制和轻口反馈，需要清楚读到底部与中层细微咬口。"
    if first in {"BFS Jighead Rig", "Small Jighead Rig", "Small Rubber Jig"}:
        return f"{basis}{spec_tail}偏轻量精细作钓；围绕 {pair_note}，重点是小饵精准落点、低阻线组出线和近中距离操控。"
    if first in {"Rubber Jig", "Texas Rig", "Light Texas Rig", "Free Rig", "Carolina Rig", "Tube Rig", "Football Jig"}:
        cover_note = "；其中 Frog/Flipping 类搭配需要额外关注近障碍控鱼" if any(item in parts for item in ["Frog", "Flipping Jig", "Punching Rig"]) else ""
        return f"{basis}{spec_tail}以底部接触和结构区作钓为主；围绕 {pair_note}，重点是控线、感知障碍和补刺力量，要能稳定贴底、穿越 cover 并完成补刺{cover_note}。"
    if first in {"Frog", "Flipping Jig", "Punching Rig"}:
        return f"{basis}{spec_tail}面向重 cover 和近障碍强力作钓；围绕 {pair_note}，重点是粗线、重饵和短距离爆发控鱼，需要快速带鱼离开障碍。"
    if first in {"Swim Jig", "Chatterbait", "Spinnerbait", "Buzzbait", "Underspin", "Blade Bait"}:
        if any(item in parts for item in ["Football Jig", "Carolina Rig", "Offshore Jig", "Rubber Jig", "Texas Rig"]):
            return f"{basis}{spec_tail}适合移动饵与结构钓组切换；围绕 {pair_note}，前段负责水层和草边搜索，后段负责结构区慢探，重点是控线稳定和节奏切换。"
        return f"{basis}{spec_tail}偏移动饵搜索；围绕 {pair_note}，重点是连续抛投、控速和震动反馈，需要稳定出线并清楚感知叶片、震动或泳姿变化。"
    if first in {"Jerkbait", "Crankbait", "Mini Crankbait", "Small Crankbait", "Squarebill Crankbait", "Shallow Crankbait", "Mid-depth Crankbait", "Deep Crankbait", "Lipless Crankbait", "Vibration", "Topwater Plug", "Small Topwater Plug", "Spybait", "Small Plug", "Small Minnow"}:
        return f"{basis}{spec_tail}以硬饵和反应饵为主；围绕 {pair_note}，重点是抛投节奏、抽停/巻物控饵和咬口瞬间的缓冲，需要稳定回弹和持续操作舒适度。"
    if first in {"Glidebait", "Swimbait", "Alabama Rig", "Umbrella Rig"}:
        return f"{basis}{spec_tail}面向大饵或高阻力饵；围绕 {pair_note}，重点是负载、抛投稳定性和搏鱼余量，需要控制饵体摆动并承受长距离抛投。"
    if first in {"Jigging Rap", "Ripping Rap", "Snap Jig", "Walleye Jig", "Bottom Bouncer Rig", "Slip Rig", "Trolling Crankbait", "Vertical Jig"}:
        return f"{basis}{spec_tail}偏 Walleye/Panfish 湖库技法；围绕 {pair_note}，重点是垂直控线、抽停节奏和拖钓稳定性，需要清楚判断贴底、停顿和轻口。"
    if first in {"Ice Rattle Bait", "Panfish Ice Jig", "Perch Spoon", "Deadstick Minnow", "Ice Spoon"}:
        return f"{basis}{spec_tail}用于冰钓短距离控线；围绕 {pair_note}，重点是小幅操饵、竿尖信号和轻口观察，Panfish/Perch/Walleye 场景下要能分辨静置、抖动和吸口变化。"
    return f"{basis}{spec_tail}适合多种钓组切换；重点是根据饵重和线号控制抛投、控线与补刺，优先按 recommended_rig_pairing 的前几项选择搭配。"


def load_spec_index():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    out = {}
    for index, item in enumerate(data):
        rod_id = f"ARKR{1000 + index}"
        for spec in item.get("specs", []):
            spec = dict(spec)
            spec["series_title"] = item.get("title", "")
            out[(rod_id, key(spec.get("sku", "")))] = spec
    return out


def ensure_field(ws):
    headers = [cell.value for cell in ws[1]]
    if FIELD in headers:
        return headers.index(FIELD) + 1
    insert_at = headers.index("hook_keeper_included") + 1 if "hook_keeper_included" in headers else len(headers) + 1
    ws.insert_cols(insert_at)
    src_col = max(1, insert_at - 1)
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row=row, column=src_col)
        dst = ws.cell(row=row, column=insert_at)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    ws.cell(row=1, column=insert_at).value = FIELD
    ws.column_dimensions[ws.cell(row=1, column=insert_at).column_letter].width = 34
    return insert_at


def main():
    spec_index = load_spec_index()
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    field_col = ensure_field(ws)
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    changed = []
    filled_rows = []
    guide_updates = []
    missing_source = []
    consistency_issues = []

    for row_idx in range(2, ws.max_row + 1):
        row = {name: ws.cell(row=row_idx, column=idx).value for name, idx in col.items()}
        spec = spec_index.get((norm(row.get("rod_id")), key(row.get("SKU"))))
        if not spec:
            missing_source.append({"row": row_idx, "id": row.get("id"), "sku": row.get("SKU")})
            continue
        pairing, source = build_pairing(row, spec)
        filled_rows.append(
            {
                "row": row_idx,
                "id": row.get("id"),
                "rod_id": row.get("rod_id"),
                "sku": row.get("SKU"),
                "recommended_rig_pairing": pairing,
                "source": source,
                "official_application": spec.get("application", ""),
            }
        )
        old = norm(ws.cell(row=row_idx, column=field_col).value)
        if old != pairing:
            ws.cell(row=row_idx, column=field_col).value = pairing
            changed.append(
                {
                    "row": row_idx,
                    "id": row.get("id"),
                    "rod_id": row.get("rod_id"),
                    "sku": row.get("SKU"),
                    "old": old,
                    "recommended_rig_pairing": pairing,
                    "source": source,
                    "official_application": spec.get("application", ""),
                }
            )
        guide = norm(row.get("guide_use_hint"))
        new_guide = guide_from_pairing(pairing, row, spec, source)
        if guide != new_guide:
            ws.cell(row=row_idx, column=col["guide_use_hint"]).value = new_guide
            guide_updates.append({"row": row_idx, "sku": row.get("SKU"), "old": guide, "new": new_guide, "pairing": pairing})
            guide = new_guide
        category = first_category(pairing)
        if category == "soft" and any(token in guide for token in ["反应饵", "连续抛投", "搜索"]):
            new_guide = "软硬饵切换：以软饵钓组为主，也能兼顾部分移动饵；重点是线弧控制、轻口反馈和节奏切换。"
            ws.cell(row=row_idx, column=col["guide_use_hint"]).value = new_guide
            guide_updates.append({"row": row_idx, "sku": row.get("SKU"), "old": guide, "new": new_guide, "pairing": pairing})
        if category == "hard" and any(token in guide for token in ["底部接触", "拖底", "jig、worm"]):
            new_guide = "移动饵与底部钓组切换：Swim Jig、Football Jig、Carolina Rig 等能兼顾搜索和贴底，重点是控线稳定和节奏切换。"
            ws.cell(row=row_idx, column=col["guide_use_hint"]).value = new_guide
            guide_updates.append({"row": row_idx, "sku": row.get("SKU"), "old": guide, "new": new_guide, "pairing": pairing})

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    wb_check = load_workbook(XLSX_PATH, data_only=True)
    ws_check = wb_check["rod_detail"]
    headers = [cell.value for cell in ws_check[1]]
    field_idx = headers.index(FIELD)
    values = [norm(row[field_idx]) for row in ws_check.iter_rows(min_row=2, values_only=True)]
    headers_map = {name: idx for idx, name in enumerate(headers)}
    guide_sync_rows = []
    for row_idx, row in enumerate(ws_check.iter_rows(min_row=2, values_only=True), start=2):
        guide = norm(row[headers_map["guide_use_hint"]])
        if guide.startswith("移动饵与底部钓组切换") or guide.startswith("软硬饵切换"):
            guide_sync_rows.append(
                {
                    "row": row_idx,
                    "id": row[headers_map["id"]],
                    "sku": row[headers_map["SKU"]],
                    "recommended_rig_pairing": row[field_idx],
                    "guide_use_hint": guide,
                }
            )
    forbidden = [value for value in values if any(bad in [part.strip() for part in value.split("/")] for bad in GENERIC_FORBIDDEN)]
    report = {
        "field": FIELD,
        "updated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "xlsx": str(XLSX_PATH),
        "normalized": str(NORMALIZED_PATH),
        "changed_cells": len(changed),
        "guide_use_hint_updates": guide_updates,
        "guide_use_hint_sync_rows": guide_sync_rows,
        "coverage": {"filled": sum(1 for value in values if value), "total": len(values)},
        "generic_forbidden_residuals": forbidden,
        "missing_source_rows": missing_source,
        "consistency_issues": consistency_issues,
        "source_counts": {},
        "source_counts_all": {},
        "filled_rows": filled_rows,
        "changed_rows": changed,
    }
    for item in changed:
        report["source_counts"][item["source"]] = report["source_counts"].get(item["source"], 0) + 1
    for item in filled_rows:
        report["source_counts_all"][item["source"]] = report["source_counts_all"].get(item["source"], 0) + 1
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: report[k] for k in ["changed_cells", "coverage", "source_counts"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
