import json
import re
from collections import defaultdict
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('megabass_rod_import.xlsx')
EVIDENCE_PATH = resolve_data_raw('megabass_rod_whitelist_player_evidence.json')
REPORT_PATH = resolve_data_raw('megabass_rod_whitelist_backfill_report.json')

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_positioning(value):
    text = n(value)
    replacements = {
        "big bait / swimbait / 强力": "大饵 / swimbait / 强力",
        "moving bait / 卷阻饵 / 搜索": "moving bait / 卷阻饵 / 搜索",
        "轻量钓组 / 精细控饵": "轻量钓组 / 精细控饵",
        "重障碍 / 大饵 / 强控鱼": "重障碍 / 大饵 / 强控鱼",
    }
    return replacements.get(text, text)


def selling_points(positioning, specs):
    pos = normalize_positioning(positioning)
    power = n(specs.get("power"))
    taper = n(specs.get("taper"))
    lure = n(specs.get("lure_wt"))
    line = n(specs.get("line_wt"))
    parts = []

    if pos == "轻量钓组 / 精细控饵":
        parts = ["轻线和轻饵规格更清楚", "适合精细控饵与咬口反馈"]
    elif pos == "重障碍 / 大饵 / 强控鱼":
        parts = ["中重负荷规格明确", "适合障碍区控鱼和强力扬竿"]
    elif pos == "大饵 / swimbait / 强力":
        parts = ["大饵负荷范围明确", "适合 swimbait、粗线和高负荷搏鱼"]
    elif pos == "moving bait / 卷阻饵 / 搜索":
        parts = ["调性更适合连续搜索", "卷阻饵节奏和抛投稳定性更清楚"]
    else:
        parts = ["第三方规格表可辅助判断具体技法定位"]

    spec_bits = [bit for bit in [power, taper, lure, line] if bit]
    if spec_bits:
        return f"{' / '.join(parts)} / 参考规格：{' / '.join(spec_bits[:4])}"
    return " / ".join(parts)


def guide_layout_value(specs, hut_hint=False):
    guides = n(specs.get("guides"))
    if not guides:
        return ""
    if hut_hint:
        return f"Fuji Alconite / {guides} 导环：耐磨并兼顾出线稳定，适合高频淡水抛投和细线控线。"
    return f"{guides} 导环配置：导环数量清楚，便于判断出线稳定、线组支撑和后期维护。"


def guide_use_hint_from_hut():
    return "Fuji Alconite 导环：耐磨且出线稳定，适合高频淡水抛投和细线控线。"


def grip_type_value(value):
    text = n(value)
    if re.search(r"\bEVA\b", text, re.I):
        return "EVA casting grip"
    if re.search(r"\bSplit\b", text, re.I):
        return "Split casting grip"
    if re.search(r"\bFull\b", text, re.I):
        return "Full casting grip"
    return ""


def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def ensure_headers(ws, headers, required):
    col = {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
    for field in required:
        if field in col:
            continue
        next_col = ws.max_column + 1
        header_cell = ws.cell(row=1, column=next_col)
        copy_cell_style(ws.cell(row=1, column=ws.max_column), header_cell)
        header_cell.value = field
        col[field] = next_col
        headers.append(field)
    return col


def set_value(ws, row_idx, col, field, value, changed, source_id, overwrite=False):
    if not n(value) or field not in col:
        return False
    cell = ws.cell(row=row_idx, column=col[field])
    current = n(cell.value)
    if current == n(value):
        return False
    if current and not overwrite:
        return False
    cell.value = value
    cell.fill = copy(YELLOW_FILL)
    changed.append(
        {
            "row": row_idx,
            "id": source_id,
            "field": field,
            "old_value": current,
            "new_value": value,
        }
    )
    return True


def main():
    evidence_data = json.loads(EVIDENCE_PATH.read_text())
    evidence_by_id = defaultdict(list)
    for item in evidence_data.get("evidence", []):
        evidence_by_id[n(item.get("id"))].append(item)

    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [n(cell.value) for cell in ws[1]]
    required = [
        "Handle Length",
        "Grip Type",
        "guide_layout_type",
        "guide_use_hint",
        "hook_keeper_included",
        "player_positioning",
        "player_selling_points",
    ]
    col = ensure_headers(ws, headers, required)

    changed = []
    rows_seen = set()

    for row_idx in range(2, ws.max_row + 1):
        source_id = n(ws.cell(row=row_idx, column=col["id"]).value)
        if not source_id or source_id not in evidence_by_id:
            continue

        items = evidence_by_id[source_id]
        tw_items = [item for item in items if item.get("source_site") == "tacklewarehouse.com" and item.get("source_specs")]
        hut_items = [item for item in items if item.get("source_site") == "thehookuptackle.com"]
        tw = tw_items[0] if tw_items else None
        specs = tw.get("source_specs", {}) if tw else {}

        if tw:
            set_value(ws, row_idx, col, "Handle Length", n(specs.get("handle_length")), changed, source_id)
            set_value(ws, row_idx, col, "Grip Type", grip_type_value(specs.get("handle_type")), changed, source_id)

            layout = guide_layout_value(specs, bool(hut_items))
            set_value(ws, row_idx, col, "guide_layout_type", layout, changed, source_id)

            positioning = normalize_positioning(tw.get("suggested_player_positioning"))
            if positioning:
                set_value(ws, row_idx, col, "player_positioning", positioning, changed, source_id, overwrite=True)
                set_value(ws, row_idx, col, "player_selling_points", selling_points(positioning, specs), changed, source_id, overwrite=True)

        if hut_items:
            set_value(ws, row_idx, col, "guide_use_hint", guide_use_hint_from_hut(), changed, source_id, overwrite=True)
            set_value(ws, row_idx, col, "hook_keeper_included", "1", changed, source_id)

        if any(item.get("source_site") in {"tacklewarehouse.com", "thehookuptackle.com"} for item in items):
            rows_seen.add(source_id)

    wb.save(XLSX_PATH)

    by_field = defaultdict(int)
    for item in changed:
        by_field[item["field"]] += 1

    report = {
        "xlsx_file": str(XLSX_PATH),
        "evidence_file": str(EVIDENCE_PATH),
        "changed_cells": len(changed),
        "changed_rows": len({item["id"] for item in changed}),
        "evidence_rows_seen": len(rows_seen),
        "yellow_fill": "FFFFFF00",
        "by_field": dict(sorted(by_field.items())),
        "changes": changed,
        "not_applied": {
            "AdminCode": "PLAT product_code is retailer context, not stable official AdminCode/JAN coverage.",
        },
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({k: report[k] for k in ["changed_cells", "changed_rows", "evidence_rows_seen", "by_field"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
