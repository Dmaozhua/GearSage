from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')

HINT_BY_ID = {
    "DRD10026": "輕量泛用：細線與輕量釣組出線更順，軟蟲、中小型硬餌和障礙邊精細作釣切換更自然。",
    "DRD10027": "全能泛用：Crankbait、Spinnerbait、Down Shot、Texas 等軟硬餌切換時，拋投、控線和回收節奏更穩。",
    "DRD10028": "強力泛用：Texas、Jig 等打擊系和 Swim Jig、Spinnerbait 等中重型移動餌都能兼顧，控線與刺魚更有餘量。",
    "DRD10029": "精細泛用：輕量釣組、小型硬餌和 I 字系等細膩操作時，線弧、竿尖回饋和小口判斷更穩。",
    "DRD10030": "全能輕量泛用：輕量釣組遠投、揚竿刺魚，以及 Twitch/Jerk 小型硬餌操作都能兼顧，移動釣行切換更順。",
}


def n(value):
    return " ".join(str(value or "").split())


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["guide_use_hint"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        if detail_id not in HINT_BY_ID:
            continue
        if n(ws.cell(row, col["rod_id"]).value) != "DR1006":
            raise RuntimeError(f"unexpected rod_id for {detail_id}")
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        new = HINT_BY_ID[detail_id]
        if old != new:
            cell.value = new
            changed.append({"row": row, "id": detail_id, "old": old, "new": new})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-guide_use_hint changes")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed": len(changed),
        "changed_ids": [item["id"] for item in changed],
    })


if __name__ == "__main__":
    main()
