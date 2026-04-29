from copy import copy
from pathlib import Path
import json
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = ROOT / "GearSage-client/pkgGear/data_raw/evergreen_rod_import.xlsx"
REPORT_PATH = ROOT / "GearSage-client/pkgGear/data_raw/evergreen_rod_line_wt_cleanup_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_line_wt(value):
    text = n(value)
    text = re.sub(r"^[／/]+", "", text)
    text = re.sub(r"[／/]+$", "", text)
    return text


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    col = {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
    required = ["id", "SKU", "Line Wt N F"]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    changes = []
    for row_idx in range(2, ws.max_row + 1):
        source_id = n(ws.cell(row_idx, col["id"]).value)
        old_value = n(ws.cell(row_idx, col["Line Wt N F"]).value)
        new_value = clean_line_wt(old_value)
        if old_value == new_value:
            continue
        cell = ws.cell(row_idx, col["Line Wt N F"])
        cell.value = new_value
        cell.fill = copy(YELLOW_FILL)
        changes.append({
            "row": row_idx,
            "id": source_id,
            "sku": n(ws.cell(row_idx, col["SKU"]).value),
            "old_value": old_value,
            "new_value": new_value,
        })

    wb.save(XLSX_PATH)

    import subprocess

    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "xlsx_file": str(XLSX_PATH),
        "changed_rows": len(changes),
        "rule": "Trim leading/trailing ASCII slash and full-width slash from Line Wt N F only.",
        "changes": changes,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "changed_rows": len(changes),
        "report_file": str(REPORT_PATH),
        "samples": changes[:10],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
