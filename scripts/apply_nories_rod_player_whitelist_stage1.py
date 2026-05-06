#!/usr/bin/env python3
import json
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


OUTPUT_DIR = DATA_RAW_DIR
IMPORT_FILE = OUTPUT_DIR / "nories_rod_import.xlsx"
EVIDENCE_FILE = OUTPUT_DIR / "nories_rod_whitelist_player_evidence.json"

PLAYER_OVERRIDES = {
    "LTT620PMH": {
        "player_positioning": "Versatile / Parabolic multi bait",
        "player_selling_points": "多用途拋投和控餌取向明確 / topwater、jig spinner、swimbait、free rig 等中量級路亞切換更自然",
        "guide_use_hint": "Bass 泛用：中量級線組出線穩定，硬餌、swimbait 和軟餌切換時更容易保持拋投節奏與控線角度。",
    },
    "STN680MS": {
        "player_positioning": "PE spin / Structure rig",
        "player_selling_points": "PE + leader 遠投結構攻略取向明確 / cover neko、down shot、jighead wacky，也能支撐 Carolina、Texas、heavy drop shot",
        "guide_use_hint": "PE spin：細 PE + leader 出線和 slack 管理更穩，遠投、深場 structure、Neko / down shot / Carolina 類釣組更容易讀底和補刺。",
    },
    "STN580ML": {
        "player_positioning": "Tsuru-stroll / PE cover finesse",
        "player_selling_points": "PE 2 号級吊るし與 cover finesse 取向清楚 / 小型 swisher、surface bait 和障礙邊定點操作更好控線",
        "guide_use_hint": "Tsuru-stroll：PE 線出線和短距離 pitching 更穩，吊るし水面下誘魚與障礙邊補刺更容易維持角度。",
    },
    "STN650M": {
        "player_positioning": "Neko / Heavy drop shot",
        "player_selling_points": "Neko、heavy drop shot 與輕 cover 精細操作取向明確 / 杭、沈木、障礙邊定點 shake 更容易讀底和補刺",
        "guide_use_hint": "Neko / drop shot：細線出線與 slack 處理更順，定點 shake、咬口判斷和短距離補刺更直接。",
    },
    "STN640MLS-Md": {
        "player_positioning": "Insect / Surface finesse",
        "player_selling_points": "蟲系、水面系與 PE twitch 取向清楚 / 低彈性竿尖讓高速 twitch 後的短咬口更容易留餌",
        "guide_use_hint": "Surface finesse：PE 出線順暢，細小 twitch、虫系停頓和水面咬口後的線路控制更穩。",
    },
    "680JMHS": {
        "player_positioning": "Power finesse / Heavy cover",
        "player_selling_points": "Power finesse 專用感更強 / PE + leader、smolaba、cover pitching 和 heavy cover 抽魚更有支撐",
        "guide_use_hint": "Power finesse：PE 出線和竿身回正更穩，cover 內 shake、短距離 pitching 和強補刺更容易連貫。",
    },
    "700JHS": {
        "player_positioning": "Long power finesse / Heavy cover",
        "player_selling_points": "長尺 power finesse / heavy cover 取向明確 / 遠距吊るし、高處補刺和 monster bass 控魚餘量更高",
        "guide_use_hint": "Long power finesse：PE 出線和長竿線路控制更穩，高處吊るし、遠距補刺和障礙邊控魚更有餘量。",
    },
}

WHITELIST_EVIDENCE = {
    "LTT620PMH": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/search?rod=LTT620PMH",
        "confidence": "medium",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "effective_info": {
            "places": ["二川ダム", "横山池", "田溝池"],
            "lures": ["ハガー", "ジグスピナーSS 10g", "フラテリス 3.5インチ", "デッドスロウル 5インチ", "LFT 145"],
            "lines": ["フロロ 14-16lb", "ナイロン 12lb"],
        },
        "field_decision": "支持改为泛用多路亚，而不是单纯 cover finesse。",
    },
    "STN680MS": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/collection?rod=%E3%83%AD%E3%83%BC%E3%83%89%E3%83%A9%E3%83%B3%E3%83%8A%E3%83%BC+%E3%82%B9%E3%83%88%E3%83%A9%E3%82%AF%E3%83%81%E3%83%A3%E3%83%BC+NXS+STN680MS",
        "confidence": "high",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "effective_info": {
            "places": ["野池", "乙戸沼"],
            "lures": ["野良ネズミ", "ISワスプ", "スグリ 60", "アンクルミノー"],
            "lines": ["PE 0.6-0.8号 + leader", "FC 8lb"],
            "rigs": ["Neko rig", "Carolina rig", "Texas rig", "heavy downshot"],
        },
        "field_decision": "官网说明和白名单共同支持 PE spin / structure rig，原 cover finesse 过窄。",
    },
    "HB6100ML-Gc": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/collection?rod=%E3%83%AD%E3%83%BC%E3%83%89%E3%83%A9%E3%83%B3%E3%83%8A%E3%83%BC+%E3%83%B4%E3%82%A9%E3%82%A4%E3%82%B9+HB6100ML",
        "confidence": "high",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "effective_info": {
            "places": ["榛名湖", "霞ヶ浦・利根川水系", "リザーバー"],
            "lures": ["ショット オーバー 5", "タダマキ 112", "コンプリートフラット 68", "デカピーナッツ II SSR", "コンバットクランク 320", "ラピッズブレード"],
            "lines": ["フロロ 12-16lb"],
        },
        "field_decision": "现有 Hard bait / 巻物搜索 判断成立，不需要额外改写。",
    },
    "STN650LS": {
        "source_site": "rods.jp",
        "source_url": "https://www.rods.jp/brand-item.php?brand2_id=105",
        "confidence": "low",
        "matched_category": "バスルアー / ロッド仕様",
        "target_fish": ["ブラックバス"],
        "effective_info": {
            "specs": ["スピニング", "L", "ファースト", "6′0″", "1.80~7.00g"],
        },
        "field_decision": "只作为规格和淡水 Bass 类别旁证，不直接写玩家文案。",
    },
    "STN610LLS": {
        "source_site": "rods.jp",
        "source_url": "https://www.rods.jp/brand-item.php?brand2_id=105",
        "confidence": "low",
        "matched_category": "バスルアー / ロッド仕様",
        "target_fish": ["ブラックバス"],
        "effective_info": {
            "specs": ["スピニング", "UL", "エクストラファースト", "6′0″", "0.90~5.30g"],
        },
        "field_decision": "只作为规格和淡水 Bass 类别旁证，不直接写玩家文案。",
    },
}


def header_map(ws):
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def shade_detail_groups(ws):
    fill_a = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
    fill_b = PatternFill(fill_type="solid", fgColor="FFE8F1FB")
    headers = header_map(ws)
    rod_col = headers["rod_id"]
    last_rod_id = None
    group = -1
    for row in range(2, ws.max_row + 1):
        rod_id = ws.cell(row=row, column=rod_col).value
        if rod_id != last_rod_id:
            group += 1
            last_rod_id = rod_id
        fill = fill_a if group % 2 == 0 else fill_b
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill


def collect_row_meta(ws_detail):
    headers = header_map(ws_detail)
    meta = {}
    for row in range(2, ws_detail.max_row + 1):
        sku = ws_detail.cell(row=row, column=headers["SKU"]).value
        if not sku:
            continue
        meta[sku] = {
            "rod_id": ws_detail.cell(row=row, column=headers["rod_id"]).value,
            "detail_id": ws_detail.cell(row=row, column=headers["id"]).value,
            "row": row,
            "current_player_fields": {
                "player_environment": ws_detail.cell(row=row, column=headers["player_environment"]).value,
                "player_positioning": ws_detail.cell(row=row, column=headers["player_positioning"]).value,
                "player_selling_points": ws_detail.cell(row=row, column=headers["player_selling_points"]).value,
                "guide_use_hint": ws_detail.cell(row=row, column=headers["guide_use_hint"]).value,
            },
        }
    return meta


def apply_overrides(ws_detail):
    headers = header_map(ws_detail)
    changes = []
    for row in range(2, ws_detail.max_row + 1):
        sku = ws_detail.cell(row=row, column=headers["SKU"]).value
        detail_id = ws_detail.cell(row=row, column=headers["id"]).value
        admin_code = ws_detail.cell(row=row, column=headers["AdminCode"]).value
        if detail_id and not admin_code:
            ws_detail.cell(row=row, column=headers["AdminCode"]).value = detail_id
            changes.append(
                {
                    "sku": sku,
                    "row": row,
                    "changed_fields": [{"field": "AdminCode", "old": admin_code, "new": detail_id}],
                }
            )
        override = PLAYER_OVERRIDES.get(sku)
        if not override:
            continue
        changed_fields = []
        for field, value in override.items():
            col = headers[field]
            old = ws_detail.cell(row=row, column=col).value
            if old != value:
                ws_detail.cell(row=row, column=col).value = value
                changed_fields.append({"field": field, "old": old, "new": value})
        if changed_fields:
            changes.append({"sku": sku, "row": row, "changed_fields": changed_fields})
    return changes


def write_evidence(row_meta, changes):
    changed_by_sku = {item["sku"]: item["changed_fields"] for item in changes}
    evidence = []
    for sku, item in WHITELIST_EVIDENCE.items():
        meta = row_meta.get(sku, {})
        evidence.append(
            {
                "brand": "NORIES",
                "sku": sku,
                "rod_id": meta.get("rod_id", ""),
                "detail_id": meta.get("detail_id", ""),
                "row": meta.get("row", ""),
                "current_player_fields": meta.get("current_player_fields", {}),
                "target_player_fields": PLAYER_OVERRIDES.get(sku, {}),
                **item,
                "applied_changes": changed_by_sku.get(sku, []),
                "write_policy": "白名单只辅助 player_* 和 guide_use_hint，不覆盖官网 official/spec 字段。",
            }
        )
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "import_file": str(IMPORT_FILE),
        "policy": {
            "official_source": "Nories official pages remain authoritative for specs and descriptions.",
            "whitelist_role": "Whitelist sources are used only when SKU/series matches clearly and the result supports player-facing context.",
            "non_write_fields": ["official_environment", "hook_keeper_included", "sweet_spot_lure_weight_real"],
        },
        "source_summaries": [
            {
                "source_site": "tackledb.uosoku.com",
                "useful_for": ["player_environment", "player_positioning", "player_selling_points", "guide_use_hint"],
                "effective_info": ["カテゴリー", "対象魚", "釣り場", "ロッド", "ライン", "ルアー", "インプレ・メモ"],
                "limits": "Contains AI/virtual tackle records; treat as supporting player-context evidence, not official product facts.",
            },
            {
                "source_site": "rods.jp",
                "useful_for": ["spec/category cross-check"],
                "effective_info": ["series category", "type", "power", "action", "length", "lure weight", "price"],
                "limits": "Sparse Nories coverage in this batch; not enough to write player copy directly.",
            },
            {
                "source_site": "rodsearch.com",
                "useful_for": [],
                "effective_info": [],
                "limits": "No stable SKU-level Nories result page was found during this pass.",
            },
        ],
        "evidence": evidence,
        "changes": changes,
    }
    EVIDENCE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main():
    wb = load_workbook(IMPORT_FILE)
    ws_detail = wb["rod_detail"]
    changes = apply_overrides(ws_detail)
    row_meta = collect_row_meta(ws_detail)
    shade_detail_groups(ws_detail)
    wb.save(IMPORT_FILE)
    write_evidence(row_meta, changes)
    print(json.dumps({"changes": len(changes), "evidence": str(EVIDENCE_FILE)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
