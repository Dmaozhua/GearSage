import json
import re
import subprocess
from collections import Counter
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('megabass_rod_import.xlsx')
REPORT_PATH = resolve_data_raw('megabass_rod_player_selling_points_stage12_report.json')
SHADE_SCRIPT = ROOT / "scripts/shade_megabass_rod_detail_groups_stage2.py"
FIELD = "player_selling_points"


ZH = {
    "BFS Jighead Rig": "BFS 轻量铅头",
    "Small Rubber Jig": "小型橡胶铅头钩",
    "Rubber Jig": "橡胶铅头钩",
    "Football Jig": "Football Jig",
    "Jighead Rig": "铅头钩",
    "Light Texas Rig": "轻德州",
    "Heavy Texas Rig": "重德州",
    "Texas Rig": "德州",
    "Free Rig": "Free Rig",
    "Neko Rig": "Neko",
    "Down Shot": "Down Shot",
    "No Sinker": "无铅软饵",
    "Wacky Rig": "Wacky",
    "Nose Hook Rig": "鼻挂软饵",
    "Frog": "Frog",
    "Punching": "Punching",
    "Jerkbait": "Jerkbait",
    "Minnow": "Minnow",
    "Trout Minnow": "鳟鱼 Minnow",
    "Small Minnow": "小型 Minnow",
    "Shad": "Shad",
    "Deep Crankbait": "Deep Crankbait",
    "Crankbait": "Crankbait",
    "Vibration": "Vibration",
    "Spinnerbait": "Spinnerbait",
    "Chatterbait": "Chatterbait",
    "Swim Jig": "Swim Jig",
    "Swimbait": "Swimbait",
    "Big Bait": "Big Bait",
    "Topwater Plug": "水面系 Plug",
    "Pencil Bait": "Pencil",
    "Popper": "Popper",
    "Bug Lure": "虫系水面饵",
    "I-shaped Plug": "I 字系 Plug",
    "Small Plug": "小型 Plug",
    "Spoon": "Spoon",
    "Magnum Spoon": "Magnum Spoon",
    "Metal Jig": "Metal Jig",
    "Shore Plug": "岸投 Plug",
    "Lead Core Trolling": "Lead Core 拖钓",
    "Downrigger Trolling": "Downrigger 拖钓",
}

SOFT = {
    "BFS Jighead Rig", "Small Rubber Jig", "Rubber Jig", "Football Jig", "Jighead Rig",
    "Light Texas Rig", "Heavy Texas Rig", "Texas Rig", "Free Rig", "Neko Rig",
    "Down Shot", "No Sinker", "Wacky Rig", "Nose Hook Rig",
}
MOVING = {"Spinnerbait", "Chatterbait", "Vibration", "Crankbait", "Deep Crankbait", "Shad"}
TWITCH = {"Jerkbait", "Minnow", "Topwater Plug", "Pencil Bait", "Popper", "Bug Lure", "Small Plug"}
POWER = {"Frog", "Punching", "Big Bait", "Swimbait"}
TROUT = {"Trout Minnow", "Spoon", "Small Plug", "Minnow", "Jerkbait"}
TROLLING = {"Lead Core Trolling", "Downrigger Trolling"}

FORBIDDEN = [
    "不发散", "不发虚", "不拖泥带水", "有底气", "余量高", "性能优秀", "适合多场景",
    "官网", "白名单", "tackledb", "来源", "证据", "参考规格", "按型号覆盖", "常规鲈钓",
]


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def split_items(value):
    return [n(item) for item in n(value).split("/") if n(item)]


def zh(label):
    return ZH.get(label, label)


def has_any(items, names):
    return bool(set(items) & set(names))


def is_trout_model(model):
    return bool(re.search(r"greathunting|great hunting", lower(model)))


def is_valkyrie(model):
    return "valkyrie" in lower(model)


def is_travel(model, sku):
    return bool(re.search(r"triza|valkyrie|huntsman|mountain stream|world expedition", lower(model))) or bool(re.search(r"-[234]p?$|-[234][a-z]?$", lower(sku)))


def cues(desc):
    text = lower(desc)
    result = []
    if re.search(r"pinpoint|accuracy|accurate|pitch|tight spots?|close quarters|精准", text):
        result.append("pinpoint")
    if re.search(r"long[- ]distance|distance|moon-shot|遠投|ロングディスタンス|wide|shore", text):
        result.append("distance")
    if re.search(r"fall|line tension|subtle bites?|light bites?|sensitivity|感度|喰わせ", text):
        result.append("sensitivity")
    if re.search(r"glass|グラス|low elasticity|低弾性", text):
        result.append("glass")
    if re.search(r"solid carbon|stinger tip|ソリッド|stinger", text):
        result.append("solid_tip")
    if re.search(r"torque|lifting|power|monster|heavy cover|mats?|vegetation|bush", text):
        result.append("torque")
    return result


def secondary_text(items):
    rest = [zh(item) for item in items[1:4]]
    if not rest:
        return ""
    if len(rest) == 1:
        return rest[0]
    return "、".join(rest)


def prepend_cue_sentence(parts, row, items, model):
    clue = cues(row.get("Description"))
    primary = items[0] if items else ""
    if "pinpoint" in clue and primary not in {"Lead Core Trolling", "Downrigger Trolling"}:
        parts.append("落点控制和低角度送饵便于贴近结构边缘操作")
    if "distance" in clue:
        if is_trout_model(model):
            parts.append("远投后的线弧管理和漂流角度更便于维持")
        else:
            parts.append("长距离抛投后仍能维持清晰的收线节奏")
    if "sensitivity" in clue and not has_any(items, TROLLING):
        parts.append("落下段线张力变化和轻触底信号有助于分辨")
    if "glass" in clue:
        parts.append("低弹性追随感有利于降低硬饵短咬脱钩率")
    if "solid_tip" in clue:
        parts.append("实心竿梢有助于放大小饵吸入和轻触底反馈")
    if "torque" in clue and not has_any(items, TROLLING):
        parts.append("中后段支撑力能稳定承接负载变化")


def selling_points(row, model):
    items = split_items(row.get("recommended_rig_pairing"))
    if not items:
        return n(row.get(FIELD))
    primary = items[0]
    secondary = secondary_text(items)
    parts = []

    if has_any(items, TROLLING):
        if primary == "Downrigger Trolling":
            parts.extend([
                "Downrigger 深水线组拖曳时竿身负载变化更线性",
                "长时间巡航中控线稳定，咬口判断清楚",
                "面向湖泊大型鳟鱼的分层巡航搜索",
            ])
        else:
            parts.extend([
                "Lead Core 线组的持续阻力承接更稳定",
                "长线距离下仍保留足够的控线反馈",
                "面向开阔湖面按速度和深度做巡航搜索",
            ])
    elif is_trout_model(model):
        if re.search(r"river&lake|river|lake|本流|湖", lower(model) + " " + lower(row.get("Description"))):
            parts.extend([
                f"{zh(primary)}和{secondary or 'Spoon'}的远投搜索边界清晰",
                "本流漂流、湖泊回收和大鱼转向时控线稳定",
                "服务于覆盖距离和维持泳层的鳟鱼场景",
            ])
        elif "x-glass" in lower(model) or "glass" in lower(row.get("Description")):
            parts.extend([
                "玻璃复合调性让小型硬饵咬口衔接更自然",
                f"{zh(primary)}的短距抛投和连续 twitch 节奏更便于维持",
                "溪流近中距离控饵时容错更高",
            ])
        else:
            parts.extend([
                f"{zh(primary)}的连续 twitch 和贴流控线反馈直接",
                "短尺或多节设定利于移动钓行中的精准落点",
                "面对急流压线和小范围转向时反馈清楚",
            ])
    elif is_valkyrie(model):
        if primary in {"Shore Plug", "Metal Jig"}:
            parts.extend([
                "岸投 Plug 与轻型金属饵之间的负荷衔接更明确",
                "港湾、河口和开放水面搜索时抛投距离可控",
                "多节远征结构便于跨淡水和轻型海水场景切换",
            ])
        elif primary in {"Big Bait", "Swimbait"}:
            parts.extend([
                "100g 级大型饵抛投和回收负荷承接稳定",
                "短距离船边结构或开阔水域中能保持足够控鱼支撑",
                "多节结构同时服务远行携带和高负荷实钓",
            ])
        elif primary == "Small Plug":
            parts.extend([
                "小型 Plug 的低负荷启动和近距落点控制细腻",
                f"可覆盖{secondary or '水面系 Plug'}等轻量硬饵搜索",
                "多鱼种远征中对轻量饵和突发大鱼都有余裕",
            ])
        else:
            parts.extend([
                "低弹性复合空白让硬饵负荷变化更平顺",
                f"{zh(primary)}与{secondary or '同级搜索饵'}的节奏切换边界明确",
                "多节远征场景中保留携带性和实钓强度",
            ])
    elif primary in {"Topwater Plug", "Pencil Bait", "Popper", "Bug Lure"}:
        parts.extend([
            f"{zh(primary)}走姿、停顿和短抽动作便于做细",
            f"可覆盖{secondary or '轻量硬饵'}，面向浅场连续搜索",
            "远投后的线弧控制有助于维持水面饵动作完整度",
        ])
    elif primary in {"Jerkbait", "Minnow"}:
        parts.extend([
            f"{zh(primary)}抽停后的回弹和停顿线弧控制清楚",
            f"搭配{secondary or '同级硬饵'}时节奏变化更清晰",
            "服务于连续控饵和反复搜索的浅中层场景",
        ])
    elif primary == "Small Plug":
        parts.extend([
            "小型 Plug 在低负荷下仍能保持稳定抛投和启动",
            f"辅以{secondary or '轻量软饵'}，不牺牲硬饵搜索效率",
            "近中距离控饵和轻障碍边缘搜索更有针对性",
        ])
    elif primary in {"Spinnerbait", "Chatterbait", "Vibration"}:
        if has_any(items, SOFT):
            parts.extend([
                f"{zh(primary)}搜索与{secondary or '软饵'}补充之间切换明确",
                "草边和浅障碍触碰后的线张力变化处理更从容",
                "可先以移动饵快速找鱼，再用 contact bait 精细补位",
            ])
        else:
            parts.extend([
                f"{zh(primary)}连续搜索时泳层和速度稳定",
                f"搭配{secondary or '同级移动饵'}能覆盖浅滩、草边和开阔水域",
                "长时间平收下负荷感更易管理",
            ])
    elif primary in {"Deep Crankbait", "Crankbait", "Shad"}:
        if has_any(items, SOFT):
            parts.extend([
                f"{zh(primary)}的卷收负荷与{secondary or '底操软饵'}形成互补",
                "硬饵触底、碰障和软饵补位之间衔接更清楚",
                "用于深浅结构交错区域的搜索和确认",
            ])
        else:
            parts.extend([
                f"{zh(primary)}卷收负荷承接平顺，泳层维持稳定",
                f"搭配{secondary or '同级 crank'}时搜索节奏保持连贯",
                "长时间抛投搜索中手部负担更可控",
            ])
    elif primary in {"Frog", "Punching"}:
        parts.extend([
            f"{zh(primary)}在草洞、草垫和重障碍中的刺鱼支撑明确",
            f"辅以{secondary or '重障碍软饵'}，覆盖打点和强制控鱼",
            "粗线高负荷下竿身稳定性和导线通过性更关键",
        ])
    elif primary in {"Big Bait", "Swimbait"}:
        parts.extend([
            f"{zh(primary)}抛投、回收和停顿阶段的负荷承接稳定",
            f"搭配{secondary or '大型移动饵'}时可覆盖开阔水域和结构边缘搜索",
            "中鱼后能更平顺地处理大饵惯性和鱼体冲击",
        ])
    elif primary in {"Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Rubber Jig", "Football Jig"}:
        parts.extend([
            f"{zh(primary)}触底、越障和短距离跳动反馈清楚",
            f"辅以{secondary or '底操软饵'}，用于木桩、岩石和草边慢速确认",
            "扬竿刺鱼和贴障碍控线更直接",
        ])
    elif primary in {"BFS Jighead Rig", "Small Rubber Jig", "No Sinker"}:
        parts.extend([
            f"{zh(primary)}低负荷抛投和落点控制更细腻",
            f"辅以{secondary or '轻量软饵'}，用于轻障碍边缘诱食",
            "落下段线张力与小口吸入反馈便于读取",
        ])
    elif primary in {"Down Shot", "Neko Rig", "Jighead Rig", "Wacky Rig", "Nose Hook Rig"}:
        parts.extend([
            f"{zh(primary)}慢速操作中的触底和轻咬信号清楚",
            f"搭配{secondary or '轻线软饵'}时覆盖开放水域和轻障碍边缘",
            "细线控线与小幅抖动动作便于保持一致",
        ])
    else:
        parts.extend([
            f"{zh(primary)}作为主线时用途边界清楚",
            f"可按{secondary or '相近饵型'}做补充搭配",
            "可根据水层、障碍和搜索节奏进行切换",
        ])

    prepend_cue_sentence(parts, row, items, model)

    deduped = []
    for part in parts:
        part = n(part)
        if part and part not in deduped:
            deduped.append(part)
    # Keep the field readable; four clauses is enough for audit and user display.
    return "；".join(deduped[:4]) + "。"


def main():
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_col = {n(cell.value): idx + 1 for idx, cell in enumerate(rod_ws[1])}
    detail_col = {n(cell.value): idx + 1 for idx, cell in enumerate(detail_ws[1])}
    required = ["id", "rod_id", "SKU", "Description", "recommended_rig_pairing", FIELD]
    missing = [field for field in required if field not in detail_col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")
    model_by_rod_id = {
        n(rod_ws.cell(row_idx, rod_col["id"]).value): n(rod_ws.cell(row_idx, rod_col["model"]).value)
        for row_idx in range(2, rod_ws.max_row + 1)
    }

    changes = []
    values = Counter()
    for row_idx in range(2, detail_ws.max_row + 1):
        row = {field: detail_ws.cell(row_idx, col_idx).value for field, col_idx in detail_col.items()}
        model = model_by_rod_id.get(n(row.get("rod_id")), "")
        next_value = selling_points(row, model)
        values[next_value] += 1
        cell = detail_ws.cell(row_idx, detail_col[FIELD])
        old = n(cell.value)
        if old != next_value:
            cell.value = next_value
            changes.append({
                "row": row_idx,
                "id": n(row.get("id")),
                "rod_id": n(row.get("rod_id")),
                "sku": n(row.get("SKU")),
                "model": model,
                "old_value": old,
                "new_value": next_value,
            })

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "xlsx_file": str(XLSX_PATH),
        "field": FIELD,
        "changed_rows": len(changes),
        "coverage": f"{sum(bool(v) for v in values.elements())}/{detail_ws.max_row - 1}",
        "unique_count": len(values),
        "top_repeats": [{"value": value, "count": count} for value, count in values.most_common(12)],
        "forbidden_terms": FORBIDDEN,
        "changes": changes,
        "policy": {
            "write_scope": "Only rod_detail.player_selling_points was written.",
            "style": "Professional user-facing clauses; avoid slang and source-chain wording.",
        },
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "changed_rows": len(changes),
        "coverage": report["coverage"],
        "unique_count": len(values),
        "report": str(REPORT_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
