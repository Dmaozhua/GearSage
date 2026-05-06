from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')

HINT_BY_ID = {
    "DRD10031": "BF 輕量多用途：No Sinker、Neko、Jighead 與小型硬餌切換時，低彈道拋投、控線和小餌操作更穩。",
    "DRD10033": "輕量軟硬餌：Soft Plastic、Topwater、小中型硬餌切換時，精細拋投、障礙邊控線和誘魚節奏更穩。",
    "DRD10035": "全能泛用：Crankbait、Spinnerbait、Down Shot、Texas 和高比重軟蟲切換時，拋投、控線與回收節奏更穩。",
    "DRD10037": "強力多用途：Texas、Jighead 打點和 Swim Jig、Spinnerbait 平收都能兼顧，遠投控線與刺魚餘量更穩。",
    "DRD10046": "BF 輕量多用途：No Sinker、Neko、Jighead 與小型硬餌切換時，低彈道拋投、控線和小餌操作更穩。",
    "DRD10048": "輕量軟硬餌：Soft Plastic、Topwater、小中型硬餌切換時，精細拋投、障礙邊控線和誘魚節奏更穩。",
    "DRD10050": "全能泛用：Crankbait、Spinnerbait、Down Shot、Texas 和高比重軟蟲切換時，拋投、控線與回收節奏更穩。",
    "DRD10061": "操作型硬餌兼精細：Jerkbait、Spinnerbait、Topwater 的動作回饋更清楚，也能兼顧 Neko、小型 Rubber Jig 控線。",
    "DRD10084": "輕量多用途：No Sinker、Down Shot、Light Texas 和小型 Crankbait 切換時，細線出線、控線與小餌節奏更穩。",
    "DRD10085": "輕量多用途：No Sinker、Down Shot、Light Texas 和小型 Crankbait 切換時，細線出線、控線與小餌節奏更穩。",
    "DRD10087": "中量泛用：Texas、Free Rig、Rubber Jig 與 Spinnerbait、Chatterbait 切換時，打點和移動餌回收節奏更穩。",
    "DRD10088": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10089": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10110": "高感輕量泛用：No Sinker、Down Shot、Light Texas 與小型 Crankbait 切換時，實心竿尖控線和輕咬判斷更穩。",
    "DRD10112": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10117": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10118": "高感強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收切換時，實心竿尖回饋與控魚更穩。",
    "DRD10121": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10122": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10123": "高感強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收切換時，實心竿尖回饋與控魚更穩。",
    "DRD10124": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10297": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10299": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10145": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10146": "中量泛用：Texas、Free Rig、Rubber Jig 與 Spinnerbait、Chatterbait 切換時，打點和移動餌回收節奏更穩。",
    "DRD10147": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10148": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10150": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10151": "強力泛用：Texas、Rubber Jig 打點和 Swim Jig、Spinnerbait 平收都能兼顧，障礙區控線與刺魚更有餘量。",
    "DRD10155": "輕量多用途：No Sinker、Down Shot、Light Texas 和小型 Crankbait 切換時，細線出線、控線與小餌節奏更穩。",
}


def n(value):
    return " ".join(str(value or "").split())


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["guide_use_hint"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        if detail_id not in HINT_BY_ID:
            continue
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        new = HINT_BY_ID[detail_id]
        if old != new:
            cell.value = new
            changed.append({"row": row, "id": detail_id, "old": old, "new": new})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-guide_use_hint changes")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed": len(changed),
        "changed_ids": [item["id"] for item in changed],
    })


if __name__ == "__main__":
    main()
