#!/usr/bin/env python3
import hashlib
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from bs4 import BeautifulSoup
from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
NORMALIZED_PATH = DATA_DIR / "evergreen_rod_normalized.json"
CACHE_DIR = DATA_DIR / "evergreen_rod_product_technical_cache"
REPORT_PATH = DATA_DIR / "evergreen_rod_product_technical_report.json"
EVIDENCE_PATH = DATA_DIR / "evergreen_rod_product_technical_official_evidence.json"
FIELD = "product_technical"
DELIMITER = " / "


def norm(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def headers(ws):
    return [cell.value for cell in ws[1]]


def col_map(ws):
    return {name: idx + 1 for idx, name in enumerate(headers(ws))}


def sheet_values_without_field(ws):
    h = headers(ws)
    cols = [idx + 1 for idx, name in enumerate(h) if name != FIELD]
    return [
        tuple(ws.cell(row=row, column=col).value for col in cols)
        for row in range(1, ws.max_row + 1)
    ]


def cache_path(url):
    return CACHE_DIR / f"{hashlib.sha1(url.encode('utf-8')).hexdigest()}.html"


def page_text(item):
    url = item.get("source_url") or ""
    path = cache_path(url)
    if path.exists() and path.stat().st_size > 1000:
        soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
        feature_blocks = soup.select("li.item-feature")
        parts = []
        title_area = soup.select_one(".titleArea")
        if title_area:
            parts.append(title_area.get_text(" ", strip=True))
        if feature_blocks:
            parts.extend(block.get_text(" ", strip=True) for block in feature_blocks)
            source_mode = "official_html_title_and_feature_blocks"
        else:
            box = soup.select_one(".productsBox") or soup.body
            if box:
                parts.append(box.get_text(" ", strip=True))
            source_mode = "official_html_product_page"
        return norm("\n".join(parts)), source_mode
    return norm(item.get("description") or ""), "normalized_official_description"


def merge_terms(terms):
    normalized_terms = [norm(term).strip(" /、。，.・") for term in terms]
    term_set = {term for term in normalized_terms if term}
    merged = []
    for value in normalized_terms:
        if value and value not in merged:
            if value == "Kガイド" and any(term != value and "Kガイド" in term for term in term_set):
                continue
            if value == "LGトップガイド" and "チタンLGトップ" in term_set:
                continue
            if value == "セミマイクロKガイド" and "スピニングセミマイクロKガイド" in term_set:
                continue
            if value == "トルザイトリングガイド" and any("トルザイト" in term and term != value for term in term_set):
                continue
            if value == "オールトルザイトリング" and "オールチタンフレーム・オールトルザイトリングガイド" in term_set:
                continue
            if value == "カーボンパイプ" and "カーボンパイプスペーサー" in term_set:
                continue
            if value == "チタンフレームLKWガイド" and "チタンフレーム・トルザイトリングLKWガイド" in term_set:
                continue
            if value == "チタンフレームSiCガイド" and any(term in term_set for term in [
                "Fuji チタンフレームSiCリングKガイド",
                "チタンフレームSiCリングKガイド",
            ]):
                continue
            if value == "ステンレスフレームSiCガイド" and any(term in term_set for term in [
                "Fuji ステンレスフレームKガイド",
                "Fuji ステンレスフレームSiCリング・シングルフットKガイド",
                "ステンレスフレームSiCリングKガイド",
            ]):
                continue
            if value == "ヘラクレスクロス製法" and "4軸ヘラクレスクロス" in term_set:
                continue
            if value == "4軸クロス" and any(term in term_set for term in [
                "4軸ヘラクレスクロス",
                "±45°オリジナル4軸クロス",
                "±30°狭角オリジナル4軸クロス",
            ]):
                continue
            if (
                value == "チタンフレームSiCリング・ダブルフットガイド"
                and "Fuji チタンフレームSiCリング・ダブルフットガイド" in term_set
            ):
                continue
            merged.append(value)
    return DELIMITER.join(merged)


def add_if(pattern, text, terms, value, flags=re.I):
    if re.search(pattern, text, flags):
        terms.append(value)


def extract_product_technical(text):
    terms = []

    add_if(r"トレカ\s*®?\s*[「『]?\s*M40X", text, terms, "トレカ®M40X")
    add_if(r"トレカ\s*®?\s*[「『]?\s*T1100G", text, terms, "トレカ®T1100G")
    add_if(r"ナノアロイ\s*®?\s*技術", text, terms, "ナノアロイ®技術")

    for ton in ["50", "46", "40", "33", "30", "24", "10"]:
        add_if(rf"{ton}\s*(?:トン|[tｔTＴ])カーボン", text, terms, f"{ton}トンカーボン")

    add_if(r"超極薄カーボンプリプレグ", text, terms, "超極薄カーボンプリプレグ")
    add_if(r"高レジンカーボン", text, terms, "高レジンカーボン")
    add_if(r"低レジンピュアカーボン", text, terms, "低レジンピュアカーボン")
    add_if(r"低レジン素材", text, terms, "低レジン素材")
    add_if(r"高強度・高弾性率炭素繊維", text, terms, "高強度・高弾性率炭素繊維")
    add_if(r"超高弾性カーボン", text, terms, "超高弾性カーボン")
    add_if(r"(?<!超)高弾性カーボン", text, terms, "高弾性カーボン")
    add_if(r"中弾性(?:率)?(?:24トン)?カーボン", text, terms, "中弾性カーボン")
    add_if(r"低弾性(?:率)?カーボン|Low\s*Modulus", text, terms, "低弾性カーボン")
    add_if(r"グラス(?:コンポジット|素材|ブランク)", text, terms, "グラスコンポジット")

    add_if(r"(?:±30°)?4軸補強", text, terms, "4軸補強")
    add_if(r"4軸ヘラクレスクロス", text, terms, "4軸ヘラクレスクロス")
    add_if(r"ヘラクレスクロス製法", text, terms, "ヘラクレスクロス製法")
    add_if(r"±45°オリジナル4軸クロス", text, terms, "±45°オリジナル4軸クロス")
    add_if(r"±30°狭角オリジナル4軸クロス", text, terms, "±30°狭角オリジナル4軸クロス")
    add_if(r"4軸クロス", text, terms, "4軸クロス")
    add_if(r"カレイド[・\s]*スーパークワトロクロス", text, terms, "カレイド・スーパークワトロクロス")
    add_if(r"4軸製法", text, terms, "4軸製法")
    add_if(r"4軸カーボンスリーブナット", text, terms, "4軸カーボンスリーブナット")
    add_if(r"4軸カーボン", text, terms, "4軸カーボン")

    add_if(r"ソリッドティップ|カーボンソリッド", text, terms, "カーボンソリッドティップ")
    add_if(r"チューブラーティップ|フルチューブラー", text, terms, "チューブラー構造")
    add_if(r"オールチューブラー設計", text, terms, "オールチューブラー設計")
    add_if(r"完全ブランクスルー構造", text, terms, "完全ブランクスルー構造")
    add_if(r"アンサンドフィニッシュ", text, terms, "アンサンドフィニッシュ")
    add_if(r"ウーブンクロス補強", text, terms, "ウーブンクロス補強")
    add_if(r"高品位スピゴットジョイント|スピゴットジョイント", text, terms, "高品位スピゴットジョイント")
    add_if(r"スレッドレス仕様", text, terms, "スレッドレス仕様")

    guide_patterns = [
        (r"Fuji[・\s]*Kガイド[（(]トルザイトリング[・\s]*(?:チタンフレーム)?", "Fuji チタンフレーム・トルザイトリングKガイド"),
        (r"Fuji[・\s]*チタンフレーム[・\s]*トルザイトリングK?ガイド", "Fuji チタンフレーム・トルザイトリングガイド"),
        (r"Fuji[・\s]*チタンフレームSiCリングKガイド", "Fuji チタンフレームSiCリングKガイド"),
        (r"チタンフレームSiCリングKガイド", "チタンフレームSiCリングKガイド"),
        (r"Fujiステンレスフレーム\s*Kガイド", "Fuji ステンレスフレームKガイド"),
        (r"Fuji\s*ステンレスフレーム\s*Kガイド", "Fuji ステンレスフレームKガイド"),
        (r"SiCリング[・\s]*ステンレスフレーム[・\s]*シングルフットのFuji\s*Kガイド", "Fuji ステンレスフレームSiCリング・シングルフットKガイド"),
        (r"Kガイドをオリジナルセッティング。?高強度ステンレスフレーム、?SiCリング", "ステンレスフレームSiCリングKガイド"),
        (r"ステンレスフレームSiCリングKガイド|オールステンレスフレームSiCリングKガイド", "ステンレスフレームSiCリングKガイド"),
        (r"高強度ステンレスフレーム、?SiCリング", "ステンレスフレームSiCガイド"),
        (r"セミマイクロKガイド|セミマイクロガイドセッティング", "セミマイクロKガイド"),
        (r"スピニングセミマイクロKガイド", "スピニングセミマイクロKガイド"),
        (r"ハイフットKガイド", "ハイフットKガイド"),
        (r"ダブルフットKガイド", "ダブルフットKガイド"),
        (r"オールダブルフットKガイド", "オールダブルフットKガイド"),
        (r"チタンKガイド", "チタンKガイド"),
        (r"Kガイドをオリジナルセッティング|Kガイド採用|Kガイドを採用|Kガイドセッティング", "Kガイド"),
        (r"チタンフレーム[・の、 ]*トルザイトリング[・の、 ]*ダブルフット(?:タイプ)?ガイド", "チタンフレーム・トルザイトリング・ダブルフットガイド"),
        (r"チタンフレーム[・の、 ]*トルザイトリング[・の、 ]*(?:強化)?シングルフット(?:タイプ)?(?:・)?(?:トルザイトリング)?ガイド", "チタンフレーム・トルザイトリング・シングルフットガイド"),
        (r"トルザイトリング[・の、 ]*チタンフレーム(?:強化)?シングルフット(?:タイプ)?(?:・)?(?:トルザイトリング)?ガイド", "チタンフレーム・トルザイトリング・シングルフットガイド"),
        (r"チタンフレームLKWガイド[（(]トルザイトリング[）)]|トルザイトリングLKWガイド|LKWガイド[（(]トルザイトリング[）)]", "チタンフレーム・トルザイトリングLKWガイド"),
        (r"LKWガイド[（(]SiCリング[）)]", "LKWガイド・SiCリング"),
        (r"チタンフレームLKWガイド", "チタンフレームLKWガイド"),
        (r"チタンフレームLNガイド", "チタンフレームLNガイド"),
        (r"ZIGZAGガイドシステム|ジグザグガイドシステム", "ZIGZAGガイドシステム"),
        (r"オールチタンフレーム[・\s]*オールトルザイトリング", "オールチタンフレーム・オールトルザイトリングガイド"),
        (r"チタンLGトップ", "チタンLGトップ"),
        (r"LGトップガイド", "LGトップガイド"),
        (r"チタンKGトップ", "チタンKGトップ"),
        (r"RVガイド", "RVガイド"),
        (r"ステンレスフレーム[・の、 ]*ダブルフット[・の、 ]*SiCリングガイド", "ステンレスフレーム・ダブルフット・SiCリングガイド"),
        (r"ステンレスフレーム[・の、 ]*SiCリング[・の、 ]*ダブルフット(?:タイプ)?ガイド", "ステンレスフレーム・ダブルフット・SiCリングガイド"),
        (r"ステンレスフレーム[・の、 ]*SiC(?:リング)?ガイド", "ステンレスフレームSiCガイド"),
        (r"チタンフレーム[・の、 ]*SiC(?:リング)?ガイド", "チタンフレームSiCガイド"),
        (r"Fuji\s*チタンフレームSiCリングダブルフットガイド", "Fuji チタンフレームSiCリング・ダブルフットガイド"),
        (r"チタンフレームSiCリングダブルフットガイド", "チタンフレームSiCリング・ダブルフットガイド"),
        (r"オールステンレスフレームSiCリング", "オールステンレスフレームSiCリング"),
        (r"ステンレス強化シングルフットフレームSiCリング", "ステンレス強化シングルフットフレームSiCリング"),
        (r"ステンレスダブルフットフレームSiCリング", "ステンレスダブルフットフレームSiCリング"),
        (r"チタンフレーム[・の、 ]*シングルフット[・の、 ]*トルザイトリング", "チタンフレーム・シングルフット・トルザイトリング"),
        (r"ステン(?:レス)?フレーム[・の、 ]*ハイフット[・の、 ]*SiCリング", "ステンレスフレーム・ハイフット・SiCリング"),
        (r"LRVガイド", "LRVガイド"),
        (r"KTガイド", "KTガイド"),
        (r"MNガイド", "MNガイド"),
        (r"SiC-S", "SiC-S"),
        (r"SiC-J", "SiC-J"),
        (r"トルザイトリングガイド", "トルザイトリングガイド"),
        (r"トルザイトリングを全ガイド|全ガイドにトルザイトリング|オールトルザイトリング", "オールトルザイトリング"),
        (r"オールダブルフットガイド", "オールダブルフットガイド"),
    ]
    for pattern, value in guide_patterns:
        add_if(pattern, text, terms, value)

    for seat in re.findall(r"Fuji[・\s]*[A-Z0-9-]+\s*リールシート", text, flags=re.I):
        terms.append(re.sub(r"Fuji[・\s]*([A-Z0-9-]+)\s*リールシート", r"Fuji \1リールシート", norm(seat), flags=re.I))
    for seat in re.findall(r"富士工業製[A-Z0-9-]+\s*(?:リール)?シート", text, flags=re.I):
        terms.append(norm(seat).replace(" ", ""))

    add_if(r"Fuji\s*T-DPS22", text, terms, "Fuji T-DPS22")
    add_if(r"(?<!Fuji[・\s])ECSリールシート", text, terms, "ECSリールシート")
    add_if(r"EGスーパーバスシート・スリム", text, terms, "EGスーパーバスシート・スリム")
    add_if(r"エバーグリーンオリジナルアーバー", text, terms, "エバーグリーンオリジナルアーバー")
    add_if(r"カーボンスリーブ・フォアグリップ", text, terms, "カーボンスリーブ・フォアグリップ")
    add_if(r"カーボンパイプスペーサー", text, terms, "カーボンパイプスペーサー")
    add_if(r"セパレートとエンドグリップをつなぐカーボンパイプ|カーボンパイプ", text, terms, "カーボンパイプ")
    add_if(r"AAAシャンパンコルク", text, terms, "AAAシャンパンコルク")
    add_if(r"ブランクタッチ方式", text, terms, "ブランクタッチ方式")

    return merge_terms(terms)


def main():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    values = {}
    evidence = []
    for idx, item in enumerate(data):
        rod_id = f"ER{1000 + idx}"
        sku = norm((item.get("variants") or [{}])[0].get("sku") or "")
        text, source_mode = page_text(item)
        value = extract_product_technical(text)
        values[(rod_id, sku)] = value
        evidence.append({
            "row_key": {"rod_id": rod_id, "sku": sku},
            "model": item.get("model"),
            "source_url": item.get("source_url"),
            "source_mode": source_mode,
            FIELD: value,
        })

    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    before_rod = sheet_values_without_field(rod_ws)
    before_detail = sheet_values_without_field(detail_ws)

    rod_headers = headers(rod_ws)
    if FIELD in rod_headers:
        raise RuntimeError("evergreen rod master unexpectedly contains product_technical")
    detail_headers = headers(detail_ws)
    if FIELD not in detail_headers:
        raise RuntimeError("evergreen rod_detail missing product_technical")
    idx = detail_headers.index(FIELD)
    if idx == 0 or detail_headers[idx - 1] != "Description":
        raise RuntimeError("product_technical must be after Description")
    if idx + 1 >= len(detail_headers) or detail_headers[idx + 1] != "Extra Spec 1":
        raise RuntimeError("product_technical must be before Extra Spec 1")

    c = col_map(detail_ws)
    tech_col = c[FIELD]
    changes = []
    missing_keys = []
    for row_idx in range(2, detail_ws.max_row + 1):
        key = (
            norm(detail_ws.cell(row=row_idx, column=c["rod_id"]).value),
            norm(detail_ws.cell(row=row_idx, column=c["SKU"]).value),
        )
        if key not in values:
            missing_keys.append({"row": row_idx, "key": key})
            value = ""
        else:
            value = values[key]
        old = norm(detail_ws.cell(row=row_idx, column=tech_col).value)
        if old != value:
            detail_ws.cell(row=row_idx, column=tech_col).value = value
            changes.append({"row": row_idx, "rod_id": key[0], "sku": key[1], "old": old, "new": value})

    if sheet_values_without_field(rod_ws) != before_rod:
        raise RuntimeError("rod sheet changed outside product_technical")
    if sheet_values_without_field(detail_ws) != before_detail:
        raise RuntimeError("rod_detail changed outside product_technical")
    bad_delimiter = [
        change for change in changes
        if "/" in change["new"] and DELIMITER not in change["new"]
    ]
    if missing_keys or bad_delimiter:
        raise RuntimeError({"missing_keys": missing_keys[:5], "bad_delimiter": bad_delimiter[:5]})

    wb.save(XLSX_PATH)
    shade_script = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"
    if shade_script.exists():
        import subprocess
        subprocess.run(["python3", str(shade_script)], check=True)

    nonempty = [row[FIELD] for row in evidence if row[FIELD]]
    report = {
        "field": FIELD,
        "scope": "evergreen_rod_import.xlsx rod_detail.product_technical only",
        "source_policy": "Evergreen official product pages only; one source_url per local child SKU; no whitelist/player/spec inference",
        "delimiter": DELIMITER,
        "updated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "rows": len(evidence),
        "changed_cells": len(changes),
        "nonempty": len(nonempty),
        "blank": len(evidence) - len(nonempty),
        "unique_values": len(set(nonempty)),
        "source_modes": dict(Counter(row["source_mode"] for row in evidence)),
        "top_values": Counter(nonempty).most_common(30),
        "blank_examples": [row for row in evidence if not row[FIELD]][:30],
        "changed_samples": changes[:30],
    }
    EVIDENCE_PATH.write_text(json.dumps({"items": evidence}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
