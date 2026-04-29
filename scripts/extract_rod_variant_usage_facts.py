import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_XLSX = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")
DEFAULT_OUTPUT = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_variant_usage_facts.json")

RIG_PATTERNS = [
    ("Down Shot", [r"down\s*shot", "倒吊", "落下釣組"]),
    ("Neko Rig", [r"\bneko\b", "ネコ", "貓釣組"]),
    ("No Sinker", [r"no\s*sinker", "無鉛"]),
    ("Light Texas", ["light texas", "輕德州"]),
    ("Texas Rig", [r"\btexas\b", "德州"]),
    ("Free Rig", [r"free\s*rig", "自由釣組"]),
    ("Leaderless Down Shot", [r"leaderless\s*down\s*shot"]),
    ("Heavy Down Shot", [r"heavy\s*down\s*shot", "重型.*倒吊"]),
    ("Rubber Jig", [r"rubber\s*jig", "膠裙", "橡膠鉛頭鉤", "軟膠鉛頭鉤", r"\bjig\b"]),
    ("Small Rubber Jig", [r"small\s*rubber", "小型.*rubber", "小型.*軟膠鉛頭鉤", "小型.*橡膠鉛頭鉤"]),
    ("Jighead Rig", [r"jig\s*head", r"jighead", "汲鉤頭", r"(?<!刀片)(?<!刀片式)(?<!橡膠)(?<!軟膠)(?<!泳餌)(?<!無)鉛頭鉤"]),
    ("Wacky Rig", [r"wacky"]),
    ("Frog", [r"\bfrog\b"]),
    ("Punching", [r"\bpunch", "重障礙"]),
    ("Swimbait", [r"swim\s*bait", r"swimbait", "泳餌"]),
    ("Big Bait", [r"big\s*bait", "大型餌", "大餌"]),
    ("Swim Jig", [r"swim\s*jig", "泳餌鉛頭鉤", "泳鉛鉤"]),
    ("Spinnerbait", [r"spinnerbait", "旋轉亮片", "複合式亮片", "復合式亮片", "鋼絲餌"]),
    ("Chatterbait", [r"chatter", "bladed jig", "顫泳", "刀片鉛頭鉤", "刀片式鉛頭鉤"]),
    ("Crankbait", [r"crankbait", r"\bcrank\b"]),
    ("Shad", [r"\bshad\b"]),
    ("Minnow", [r"minnow", "米諾"]),
    ("Jerkbait", [r"jerkbait", r"\bjerk\b", "急抽"]),
    ("Vibration", [r"vibration", r"\bvib\b"]),
    ("Metal Vibration", [r"metal\s*vib"]),
    ("Topwater Plug", [r"topwater", "表層", "水面"]),
    ("I-shaped Plug", [r"i字", r"i-shaped"]),
    ("Small Hardbait", ["小型硬餌", "中小型硬餌"]),
    ("Plug", [r"\bplug\b", "栓型路亞"]),
    ("Soft Plastic", ["軟蟲", "軟餌", r"\bworm\b", r"soft plastic"]),
    ("High-density Soft Plastic", ["高比重軟蟲"]),
    ("Eging", ["木蝦", r"\beging\b", "エギ"]),
    ("Tip-run Eging", [r"tip[- ]?run", "船釣木蝦"]),
    ("Sinker Rig", ["木蝦\\+鉛", "仮面シンカー"]),
    ("SLJ", [r"\bslj\b", "超級輕量", "超輕", "輕量鐵板"]),
    ("SLSJ", [r"\bslsj\b"]),
    ("Light Jigging", ["light jigging", "輕鐵"]),
    ("Slow Jigging", ["slow jigging", "慢鐵"]),
    ("Offshore Jigging", ["offshore jigging"]),
    ("Metal Jig", ["metal jig", "金屬jig", "金屬 jig", "小型鐵板", "鐵板路亞"]),
    ("WIND", [r"\bwind\b", "白帶魚"]),
]

SOFT_RIGS = {
    "Down Shot", "Neko Rig", "No Sinker", "Light Texas", "Texas Rig", "Free Rig",
    "Leaderless Down Shot", "Heavy Down Shot", "Rubber Jig", "Small Rubber Jig",
    "Jighead Rig", "Wacky Rig", "Soft Plastic", "High-density Soft Plastic",
}
HARD_RIGS = {
    "Crankbait", "Shad", "Minnow", "Jerkbait", "Vibration", "Metal Vibration",
    "Topwater Plug", "I-shaped Plug", "Small Hardbait", "Plug", "Spinnerbait",
    "Chatterbait", "Swim Jig", "Swimbait", "Big Bait",
}
POWER_RIGS = {"Frog", "Punching", "Big Bait", "Swimbait", "Heavy Texas"}
SALT_RIGS = {"Eging", "Tip-run Eging", "Sinker Rig", "SLJ", "SLSJ", "Light Jigging", "Slow Jigging", "Offshore Jigging", "Metal Jig", "WIND"}


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def match_positions(text, patterns):
    positions = []
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            positions.append(match.start())
    return positions


def append_unique(items, value):
    if value and value not in items:
        items.append(value)


def split_pairing(value):
    return [n(part) for part in n(value).split("/") if n(part)]


def extract_rigs_from_text(text):
    value = lower(text)
    found = []
    for label, patterns in RIG_PATTERNS:
        positions = match_positions(value, patterns)
        if positions:
            found.append((min(positions), label))
    return [label for _, label in sorted(found)]


def infer_versatility(text, rigs):
    value = lower(text)
    soft = any(rig in SOFT_RIGS for rig in rigs)
    hard = any(rig in HARD_RIGS for rig in rigs)
    if re.search("泛用|多用途|兼具|廣泛|全能|萬用|versatile", value, re.I):
        return "versatile"
    if soft and hard:
        return "mixed"
    if soft:
        return "soft_rig"
    if hard:
        return "moving_or_hardbait"
    if any(rig in SALT_RIGS for rig in rigs):
        return "specific_technique"
    return "unknown"


def infer_guide_hint_family(text, rigs, sku):
    value = lower(f"{sku} {text}")
    versatility = infer_versatility(text, rigs)
    if versatility in {"versatile", "mixed"}:
        return "versatile_mixed"
    if any(rig in {"Frog", "Punching"} for rig in rigs):
        return "frog_cover"
    if any(rig in {"Big Bait", "Swimbait"} for rig in rigs):
        return "big_bait"
    if any(rig in SOFT_RIGS for rig in rigs):
        if re.search(r"[-•.\s]st|smt|solid|實心|mega\s*top", value, re.I):
            return "high_sensitivity_soft_rig"
        if re.search("ul|xul|bf|finesse|精細|輕量", value, re.I):
            return "finesse_rig"
        return "soft_rig_bottom"
    if any(rig in HARD_RIGS for rig in rigs):
        return "moving_or_hardbait"
    if any(rig in {"Eging", "Tip-run Eging", "Sinker Rig"} for rig in rigs):
        return "eging"
    if any(rig in {"SLJ", "SLSJ", "Light Jigging", "Slow Jigging", "Offshore Jigging", "Metal Jig"} for rig in rigs):
        return "jigging"
    if "wind" in value:
        return "wind"
    return "general"


def infer_line_control_need(family):
    return {
        "versatile_mixed": "軟硬餌切換時維持拋投、線弧、控線與回收節奏一致",
        "finesse_rig": "細線小餌低負荷出線、鬆線控制與微弱咬口判斷",
        "high_sensitivity_soft_rig": "實心竿尖/高感配置下的張力變化、讀底與輕咬判斷",
        "soft_rig_bottom": "貼底釣組的線張力、障礙觸感與刺魚時機判斷",
        "moving_or_hardbait": "移動餌平收、抽停、泳層維持與連續搜索節奏",
        "frog_cover": "粗線出線、重障礙落點控制與中魚後拉離覆蓋物",
        "big_bait": "重餌拋投、回收負荷與長距離操餌控線",
        "eging": "PE 出線、抽竿後回線、看線與補刺節奏",
        "jigging": "垂直控線、抽停負荷、下沉咬口與長時間搏魚",
        "wind": "連續抽動後線弧控制、下沉咬口感知與白帶魚節奏操作",
    }.get(family, "日常路亞拋投、出線與控線穩定")


def confidence(description, source):
    if description and source == "official_description":
        return "high"
    if source == "recommended_rig_pairing":
        return "medium"
    return "low"


def row_value(row, col, field):
    if field not in col:
        return ""
    return n(row[col[field] - 1].value)


def build_fact(row, col, model_by_rod_id, source_label):
    detail_id = row_value(row, col, "id")
    rod_id = row_value(row, col, "rod_id")
    sku = row_value(row, col, "SKU")
    description = row_value(row, col, "Description")
    recommended = row_value(row, col, "recommended_rig_pairing")
    model = model_by_rod_id.get(rod_id, "")

    official_rigs = extract_rigs_from_text(description)
    pairing_rigs = split_pairing(recommended)
    source = "official_description" if official_rigs else "recommended_rig_pairing" if pairing_rigs else "spec_or_family"

    rigs = []
    for rig in official_rigs:
        append_unique(rigs, rig)
    for rig in pairing_rigs:
        append_unique(rigs, rig)

    family = infer_guide_hint_family(f"{model} {sku} {description} {recommended}", rigs, sku)
    primary = rigs[:3]
    secondary = rigs[3:8]
    return {
        "detail_id": detail_id,
        "rod_id": rod_id,
        "model": model,
        "sku": sku,
        "source_label": source_label,
        "fact_source": source,
        "confidence": confidence(description, source),
        "description_present": bool(description),
        "primary_rigs": primary,
        "secondary_rigs": secondary,
        "all_rigs": rigs,
        "versatility": infer_versatility(description or recommended, rigs),
        "guide_hint_family": family,
        "line_control_need": infer_line_control_need(family),
        "signals": {
            "description_keywords": [rig for rig in official_rigs],
            "recommended_rig_pairing": pairing_rigs,
            "power": row_value(row, col, "POWER"),
            "action": row_value(row, col, "Action"),
            "type": row_value(row, col, "TYPE"),
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Extract rod variant usage facts from an import workbook.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--source-label", default="daiwa_tw_rod_import")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_headers = [cell.value for cell in rod_ws[1]]
    detail_headers = [cell.value for cell in detail_ws[1]]
    rod_col = {name: idx + 1 for idx, name in enumerate(rod_headers)}
    detail_col = {name: idx + 1 for idx, name in enumerate(detail_headers)}

    model_by_rod_id = {}
    for row in rod_ws.iter_rows(min_row=2):
        model_by_rod_id[n(row[rod_col["id"] - 1].value)] = n(row[rod_col["model"] - 1].value)

    required = ["id", "rod_id", "SKU", "Description", "guide_use_hint"]
    missing = [field for field in required if field not in detail_col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")

    facts = [build_fact(row, detail_col, model_by_rod_id, args.source_label) for row in detail_ws.iter_rows(min_row=2)]
    output = {
        "schema": "rod_variant_usage_facts_v1",
        "source_xlsx": str(args.xlsx),
        "count": len(facts),
        "facts": facts,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n")
    counts = {}
    for fact in facts:
        counts[fact["confidence"]] = counts.get(fact["confidence"], 0) + 1
    print({"output": str(args.output), "count": len(facts), "confidence_counts": counts})


if __name__ == "__main__":
    main()
