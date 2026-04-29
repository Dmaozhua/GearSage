from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")

HINT_BY_ID = {
    "DRD10206": "岸投遠投：Metal Jig、Minnow、Sinking Pencil 與 Surf Plug 切換時，PE 出線、線弧和迎風控線更穩。",
    "DRD10208": "岸投遠投：Metal Jig、Minnow、Sinking Pencil 與 Surf Plug 切換時，PE 出線、線弧和迎風控線更穩。",
    "DRD10265": "便攜岸投小型餌：Small Metal Jig、Minnow、Surf Plug 拋投與抽竿控線更穩，適合海水岸拋搜索。",
    "DRD10289": "強力泛用：Texas、Soft Plastic、Rubber Jig 與 Spinnerbait 都能兼顧，打點、慢捲和障礙迴避切換更自然。",
}


def n(value):
    return " ".join(str(value or "").split())


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["guide_use_hint"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        if detail_id not in HINT_BY_ID:
            continue
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        new = HINT_BY_ID[detail_id]
        if old != new:
            cell.value = new
            changed.append({"row": row, "id": detail_id, "old": old, "new": new})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-guide_use_hint changes")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed": len(changed),
        "changed_ids": [item["id"] for item in changed],
    })


if __name__ == "__main__":
    main()
