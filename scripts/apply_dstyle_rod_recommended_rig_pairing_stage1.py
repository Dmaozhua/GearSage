import json
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "dstyle_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "dstyle_rod_recommended_rig_pairing_report.json"

FIELD = "recommended_rig_pairing"
AFTER_FIELD = "guide_use_hint"


PAIRINGS = {
    "DHRS-GD-64L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "source": "official_description",
        "notes": "RIKU spinning finesse positioning; ordered for light-rig use before small plugs.",
    },
    "DHRC-GD-68M-LM": {
        "pairing": "Mid Strolling Jighead / Hover Strolling / No Sinker / Down Shot",
        "source": "official_description",
        "notes": "THE SUTO bait mid-strolling positioning.",
    },
    "DHRS-511UL": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "source": "official_description",
        "notes": "Official variant description names finesse shake and light-rig use.",
    },
    "DHRS-60L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "source": "official_description",
        "notes": "Official variant description names finesse shake and light-rig use.",
    },
    "DHRS-63L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "source": "official_description",
        "notes": "Finesse versatile spinning model; specific finesse rigs before small plugs.",
    },
    "DHRS-63UL-FS": {
        "pairing": "Mid Strolling Jighead / Hover Strolling / Down Shot / No Sinker",
        "source": "official_description",
        "notes": "Light Midstroll/Finesse Swim code name and official use positioning.",
    },
    "DHRS-66ML": {
        "pairing": "Power Finesse / Small Rubber Jig / Cover Neko / Elastomer Softbait / Down Shot",
        "source": "official_description",
        "notes": "MAX Finesse power-spinning description supports cover/light power finesse.",
    },
    "DHRS-66L/ML-FS": {
        "pairing": "Mid Strolling Jighead / Hover Strolling / Power Finesse / Cover Neko",
        "source": "official_description",
        "notes": "Power Finesse Swim code name and official middle-swim positioning.",
    },
    "DHRS-68M": {
        "pairing": "Power Finesse / Cover Neko / Small Rubber Jig / Heavy Down Shot / No Sinker",
        "source": "official_description",
        "notes": "Power Finesse Shake positioning; power finesse before secondary finesse rigs.",
    },
    "DHRC-66L/ML": {
        "pairing": "Bait Finesse Jighead / Small Rubber Jig / Neko Rig / Down Shot / Small Plug",
        "source": "official_description",
        "notes": "Bait Finesse Versatile model; light baitcasting rigs before small plugs.",
    },
    "DHRC-69L+": {
        "pairing": "Cover Neko / Small Rubber Jig / Power Finesse / Light Texas",
        "source": "official_description",
        "notes": "Cover Finesse model; cover finesse soft-rig use.",
    },
    "DHRC-68M": {
        "pairing": "Texas Rig / Free Rig / Spinnerbait / Chatterbait / Crankbait",
        "source": "official_description",
        "notes": "FX versatile bait model; soft-rig operation before moving baits.",
    },
    "DHRC-610MH": {
        "pairing": "Texas Rig / Free Rig / Rubber Jig / Spinnerbait / Chatterbait / Crankbait",
        "source": "official_description",
        "notes": "FXR middle-versatile description supports worm rigs, jigs, plugs, and wire baits.",
    },
    "DHRC-70H": {
        "pairing": "Heavy Texas / Rubber Jig / Frog / Swim Jig / Heavy Cover No Sinker",
        "source": "official_description",
        "notes": "Power Fishing Versatile; no Big Bait without explicit evidence.",
    },
    "DHRS-E-62UL-S": {
        "pairing": "Down Shot / Neko Rig / Jighead Rig / No Sinker / Small Plug",
        "source": "official_description",
        "notes": "Finesse Shooting model; shooting finesse rigs before small plugs.",
    },
    "DHRS-E-510XUL-S": {
        "pairing": "Down Shot / Neko Rig / No Sinker / Small Rubber Jig",
        "source": "official_description",
        "notes": "Finesse Shake KIWAMI ultralight model.",
    },
    "DBTS-SS-63ML": {
        "pairing": "Neko Rig / Down Shot / Small Rubber Jig / Power Finesse",
        "source": "official_description",
        "notes": "SABER Lightning spec and official detail page support finesse/power-finesse use.",
    },
    "DBTS-SS-68L": {
        "pairing": "Mid Strolling Jighead / Down Shot / No Sinker / Small Rubber Jig",
        "source": "official_description",
        "notes": "SABER BM1 detail page supports mid-strolling and finesse rigs.",
    },
    "DBTC-SS-64MH": {
        "pairing": "High-density No Sinker / Jerk Worm / Texas Rig / Free Rig",
        "source": "official_description",
        "notes": "SABER Dagger detail page emphasizes high-density worm no-sinker jerking.",
    },
    "DBTC-SS-69MH+": {
        "pairing": "Texas Rig / Free Rig / Spinnerbait / Chatterbait / Swimbait / Big Bait",
        "source": "official_description",
        "notes": "SABER versatile SP official detail supports worm rigs, wire baits, swimbait, and big bait.",
    },
    "DBTS-612UL+-S": {
        "pairing": "No Sinker / Down Shot / Neko Rig / Jighead Rig / Small Rubber Jig / Hover Strolling",
        "source": "official_description",
        "notes": "BLUE TREK two-piece table/detail supports finesse soft-rig use.",
    },
    "DBTS-632UL/L-S": {
        "pairing": "No Sinker / Down Shot / Neko Rig / Jighead Rig / Small Rubber Jig / Elastomer Softbait",
        "source": "official_description",
        "notes": "All-mighty finesse spinning model; explicit finesse-rig family.",
    },
    "DBTS-642UL+-MIDSP": {
        "pairing": "Mid Strolling Jighead / Bottom Strolling / Down Shot / Jighead Rig",
        "source": "official_description",
        "notes": "MIDSP model; mid/bottom strolling before secondary finesse rigs.",
    },
    "DBTS-662L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Plug / Shad",
        "source": "official_description",
        "notes": "Finesse-to-plug versatile spinning use.",
    },
    "DBTS-662M": {
        "pairing": "Power Finesse / Cover Neko / Guarded Small Rubber Jig / High-density No Sinker",
        "source": "official_description",
        "notes": "PE power finesse and cover finesse positioning.",
    },
    "DBTS-6102ML-S": {
        "pairing": "Long Cast Down Shot / I-shaped Plug / Bug Lure / Surface Plug / Neko Rig",
        "source": "official_description",
        "notes": "Long-shooting finesse and surface/I-shaped plug use.",
    },
    "DBTC-662ML-BF": {
        "pairing": "Bait Finesse Small Plug / Small Rubber Jig / Neko Rig / Down Shot / Small Spinnerbait",
        "source": "official_description",
        "notes": "Bait-finesse model; specific small plug and finesse-rig use.",
    },
    "DBTC-672MH-S": {
        "pairing": "Cover Neko / Heavy Down Shot / Texas Rig / Rubber Jig",
        "source": "official_description",
        "notes": "Solid-tip tough cover model; cover soft-rig use.",
    },
    "DBTC-6102M": {
        "pairing": "Texas Rig / Free Rig / Crankbait / Spinnerbait / Chatterbait / Jerkbait",
        "source": "official_description",
        "notes": "Versatile bait model; worm rigs plus moving baits.",
    },
    "DBTC-6102MH": {
        "pairing": "Texas Rig / Free Rig / Rubber Jig / Spinnerbait / Chatterbait / Crankbait",
        "source": "official_description",
        "notes": "Middle-versatile bait model; soft-rig and moving-bait range.",
    },
    "DBTC-6102XH": {
        "pairing": "Big Bait / Swimbait / Crawler Bait / Rubber Jig / Heavy Texas",
        "source": "official_description",
        "notes": "Official extra-heavy model names big bait/swimbait/crawler-class use.",
    },
    "DBTC-702H": {
        "pairing": "Heavy Texas / Rubber Jig / Frog / Big Bait / Swimbait",
        "source": "official_description",
        "notes": "Heavy versatile model; heavy soft-rig/frog before large moving baits.",
    },
    "DBTS-60XUL-S": {
        "pairing": "Down Shot / Neko Rig / Hover Strolling / No Sinker",
        "source": "official_description",
        "notes": "10th Anniversary finesse spinning detail supports these finesse rigs.",
    },
    "DBTS-67UL-S": {
        "pairing": "Light Carolina Rig / Long Leader Down Shot / No Sinker / I-shaped Plug",
        "source": "official_description",
        "notes": "10th Anniversary long finesse model supports light Carolina and I-shaped use.",
    },
    "DBTS-60UL-S": {
        "pairing": "Down Shot / Neko Rig / No Sinker / Small Rubber Jig",
        "source": "official_description",
        "notes": "BLUE TREK super-finesse spinning model.",
    },
    "DBTS-61UL+-S": {
        "pairing": "Down Shot / Neko Rig / Jighead Rig / No Sinker",
        "source": "official_description",
        "notes": "BLUE TREK ultralight-plus finesse spinning model.",
    },
    "DBTS-61L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "source": "official_plus_whitelist",
        "notes": "Whitelist exact match confirms Neko Rig field use; official line remains finesse oriented.",
    },
    "DBTS-63UL-MIDSP": {
        "pairing": "Mid Strolling Jighead / Hover Strolling / Down Shot / No Sinker",
        "source": "official_description",
        "notes": "MIDSP model; mid-strolling first.",
    },
    "DBTS-65L+": {
        "pairing": "Neko Rig / Down Shot / Small Plug / No Sinker",
        "source": "official_description",
        "notes": "Light-plus versatile spinning model; finesse rigs before small plug.",
    },
    "DBTS-66ML-S-MIDSP": {
        "pairing": "Mid Strolling Jighead / No Sinker / Hover Strolling / Small Plug",
        "source": "official_plus_whitelist",
        "notes": "Whitelist exact match confirms No Sinker; official MIDSP supports mid-strolling.",
    },
    "DBTS-66M": {
        "pairing": "Power Finesse / Cover Neko / Small Rubber Jig / No Sinker",
        "source": "official_description",
        "notes": "Power-finesse spinning model.",
    },
    "DBTS-68H-S-PF": {
        "pairing": "Power Finesse / Cover Neko / Small Rubber Jig / High-density No Sinker",
        "source": "official_description",
        "notes": "PF model; heavy power-finesse cover use.",
    },
    "DBTS-610L-S": {
        "pairing": "Long Cast Neko Rig / I-shaped Plug / Surface Plug / Small Rubber Jig",
        "source": "official_plus_whitelist",
        "notes": "Whitelist collection supports surface plug and finesse examples; official long-cast light spinning use.",
    },
    "DBTC-64ML-FM": {
        "pairing": "Crankbait / Shad / Minnow / Topwater Plug",
        "source": "official_description",
        "notes": "FM/glass moving-bait model; hard moving baits only.",
    },
    "DBTC-65ML-FM": {
        "pairing": "Minnow / Shad / Crankbait / Topwater Plug",
        "source": "official_description",
        "notes": "Light plugging/moving-bait model.",
    },
    "DBTC-65M+-PF": {
        "pairing": "Bait Power Finesse / Small Rubber Jig / Bait Neko / Heavy Down Shot / Free Rig / Texas Rig / Football Jig",
        "source": "official_description",
        "notes": "PF bait model; PE cover finesse before fluorocarbon bottom-rig uses.",
    },
    "DBTC-67ML-BF": {
        "pairing": "Bait Finesse Small Plug / Small Rubber Jig / Neko Rig / Down Shot / Small Spinnerbait",
        "source": "official_description",
        "notes": "Bait-finesse model; light baitcasting rig and small plug use.",
    },
    "DBTC-68M": {
        "pairing": "Crankbait / Shad / Jerkbait / Spinnerbait / Chatterbait / Texas Rig",
        "source": "official_plus_whitelist",
        "notes": "Whitelist exact match confirms hardbait use; official positioning remains versatile.",
    },
    "DBTC-610MH": {
        "pairing": "Chatterbait / Spinnerbait / Texas Rig / Rubber Jig / Crankbait",
        "source": "official_plus_whitelist",
        "notes": "Whitelist exact match confirms bladed jig; Chatterbait kept first.",
    },
    "DBTC-70M+-FM": {
        "pairing": "Crankbait / Big Minnow / Topwater Plug / Spinnerbait / Buzzbait / Chatterbait / Shad Tail",
        "source": "official_description",
        "notes": "Fast-moving model; official use supports plugs, wire baits, bladed jig, and shad-tail moving bait.",
    },
    "DBTC-70H-S": {
        "pairing": "Cover Neko / Texas Rig / Rubber Jig / Heavy Down Shot",
        "source": "official_description",
        "notes": "Heavy solid-tip cover finesse model.",
    },
    "DBTC-71MH-PF": {
        "pairing": "Frog / Cover Power Finesse / Heavy Texas / Rubber Jig",
        "source": "official_plus_whitelist",
        "notes": "Whitelist exact match confirms frog; official PF model supports cover power use.",
    },
    "DBTC-73H": {
        "pairing": "Big Bait / Crawler Bait / Swimbait / Swim Jig / Heavy Texas / Flipping",
        "source": "official_description",
        "notes": "Official heavy model supports big bait/crawler/swimbait/swim jig and heavy cover pitching.",
    },
}

FORBIDDEN_GENERIC = ["General Lure", "Light Rig", "Hardbait", "Soft Bait"]


def normalize(value):
    return " ".join(str(value or "").split())


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)


def ensure_field(ws, headers):
    if FIELD in headers:
        return headers.index(FIELD) + 1, False
    if AFTER_FIELD not in headers:
        raise RuntimeError(f"missing anchor column: {AFTER_FIELD}")
    insert_at = headers.index(AFTER_FIELD) + 2
    ws.insert_cols(insert_at)
    for row in range(1, ws.max_row + 1):
        copy_cell_style(ws.cell(row=row, column=insert_at - 1), ws.cell(row=row, column=insert_at))
    ws.cell(row=1, column=insert_at).value = FIELD
    return insert_at, True


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    field_col, inserted = ensure_field(ws, headers)
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    required = ["id", "rod_id", "SKU", "Description", "guide_use_hint", FIELD]
    missing = [name for name in required if name not in col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")

    changed_rows = []
    missing_skus = []
    source_counts = {}
    forbidden_residuals = []
    for row in range(2, ws.max_row + 1):
        sku = normalize(ws.cell(row=row, column=col["SKU"]).value)
        detail_id = normalize(ws.cell(row=row, column=col["id"]).value)
        rod_id = normalize(ws.cell(row=row, column=col["rod_id"]).value)
        item = PAIRINGS.get(sku)
        if not item:
            missing_skus.append({"row": row, "id": detail_id, "rod_id": rod_id, "SKU": sku})
            continue
        previous = normalize(ws.cell(row=row, column=field_col).value)
        pairing = item["pairing"]
        ws.cell(row=row, column=field_col).value = pairing
        source_counts[item["source"]] = source_counts.get(item["source"], 0) + 1
        residual = [term for term in FORBIDDEN_GENERIC if term.lower() in pairing.lower()]
        if residual:
            forbidden_residuals.append({"row": row, "id": detail_id, "SKU": sku, "terms": residual})
        changed_rows.append({
            "xlsx_row": row,
            "id": detail_id,
            "rod_id": rod_id,
            "SKU": sku,
            "previous": previous,
            "recommended_rig_pairing": pairing,
            "source": item["source"],
            "notes": item["notes"],
            "changed": previous != pairing,
        })

    if missing_skus:
        raise RuntimeError(f"missing recommended_rig_pairing mappings: {missing_skus}")
    if forbidden_residuals:
        raise RuntimeError(f"forbidden generic terms remain: {forbidden_residuals}")

    wb.save(XLSX_PATH)

    report = {
        "schema": "dstyle_rod_recommended_rig_pairing_stage1",
        "source_xlsx": str(XLSX_PATH),
        "field": FIELD,
        "inserted_column": inserted,
        "total_detail_rows": ws.max_row - 1,
        "changed_row_count": sum(1 for row in changed_rows if row["changed"]),
        "filled_row_count": len(changed_rows),
        "source_counts": source_counts,
        "forbidden_generic_terms": FORBIDDEN_GENERIC,
        "forbidden_residuals": forbidden_residuals,
        "rows": changed_rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "xlsx": str(XLSX_PATH),
        "report": str(REPORT_PATH),
        "inserted_column": inserted,
        "filled_row_count": len(changed_rows),
        "changed_row_count": report["changed_row_count"],
        "source_counts": source_counts,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
