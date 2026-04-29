import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_XLSX = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")
DEFAULT_FACTS = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_variant_usage_facts.json")
DEFAULT_REPORT = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_usage_consistency_report.json")

SOFT_TERMS = [
    "Down Shot", "Neko", "No Sinker", "Light Texas", "Texas", "Free Rig",
    "Leaderless", "Heavy Down Shot", "Rubber Jig", "Small Rubber", "Jighead",
    "Wacky", "Soft Plastic", "High-density Soft Plastic",
]
HARD_TERMS = [
    "Crankbait", "Shad", "Minnow", "Jerkbait", "Vibration", "Metal Vibration",
    "Topwater", "I-shaped", "Small Hardbait", "Plug", "Spinnerbait", "Chatterbait",
    "Swim Jig", "Swimbait", "Big Bait",
]
SALT_TERMS = ["Eging", "Tip-run", "Sinker Rig", "SLJ", "SLSJ", "Light Jigging", "Slow Jigging", "Offshore Jigging", "Metal Jig", "WIND"]

HARD_ONLY_HINTS = ["硬餌搜索", "卷阻/搜索硬餌", "操作型硬餌", "輕硬餌泛用"]
SOFT_ONLY_HINTS = ["軟餌底操", "底操軟餌", "高感底操軟餌", "強力底操軟餌"]
VERSATILE_HINTS = ["泛用", "多用途", "軟硬餌", "切換", "兼顧"]


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def contains_any(text, terms):
    value = n(text).lower()
    return any(term.lower() in value for term in terms)


def split_pairing(value):
    return [n(part) for part in n(value).split("/") if n(part)]


def has_soft(pairing):
    return contains_any(pairing, SOFT_TERMS)


def has_hard(pairing):
    return contains_any(pairing, HARD_TERMS)


def has_salt(pairing):
    return contains_any(pairing, SALT_TERMS)


def first_group(pairing):
    parts = split_pairing(pairing)
    if not parts:
        return "none"
    first = parts[0]
    if contains_any(first, SOFT_TERMS):
        return "soft"
    if contains_any(first, HARD_TERMS):
        return "hard"
    if contains_any(first, SALT_TERMS):
        return "salt"
    return "other"


def description_is_versatile(description):
    return bool(re.search("泛用|多用途|兼具|廣泛|全能|萬用|versatile", n(description), re.I))


def description_lists_specific_rig_missing(description, pairing):
    desc = n(description)
    if not desc:
        return []
    checks = [
        ("Crankbait", ["crankbait", "crank"]),
        ("Spinnerbait", ["spinnerbait", "旋轉亮片", "複合式亮片", "復合式亮片"]),
        ("Down Shot", ["down shot", "倒吊", "落下釣組"]),
        ("Texas", ["texas", "德州"]),
        ("Neko", ["neko"]),
        ("No Sinker", ["no sinker", "無鉛"]),
        ("Jighead", ["jighead", "jig head", "汲鉤頭", r"(?<!刀片)(?<!刀片式)(?<!橡膠)(?<!軟膠)(?<!泳餌)(?<!無)鉛頭鉤"]),
        ("Minnow", ["minnow", "米諾"]),
        ("Jerkbait", ["jerkbait", "jerk", "twitch"]),
        ("Metal Jig", ["metal jig", "金屬jig", "金屬 jig", "小型鐵板", "鐵板路亞"]),
        ("Eging", ["木蝦", "eging", "エギ"]),
    ]
    missing = []
    lower_desc = desc.lower()
    lower_pairing = n(pairing).lower()
    for label, patterns in checks:
        if any(re.search(pattern, lower_desc, re.I) for pattern in patterns) and label.lower() not in lower_pairing:
            missing.append(label)
    return missing


def issue(row, code, severity, message, evidence):
    return {
        "detail_id": row["id"],
        "rod_id": row["rod_id"],
        "sku": row["SKU"],
        "code": code,
        "severity": severity,
        "message": message,
        "evidence": evidence,
    }


def validate_row(row, fact=None):
    issues = []
    description = row["Description"]
    pairing = row["recommended_rig_pairing"]
    hint = row["guide_use_hint"]
    first = first_group(pairing)
    soft = has_soft(pairing)
    hard = has_hard(pairing)
    salt = has_salt(pairing)
    hint_hard_only = contains_any(hint, HARD_ONLY_HINTS)
    hint_soft_only = contains_any(hint, SOFT_ONLY_HINTS)
    hint_versatile = contains_any(hint, VERSATILE_HINTS)
    desc_versatile = description_is_versatile(description)

    if desc_versatile and soft and hard and (hint_hard_only or hint_soft_only) and not hint_versatile:
        issues.append(issue(
            row,
            "versatile_description_single_route_hint",
            "error",
            "Description says this variant is versatile/multi-use, but guide_use_hint is a single soft- or hard-bait route.",
            {"Description": description, "recommended_rig_pairing": pairing, "guide_use_hint": hint},
        ))

    if first == "soft" and hint_hard_only and soft:
        issues.append(issue(
            row,
            "soft_primary_pairing_hard_hint",
            "error",
            "recommended_rig_pairing starts with a soft-rig technique, but guide_use_hint is hardbait/search oriented.",
            {"recommended_rig_pairing": pairing, "guide_use_hint": hint},
        ))

    if first == "hard" and hint_soft_only and hard and not soft:
        issues.append(issue(
            row,
            "hard_primary_pairing_soft_hint",
            "error",
            "recommended_rig_pairing is hardbait/search oriented, but guide_use_hint is bottom soft-rig oriented.",
            {"recommended_rig_pairing": pairing, "guide_use_hint": hint},
        ))

    if soft and hard and not hint_versatile and not desc_versatile and (hint_hard_only or hint_soft_only):
        issues.append(issue(
            row,
            "mixed_pairing_single_route_hint",
            "warning",
            "recommended_rig_pairing mixes soft-rig and hardbait techniques, but guide_use_hint does not explain the mixed-use nature.",
            {"recommended_rig_pairing": pairing, "guide_use_hint": hint},
        ))

    missing = description_lists_specific_rig_missing(description, pairing)
    if missing:
        issues.append(issue(
            row,
            "description_rig_missing_from_pairing",
            "warning",
            "Description explicitly names techniques that are missing from recommended_rig_pairing.",
            {"missing_terms": missing, "Description": description, "recommended_rig_pairing": pairing},
        ))

    if not n(description) and fact and fact.get("confidence") == "low" and not n(pairing):
        issues.append(issue(
            row,
            "low_confidence_without_description_or_pairing",
            "warning",
            "No variant Description and no recommended_rig_pairing; guide_use_hint cannot be trusted as a fine-grained claim.",
            {"guide_use_hint": hint},
        ))

    if salt and (contains_any(hint, HARD_ONLY_HINTS) or contains_any(hint, SOFT_ONLY_HINTS)) and not contains_any(hint, ["木蝦", "鐵板", "岸投", "船釣", "WIND", "輕海水"]):
        issues.append(issue(
            row,
            "salt_pairing_bass_hint",
            "error",
            "recommended_rig_pairing is saltwater-specific, but guide_use_hint looks like a bass soft/hardbait hint.",
            {"recommended_rig_pairing": pairing, "guide_use_hint": hint},
        ))

    return issues


def read_facts(path):
    if not path or not path.exists():
        return {}
    data = json.loads(path.read_text())
    return {item["detail_id"]: item for item in data.get("facts", [])}


def main():
    parser = argparse.ArgumentParser(description="Validate Description/recommended_rig_pairing/guide_use_hint consistency.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--facts", type=Path, default=DEFAULT_FACTS)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--fail-on-error", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}
    required = ["id", "rod_id", "SKU", "Description", "guide_use_hint"]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")
    optional = ["recommended_rig_pairing"]
    missing_optional = [field for field in optional if field not in col]

    facts = read_facts(args.facts)
    issues = []
    if missing_optional:
        issues.append({
            "detail_id": "",
            "rod_id": "",
            "sku": "",
            "code": "missing_recommended_rig_pairing_column",
            "severity": "warning",
            "message": "rod_detail has no recommended_rig_pairing column; validator can only use Description and guide_use_hint until the field is added.",
            "evidence": {"missing_columns": missing_optional},
        })
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {field: n(values[col[field]]) for field in required}
        row["recommended_rig_pairing"] = n(values[col["recommended_rig_pairing"]]) if "recommended_rig_pairing" in col else ""
        issues.extend(validate_row(row, facts.get(row["id"])))

    counts = {}
    for item in issues:
        counts[item["severity"]] = counts.get(item["severity"], 0) + 1
    report = {
        "schema": "rod_usage_consistency_report_v1",
        "source_xlsx": str(args.xlsx),
        "facts_file": str(args.facts) if args.facts else "",
        "issue_count": len(issues),
        "severity_counts": counts,
        "issues": issues,
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print({"report": str(args.report), "issue_count": len(issues), "severity_counts": counts})
    if args.fail_on_error and counts.get("error", 0):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
