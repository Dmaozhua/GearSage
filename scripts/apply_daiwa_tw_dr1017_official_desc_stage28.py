from copy import copy
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel
import re
import ssl
from urllib.request import Request, urlopen

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')
OFFICIAL_URL = "https://www.daiwa.com/tw/product/t32myhm"
TARGET_ROD_ID = "DR1017"
TARGET_FIELDS = ["Description", "recommended_rig_pairing", "guide_use_hint"]


RIG_AND_HINT_BY_CODE = {
    "SC C66ML-G": (
        "Shad Crankbait / Small Crankbait / Medium Crankbait / Flat Side Crankbait / Shad / Minnow / Popper / Pencil Bait / Crawler Bait",
        "玻璃纖維卷阻硬餌：Shad Crankbait、Small/Medium Crank、Flat Side 和 Minnow 為主，竿尖追隨性也能承接 Popper、Pencil、Crawler 的水面節奏。",
    ),
    "SC C66M/ML-SV-ST": (
        "Neko Rig / Small Rubber Jig / No Sinker / Down Shot",
        "WEREWOLF 高感精細底操：Neko、Small Rubber Jig、No Sinker 和 Down Shot 為主，Mega Top R 實心竿尖強化讀底、輕咬與刺魚時機。",
    ),
    "SC C68H-ST-SB": (
        "Swimbait / Big Bait / Crawler Bait / Alabama Rig / Rubber Jig / Texas Rig / Leaderless Down Shot / No Sinker / Spinnerbait",
        "SB 大餌/泳餌：Swimbait、Big Bait、Crawler 和 Alabama Rig 為主，實心竿尖兼顧近距離拋準、咬餌追隨與重鉤貫穿。",
    ),
    "SC C69M+-ST": (
        "Neko Rig / No Sinker / Down Shot / Texas Rig / Leaderless Down Shot / Free Rig / Rubber Jig / Guarded Jighead / Spinnerbait / Vibration / Crankbait / Craw",
        "FIRE WOLF 全能底操：Neko、No Sinker、Down Shot、Texas 和 Free Rig 為主，實心竿尖也能兼顧 Guarded Jighead、Spinnerbait、Vibration、Crankbait 的鬆線回收與觸感。",
    ),
    "SC C69M+-2-ST": (
        "Neko Rig / No Sinker / Down Shot / Texas Rig / Leaderless Down Shot / Free Rig / Rubber Jig / Spinnerbait / Vibration / Crankbait / Craw",
        "二節 FIRE WOLF 全能底操：Neko、No Sinker、Down Shot、Texas 和 Free Rig 為主，中心切割設計保留鬆線回收、觸感與攜帶性。",
    ),
    "SC C69MH": (
        "Rubber Jig / Texas Rig / Leaderless Down Shot / No Sinker / Spinnerbait / Soft Plastic / Swimbait / Big Bait / Big Crankbait / Frog",
        "KING VIPER 重型全能：Rubber Jig、Texas、No Sinker 和 Spinnerbait 為主，可擴展到 Swimbait、Big Bait、Big Crankbait 與 Frog。",
    ),
}

LOCAL_CODE_ALIASES = {
    "SC C68H-SB": "SC C68H-ST-SB",
}


class TextParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    def handle_starttag(self, tag, attrs):
        if tag in {"br", "p", "div", "section", "h2", "h3", "li"}:
            self.parts.append(" ")

    def handle_endtag(self, tag):
        if tag in {"p", "div", "section", "h2", "h3", "li"}:
            self.parts.append(" ")

    def handle_data(self, data):
        if data:
            self.parts.append(data)

    def text(self):
        return clean_text("".join(self.parts))


def clean_text(value):
    return re.sub(r"\s+", " ", unescape(str(value or ""))).strip()


def strip_html(fragment):
    parser = TextParser()
    parser.feed(fragment)
    return parser.text()


def normalize_code(value):
    text = clean_text(value).upper()
    text = re.sub(r"[【〖].*?[】〗]", "", text)
    text = text.replace("＋", "+").replace("－", "-")
    text = text.replace("・", "-").replace("•", "-").replace("／", "/")
    text = text.replace(" ", "")
    text = re.sub(r"^STEEZ", "", text)
    return text


def fetch_html():
    req = Request(OFFICIAL_URL, headers={"User-Agent": "Mozilla/5.0"})
    context = ssl._create_unverified_context()
    return urlopen(req, timeout=20, context=context).read().decode("utf-8", "ignore")


def fetch_official_descriptions():
    html = fetch_html()
    headings = list(re.finditer(r"<h2[^>]*>\s*■([^<]+)</h2>", html, re.I))
    sections = {}
    for idx, heading in enumerate(headings):
        title = strip_html(heading.group(1))
        code = normalize_code(title)
        start = heading.end()
        end = headings[idx + 1].start() if idx + 1 < len(headings) else len(html)
        body = strip_html(html[start:end])
        body = re.sub(r"\s+(彎曲曲線|調性比較|MOVIE|發售月份|產品規格)\b.*$", "", body)
        description = clean_text(f"{title} {body}")
        for target_code in RIG_AND_HINT_BY_CODE:
            if code == normalize_code(target_code):
                sections[target_code] = description
                break

    missing = sorted(set(RIG_AND_HINT_BY_CODE) - set(sections))
    if missing:
        raise RuntimeError(f"missing official descriptions: {missing}")
    return sections


def snapshot_without_targets(ws, target_cols):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col not in target_cols
        })
    return rows


def copy_fills(src_row, dst_row):
    for src, dst in zip(src_row, dst_row):
        dst.fill = copy(src.fill)


def main():
    descriptions = fetch_official_descriptions()
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    ws = wb["rod_detail"]
    rod_headers = [cell.value for cell in rod_ws[1]]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    for field in ["id", "rod_id", "SKU", *TARGET_FIELDS]:
        if field not in col:
            raise RuntimeError(f"missing rod_detail column: {field}")

    target_cols = {col[field] for field in TARGET_FIELDS}
    before_detail = snapshot_without_targets(ws, target_cols)
    before_rod = [[rod_ws.cell(row, c).value for c in range(1, rod_ws.max_column + 1)] for row in range(1, rod_ws.max_row + 1)]

    changed = []
    matched = []
    for row in range(2, ws.max_row + 1):
        if clean_text(ws.cell(row, col["rod_id"]).value) != TARGET_ROD_ID:
            continue
        sku = clean_text(ws.cell(row, col["SKU"]).value)
        local_code = re.sub(r"^STEEZ\s+", "", sku, flags=re.I)
        official_code = LOCAL_CODE_ALIASES.get(local_code, local_code)
        if normalize_code(official_code) not in {normalize_code(code) for code in RIG_AND_HINT_BY_CODE}:
            continue
        for code in RIG_AND_HINT_BY_CODE:
            if normalize_code(code) == normalize_code(official_code):
                official_code = code
                break
        matched.append({"id": clean_text(ws.cell(row, col["id"]).value), "sku": sku, "official_code": official_code})
        new_values = {
            "Description": descriptions[official_code],
            "recommended_rig_pairing": RIG_AND_HINT_BY_CODE[official_code][0],
            "guide_use_hint": RIG_AND_HINT_BY_CODE[official_code][1],
        }
        row_changes = {}
        for field, new_value in new_values.items():
            cell = ws.cell(row, col[field])
            old = clean_text(cell.value)
            if old != new_value:
                cell.value = new_value
                row_changes[field] = {"old": old, "new": new_value}
        if row_changes:
            changed.append({
                "id": clean_text(ws.cell(row, col["id"]).value),
                "sku": sku,
                "official_code": official_code,
                "fields": sorted(row_changes),
            })

    expected_codes = sorted(normalize_code(code) for code in RIG_AND_HINT_BY_CODE)
    matched_codes = sorted(normalize_code(item["official_code"]) for item in matched)
    if matched_codes != expected_codes:
        raise RuntimeError(f"DR1017 matched code mismatch: expected={expected_codes}, matched={matched_codes}, rows={matched}")

    after_detail = snapshot_without_targets(ws, target_cols)
    if before_detail != after_detail:
        raise RuntimeError("unexpected non-target field changes in rod_detail")
    after_rod = [[rod_ws.cell(row, c).value for c in range(1, rod_ws.max_column + 1)] for row in range(1, rod_ws.max_row + 1)]
    if before_rod != after_rod:
        raise RuntimeError("unexpected changes in rod sheet")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "official_url": OFFICIAL_URL,
        "target_rod_id": TARGET_ROD_ID,
        "matched_rows": len(matched),
        "changed_rows": len(changed),
        "changed_cells": sum(len(item["fields"]) for item in changed),
        "matched": matched,
        "changed": changed,
    })


if __name__ == "__main__":
    main()
