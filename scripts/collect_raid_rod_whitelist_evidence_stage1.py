import json
import ssl
import time
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook

NORMALIZED_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/raid_rod_normalized.json")
XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/raid_rod_import.xlsx")
OUTPUT_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/raid_rod_whitelist_player_evidence.json")

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
)
SSL_CONTEXT = ssl._create_unverified_context()


def n(value):
    return str(value or "").replace("\xa0", " ").strip()


def unique(values, limit=None):
    out = []
    for value in values:
        value = n(value)
        if value and value not in out:
            out.append(value)
    return out[:limit] if limit else out


def fetch(url, timeout=30):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout, context=SSL_CONTEXT) as res:
        return res.read()


def sku_base(sku):
    return n(sku).split(" 【")[0]


def query_candidates(sku):
    base = sku_base(sku)
    candidates = [base]
    if "+" in base:
        candidates.append(base.replace("+", ""))
    if "⁺" in base:
        candidates.append(base.replace("⁺", "+"))
        candidates.append(base.replace("⁺", ""))
    return unique(candidates)


def parse_tackledb_block(block):
    fields = {}
    nodes = block.select("dl dt, dl dd")
    i = 0
    while i < len(nodes) - 1:
        if nodes[i].name == "dt" and nodes[i + 1].name == "dd":
            fields[nodes[i].get_text(" ", strip=True)] = nodes[i + 1].get_text(" ", strip=True)
            i += 2
        else:
            i += 1
    return fields


def parse_tackledb_page(html):
    soup = BeautifulSoup(html, "html.parser")
    blocks = []
    for block in soup.select(".tackle-block"):
        fields = parse_tackledb_block(block)
        if fields:
            blocks.append(fields)

    summary = {}
    for heading in soup.select("h3"):
        key = heading.get_text(" ", strip=True)
        value_node = heading.find_next("div")
        if key and value_node:
            summary[key] = value_node.get_text(" ", strip=True)

    return blocks, summary


def collect_tackledb(sku):
    errors = []
    for candidate in query_candidates(sku):
        url = "https://tackledb.uosoku.com/collection?rod=" + urllib.parse.quote(candidate)
        try:
            html = fetch(url)
            blocks, summary = parse_tackledb_page(html)
            if blocks:
                return {
                    "source_site": "tackledb.uosoku.com",
                    "source_url": url,
                    "query": candidate,
                    "hit_count": len(blocks),
                    "blocks": blocks[:12],
                    "summary": summary,
                    "status": "matched",
                }
            errors.append({"query": candidate, "url": url, "error": "no tackle blocks"})
        except Exception as exc:
            errors.append({"query": candidate, "url": url, "error": str(exc)})
        time.sleep(0.2)

    return {
        "source_site": "tackledb.uosoku.com",
        "source_url": errors[-1]["url"] if errors else "",
        "query": query_candidates(sku)[0],
        "hit_count": 0,
        "blocks": [],
        "summary": {},
        "status": "not_matched",
        "errors": errors,
    }


def summarize_tackledb(evidence):
    blocks = evidence.get("blocks") or []
    categories = unique(block.get("カテゴリー") for block in blocks)
    target_fish = unique(block.get("対象魚") for block in blocks)
    places = unique(block.get("釣り場") for block in blocks)
    lures = unique((block.get("ルアー") for block in blocks), 8)
    lines = unique((block.get("ライン") for block in blocks), 8)

    lure_type_text = n((evidence.get("summary") or {}).get("ルアーのタイプ"))
    lure_types = unique(lure_type_text.split())
    all_text = " ".join([lure_type_text, " ".join(lures)]).upper()

    supported_fields = []
    positioning_signals = []
    caution_notes = []
    confidence = "low"
    if blocks:
        supported_fields.extend(["player_environment", "player_positioning", "player_selling_points"])
        confidence = "medium" if len(blocks) < 3 else "high"
        if any("バス" in value for value in categories + target_fish):
            suggested_environment = "淡水 / Bass"
        else:
            suggested_environment = ""

        if "フロッグ" in all_text:
            positioning_signals.append("Frog / topwater")
        if any(k in all_text for k in ["ビッグベイト", "ジャイアントベイト", "スイムベイト", "BIG"]):
            positioning_signals.append("big bait / swimbait")
        if any(k in all_text for k in ["クランク", "バイブ", "ミノー", "ジャーク", "スピナー", "シャッド", "トップ"]):
            positioning_signals.append("moving bait / hard bait")
        if any(k in all_text for k in ["ワーム", "スモラバ", "虫系"]):
            positioning_signals.append("worm / finesse")
        if len(positioning_signals) > 1:
            caution_notes.append(
                "TackleDB is user tackle-log evidence; mixed lure records should support, not override, official positioning."
            )
    else:
        suggested_environment = ""

    return {
        "categories": categories,
        "target_fish": target_fish,
        "places": places[:8],
        "lure_types": lure_types[:12],
        "sample_lures": lures,
        "sample_lines": lines,
        "suggested_player_environment": suggested_environment,
        "positioning_signals": positioning_signals,
        "caution_notes": caution_notes,
        "supported_fields": supported_fields,
        "confidence": confidence,
    }


def read_detail_ids():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    columns = {name: idx for idx, name in enumerate(headers)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        out[sku_base(row[columns["SKU"]])] = {
            "detail_id": row[columns["id"]],
            "rod_id": row[columns["rod_id"]],
            "sku": row[columns["SKU"]],
        }
    return out


def main():
    normalized = json.loads(NORMALIZED_PATH.read_text())
    detail_map = read_detail_ids()
    records = []
    coverage = Counter()

    for category in normalized:
        for variant in category["variants"]:
            base = sku_base(variant["sku"])
            detail = detail_map.get(base, {})
            tackledb = collect_tackledb(variant["sku"])
            summary = summarize_tackledb(tackledb)
            coverage[tackledb["status"]] += 1
            records.append(
                {
                    "detail_id": detail.get("detail_id"),
                    "rod_id": detail.get("rod_id"),
                    "sku": detail.get("sku") or variant["sku"],
                    "source_official_url": variant.get("url"),
                    "whitelist": {
                        "tackledb": tackledb,
                        "rodsjp": {
                            "status": "not_useful",
                            "note": "Rods.JP search for Raid Japan / representative SKU returned 0 product results.",
                            "sample_url": "https://rods.jp/item.php?search=&keywords=GA-74XHC",
                        },
                        "rodsearch": {
                            "status": "not_useful",
                            "note": "Rodsearch did not expose usable Raid Japan product records through text search during this pass.",
                            "sample_url": "https://rodsearch.com/?s=GA-74XHC",
                        },
                    },
                    "summary": summary,
                    "apply_policy": "Use whitelist evidence as auxiliary player-field support only. Prefer official Raid Japan product text for model-specific positioning.",
                }
            )
            time.sleep(0.15)

    OUTPUT_PATH.write_text(
        json.dumps(
            {
                "source": "Raid Japan rod whitelist player evidence",
                "coverage": dict(coverage),
                "records": records,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )
    print({"output": str(OUTPUT_PATH), "records": len(records), "coverage": dict(coverage)})


if __name__ == "__main__":
    main()
