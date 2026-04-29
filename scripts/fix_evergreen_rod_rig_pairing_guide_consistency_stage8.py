from copy import copy
from datetime import datetime, timezone
from pathlib import Path
import json
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = ROOT / "GearSage-client/pkgGear/data_raw"
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "evergreen_rod_rig_pairing_guide_consistency_fix_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"

PAIR_FIELD = "recommended_rig_pairing"
GUIDE_FIELD = "guide_use_hint"
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")

SOFT_TERMS = {
    "Texas Rig", "Heavy Texas", "Light Texas", "Free Rig", "Leaderless Down Shot",
    "Heavy Down Shot", "Down Shot", "Neko Rig", "No Sinker", "High-density No Sinker",
    "Small Rubber Jig", "Football Jig", "Rubber Jig", "Guarded Rubber Jig", "Jighead Rig",
    "Wacky Rig", "Mid Strolling Jighead", "Hover Strolling", "Bottom Strolling",
}

HARD_TERMS = {
    "Chatterbait", "Swim Jig", "Spinnerbait", "Buzzbait", "Crankbait", "Deep Crankbait",
    "Shallow Crankbait", "Shad", "Jerkbait", "Minnow", "Vibration", "Metal Vibration",
    "Topwater Plug", "Pencil Bait", "Popper", "Crawler Bait", "Spoon", "Big Bait",
    "Swimbait", "Alabama Rig",
}


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def split_pairing(value):
    return [part.strip() for part in n(value).split(" / ") if part.strip()]


def snapshot_without_guide(ws, guide_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != guide_col
        })
    return rows


def is_conflict(pairing, guide):
    parts = split_pairing(pairing)
    if not parts:
        return False
    first = parts[0]
    guide_text = n(guide)
    if first in HARD_TERMS and re.search(r"軟餌底操|软饵底操|底操", guide_text) and not re.search(r"泛用|切換|切换|兼顧|兼顾", guide_text):
        return True
    if first in SOFT_TERMS and re.search(r"硬餌|硬饵|搜索", guide_text) and not re.search(r"泛用|切換|切换|兼顧|兼顾", guide_text):
        return True
    return False


def build_hint(pairing):
    parts = split_pairing(pairing)
    first_terms = parts[:3]
    soft_terms = [term for term in parts if term in SOFT_TERMS][:2]
    lead = " / ".join(first_terms)
    if parts and parts[0] in {"Big Bait", "Swimbait", "Alabama Rig"}:
        prefix = "大餌/泳餌主軸"
    elif parts and parts[0] in {"Crankbait", "Deep Crankbait", "Shallow Crankbait", "Shad", "Jerkbait", "Minnow"}:
        prefix = "硬餌搜索主軸"
    else:
        prefix = "移動餌主軸"
    if soft_terms:
        soft = " / ".join(soft_terms)
        return f"{prefix}：{lead} 為主，平收、抽停與泳層控制更清楚；需要放慢時可切換 {soft} 做軟餌底操。"
    return f"{prefix}：{lead} 為主，平收、抽停與泳層控制更清楚，適合連續搜索與節奏變化。"


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    for field in ["id", "SKU", PAIR_FIELD, GUIDE_FIELD]:
        if field not in col:
            raise RuntimeError(f"missing column: {field}")

    before = snapshot_without_guide(ws, col[GUIDE_FIELD])
    changes = []
    for row in range(2, ws.max_row + 1):
        pairing = n(ws.cell(row, col[PAIR_FIELD]).value)
        old = n(ws.cell(row, col[GUIDE_FIELD]).value)
        if not is_conflict(pairing, old):
            continue
        new = build_hint(pairing)
        cell = ws.cell(row, col[GUIDE_FIELD])
        cell.value = new
        cell.fill = copy(YELLOW_FILL)
        changes.append({
            "row": row,
            "id": n(ws.cell(row, col["id"]).value),
            "sku": n(ws.cell(row, col["SKU"]).value),
            "recommended_rig_pairing": pairing,
            "old_guide_use_hint": old,
            "new_guide_use_hint": new,
        })

    after = snapshot_without_guide(ws, col[GUIDE_FIELD])
    if before != after:
        raise RuntimeError("unexpected changes outside guide_use_hint")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)
    REPORT_PATH.write_text(json.dumps({
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_file": str(XLSX_PATH),
        "changed_guide_use_hint_rows": len(changes),
        "reason": "Only rows where recommended_rig_pairing starts with hard/moving bait but guide_use_hint still described soft-bottom operation were updated.",
        "changes": changes,
    }, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "changed_guide_use_hint_rows": len(changes),
        "changed_ids": [item["id"] for item in changes],
        "report_file": str(REPORT_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
