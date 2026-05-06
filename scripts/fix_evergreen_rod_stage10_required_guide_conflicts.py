from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel
import json
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "evergreen_rod_stage10_required_guide_conflicts_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")


GUIDE_FIXES = {
    "ERD10193": "船釣拋投鯛膠：Tairaba 優先，船上低彈道側拋、落底後斜向平收和遠端掛口更穩，適合 40～150g、PE0.5～1.2号 的淺場障礙邊搜索。",
    "ERD10194": "船釣拋投鯛膠：Tairaba 優先，船上低彈道側拋、落底後斜向平收和遠端掛口更穩，適合 40～80g、PE0.5～1.2号 的淺場精細搜索。",
    "ERD10195": "船釣拋投鯛膠：Casting Tairaba / Tairaba 優先，船上低彈道側拋、落底後斜向平收和遠端掛口更穩，Eging 只作同類抽停操作的副用途邊界。",
    "ERD10196": "船釣拋投鯛膠：Tairaba 優先，船上低彈道側拋、落底後斜向平收和遠端掛口更穩，適合 40～80g、PE0.5～1.2号 的淺場精細搜索。",
    "ERD10264": "尺級 Mebaring：Mebaring Jighead / Ajing Jighead 優先，近距離結構邊精準打點、偏重釣組操作和障礙邊控魚更直接，Micro Jigging / Minnow 作副用途搜索。",
    "ERD10276": "輕海水多用：Mebaring Jighead / Light Game Jighead / Jighead Rig 優先，小型 Plug、Metal Jig 與 Texas/Free Rig 可按標點切換，Eging 只作輕量副用途。",
    "ERD10279": "強力輕海水：Mebaring Jighead / Light Game Jighead / Jighead Rig / Carolina Rig 優先，遠投浮漂、重 Jighead 和 Rockfish 底操更穩，Offshore Jigging 只作副用途邊界。",
}


def snapshot_except(ws, excluded_col):
    headers = [cell.value for cell in ws[1]]
    return [
        {
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != excluded_col
        }
        for row in range(2, ws.max_row + 1)
    ]


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    for required in ["id", "SKU", "recommended_rig_pairing", "guide_use_hint"]:
        if required not in col:
            raise RuntimeError(f"missing column: {required}")

    guide_col = col["guide_use_hint"]
    before = snapshot_except(ws, guide_col)
    changes = []
    seen = set()
    for row in range(2, ws.max_row + 1):
        detail_id = str(ws.cell(row, col["id"]).value or "").strip()
        if detail_id not in GUIDE_FIXES:
            continue
        seen.add(detail_id)
        old = str(ws.cell(row, guide_col).value or "").strip()
        new = GUIDE_FIXES[detail_id]
        if old != new:
            cell = ws.cell(row, guide_col)
            cell.value = new
            cell.fill = copy(YELLOW_FILL)
            changes.append({
                "row": row,
                "id": detail_id,
                "sku": ws.cell(row, col["SKU"]).value,
                "recommended_rig_pairing": ws.cell(row, col["recommended_rig_pairing"]).value,
                "old": old,
                "new": new,
            })

    missing = sorted(set(GUIDE_FIXES) - seen)
    if missing:
        raise RuntimeError(f"target ids not found: {missing}")
    if before != snapshot_except(ws, guide_col):
        raise RuntimeError("unexpected changes outside guide_use_hint")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_file": str(XLSX_PATH),
        "changed_rows": len(changes),
        "field": "guide_use_hint",
        "reason": "Only rows with obvious conflict between guide_use_hint, recommended_rig_pairing, Description, and player fields were corrected during player-field QA.",
        "changes": changes,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({"changed_rows": len(changes), "report_file": str(REPORT_PATH)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
