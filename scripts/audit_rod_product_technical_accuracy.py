#!/usr/bin/env python3
import json
import importlib.util
from collections import defaultdict
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
REPORT_PATH = DATA_DIR / "rod_product_technical_accuracy_audit_report.json"
FIELD = "product_technical"
DAIWA_STAGE47_SCRIPT = ROOT / "scripts/fix_daiwa_rod_stage47_product_technical_from_variant_description.py"


def n(value):
    return str(value or "").strip()


def split_terms(value):
    return [part.strip() for part in n(value).split(" / ") if part.strip()]


def headers(ws):
    return [cell.value for cell in ws[1]]


def load_sheet_rows(path):
    wb = load_workbook(path, read_only=False, data_only=True)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_headers = headers(rod_ws)
    detail_headers = headers(detail_ws)
    rod_col = {name: idx + 1 for idx, name in enumerate(rod_headers)}
    detail_col = {name: idx + 1 for idx, name in enumerate(detail_headers)}

    rods = {}
    for row_index, values in enumerate(rod_ws.iter_rows(min_row=2, values_only=True), start=2):
        row_values = {header: values[idx] if idx < len(values) else "" for idx, header in enumerate(rod_headers)}
        rod_id = n(row_values.get("id"))
        rods[rod_id] = {
            "row": row_index,
            "id": rod_id,
            "model": n(row_values.get("model")),
        }

    rows = []
    for row_index, values in enumerate(detail_ws.iter_rows(min_row=2, values_only=True), start=2):
        row_values = {header: values[idx] if idx < len(values) else "" for idx, header in enumerate(detail_headers)}
        rows.append(
            {
                "row": row_index,
                "id": n(row_values.get("id")),
                "rod_id": n(row_values.get("rod_id")),
                "sku": n(row_values.get("SKU")),
                "code_name": n(row_values.get("Code Name")),
                "description": n(row_values.get("Description")),
                "value": n(row_values.get(FIELD)),
            }
        )
    return rod_headers, detail_headers, rods, rows


def check_common(path, rod_headers, detail_headers, rows):
    issues = []
    if FIELD in rod_headers:
        issues.append({"severity": "error", "issue": "rod master contains product_technical"})
    if FIELD not in detail_headers:
        issues.append({"severity": "error", "issue": "rod_detail missing product_technical"})
    else:
        idx = detail_headers.index(FIELD)
        prev_header = detail_headers[idx - 1] if idx else ""
        next_header = detail_headers[idx + 1] if idx + 1 < len(detail_headers) else ""
        if prev_header != "Description" or next_header != "Extra Spec 1":
            issues.append(
                {
                    "severity": "error",
                    "issue": "wrong column position",
                    "previous": prev_header,
                    "next": next_header,
                }
            )
    forbidden = ["白名单", "玩家", "钓组", "饵型", "来源", "官网确认", "General Lure", "Light Rig"]
    for row in rows:
        value = row["value"]
        if not value:
            continue
        if "/" in value and " / " not in value:
            issues.append({"severity": "error", "issue": "bad separator", "row": row["row"], "value": value})
        for fragment in forbidden:
            if fragment in value:
                issues.append(
                    {"severity": "error", "issue": "forbidden fragment", "fragment": fragment, "row": row["row"], "value": value}
                )
    return issues


def group_same_value_risks(rods, rows, brand):
    by_rod = defaultdict(list)
    for row in rows:
        by_rod[row["rod_id"]].append(row)
    risks = []
    for rod_id, group in by_rod.items():
        values = {row["value"] for row in group}
        nonempty = [value for value in values if value]
        if len(group) <= 1 or not nonempty:
            continue
        if len(nonempty) == 1 and "" not in values:
            value = nonempty[0]
            risk = "same_value_multi_sku_needs_evidence_review"
            if brand in {"abu", "ark", "daiwa"}:
                risk = "same_value_multi_sku_series_level_source"
            risks.append(
                {
                    "rod_id": rod_id,
                    "model": rods.get(rod_id, {}).get("model", ""),
                    "variant_count": len(group),
                    "risk": risk,
                    "value": value,
                }
            )
    return risks


def audit_abu(rows):
    allowed_terms = {
        "ROCS™",
        "Powerlux® 100",
        "Powerlux® 200",
        "Powerlux® 500",
        "Powerlux® 1000",
        "IntraCarbon™",
        "CCRS™",
        "2 piece ferrule locking mechanism",
    }
    issues = []
    for row in rows:
        for term in split_terms(row["value"]):
            if term not in allowed_terms:
                issues.append({"severity": "error", "issue": "unexpected Abu technical term", **row, "term": term})
    return issues


def audit_daiwa(rows, stage46_report, stage47_report, stage47_module):
    if not stage46_report:
        return [
            {
                "severity": "warning",
                "issue": "Daiwa SKU applicability report is missing; current values need SKU-level evidence review",
            }
        ]
    expected_by_id = {
        n(row.get("id")): n(row.get("after"))
        for row in stage46_report.get("evidence", [])
        if n(row.get("id"))
    }
    if stage47_report:
        for change in stage47_report.get("changes", []):
            row_id = n(change.get("id"))
            if row_id:
                expected_by_id[row_id] = n(change.get("after"))
    if stage47_module:
        for row in rows:
            terms = stage47_module.extract_variant_technology(row["description"])
            if terms:
                expected_by_id[row["id"]] = stage47_module.DELIMITER.join(terms)
    mismatches = []
    for row in rows:
        if row["id"] not in expected_by_id:
            continue
        expected = expected_by_id[row["id"]]
        if row["value"] != expected:
            mismatches.append(
                {
                    "row": row["row"],
                    "id": row["id"],
                    "rod_id": row["rod_id"],
                    "sku": row["sku"],
                    "current": row["value"],
                    "expected": expected,
                }
            )
    if mismatches:
        return [
            {
                "severity": "error",
                "issue": "Daiwa product_technical differs from SKU applicability evidence",
                "mismatches": mismatches[:20],
                "mismatch_count": len(mismatches),
            }
        ]
    return []


def load_json(path):
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def load_daiwa_stage47_extractor():
    if not DAIWA_STAGE47_SCRIPT.exists():
        return None
    spec = importlib.util.spec_from_file_location("daiwa_stage47_product_technical", DAIWA_STAGE47_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def main():
    report = {
        "field": FIELD,
        "scope": "rod_detail product_technical accuracy and SKU matching audit",
        "files": [],
        "summary": {},
        "global_issues": [],
    }

    shimano_gate = load_json(DATA_DIR / "shimano_rod_product_technical_gate_report.json")
    daiwa_stage46 = load_json(DATA_DIR / "daiwa_rod_product_technical_detail_match_stage46_report.json")
    daiwa_stage47 = load_json(DATA_DIR / "daiwa_rod_product_technical_variant_desc_stage47_report.json")
    daiwa_stage47_module = load_daiwa_stage47_extractor()
    if shimano_gate:
        report["summary"]["shimano_gate"] = {
            "apply_allowed": shimano_gate.get("apply_allowed"),
            "risk_count": shimano_gate.get("risk_count"),
            "items_with_risk": shimano_gate.get("summary", {}).get("items_with_risk"),
            "all_same_risk_count": shimano_gate.get("summary", {}).get("all_same_risk_count"),
            "xlsx_mismatch_count": shimano_gate.get("summary", {}).get("xlsx_mismatch_count"),
        }
    if daiwa_stage46:
        report["summary"]["daiwa_stage46"] = {
            "rows": daiwa_stage46.get("rows"),
            "nonempty_after": daiwa_stage46.get("nonempty_after"),
            "changed_import": daiwa_stage46.get("changed_import"),
            "evidence_rows": len(daiwa_stage46.get("evidence", [])),
            "source_policy": daiwa_stage46.get("source_policy"),
        }
    if daiwa_stage47:
        report["summary"]["daiwa_stage47"] = {
            "changed_import": daiwa_stage47.get("changed_import"),
            "changed_rate": daiwa_stage47.get("changed_rate"),
            "source_policy": daiwa_stage47.get("source_policy"),
        }

    total_errors = 0
    total_warnings = 0
    for path in sorted(DATA_DIR.glob("*_rod_import.xlsx")):
        brand = path.name.split("_rod_import.xlsx")[0]
        rod_headers, detail_headers, rods, rows = load_sheet_rows(path)
        values = [row["value"] for row in rows if row["value"]]
        issues = check_common(path, rod_headers, detail_headers, rows)
        if brand == "abu":
            issues.extend(audit_abu(rows))
        if brand == "daiwa":
            issues.extend(audit_daiwa(rows, daiwa_stage46, daiwa_stage47, daiwa_stage47_module))

        same_value_risks = group_same_value_risks(rods, rows, brand)
        if brand == "shimano" and shimano_gate and shimano_gate.get("apply_allowed") is False:
            issues.append(
                {
                    "severity": "warning",
                    "issue": "Shimano gate is not passed; same-value multi-SKU rows need SKU-level evidence before being considered final",
                    "risk_count": shimano_gate.get("risk_count"),
                }
            )
        errors = [issue for issue in issues if issue.get("severity") == "error"]
        warnings = [issue for issue in issues if issue.get("severity") != "error"]
        total_errors += len(errors)
        total_warnings += len(warnings)
        report["files"].append(
            {
                "file": path.name,
                "rows": len(rows),
                "nonempty": len(values),
                "unique_values": len(set(values)),
                "rod_has_product_technical": FIELD in rod_headers,
                "detail_column_position": {
                    "previous": detail_headers[detail_headers.index(FIELD) - 1] if FIELD in detail_headers and detail_headers.index(FIELD) else "",
                    "next": detail_headers[detail_headers.index(FIELD) + 1] if FIELD in detail_headers and detail_headers.index(FIELD) + 1 < len(detail_headers) else "",
                },
                "same_value_multi_sku_groups": len(same_value_risks),
                "same_value_examples": same_value_risks[:10],
                "errors": errors[:20],
                "warnings": warnings[:20],
            }
        )

    report["summary"].update({"error_count": total_errors, "warning_count": total_warnings})
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
