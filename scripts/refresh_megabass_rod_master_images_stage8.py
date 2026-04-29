import json
import re
import shutil
import subprocess
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = ROOT / "GearSage-client/pkgGear/data_raw/megabass_rod_import.xlsx"
RAW_PATH = ROOT / "GearSage-client/pkgGear/data_raw/megabass_rod_raw.json"
REPORT_PATH = ROOT / "GearSage-client/pkgGear/data_raw/megabass_rod_master_images_refresh_report.json"

BRAND_DIR = "megabass_rods"
OLD_DIR = Path("/Users/tommy/Pictures/images_old_copy") / BRAND_DIR
TARGET_DIR = Path("/Users/tommy/Pictures/images") / BRAND_DIR
STATIC_PREFIX = f"https://static.gearsage.club/gearsage/Gearimg/images/{BRAND_DIR}"


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def ext_from_url(url):
    suffix = Path(urlparse(url).path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
        return ".jpg" if suffix == ".jpeg" else suffix
    return ".jpg"


def safe_filename(model_name, image_url):
    text = n(model_name)
    text = text.replace("/", "_")
    text = re.sub(r"[^\w.\-+ &]+", "_", text, flags=re.ASCII)
    text = re.sub(r"_+", "_", text).strip("_")
    return f"{text}_main{ext_from_url(image_url)}"


def basename_from_current(value):
    text = n(value)
    if not text:
        return ""
    parsed = urlparse(text)
    name = Path(parsed.path if parsed.scheme else text).name
    return name if name else ""


def is_usable_image(path):
    if not path.exists() or path.stat().st_size < 1024:
        return False
    try:
        output = subprocess.check_output(["file", str(path)], text=True)
    except Exception:
        return False
    return "image" in output.lower()


def download(url, dest):
    subprocess.check_call([
        "curl",
        "-L",
        "--fail",
        "--silent",
        "--show-error",
        "--max-time",
        "45",
        "-o",
        str(dest),
        url,
    ])


def main():
    TARGET_DIR.mkdir(parents=True, exist_ok=True)

    raw = json.loads(RAW_PATH.read_text())
    raw_by_series = {}
    for item in raw:
        series = n(item.get("series_name"))
        if series and series not in raw_by_series:
            raw_by_series[series] = item

    wb = load_workbook(XLSX_PATH)
    ws = wb["rod"]
    col = {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
    required = ["id", "model", "images"]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing rod columns: {missing}")

    rows = []
    changed_cells = 0
    copied = 0
    downloaded = 0
    reused_target = 0
    missing_sources = []

    for row_idx in range(2, ws.max_row + 1):
        row_id = n(ws.cell(row_idx, col["id"]).value)
        model = n(ws.cell(row_idx, col["model"]).value)
        current_image = n(ws.cell(row_idx, col["images"]).value)
        raw_item = raw_by_series.get(model, {})
        source_url = n((raw_item.get("images") or [""])[0])
        source_model = n(raw_item.get("model_name")) or model

        filename = basename_from_current(current_image)
        if not filename:
            filename = safe_filename(source_model, source_url)

        target = TARGET_DIR / filename
        old = OLD_DIR / filename
        action = ""

        if is_usable_image(target):
            reused_target += 1
            action = "reuse_target"
        elif is_usable_image(old):
            shutil.copy2(old, target)
            copied += 1
            action = "copy_old"
        elif source_url:
            download(source_url, target)
            downloaded += 1
            action = "download"
        else:
            missing_sources.append({"id": row_id, "model": model, "filename": filename})
            action = "missing_source"

        if action != "missing_source" and not is_usable_image(target):
            raise RuntimeError(f"image not usable after {action}: {target}")

        next_url = f"{STATIC_PREFIX}/{filename}"
        cell = ws.cell(row_idx, col["images"])
        old_value = n(cell.value)
        if old_value != next_url:
            cell.value = next_url
            changed_cells += 1

        rows.append({
            "id": row_id,
            "model": model,
            "source_model": source_model,
            "source_url": source_url,
            "filename": filename,
            "target_path": str(target),
            "images_value": next_url,
            "action": action,
        })

    wb.save(XLSX_PATH)

    report = {
        "xlsx_file": str(XLSX_PATH),
        "old_dir": str(OLD_DIR),
        "target_dir": str(TARGET_DIR),
        "static_prefix": STATIC_PREFIX,
        "total_master_rows": len(rows),
        "changed_image_cells": changed_cells,
        "copied_from_old": copied,
        "downloaded": downloaded,
        "reused_target": reused_target,
        "missing_sources": missing_sources,
        "rows": rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "total_master_rows": len(rows),
        "changed_image_cells": changed_cells,
        "copied_from_old": copied,
        "downloaded": downloaded,
        "reused_target": reused_target,
        "missing_sources": len(missing_sources),
        "target_dir": str(TARGET_DIR),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
