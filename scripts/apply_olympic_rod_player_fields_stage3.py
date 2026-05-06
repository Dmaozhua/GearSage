import json
import re
import subprocess
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('olympic_rod_import.xlsx')
REPORT_PATH = resolve_data_raw('olympic_rod_player_fields_stage3_report.json')
SHADE_SCRIPT = ROOT / "scripts/shade_olympic_rod_detail_groups.py"

TARGET_FIELDS = ["player_environment", "player_positioning", "player_selling_points"]
READ_FIELDS = ["SKU", "Description", "recommended_rig_pairing", "guide_use_hint", "official_environment"]


PLAYER_FIELDS = {
    "20GVIGC-71H": {
        "player_environment": "淡水 Bass / 岸邊與船上 cover 打撃 / weed edge",
        "player_positioning": "Texas 與 Punch Shot 主軸的障礙打撃竿",
        "player_selling_points": "3/16-3/8oz Texas 和 1/4-1oz Punch Shot 落點控制清楚，能在 weed cover 內補刺並把魚拉離障礙；Wacky、Neko 可作慢速誘食補充。",
        "basis": "whitelist_plus_pairing_description",
    },
    "20GVIGC-75M": {
        "player_environment": "淡水 Bass / 開放水域與岸投硬餌搜索 / 冬季金屬餌",
        "player_positioning": "Vibration、Jerkbait 主軸的巻物與操作硬餌竿",
        "player_selling_points": "1/2-5/8oz Vibration、Jerkbait 連續搜索時保留操作性，Scone Rig 可承接吸い込み bite；冬季 Metal Vibration / Metal Jig 也能做反應搜索。",
        "basis": "pairing_description",
    },
    "20GVIGC-76MH": {
        "player_environment": "淡水 Bass / 中重型巻物搜索 / lake grass edge",
        "player_positioning": "Spinnerbait Slow Roll 與中重型 moving bait 專用",
        "player_selling_points": "1/2-1.5oz 巻物負荷下竿身彎曲過渡穩，能讀出 blade 振動；太軸單鉤補刺速度和力量足，適合 slow roll、Chatterbait、Swim Jig 搜索。",
        "basis": "whitelist_plus_pairing_description",
    },
    "20GVIGC-77XH": {
        "player_environment": "淡水 Bass / 大型餌拋投 / 開放水域與精準打點",
        "player_positioning": "Big Bait、Crawler Bait、Swimbait 高負荷竿",
        "player_selling_points": "XH blank 承接 Big Bait 和 Crawler Bait 的拋投負荷，4 軸組布和 T1100G 幫助減少 cast wobble；適合重餌精準落點和穩定泳姿回收。",
        "basis": "whitelist_plus_pairing_description",
    },
    "20GVIGS-610ML": {
        "player_environment": "淡水 Bass / 岸邊 light cover / 直柄精細與小型 plug",
        "player_positioning": "No Sinker、Neko、虫系兼小型 plug 的 power finesse spin",
        "player_selling_points": "細線下能操作 No Sinker、Neko 和虫系，也能切到小型 plug 搜索；ML power 留有 cover 內控魚餘量，不只是純開放水域精細竿。",
        "basis": "whitelist_plus_pairing_description",
    },
    "GVIGS-6102ML": {
        "player_environment": "淡水 Bass / 岸釣移動作戰 / light rig 全般",
        "player_positioning": "攜行取向的 light rig 泛用兩節直柄",
        "player_selling_points": "中心切 2 piece 方便岸釣移動，Regular-Fast 設定讓 Neko、No Sinker、Down Shot 的拋投、誘い、landing 動作連貫；口徑保守，不擴寫到重 cover。",
        "basis": "pairing_description_conservative",
    },
    "20GVIGS-742M": {
        "player_environment": "淡水 Bass / PE 遠投 / 開放水域 light rig 與 plug",
        "player_positioning": "PE light rig 遠投與 plug 操作兼用直柄",
        "player_selling_points": "7'4\" 長度和 M power 支撐 Long Cast Neko、PE Down Shot 的遠距控線，遠端細微違和感更容易讀取；Small Crankbait、Topwater 可做遠投搜索補充。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUC-65ML": {
        "player_environment": "淡水 Bass / shallow cover edge / 小中型硬餌與 cover finesse",
        "player_positioning": "小中型 Topwater / Shallow Crank 加 cover bait finesse",
        "player_selling_points": "Topwater Plug、Shallow Crankbait 做近中距離搜索，Cover Neko 和 Small Rubber Jig 負責輕 cover 內精細打點；ML power 讓拋投準度和彎曲承接更友好。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUC-69MH": {
        "player_environment": "淡水 Bass / cover 邊緣與開放水域切換 / bait versatile",
        "player_positioning": "Texas、Light Rubber Jig 主軸的打撃搜索兩用",
        "player_selling_points": "3/16-3/8oz Texas、Light Rubber Jig 和高比重 No Sinker 可精準打點，3/8-1/2oz Spinnerbait / Chatterbait 能快速搜索；一支竿覆蓋軟餌與 wire bait 切換。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUC-610M": {
        "player_environment": "淡水 Bass / 開放水域中型硬餌 / 常規岸船泛用",
        "player_positioning": "Medium Crankbait、Spinnerbait 與高比重 No Sinker 泛用 bait",
        "player_selling_points": "中型 Crankbait / Minnow 和 1/2oz 內 Spinnerbait 負責搜索，高比重 No Sinker、1/4oz Texas 負責慢速打點；彎曲曲線寬容，拋投手感和準度較友好。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUC-70H": {
        "player_environment": "淡水 Bass / heavy cover / vegetation frog game",
        "player_positioning": "Texas、Rubber Jig、Heavy Spinnerbait、Frog 的重障礙泛用",
        "player_selling_points": "3/8oz 以上 Texas 和 Rubber Jig 可攻 cover，Heavy Spinnerbait slow roll 能沿草邊搜索，Frog 可處理 vegetation；H power 更重視控魚和脫離障礙。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUC-74X": {
        "player_environment": "淡水 Bass / heavy cover 與大型餌 / 遠距高負荷",
        "player_positioning": "Heavy Texas 與 MAX3oz Big Bait / Swimbait 強力竿",
        "player_selling_points": "Heavy Texas 用於 cover 打點，Big Bait / Swimbait 可到 MAX3oz；不是單純硬竿，粘り能承接重餌拋投與咬口跟隨，grip joint 也利於攜行。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUS-64L": {
        "player_environment": "淡水 Bass / 近中距離 finesse / 小型 plug 搜索",
        "player_positioning": "短尺 light rig 與 small plug 直柄泛用",
        "player_selling_points": "6'4\" 長度提升近距離落點和竿尖操作，Neko、Down Shot、No Sinker 做食わせ，小型 Crankbait / Minnow 可搜索；較 regular 的調性讓拋投手感更順。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUS-610ML": {
        "player_environment": "淡水 Bass / 岸投遠投 finesse / 開放水域長距離控線",
        "player_positioning": "Long Cast Neko、Down Shot 與小型 plug 的長尺直柄",
        "player_selling_points": "6'10\" 長度放大 light rig 投送距離，輕量平衡保留細操作；Neko、Down Shot、No Sinker 是主軸，小型 Crankbait / Minnow 用於遠處搜索。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUS-611H-PF": {
        "player_environment": "淡水 Bass / heavy cover power finesse / PE 直柄強攻",
        "player_positioning": "Cover Neko 與 Small Rubber Jig 的 heavy cover power spin",
        "player_selling_points": "PE 線下可把 Cover Neko、Small Rubber Jig 和高比重 No Sinker 打進難攻 cover，短柄提升操作和起魚節奏；重點是精細餌進 cover 後的強制控魚。",
        "basis": "whitelist_plus_pairing_description",
    },
    "21GVELUS-742M": {
        "player_environment": "淡水 Bass / PE 遠投 / light rig 到中量 moving bait",
        "player_positioning": "PE 遠投 light rig 兼 1/2oz moving bait 直柄",
        "player_selling_points": "Long Cast Neko、PE Down Shot、No Sinker 是核心，7'4\" 長度提升遠投與回線效率；Spinnerbait、Vibration 可作較重搜索端，混合用途邊界清楚。",
        "basis": "whitelist_plus_pairing_description",
    },
    "GSBS-602XUL": {
        "player_environment": "淡水 Area Trout / 小池高壓場 / micro spoon-crank",
        "player_positioning": "XUL micro spoon 與 micro crank 精細型",
        "player_selling_points": "0.4-3.5g 與 1-3lb 線組適合低負荷 Micro Spoon、Micro Crankbait；高感 cork 與輕量 guide 對短咬口和細線控魚更有幫助。",
        "basis": "pairing_spec_conservative",
    },
    "GSBS-642UL": {
        "player_environment": "淡水 Area Trout / 標準池 spoon-crank / 中距離控魚",
        "player_positioning": "UL spoon、crank、small minnow 中量泛用",
        "player_selling_points": "0.5-5g、1.5-4lb 覆蓋標準 Spoon、Crankbait 和 Small Minnow，較 602XUL 更能應付中距離和稍大魚，仍保留細線操作感。",
        "basis": "pairing_spec_conservative",
    },
    "GSBS-672L": {
        "player_environment": "淡水 Area Trout / 大池遠投 / full-size lure",
        "player_positioning": "L power full-size spoon / crank / minnow 強化型",
        "player_selling_points": "0.6-7g、2-5lb 可承接 Full-size Spoon、Crankbait、Minnow，長度和 L power 更適合大池遠投與較大型 trout 控魚。",
        "basis": "pairing_spec_conservative",
    },
    "26GBELPS-572XUL-S": {
        "player_environment": "淡水 Area Trout / 高壓慢魚 / ester line micro lure",
        "player_positioning": "solid tip micro spoon / micro crank 掛け型精細",
        "player_selling_points": "Micro Spoon、Micro Crankbait 在シビア場更容易讓魚口殘留，solid tip 吸收短咬，tubular 部分張力可補縦釣り操作。",
        "basis": "pairing_description",
    },
    "26GBELPS-582SUL-T": {
        "player_environment": "淡水 Area Trout / 標準放流後節奏 / spoon-crank 巻き",
        "player_positioning": "1.6g Spoon 與 Crankbait 的掛け型巻き竿",
        "player_selling_points": "1.6g 前後 Spoon、Crankbait 平收時能主動掛けに行く，張りを活かして操作系 plug 也能處理；適合 ester line 主流配置。",
        "basis": "pairing_description",
    },
    "26GBELPS-612UL-T": {
        "player_environment": "淡水 Area Trout / 放流狩り / 大型 plug 與色物",
        "player_positioning": "Heavy Spoon、Full-size Crank、Minnow Twitching 攻擊型",
        "player_selling_points": "重め Spoon 和 Full-size Crank 用於放流魚快速搜索，Fast taper 可做 Minnow Twitching；bat power 對大型 trout 和色物更安心。",
        "basis": "pairing_description",
    },
    "26GBELPS-642L-T": {
        "player_environment": "淡水 Area Trout / 大規模 pond / 超大型 trout",
        "player_positioning": "Large Plug、Full-size Spoon 的大型魚 power 型",
        "player_selling_points": "L power 對大型 plug、Full-size Spoon 負荷更穩，適合大規模 pond 搜索和超大型 trout 控魚；不是 micro lure 精細路線。",
        "basis": "pairing_description",
    },
    "26GBELPS-672SUL-T": {
        "player_environment": "淡水 Area Trout / 大池沖目遠投 / 未被打擾魚群",
        "player_positioning": "Long Cast Spoon / Crank 遠投搜索型",
        "player_selling_points": "6'7\" 長度把 Spoon、Crankbait 投到大池沖目，SUL 調性在遠距離中魚後更不易脫鉤；適合找未被釣壓干擾的魚。",
        "basis": "pairing_description",
    },
    "24GBELUS-582XUL-T": {
        "player_environment": "淡水 Area Trout / micro spoon-crank 定速巻き",
        "player_positioning": "Micro Spoon、Micro Crankbait short-bite 對應型",
        "player_selling_points": "定速 retrieve micro lure 時，XUL tubular 和柔軟性降低彈口；OP-01 碳纖 reel seat 提升短咬口感知。",
        "basis": "pairing_description",
    },
    "25GBELUS-612SUL-S": {
        "player_environment": "淡水 Area Trout / slow retrieve / 高壓短咬",
        "player_positioning": "Slow Spoon 與 Micro Crank solid tip 乗せ掛け型",
        "player_selling_points": "Slow Spoon、Micro Crankbait 是主軸，solid tip 讓高壓短咬時 hook 留在口內，掛けた後 belly 順暢彎曲降低脫鉤。",
        "basis": "pairing_description",
    },
    "24GBELUS-622SUL-T": {
        "player_environment": "淡水 Area Trout / bottom-twitch 操作系 plug / spooning 兼用",
        "player_positioning": "Vibration、Minnow Twitching、Bottom Bump 操作型",
        "player_selling_points": "Fast-ish action 讓 Vibration、Minnow Twitching 和 Bottom Bump 更清楚，遇到穩定魚情可切回 Spoon；不是單純慢巻き型。",
        "basis": "pairing_description",
    },
    "24GBELUS-652UL-T": {
        "player_environment": "淡水 Area Trout / 沖目遠投 / 大型 trout",
        "player_positioning": "Long Cast Crankbait / Spoon 遠投泛用型",
        "player_selling_points": "較長竿身和系列最強 bat power 讓 Long Cast Crankbait、Spoon、Minnow 能覆蓋沖目魚，對大型 trout 補刺和控魚更穩。",
        "basis": "pairing_description",
    },
    "25GBELUS-682SUL-T": {
        "player_environment": "淡水 Area Trout / 廣大場地最遠投 / plug-spoon 操作回收",
        "player_positioning": "最長尺 Long Cast Crank / Spoon / Minnow Twitching 型",
        "player_selling_points": "6'8\" 最長尺用來觸及大場未被打擾的沖目魚，Crankbait、Spoon 和 Minnow Twitching 都能處理，遠距離回收和控魚距離更有優勢。",
        "basis": "pairing_description",
    },
}


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def header_map(ws):
    return {n(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if n(cell.value)}


def snapshot_without_targets(ws, col, target_fields):
    target_cols = {col[field] for field in target_fields}
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            ws.cell(1, c).value: ws.cell(row, c).value
            for c in range(1, ws.max_column + 1)
            if c not in target_cols
        })
    return rows


def classify_field(value):
    text = n(value)
    if "Bass" in text or "cover" in text or "frog" in text.lower():
        return "bass"
    if "管釣" in text or "Trout" in text or "Area" in text:
        return "trout"
    return ""


def validate_rows(ws, col):
    problems = []
    generic_terms = [
        "性能优秀",
        "適合多場景",
        "多场景",
        "很好搭配",
        "用途广",
        "泛用覆蓋較廣",
        "官网",
        "白名单",
        "tackledb",
        "evidence",
        "source",
        "證據",
        "证据",
    ]
    env_values, pos_values, sell_values = set(), set(), set()

    for row in range(2, ws.max_row + 1):
        sku = n(ws.cell(row, col["SKU"]).value)
        official = n(ws.cell(row, col["official_environment"]).value)
        pair = n(ws.cell(row, col["recommended_rig_pairing"]).value)
        desc = n(ws.cell(row, col["Description"]).value)
        env = n(ws.cell(row, col["player_environment"]).value)
        pos = n(ws.cell(row, col["player_positioning"]).value)
        sell = n(ws.cell(row, col["player_selling_points"]).value)
        env_values.add(env)
        pos_values.add(pos)
        sell_values.add(sell)

        for field_name, value in [
            ("player_environment", env),
            ("player_positioning", pos),
            ("player_selling_points", sell),
        ]:
            if not value:
                problems.append({"sku": sku, "type": "blank", "field": field_name})
            for term in generic_terms:
                if term.lower() in value.lower():
                    problems.append({"sku": sku, "type": "generic_or_source_term", "field": field_name, "term": term})

        if official == "Black Bass":
            if any(token in env + pos + sell for token in ["海水", "船釣", "木虾", "木蝦", "管釣"]):
                problems.append({"sku": sku, "type": "bass_misclassified_as_non_bass"})
            if any(token in pair for token in ["Big Bait", "Swimbait", "Crawler"]) and "大" not in pos and "Big" not in pos:
                problems.append({"sku": sku, "type": "big_bait_pairing_not_reflected", "pairing": pair, "positioning": pos})
            if any(token in pair for token in ["Punch", "Cover", "Heavy Texas", "Heavy Cover"]) and not any(token in env + pos + sell for token in ["cover", "障礙", "打撃", "Heavy", "強力"]):
                problems.append({"sku": sku, "type": "cover_soft_pairing_not_reflected", "pairing": pair})
        elif official == "Area Trout":
            if any(token in env + pos + sell for token in ["Bass", "海水", "船釣", "木虾", "木蝦", "frog", "Frog"]):
                problems.append({"sku": sku, "type": "trout_misclassified"})
            if any(token in pair for token in ["Long Cast", "Full-size", "Large", "Heavy"]) and not any(token in env + pos + sell for token in ["遠投", "大型", "大池", "大規模", "沖", "Full-size", "Heavy"]):
                problems.append({"sku": sku, "type": "heavy_trout_pairing_not_reflected", "pairing": pair})

        if "ビッグベイト" in desc and "大" not in pos + env + sell and "Big" not in pair:
            problems.append({"sku": sku, "type": "description_bigbait_not_reflected"})
        if "マイクロ" in desc and "micro" not in (env + pos + sell + pair).lower():
            problems.append({"sku": sku, "type": "description_micro_not_reflected"})

    return problems, {
        "player_environment_unique": len(env_values),
        "player_positioning_unique": len(pos_values),
        "player_selling_points_unique": len(sell_values),
    }


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    col = header_map(ws)
    for field in [*READ_FIELDS, *TARGET_FIELDS, "id"]:
        if field not in col:
            raise RuntimeError(f"missing column: {field}")

    before = snapshot_without_targets(ws, col, TARGET_FIELDS)
    changed = []
    by_field = {field: 0 for field in TARGET_FIELDS}
    basis_counts = {}

    for row in range(2, ws.max_row + 1):
        sku = n(ws.cell(row, col["SKU"]).value)
        spec = PLAYER_FIELDS.get(sku)
        if not spec:
            raise RuntimeError(f"missing player field mapping: {sku}")
        row_changes = {}
        for field in TARGET_FIELDS:
            cell = ws.cell(row, col[field])
            old = n(cell.value)
            new = n(spec[field])
            if old != new:
                cell.value = new
                row_changes[field] = {"old": old, "new": new}
                by_field[field] += 1
        if row_changes:
            basis_counts[spec["basis"]] = basis_counts.get(spec["basis"], 0) + 1
            changed.append({
                "id": n(ws.cell(row, col["id"]).value),
                "sku": sku,
                "fields": sorted(row_changes),
                "basis": spec["basis"],
                "changes": row_changes,
            })

    after = snapshot_without_targets(ws, col, TARGET_FIELDS)
    if before != after:
        raise RuntimeError("unexpected non-player-field changes in rod_detail")

    problems, unique_counts = validate_rows(ws, col)
    if problems:
        raise RuntimeError(f"player field consistency check failed: {json.dumps(problems, ensure_ascii=False, indent=2)}")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    coverage = {field: 0 for field in TARGET_FIELDS}
    for row in range(2, ws.max_row + 1):
        for field in TARGET_FIELDS:
            if n(ws.cell(row, col[field]).value):
                coverage[field] += 1

    report = {
        "generated_at": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(),
        "xlsx_file": str(XLSX_PATH.relative_to(ROOT)),
        "target_fields": TARGET_FIELDS,
        "policy": {
            "write_scope": "Only rod_detail player_environment, player_positioning, player_selling_points were edited.",
            "source_priority": [
                "whitelist/player usage context where available",
                "confirmed recommended_rig_pairing",
                "Description/spec as boundary",
                "conservative SKU/power/action/lure-weight inference",
            ],
            "table_fields_do_not_contain_evidence_chain": True,
        },
        "rows": ws.max_row - 1,
        "coverage": coverage,
        "unique_counts": unique_counts,
        "changed_rows": len(changed),
        "changed_cells": sum(len(item["fields"]) for item in changed),
        "by_field": {field: count for field, count in by_field.items() if count},
        "basis_counts": basis_counts,
        "changes": changed,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "report_file": str(REPORT_PATH.relative_to(ROOT)),
        "coverage": coverage,
        "unique_counts": unique_counts,
        "changed_rows": report["changed_rows"],
        "changed_cells": report["changed_cells"],
        "by_field": report["by_field"],
        "basis_counts": basis_counts,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
