import json
import re
import subprocess
from collections import Counter
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('megabass_rod_import.xlsx')
EVIDENCE_PATH = resolve_data_raw('megabass_rod_whitelist_player_evidence.json')
REPORT_PATH = resolve_data_raw('megabass_rod_recommended_rig_pairing_report.json')
SHADE_SCRIPT = ROOT / "scripts/shade_megabass_rod_detail_groups_stage2.py"

FIELD = "recommended_rig_pairing"
INSERT_AFTER = "guide_use_hint"
MAX_ITEMS = 8


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def add(matches, label, pos, source):
    if not label:
        return
    for item in matches:
        if item["label"] == label:
            item["pos"] = min(item["pos"], pos)
            if source not in item["sources"]:
                item["sources"].append(source)
            return
    matches.append({"label": label, "pos": pos, "sources": [source]})


def add_if(matches, text, label, patterns, source="official_description"):
    best = None
    for pattern in patterns:
        found = re.search(pattern, text, re.I)
        if found:
            best = found.start() if best is None else min(best, found.start())
    if best is not None:
        add(matches, label, best, source)


def sorted_labels(matches):
    labels = [
        item["label"]
        for item in sorted(matches, key=lambda item: (item["pos"], item["label"]))[:MAX_ITEMS]
    ]
    return normalize_labels(labels)


def normalize_labels(labels):
    result = []
    for label in labels:
        if label == "Rubber Jig" and "Small Rubber Jig" in labels:
            continue
        if label == "Crankbait" and "Deep Crankbait" in labels:
            continue
        if label not in result:
            result.append(label)
    return result[:MAX_ITEMS]


def build_evidence_index():
    if not EVIDENCE_PATH.exists():
        return {}
    data = json.loads(EVIDENCE_PATH.read_text())
    by_id = {}
    for item in data.get("evidence", []):
        by_id.setdefault(n(item.get("id")), []).append(item)
    return by_id


def explicit_pairings(row):
    text = lower(f"{row.get('SKU')} {row.get('Code Name')} {row.get('Description')} {row.get('player_positioning')} {row.get('player_selling_points')}")
    matches = []

    # Product-name cues from Megabass descriptions. These are lures/rigs users recognize.
    add_if(matches, text, "Bug Lure", [r"\bmushi\b", r"\bsiglett\b", r"虫"])
    add_if(matches, text, "Popper", [r"\bpopx\b", r"\bpopmax\b", r"\bpopper\b", r"popping", r"ポップ"])
    add_if(matches, text, "Pencil Bait", [r"\bdog-?x\b", r"\bgatta-?x\b", r"\bpencil\b", r"dog\s*walk", r"ドッグウォーク"])
    add_if(matches, text, "Topwater Plug", [r"topwater", r"surface game", r"surface", r"\bi-wing\b", r"waterside topwater"])
    add_if(matches, text, "I-shaped Plug", [r"i-shaped", r"\bi[- ]?slide\b"])
    add_if(matches, text, "Jerkbait", [r"jerkbait", r"vision oneten", r"oneten", r"jerking", r"twitch"])
    add_if(matches, text, "Minnow", [r"\bminnows?\b", r"trout minnow", r"小型ミノー", r"ミノー", r"米诺"])
    add_if(matches, text, "Shad", [r"\bshad\b", r"shading-x"])
    add_if(matches, text, "Deep Crankbait", [r"deep crank", r"super-deep crank", r"super deep crank", r"large crank", r"granade"])
    add_if(matches, text, "Crankbait", [r"crankbait", r"cranking", r"\bcranks?\b", r"small crank", r"mini crank", r"light cranking"])
    add_if(matches, text, "Vibration", [r"vibration", r"\bvib\b", r"blade bait"])
    add_if(matches, text, "Spinnerbait", [r"spinnerbait", r"spinner bait"])
    add_if(matches, text, "Chatterbait", [r"chatterbait", r"bladed jig", r"bladed swim jig", r"wild header", r"刀片铅头钩", r"刀片式铅头钩"])
    add_if(matches, text, "Swim Jig", [r"swim jig", r"泳饵铅头钩"])
    add_if(matches, text, "Small Rubber Jig", [r"small rubber jig"])
    add_if(matches, text, "Rubber Jig", [r"rubber jig", r"\bjigs?\b", r"橡胶铅头钩", r"橡膠鉛頭鉤"])
    add_if(matches, text, "Football Jig", [r"football jig"])
    add_if(matches, text, "Jighead Rig", [r"jighead", r"jig head", r"mid-?strolling", r"shakyhead", r"shaky head"])
    add_if(matches, text, "Light Texas Rig", [r"light texas"])
    add_if(matches, text, "Heavy Texas Rig", [r"heavy texas"])
    add_if(matches, text, "Texas Rig", [r"\btexas\b"])
    add_if(matches, text, "Free Rig", [r"free rig"])
    add_if(matches, text, "Neko Rig", [r"\bneko\b"])
    add_if(matches, text, "Down Shot", [r"down[- ]?shot", r"drop[- ]?shot", r"dropshot"])
    add_if(matches, text, "No Sinker", [r"no[- ]?sinker", r"weightless", r"无铅头钩组", r"無鉛"])
    add_if(matches, text, "Wacky Rig", [r"wacky"])
    add_if(matches, text, "Nose Hook Rig", [r"nose-hook", r"nose hooking"])
    add_if(matches, text, "Frog", [r"\bfrog\b", r"frogging", r"heavy mat"])
    add_if(matches, text, "Punching", [r"punching", r"punch\b"])
    add_if(matches, text, "Dark Sleeper", [r"dark sleeper"])
    add_if(matches, text, "Sleeper Craw", [r"sleeper craw"])
    add_if(matches, text, "Craw", [r"\bcraw\b"])
    add_if(matches, text, "Swimbait", [r"swimbait", r"swim bait", r"magdraft", r"vatalion", r"bottom-tracing swimbait"])
    add_if(matches, text, "Big Bait", [r"bigbait", r"big bait", r"big[- ]game", r"giant bait", r"giant,?\s*40 cm", r"40 cm\+ bait", r"30-35 cm class", r"20-30 cm"])
    add_if(matches, text, "Magnum Spoon", [r"magnum-sized spoons?", r"spoon"])
    add_if(matches, text, "Metal Jig", [r"metal jig", r"shore jigging"])
    add_if(matches, text, "Shore Plug", [r"shore-based", r"wading", r"surf games", r"shore plug", r"shore plugging"])
    add_if(matches, text, "Small Plug", [r"small plug", r"small to medium plugs?", r"tiny plug", r"lightweight plugs?", r"ライトプラッギング", r"プラグ"])
    add_if(matches, text, "Spoon", [r"\bspoon\b", r"スプーン"])
    add_if(matches, text, "Trout Minnow", [r"trout", r"yamame", r"iwana", r"amago", r"satsuki", r"masu", r"brown trout", r"rainbow", r"ミノー"])
    add_if(matches, text, "Lead Core Trolling", [r"lead core"])
    add_if(matches, text, "Downrigger Trolling", [r"downrigger"])

    return sorted_labels(matches)


def whitelist_pairings(row, evidence_items):
    if not evidence_items:
        return []
    text = lower(" ".join(
        [
            n(item.get("suggested_player_positioning"))
            + " "
            + " ".join(n(v) for v in (item.get("source_specs") or {}).values())
            for item in evidence_items
        ]
    ))
    sku_text = lower(row.get("SKU"))

    if "big bait" in text or "swimbait" in text or re.search(r"\b[2-9](?:\.\d+)?(?:-|oz|max)", text):
        if re.search(r"swimbait|magdraft|big bait", lower(row.get("Description"))):
            return ["Swimbait", "Big Bait"]
    if "moving bait" in text:
        return ["Crankbait", "Spinnerbait", "Chatterbait", "Vibration"]
    if "轻量钓组" in text or "finesse" in text or re.search(r"^f[012]", sku_text):
        return ["Down Shot", "Neko Rig", "Jighead Rig", "Small Minnow"]
    if "重障碍" in text:
        return ["Texas Rig", "Rubber Jig", "Frog", "Chatterbait"]
    return []


def conservative_pairings(row, model):
    text = lower(f"{model} {row.get('SKU')} {row.get('POWER')} {row.get('TYPE')} {row.get('LURE WEIGHT')} {row.get('player_environment')} {row.get('player_positioning')} {row.get('player_selling_points')}")
    sku = lower(row.get("SKU"))
    typ = n(row.get("TYPE"))
    power = lower(row.get("POWER"))

    if "tracking buddy" in text or "lead core" in text:
        if "downrigger" in text and "lead core" in text:
            return ["Downrigger Trolling", "Lead Core Trolling", "Lake Trout Spoon"]
        if "downrigger" in text:
            return ["Downrigger Trolling", "Lake Trout Spoon", "Trolling Plug"]
        return ["Lead Core Trolling", "Lake Trout Spoon", "Trolling Plug"]

    if "greathunting" in text or "great hunting" in text or re.search(r"^gh", sku):
        if "river&lake" in text or "river" in text or "lake" in text:
            return ["Trout Minnow", "Spoon", "Sinking Minnow", "Small Plug"]
        if "x-glass" in text:
            return ["Trout Minnow", "Small Plug", "Spoon"]
        return ["Trout Minnow", "Sinking Minnow", "Spoon", "Small Plug"]

    if "pagani" in text:
        return ["Topwater Plug", "Popper", "Pencil Bait", "Wakebait"]

    if "ts" in sku or "destroyer t.s" in text:
        if typ == "S":
            return ["Down Shot", "Deep Crankbait", "Swimbait"]
        return ["Swimbait", "Big Bait", "Deep Crankbait", "Rubber Jig"]

    if "valkyrie" in text:
        if typ == "S":
            return ["Shore Plug", "Minnow", "Metal Jig", "Sinking Pencil"]
        if "xxh" in power or "xh" in power:
            return ["Big Bait", "Swimbait", "Frog", "Punching"]
        return ["Swimbait", "Spinnerbait", "Crankbait", "Shore Plug"]

    if "triza" in text:
        if typ == "S":
            return ["Jighead Rig", "Down Shot", "Small Plug", "Minnow"]
        if re.search(r"f[67]|h|xh", power + " " + sku):
            return ["Texas Rig", "Rubber Jig", "Frog", "Swimbait"]
        if re.search(r"f[0-2]", power + " " + sku):
            return ["Topwater Plug", "Small Plug", "BFS Jighead Rig", "Shad"]
        return ["Crankbait", "Spinnerbait", "Texas Rig", "Topwater Plug"]

    if typ == "S":
        if re.search(r"f0|f1|ul", power + " " + sku):
            return ["Down Shot", "Neko Rig", "Jighead Rig", "Small Minnow"]
        return ["Down Shot", "Neko Rig", "Jighead Rig", "Shad"]

    if re.search(r"f0|f1|f2|bfs|bait finesse", text):
        return ["BFS Jighead Rig", "Small Rubber Jig", "No Sinker", "Small Plug"]
    if re.search(r"f7|f8|f9|xh|xxh", text):
        return ["Frog", "Punching", "Rubber Jig", "Swimbait"]
    if re.search(r"f5|f6|mh|h", text):
        return ["Texas Rig", "Rubber Jig", "Spinnerbait", "Chatterbait"]
    return ["Crankbait", "Spinnerbait", "Texas Rig", "Jerkbait"]


def pairing_for(row, model, evidence_index):
    explicit = explicit_pairings(row)
    if explicit:
        return explicit, "official_description"
    white = whitelist_pairings(row, evidence_index.get(n(row.get("id")), []))
    if white:
        return white[:MAX_ITEMS], "whitelist_specs"
    return conservative_pairings(row, model)[:MAX_ITEMS], "conservative_specs"


SOFT = {
    "BFS Jighead Rig", "Jighead Rig", "Small Rubber Jig", "Rubber Jig", "Football Jig",
    "Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Neko Rig",
    "Down Shot", "No Sinker", "Wacky Rig", "Nose Hook Rig", "Dark Sleeper",
    "Sleeper Craw", "Craw",
}
HARD = {
    "Bug Lure", "Popper", "Pencil Bait", "Topwater Plug", "I-shaped Plug", "Jerkbait",
    "Minnow", "Shad", "Deep Crankbait", "Crankbait", "Vibration", "Spinnerbait",
    "Chatterbait", "Swim Jig", "Swimbait", "Big Bait", "Magnum Spoon", "Metal Jig",
    "Shore Plug", "Small Plug", "Spoon", "Trout Minnow", "Sinking Minnow", "Wakebait",
}
POWER = {"Frog", "Punching", "Big Bait", "Swimbait"}
TROUT = {"Trout Minnow", "Spoon", "Sinking Minnow", "Small Plug"}
TROLLING = {"Lead Core Trolling", "Downrigger Trolling", "Lake Trout Spoon", "Trolling Plug"}


def split_pairing(value):
    return [n(part) for part in n(value).split("/") if n(part)]


def family(items):
    labels = set(items)
    if labels & TROLLING:
        return "trolling"
    if labels & TROUT and not (labels & (SOFT - TROUT)):
        return "trout"
    has_soft = bool(labels & SOFT)
    has_hard = bool(labels & HARD)
    has_power = bool(labels & POWER)
    if has_power:
        return "power"
    if has_soft and has_hard:
        return "mixed"
    if has_soft:
        return "soft"
    if has_hard:
        return "hard"
    return "other"


def guide_conflict(pairing, hint):
    items = split_pairing(pairing)
    fam = family(items)
    value = n(hint)
    hard_only = "硬饵搜索" in value or "远投搜索" in value
    soft_only = "障碍区强力" in value or "轻线精细" in value
    versatile = any(word in value for word in ["泛用", "切换", "兼容", "软饵、硬饵"])

    if fam == "mixed" and not versatile:
        return "mixed_pairing_requires_versatile_hint"
    if fam == "soft" and hard_only:
        return "soft_primary_hard_hint"
    if fam == "hard" and "软饵底操" in value:
        return "hard_primary_soft_hint"
    if fam == "power" and not any(word in value for word in ["大饵", "Frog", "重障碍", "强力", "障碍"]):
        return "power_pairing_weak_hint"
    return ""


def guide_hint_for_pairing(pairing):
    fam = family(split_pairing(pairing))
    if fam == "mixed":
        return "泛用鲈钓：兼容软饵底操、硬饵搜索和移动饵切换，导环出线稳定，方便根据水层和障碍快速调整。"
    if fam == "soft":
        return "软饵钓组：细线到中线控线更直接，利于 jighead、Texas、Free Rig、Neko、Down Shot 等贴底或障碍操作。"
    if fam == "hard":
        return "硬饵搜索：抛投后线弧更稳定，利于 crank、jerkbait、minnow、spinnerbait 等维持泳层和节奏。"
    if fam == "power":
        if "Frog" in pairing or "Punching" in pairing:
            return "重障碍/Frog：粗线通过更稳定，导环能承受高负荷控鱼，适合把鱼从草洞或障碍中拉出来。"
        return "大饵强力：大饵抛投时出线更稳，粗线通过更可靠，也能支撑高负荷搏鱼。"
    if fam == "trout":
        return "溪流鳟鱼精细：小型米诺和轻饵更容易顺畅出线，细线控线和竿尖信号也更清楚。"
    if fam == "trolling":
        return "湖泊拖钓：让 lead core / downrigger 线组出线更稳定，并能承接长时间拖曳负荷。"
    return ""


def ensure_column(ws, field, insert_after):
    headers = [n(cell.value) for cell in ws[1]]
    if field in headers:
        return {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}, False
    if insert_after not in headers:
        raise RuntimeError(f"missing insert anchor: {insert_after}")
    insert_idx = headers.index(insert_after) + 2
    ws.insert_cols(insert_idx)
    left = ws.cell(1, insert_idx - 1)
    target = ws.cell(1, insert_idx)
    if left.has_style:
        target.font = copy(left.font)
        target.border = copy(left.border)
        target.fill = copy(left.fill)
        target.number_format = left.number_format
        target.protection = copy(left.protection)
        target.alignment = copy(left.alignment)
    target.value = field
    return {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}, True


def main():
    evidence_index = build_evidence_index()
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_col = {n(cell.value): idx + 1 for idx, cell in enumerate(rod_ws[1])}
    detail_col, column_added = ensure_column(detail_ws, FIELD, INSERT_AFTER)

    model_by_rod_id = {}
    for row_idx in range(2, rod_ws.max_row + 1):
        model_by_rod_id[n(rod_ws.cell(row_idx, rod_col["id"]).value)] = n(rod_ws.cell(row_idx, rod_col["model"]).value)

    changed = []
    guide_changes = []
    source_counts = Counter()
    coverage = 0

    for row_idx in range(2, detail_ws.max_row + 1):
        row = {field: detail_ws.cell(row_idx, col_idx).value for field, col_idx in detail_col.items()}
        model = model_by_rod_id.get(n(row.get("rod_id")), "")
        items, source = pairing_for(row, model, evidence_index)
        pairing = " / ".join(items)
        source_counts[source] += 1
        if pairing:
            coverage += 1

        cell = detail_ws.cell(row_idx, detail_col[FIELD])
        old = n(cell.value)
        if old != pairing:
            cell.value = pairing
            changed.append({
                "row": row_idx,
                "id": n(row.get("id")),
                "rod_id": n(row.get("rod_id")),
                "sku": n(row.get("SKU")),
                "model": model,
                "old_value": old,
                "new_value": pairing,
                "source": source,
            })

        conflict = guide_conflict(pairing, row.get("guide_use_hint"))
        if conflict:
            next_hint = guide_hint_for_pairing(pairing)
            if next_hint and n(row.get("guide_use_hint")) != next_hint:
                hint_cell = detail_ws.cell(row_idx, detail_col["guide_use_hint"])
                guide_changes.append({
                    "row": row_idx,
                    "id": n(row.get("id")),
                    "rod_id": n(row.get("rod_id")),
                    "sku": n(row.get("SKU")),
                    "model": model,
                    "old_value": n(row.get("guide_use_hint")),
                    "new_value": next_hint,
                    "reason": conflict,
                })
                hint_cell.value = next_hint

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "xlsx_file": str(XLSX_PATH),
        "field": FIELD,
        "column_added": column_added,
        "changed_recommended_rig_pairing_cells": len(changed),
        "changed_guide_use_hint_cells": len(guide_changes),
        "coverage": f"{coverage}/{detail_ws.max_row - 1}",
        "source_counts": dict(source_counts),
        "changes": changed,
        "guide_use_hint_conflict_fixes": guide_changes,
        "source_policy": {
            "official_description": "Matched explicit rig/lure words in local Megabass official Description.",
            "whitelist_specs": "Used only when official Description did not expose concrete rigs; exact local SKU match from evidence sidecar.",
            "conservative_specs": "Used SKU/type/power/lure-weight/player context conservatively; no Big Bait from MH/H power alone.",
        },
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "coverage": report["coverage"],
        "column_added": column_added,
        "changed_recommended_rig_pairing_cells": len(changed),
        "changed_guide_use_hint_cells": len(guide_changes),
        "source_counts": dict(source_counts),
        "report": str(REPORT_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
