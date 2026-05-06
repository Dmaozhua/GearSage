#!/usr/bin/env python3
import copy
import json
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "nories_rod_normalized.json"
XLSX_PATH = DATA_DIR / "nories_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "nories_rod_product_technical_stage5_report.json"

FIELD = "product_technical"
INSERT_AFTER = "Description"


def norm(value):
    return str(value or "").strip()


def merge_terms(terms):
    merged = []
    for term in terms:
        text = norm(term)
        if text and text not in merged:
            merged.append(text)
    return " / ".join(merged)


def infer_product_technical(model, sku, type_code):
    model_upper = norm(model).upper()
    terms = []
    if "VOICE LTT" in model_upper:
        terms.extend(
            [
                "富士工業製ステンレスフレームSICガイド",
                "Kガイド",
                "MNSTトップガイド",
                "富士工業製PTSリールシート",
                "富士工業製WBCバランサー",
            ]
        )
    elif "HARD BAIT SPECIAL" in model_upper:
        terms.append("バリアブルテーパー")
        if sku.endswith("-Gc"):
            terms.extend(["グラスコンポジット（-Gc）", "バキューム"])
        if type_code == "S":
            terms.extend(
                [
                    "富士工業製チタンフレームSiC KRコンセプトガイド",
                    "SGt（シャキットグラスティップ）",
                    "富士工業製VSSリールシート",
                ]
            )
        elif sku == "HB660L-Gc":
            terms.extend(
                [
                    "チタンフレームガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
        elif sku == "HB6100ML-Gc":
            terms.extend(
                [
                    "ステンレスフレームオールダブルフットガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
        elif sku in {"HB6100M-Gc", "HB730MH-Gc"}:
            terms.extend(
                [
                    "チタンフレームガイド",
                    "シングルフットティップガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
            if sku == "HB730MH-Gc":
                terms.append("グリップジョイント")
        else:
            terms.extend(
                [
                    "富士工業製GMステンレスフレームSiC-Sガイド",
                    "Kガイド",
                    "LGSTトップガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
            if sku in {"HB710LL", "HB760L", "HB760M"}:
                terms.append("テレスコピックハンドル")
            if sku == "HB730MH-Gc":
                terms.append("グリップジョイント")
    elif "STRUCTURE NXS" in model_upper:
        terms.append("ストラクチャーNXSブランク")
        if type_code == "S":
            terms.extend(
                [
                    "富士工業製チタンフレームトルザイトガイド",
                    "KLガイド",
                    "KTガイド",
                    "富士工業製VSSシート",
                ]
            )
        else:
            terms.extend(
                [
                    "富士工業製チタンフレームトルザイトガイド",
                    "LRVガイド",
                    "KTガイド",
                    "富士工業製ECSリールシート",
                ]
            )
        if sku.endswith("-St"):
            terms.append("ショートカーボンソリッドティップ")
        if sku in {"STN720MH", "STN720H"}:
            terms.append("テレスコピック")
        if sku == "STN640MLS-Md":
            terms.append("モーメントディレイブランク")
        if sku == "STN511LLS":
            terms.append("アンサンドフィニッシュ")
    elif "VOICE JUNGLE" in model_upper:
        if type_code == "S":
            terms.extend(
                [
                    "富士工業製チタンフレームトルザイトリング",
                    "KRコンセプト",
                    "富士工業製VSSリールシート",
                ]
            )
        else:
            terms.extend(
                [
                    "テレスコピック",
                    "富士工業製ステンレスフレームSICガイド",
                    "Kガイド",
                    "MNSTトップガイド",
                    "テレスコピックストッパー",
                    "富士工業製TCSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
    return merge_terms(terms)


def sheet_values(ws, ignore_fields=None):
    ignore_fields = set(ignore_fields or [])
    headers = [cell.value for cell in ws[1]]
    keep_indexes = [i for i, header in enumerate(headers, start=1) if header not in ignore_fields]
    return [
        tuple(ws.cell(row=row, column=col).value for col in keep_indexes)
        for row in range(1, ws.max_row + 1)
    ]


def copy_column_style(ws, source_col, target_col):
    ws.column_dimensions[ws.cell(row=1, column=target_col).column_letter].width = (
        ws.column_dimensions[ws.cell(row=1, column=source_col).column_letter].width
    )
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row=row, column=source_col)
        target = ws.cell(row=row, column=target_col)
        target._style = copy.copy(source._style)
        if source.has_style:
            target.font = copy.copy(source.font)
            target.fill = copy.copy(source.fill)
            target.border = copy.copy(source.border)
            target.alignment = copy.copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy.copy(source.protection)


def delete_column_if_present(ws, field):
    headers = [cell.value for cell in ws[1]]
    if field not in headers:
        return False
    ws.delete_cols(headers.index(field) + 1)
    return True


def ensure_column_after(ws, field, after_field):
    headers = [cell.value for cell in ws[1]]
    if field in headers:
        current_col = headers.index(field) + 1
        desired_col = headers.index(after_field) + 2
        if current_col == desired_col:
            return current_col, False
        ws.delete_cols(current_col)
        headers = [cell.value for cell in ws[1]]
    if after_field not in headers:
        raise ValueError(f"{ws.title} missing anchor column: {after_field}")
    after_col = headers.index(after_field) + 1
    insert_col = after_col + 1
    ws.insert_cols(insert_col)
    copy_column_style(ws, after_col, insert_col)
    ws.cell(row=1, column=insert_col).value = field
    ws.column_dimensions[ws.cell(row=1, column=insert_col).column_letter].width = 48
    return insert_col, True


def restore_rod_detail_fills(ws):
    headers = [cell.value for cell in ws[1]]
    if "rod_id" not in headers:
        return
    fill_a = PatternFill("solid", fgColor="FFF8F3C8")
    fill_b = PatternFill("solid", fgColor="FFE8F1FB")
    rod_col = headers.index("rod_id") + 1
    last_rod_id = None
    group_index = -1
    for row in range(2, ws.max_row + 1):
        rod_id = ws.cell(row=row, column=rod_col).value
        if rod_id != last_rod_id:
            group_index += 1
            last_rod_id = rod_id
        fill = fill_a if group_index % 2 == 0 else fill_b
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill


def build_variant_values(data):
    values = []
    removed_item_fields = []
    for item in data:
        model = item.get("model", "")
        if FIELD in item:
            removed_item_fields.append({"model": model, "before": item.get(FIELD)})
            item.pop(FIELD, None)
        for variant in item.get("variants", []):
            sku = variant.get("sku") or variant.get("fields", {}).get("SKU") or ""
            type_code = variant.get("fields", {}).get("TYPE") or ""
            value = infer_product_technical(model, sku, type_code)
            variant[FIELD] = value
            variant.setdefault("fields", {})[FIELD] = value
            values.append({"model": model, "sku": sku, "type": type_code, "value": value})
    return values, removed_item_fields


def update_normalized():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    values, removed_item_fields = build_variant_values(data)
    NORMALIZED_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {
        "items": len(data),
        "variants": len(values),
        "removed_item_product_technical_fields": removed_item_fields,
        "variant_product_technical_nonempty": sum(1 for row in values if row["value"]),
        "variant_values": values,
    }


def update_workbook(variant_values):
    wb = load_workbook(XLSX_PATH)
    ws_rod = wb["rod"]
    ws_detail = wb["rod_detail"]

    rod_before = sheet_values(ws_rod, ignore_fields={FIELD})
    detail_before = sheet_values(ws_detail, ignore_fields={FIELD})

    rod_column_removed = delete_column_if_present(ws_rod, FIELD)
    detail_col, detail_column_inserted_or_moved = ensure_column_after(ws_detail, FIELD, INSERT_AFTER)

    if ws_detail.max_row - 1 != len(variant_values):
        raise RuntimeError(f"rod_detail rows ({ws_detail.max_row - 1}) != normalized variants ({len(variant_values)})")

    workbook_changes = []
    for row_idx, row in enumerate(variant_values, start=2):
        before = ws_detail.cell(row=row_idx, column=detail_col).value
        value = row["value"]
        if before != value:
            workbook_changes.append({"row": row_idx, "sku": row["sku"], "before": before, "after": value})
        ws_detail.cell(row=row_idx, column=detail_col).value = value

    if sheet_values(ws_rod, ignore_fields={FIELD}) != rod_before:
        raise RuntimeError("Unexpected rod sheet changes outside product_technical column removal")
    if sheet_values(ws_detail, ignore_fields={FIELD}) != detail_before:
        raise RuntimeError("Unexpected rod_detail changes outside product_technical")

    restore_rod_detail_fills(ws_detail)
    wb.save(XLSX_PATH)

    return {
        "rod_column_removed": rod_column_removed,
        "rod_detail_column_inserted_or_moved": detail_column_inserted_or_moved,
        "rod_detail_rows": ws_detail.max_row - 1,
        "rod_detail_product_technical_nonempty": sum(1 for row in variant_values if row["value"]),
        "workbook_changes": workbook_changes,
    }


def main():
    normalized_report = update_normalized()
    xlsx_report = update_workbook(normalized_report["variant_values"])
    report = {
        "field": FIELD,
        "scope": "rod_detail SKU level only",
        "source_policy": "Official Nories product pages only; whitelist/player/spec inference not used.",
        "delimiter": " / ",
        "column_position": "rod_detail after Description; rod master must not contain product_technical",
        "normalized": normalized_report,
        "xlsx": xlsx_report,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
