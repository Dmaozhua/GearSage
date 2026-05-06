from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')
DETAIL_ID = "DRD10038"
NEW_HINT = "Power Game 泛用：Jighead、Texas、Punching 與 Heavy Carolina 打點為主，也能兼顧 Spoon、Spinnerbait 等重型移動餌。"


def n(value):
    return " ".join(str(value or "").split())


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["guide_use_hint"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        if detail_id != DETAIL_ID:
            continue
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        if old != NEW_HINT:
            cell.value = NEW_HINT
            changed.append({"row": row, "id": detail_id, "old": old, "new": NEW_HINT})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-guide_use_hint changes")

    wb.save(XLSX_PATH)
    print({"file": str(XLSX_PATH), "changed": len(changed), "changed_ids": [item["id"] for item in changed]})


if __name__ == "__main__":
    main()
