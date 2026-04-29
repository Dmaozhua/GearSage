import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")
FIELD = "recommended_rig_pairing"
INSERT_AFTER = "guide_use_hint"
MAX_ITEMS = 6


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def has_any(text, *patterns):
    return any(re.search(pattern, text, re.I) for pattern in patterns)


def add(items, value):
    if value and value not in items:
        items.append(value)


def add_if(items, text, value, *patterns):
    if has_any(text, *patterns):
        add(items, value)


POSITION_PATTERNS = {
    "Eging": ["木蝦", "エギ", "eging"],
    "Tip-run Eging": ["tip[- ]?run", "船釣木蝦"],
    "Sinker Rig": ["木蝦\\+鉛", "仮面シンカー"],
    "SLJ": [r"\bslj\b", "超級輕量", "超輕", "輕量鐵板"],
    "One Pitch Jig": ["one pitch", "單抽"],
    "Slow Pitch Jig": ["slow", "慢鐵", "低反發", "lowresponse"],
    "Metal Jig": ["鐵板", "金屬\\s*jig", "metal jig", "小型鐵板"],
    "Tungsten Jig": ["鎢"],
    "Semi-long Jig": ["半長型"],
    "Long Jig": ["長型鐵板"],
    "Free Rig": ["free rig", "自由釣組"],
    "Texas Rig": ["texas", "德州"],
    "Heavy Texas": ["heavy texas", "重型.*德州", "輕德州"],
    "Down Shot": ["down shot", "倒吊"],
    "Heavy Down Shot": ["heavy down shot", "重型.*倒吊"],
    "Leaderless Down Shot": ["leaderless down shot"],
    "Neko Rig": ["neko"],
    "No Sinker": ["no sinker", "無鉛"],
    "Small Rubber Jig": ["small rubber", "小型.*rubber"],
    "Rubber Jig": ["rubber\\s*jig", "膠裙"],
    "Jighead Rig": ["jighead", "jig head", "鉛頭", "汲鉤頭"],
    "Wacky Rig": ["wacky"],
    "Carolina Rig": ["carolina", "卡羅萊納"],
    "Heavy Carolina Rig": ["heavy carolina", "重型卡羅萊納"],
    "Float Rig": ["float rig", "漂浮釣組"],
    "Frog": ["frog"],
    "Punching": ["punch", "重障礙"],
    "Swimbait": ["swim bait", "swimbait", "泳餌", "軟質魚型"],
    "Big Bait": ["big bait", "大型餌", "大型路亞", "大餌"],
    "Swim Jig": ["swim jig"],
    "Spinnerbait": ["spinnerbait", "旋轉亮片", "鋼絲餌"],
    "Chatterbait": ["chatter", "bladed jig", "顫泳"],
    "Crankbait": ["crank"],
    "Shad": ["shad"],
    "Minnow": ["minnow", "米諾"],
    "Jerkbait": ["jerkbait", "jerk", "twitch"],
    "Vibration": ["vib", "vibration"],
    "Metal Vibration": ["metal vib"],
    "Topwater Plug": ["topwater", "水面"],
    "Pencil Bait": ["pencil", "鉛筆", "沉水鉛筆"],
    "Popper": ["popper"],
    "Bug Lure": ["bug lure", "蟲系", "虫系"],
    "Spoon": ["spoon"],
    "Plug": ["plug", "栓型路亞"],
    "Soft Plastic": ["軟蟲", "soft plastic"],
    "High-density Soft Plastic": ["高比重軟蟲"],
    "Light Rig": ["輕量釣組", "精細釣組", "fitness spin"],
    "Small Hardbait": ["小型硬餌", "中小型硬餌"],
    "WIND": [r"\bwind\b", "白帶魚"],
}


def first_match_pos(text, patterns):
    best = None
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            best = match.start() if best is None else min(best, match.start())
    return best


def sort_by_official_order(items, text):
    indexed = []
    for idx, item in enumerate(items):
        pos = first_match_pos(text, POSITION_PATTERNS.get(item, []))
        indexed.append((pos if pos is not None else 10**9, idx, item))
    return [item for _, _, item in sorted(indexed)]


def pairings_from_official_text(model, sku, description, player_positioning):
    model_text = lower(model)
    sku_text = lower(sku)
    text = lower(f"{model} {sku} {description} {player_positioning}")
    explicit = lower(f"{sku} {description}")
    items = []

    # This is intentionally an open phrase extractor, not a closed enum.
    # Add brand-specific phrases as they appear in official text.
    add_if(items, explicit, "Eging", "木蝦", "エギ", "eging")
    add_if(items, explicit, "Tip-run Eging", "tip[- ]?run", "船釣木蝦")
    add_if(items, explicit, "Sinker Rig", "木蝦\\+鉛", "仮面シンカー")

    add_if(items, explicit, "SLJ", r"\bslj\b", "超級輕量", "超輕", "輕量鐵板")
    add_if(items, explicit, "One Pitch Jig", "one pitch", "單抽")
    add_if(items, explicit, "Slow Pitch Jig", "slow", "慢鐵", "低反發", "lowresponse")
    add_if(items, explicit, "Metal Jig", "鐵板", "金屬\\s*jig", "metal jig", "小型鐵板")
    add_if(items, explicit, "Tungsten Jig", "鎢")
    add_if(items, explicit, "Semi-long Jig", "半長型")
    add_if(items, explicit, "Long Jig", "長型鐵板")

    add_if(items, explicit, "Free Rig", "free rig", "自由釣組")
    add_if(items, explicit, "Texas Rig", "texas", "德州")
    add_if(items, explicit, "Heavy Texas", "heavy texas", "重型.*德州", "輕德州")
    add_if(items, explicit, "Down Shot", "down shot", "倒吊")
    add_if(items, explicit, "Heavy Down Shot", "heavy down shot", "重型.*倒吊")
    add_if(items, explicit, "Leaderless Down Shot", "leaderless down shot")
    add_if(items, explicit, "Neko Rig", "neko")
    add_if(items, explicit, "No Sinker", "no sinker", "無鉛")
    add_if(items, explicit, "Small Rubber Jig", "small rubber", "小型.*rubber")
    add_if(items, explicit, "Rubber Jig", "rubber\\s*jig", "膠裙")
    add_if(items, explicit, "Jighead Rig", "jighead", "jig head", "鉛頭", "汲鉤頭")
    add_if(items, explicit, "Wacky Rig", "wacky")
    add_if(items, explicit, "Carolina Rig", "carolina", "卡羅萊納")
    add_if(items, explicit, "Heavy Carolina Rig", "heavy carolina", "重型卡羅萊納")
    add_if(items, explicit, "Float Rig", "float rig", "漂浮釣組")

    add_if(items, explicit, "Frog", "frog")
    add_if(items, explicit, "Punching", "punch", "重障礙")
    add_if(items, explicit, "Swimbait", "swim bait", "swimbait", "泳餌", "軟質魚型")
    add_if(items, explicit, "Big Bait", "big bait", "大型餌", "大型路亞", "大餌")
    add_if(items, explicit, "Swim Jig", "swim jig")
    add_if(items, explicit, "Spinnerbait", "spinnerbait", "旋轉亮片", "鋼絲餌")
    add_if(items, explicit, "Chatterbait", "chatter", "bladed jig", "顫泳")
    add_if(items, explicit, "Crankbait", "crank")
    add_if(items, explicit, "Shad", "shad")
    add_if(items, explicit, "Minnow", "minnow", "米諾")
    add_if(items, explicit, "Jerkbait", "jerkbait", "jerk", "twitch")
    add_if(items, explicit, "Vibration", "vib", "vibration")
    add_if(items, explicit, "Metal Vibration", "metal vib")
    add_if(items, explicit, "Topwater Plug", "topwater", "水面")
    add_if(items, explicit, "Pencil Bait", "pencil", "鉛筆", "沉水鉛筆")
    add_if(items, explicit, "Popper", "popper")
    add_if(items, explicit, "Bug Lure", "bug lure", "蟲系", "虫系")
    add_if(items, explicit, "Spoon", "spoon")
    add_if(items, explicit, "Plug", "plug", "栓型路亞")
    add_if(items, explicit, "Soft Plastic", "軟蟲", "soft plastic")
    add_if(items, explicit, "High-density Soft Plastic", "高比重軟蟲")
    add_if(items, explicit, "Light Rig", "輕量釣組", "精細釣組", "fitness spin")
    add_if(items, explicit, "Small Hardbait", "小型硬餌", "中小型硬餌")

    add_if(items, explicit, "WIND", r"\bwind\b", "白帶魚")

    if items:
        return sort_by_official_order(items, explicit)[:MAX_ITEMS], "official_text"

    fallback = []
    if has_any(text, "landing gaff", "搭鉤"):
        return [], "not_applicable"
    if has_any(text, "emeraldas", "eging", "木蝦", "烏賊"):
        add(fallback, "Eging")
        add(fallback, "Sinker Rig")
        return fallback, "family_fallback"
    if has_any(text, "silver wolf", "黑鯛", "chining"):
        add(fallback, "Chining Free Rig")
        add(fallback, "Bottom Rig")
        return fallback, "family_fallback"
    if has_any(model_text, "morethan", "lateo", "海鱸"):
        add(fallback, "Minnow")
        add(fallback, "Sinking Pencil")
        add(fallback, "Vibration")
        add(fallback, "Seabass Plug")
        return fallback, "family_fallback"
    if has_any(text, "saltiga slj", r"\bslj\b"):
        add(fallback, "SLJ")
        add(fallback, "Metal Jig")
        add(fallback, "Tungsten Jig")
        return fallback, "family_fallback"
    if has_any(text, "lowresponse", "outrage.*lj", "outrage.*sj", "鏡牙"):
        add(fallback, "Light Jigging")
        add(fallback, "Metal Jig")
        return fallback, "family_fallback"
    if has_any(text, "saltiga c", "outrage.*br.*c", "船拋", "鮪", "黃鰭", "gt"):
        add(fallback, "Plug")
        add(fallback, "Diving Pencil")
        add(fallback, "Popper")
        return fallback, "family_fallback"
    if has_any(model_text, "overthere", "over there", "dragger") or has_any(explicit, r"\bslsj\b", "surf"):
        add(fallback, "Metal Jig")
        add(fallback, "Minnow")
        add(fallback, "Sinking Pencil")
        add(fallback, "Surf Plug")
        return fallback, "family_fallback"
    if has_any(text, "hardrock", "rockfish", "岩魚", "根魚"):
        add(fallback, "Texas Rig")
        add(fallback, "Jighead Rig")
        add(fallback, "Metal Jig")
        add(fallback, "Rockfish Rig")
        return fallback, "family_fallback"
    if has_any(text, "月下美人", "ajing", "mebaru", "竹莢", "鮶"):
        add(fallback, "Jighead Rig")
        add(fallback, "Small Plug")
        add(fallback, "Float Rig")
        return fallback, "family_fallback"
    if has_any(text, "trout", "iprimi", "鱒"):
        add(fallback, "Spoon")
        add(fallback, "Small Plug")
        add(fallback, "Trout Minnow")
        return fallback, "family_fallback"
    if has_any(model_text, "wind x"):
        add(fallback, "WIND")
        add(fallback, "Jighead Dart")
        add(fallback, "Metal Jig")
        return fallback, "family_fallback"

    is_bass = has_any(model_text, "tatula", "steez", "black label", "heartland", "airedge", "wilderness", "vertice", "crossfire")
    if has_any(text, "frog", "強障礙"):
        return ["Frog", "Punching", "Heavy Texas"], "player_fallback"
    if has_any(text, "big bait", "swim bait", "swimbait", "大餌", "大型餌") or has_any(sku_text, r"sb\b"):
        return ["Swimbait", "Big Bait", "Swim Jig", "Spinnerbait"], "player_fallback"
    if has_any(text, "軟餌", "底操", "texas", "free rig"):
        return ["Texas Rig", "Free Rig", "Rubber Jig", "Jighead Rig"], "player_fallback"
    if has_any(text, "crank", "spinnerbait", "硬餌", "搜索", "卷阻"):
        return ["Crankbait", "Shad", "Spinnerbait", "Vibration"], "player_fallback"
    if has_any(text, "精細", "輕餌", "finesse"):
        return ["Neko Rig", "Down Shot", "No Sinker", "Small Rubber Jig"], "player_fallback"
    if has_any(text, "強力泛用") or (is_bass and has_any(sku_text, r"mh", r"\bh\b", r"x[xh]")):
        return ["Texas Rig", "Rubber Jig", "Swim Jig", "Spinnerbait"], "player_fallback"
    if is_bass or has_any(text, "bass", "淡水泛用", "槍柄泛用", "直柄泛用"):
        return ["Texas Rig", "Crankbait", "Spinnerbait", "Down Shot"], "player_fallback"

    return ["General Lure"], "generic_fallback"


def ensure_column(ws, header):
    headers = [cell.value for cell in ws[1]]
    if header in headers:
        return headers.index(header) + 1
    if INSERT_AFTER not in headers:
        raise RuntimeError(f"missing insert anchor column: {INSERT_AFTER}")
    insert_at = headers.index(INSERT_AFTER) + 2
    ws.insert_cols(insert_at)
    ws.cell(1, insert_at).value = header

    src_col = insert_at - 1
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, src_col)
        dst = ws.cell(row, insert_at)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
    return insert_at


def main():
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]

    rod_headers = [cell.value for cell in rod_ws[1]]
    rod_col = {name: i + 1 for i, name in enumerate(rod_headers)}
    master_by_id = {}
    for row in range(2, rod_ws.max_row + 1):
        rid = n(rod_ws.cell(row, rod_col["id"]).value)
        master_by_id[rid] = {
            "model": n(rod_ws.cell(row, rod_col["model"]).value),
            "description": n(rod_ws.cell(row, rod_col["Description"]).value),
        }

    field_col = ensure_column(detail_ws, FIELD)
    detail_headers = [cell.value for cell in detail_ws[1]]
    detail_col = {name: i + 1 for i, name in enumerate(detail_headers)}

    required = ["id", "rod_id", "SKU", "Description", "player_positioning", "player_selling_points"]
    missing = [field for field in required if field not in detail_col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")

    changed = 0
    skipped_existing = 0
    blank = 0
    source_counts = {}
    samples = []

    for row in range(2, detail_ws.max_row + 1):
        cell = detail_ws.cell(row, field_col)
        if n(cell.value):
            skipped_existing += 1
            continue

        rid = n(detail_ws.cell(row, detail_col["rod_id"]).value)
        master = master_by_id.get(rid, {})
        sku = n(detail_ws.cell(row, detail_col["SKU"]).value)
        description = n(detail_ws.cell(row, detail_col["Description"]).value)
        player_positioning = n(detail_ws.cell(row, detail_col["player_positioning"]).value)
        player_selling = n(detail_ws.cell(row, detail_col["player_selling_points"]).value)

        items, source = pairings_from_official_text(
            master.get("model", ""),
            sku,
            description,
            f"{player_positioning} {player_selling}",
        )
        source_counts[source] = source_counts.get(source, 0) + 1
        value = " / ".join(items)
        if value:
            cell.value = value
            changed += 1
            if len(samples) < 12:
                samples.append({
                    "id": n(detail_ws.cell(row, detail_col["id"]).value),
                    "rod_id": rid,
                    "sku": sku,
                    "value": value,
                    "source": source,
                })
        else:
            blank += 1

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "field": FIELD,
        "changed_cells": changed,
        "skipped_existing": skipped_existing,
        "blank": blank,
        "source_counts": source_counts,
        "samples": samples,
    })


if __name__ == "__main__":
    main()
