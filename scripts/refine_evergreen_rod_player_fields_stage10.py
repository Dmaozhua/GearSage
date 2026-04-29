from copy import copy
from datetime import datetime, timezone
from pathlib import Path
import json
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = ROOT / "GearSage-client/pkgGear/data_raw"
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "evergreen_rod_player_fields_stage10_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"

TARGET_FIELDS = ["player_environment", "player_positioning", "player_selling_points"]
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def has(text, *patterns):
    return any(re.search(pattern, text, re.I) for pattern in patterns)


def split_pairing(value):
    return [part.strip() for part in n(value).split(" / ") if part.strip()]


def first(parts, choices):
    return next((part for part in parts if part in choices), "")


def join(items, limit=3):
    out = []
    for item in items:
        item = n(item)
        if item and item not in out:
            out.append(item)
    return " / ".join(out[:limit])


def spec_text(row):
    lure = n(row.get("LURE WEIGHT"))
    line = " / ".join(x for x in [n(row.get("Line Wt N F")), n(row.get("PE Line Size"))] if x)
    bits = []
    if lure:
        bits.append(lure)
    if line:
        bits.append(line)
    return "，".join(bits)


BASS_SOFT_LIGHT = {"Small Rubber Jig", "Jighead Rig", "Down Shot", "Neko Rig", "No Sinker", "Wacky Rig", "Power Finesse", "Hover Strolling", "Mid Strolling Jighead", "Bottom Strolling"}
BASS_BOTTOM = {"Texas Rig", "Light Texas", "Heavy Texas", "Free Rig", "Leaderless Down Shot", "Heavy Down Shot", "Rubber Jig", "Guarded Rubber Jig", "Football Jig", "Carolina Rig", "Heavy Carolina Rig", "High-density No Sinker"}
BASS_MOVING = {"Chatterbait", "Spinnerbait", "Buzzbait", "Swim Jig", "Vibration", "Metal Vibration"}
BASS_HARD = {"Crankbait", "Deep Crankbait", "Shallow Crankbait", "Shad", "Minnow", "Jerkbait", "Topwater Plug", "Pencil Bait", "Popper", "Crawler Bait", "Spoon", "Small Plug"}
BASS_BIG = {"Big Bait", "Swimbait", "Alabama Rig"}


def is_light_salt_env(row, parts):
    env = n(row.get("player_environment"))
    sku = n(row.get("SKU"))
    return (
        "轻型海水" in env
        or "輕型海水" in env
        or any(p in parts for p in ["Ajing Jighead", "Mebaring Jighead", "Light Game Jighead"])
        or ("Rockfish Rig" in parts and sku.startswith(("PSS", "SPR", "NEOS")))
    )


def build_bass(parts, row):
    desc = lower(row.get("Description"))
    sku = n(row.get("SKU"))
    spec = spec_text(row)
    lead = join(parts, 3)
    has_soft = any(p in BASS_SOFT_LIGHT or p in BASS_BOTTOM for p in parts)
    has_moving = any(p in BASS_MOVING or p in BASS_HARD or p in BASS_BIG for p in parts)

    if parts and parts[0] in BASS_BIG:
        env = "淡水 Bass / 大饵远投 / 开阔水域"
        if "Frog" in parts or "Punching" in parts:
            env = "淡水 Bass / 大饵重障碍 / 草区边缘"
        big_lead = join([p for p in parts if p in BASS_BIG or p in {"Spinnerbait", "Swim Jig", "Topwater Plug"}], 3) or lead
        pos = f"{big_lead} 主轴的强力抛投竿"
        sell = "重饵上竿和长距离出线更稳，回收负荷高的 Swimbait/Big Bait 也有足够控鱼余量"
        if spec:
            sell += f"，适配 {spec}"
        return env, pos, sell

    if parts and parts[0] in {"Frog", "Punching", "Heavy Texas"}:
        env = "淡水 Bass / 草区 Frog / 重障碍"
        cover_lead = join([p for p in parts if p in {"Frog", "Punching", "Heavy Texas", "Texas Rig", "Rubber Jig", "Guarded Rubber Jig"}], 3) or lead
        pos = f"{cover_lead} 取向的重障碍强攻竿"
        sell = "粗线近距打点、草洞穿越和中鱼后拉离覆盖物更直接，不把强力竿误用成单纯大饵竿"
        return env, pos, sell

    if parts and parts[0] in BASS_MOVING:
        env = "淡水 Bass / 移动饵搜索 / 草边开阔水"
        if "Swim Jig" in parts or "Guarded Rubber Jig" in parts:
            env = "淡水 Bass / 草边移动饵 / 结构搜索"
        moving_lead = join([p for p in parts if p in BASS_MOVING or p in BASS_HARD or p in BASS_BIG], 3) or lead
        pos = f"{moving_lead} 优先的移动饵搜索竿"
        soft = join([p for p in parts if p in BASS_BOTTOM], 2)
        sell = "连续平收、变速和碰障反应更稳定，适合用线弧维持泳层"
        if soft:
            sell += f"，放慢时还能切到 {soft} 做结构边底操"
        if spec:
            sell += f"，负荷边界参考 {spec}"
        return env, pos, sell

    if parts and parts[0] in BASS_HARD:
        env = "淡水 Bass / 硬饵搜索 / 开放水域"
        if parts[0] in {"Topwater Plug", "Pencil Bait", "Popper", "Jerkbait"}:
            env = "淡水 Bass / 操作型硬饵 / 近中距离标点"
        hard_lead = join([p for p in parts if p in BASS_HARD or p in BASS_MOVING], 3) or lead
        pos = f"{hard_lead} 取向的硬饵操作竿"
        sell = "抽停、平收和碰障后的追随性更好，连续抛投时不容易丢失泳层和咬口节奏"
        if has_soft:
            sell += "，兼顾轻软饵时以落点精度和控松线为主"
        if spec:
            sell += f"，按 {spec} 控制饵重和线径"
        return env, pos, sell

    if parts and parts[0] in BASS_SOFT_LIGHT:
        env = "淡水 Bass / 细线精细 / 结构边"
        if has(desc, "カバー", "ブッシュ", "ウィード", "cover"):
            env = "淡水 Bass / 精细打障碍 / 草木结构"
        soft_lead = join([p for p in parts if p in BASS_SOFT_LIGHT or p in BASS_BOTTOM], 3) or lead
        pos = f"{soft_lead} 优先的精细软饵竿"
        sell = "轻量钓组存在感、读底和松线咬口更清楚，适合在压力水域做慢节奏试探"
        if has_moving:
            sell += "，必要时可切小型硬饵做反应搜索"
        if spec:
            sell += f"，在 {spec} 范围内更容易维持细线控线"
        return env, pos, sell

    if parts and parts[0] in BASS_BOTTOM:
        env = "淡水 Bass / 底操软饵 / 结构障碍"
        if has(desc, "ウィード", "grass", "カバー", "cover", "ブッシュ"):
            env = "淡水 Bass / 草区底操 / 覆盖物边"
        bottom_lead = join([p for p in parts if p in BASS_BOTTOM or p in BASS_SOFT_LIGHT], 3) or lead
        pos = f"{bottom_lead} 主轴的底操软饵竿"
        bottom_examples = join([p for p in parts if p in BASS_BOTTOM or p in BASS_SOFT_LIGHT], 2) or "底操软饵"
        sell = f"触底、穿越障碍和轻口刺鱼时机更好判断，适合 {bottom_examples} 一类需要控线的结构打点"
        if has_moving:
            sell += "，后段也能接移动饵搜索扩大覆盖"
        if spec:
            sell += f"，负荷边界参考 {spec}"
        return env, pos, sell

    env = "淡水 Bass / 多技法泛用"
    pos = f"{sku} 日常泛用竿"
    sell = "按当前钓组范围在软饵、硬饵和移动饵之间保守切换，重点是落点、控线和咬口判断"
    return env, pos, sell


def build_jigging(parts, row):
    spec = spec_text(row)
    if "Long Fall Jig" in parts:
        return (
            "近海船钓 / 长落铁板 / 深场",
            "长落节奏的慢铁板竿",
            f"下沉段控线、长抽后的落速管理和深场重铁板负荷更稳定{f'，适配 {spec}' if spec else ''}",
        )
    if "Micro Jigging" in parts:
        return (
            "近海船钓 / 微铁板 / 浅场多鱼种",
            "小型 Metal Jig 的轻量船钓竿",
            f"轻抛、快慢收切换和浅场小口停顿更好控制，适合微铁板兼顾根鱼和中小型青物{f'，适配 {spec}' if spec else ''}",
        )
    if "Light Jigging" in parts:
        return (
            "近海船钓 / 轻铁板 / 中浅场",
            "轻量 Metal Jig 的船钓搜索竿",
            f"较轻铁板的抽停、变速回收和多鱼种咬口判断更清楚，长时间操作负担更低{f'，适配 {spec}' if spec else ''}",
        )
    return (
        "近海船钓 / 慢抽铁板 / 垂直水层",
        "Slow Pitch Jig 主轴的船钓铁板竿",
        f"Metal Jig 横摆、跳动和下沉咬口更容易掌握{f'，适配 {spec}' if spec else ''}",
    )


def build_tairaba(parts, row):
    spec = spec_text(row)
    return (
        "近海船钓 / Casting Tairaba / 浅场斜向搜索",
        "Casting Tairaba 专用抛投竿",
        f"船上低弹道侧抛、落底后斜向平收和远端挂口更稳定，浅场障碍边的大鲷控鱼余量更足{f'，适配 {spec}' if spec else ''}",
    )


def build_eging(parts, row):
    spec = spec_text(row)
    if "Deep Eging" in parts or "Tip-run Eging" in parts:
        return (
            "船钓木虾 / 深场乌贼 / 垂直控线",
            "深场木虾高感操作竿",
            f"水深、潮流和船漂移造成的负荷下，竿尖视觉咬口、抽停后持线和大乌贼控鱼更清楚{f'，适配 {spec}' if spec else ''}",
        )
    if parts and parts[0] == "Eging" and "Long Fall Jig" in parts:
        return (
            "岸投木虾 / 远投深场 / 长落节奏",
            "Eging / Long Fall Jig 兼顾的远投木虾竿",
            "远浅沙滩、礁岸和高脚位下更容易把木虾送远，长竿线弧管理和慢速长落节奏更稳定",
        )
    if "Slack-jerk Eging" in parts:
        return (
            "岸投木虾 / Slack Jerk / 港湾外礁",
            "Slack Jerk 木虾操作竿",
            "松线抽竿、左右 dart 和停顿抱饵更容易做出节奏，连续操作负担较低",
        )
    return (
        "岸投木虾 / 标准 Eging / 港湾礁岸",
        "木虾抽竿控线泛用竿",
        f"抽竿、控线和停顿抱饵判断更清楚，可按风流在标准木虾和加铅节奏间切换{f'，适配 {spec}' if spec else ''}",
    )


def build_seabass(parts, row):
    spec = spec_text(row)
    if parts and parts[0] in {"Jighead Rig", "Light Game Jighead"}:
        return (
            "海鲈岸投 / 轻量软饵 / 港湾河口",
            "小型 Jighead 与轻量 Plug 搜索竿",
            f"近岸结构、小型饵和短咬场景更容易控线，轻量 Jighead 也能保持触感{f'，适配 {spec}' if spec else ''}",
        )
    if "Shore Plug" in parts or "Metal Jig" in parts:
        return (
            "海鲈岸投 / 远投搜索 / 河口外海",
            "中远投 Plug 与 Metal Jig 搜索竿",
            f"长距离出线、迎风控线和泳层维持更稳，适合按风浪在 Plug 与 Metal Jig 间切换{f'，适配 {spec}' if spec else ''}",
        )
    return (
        "海鲈岸投 / 港湾河口 / 中近距离搜索",
        "Minnow / Vibration 取向的海鲈搜索竿",
        f"小中型硬饵、Vibration 和轻量软饵都能保持线弧，适合城市港湾和河口结构边{f'，适配 {spec}' if spec else ''}",
    )


def build_light_salt(parts, row):
    desc = lower(row.get("Description"))
    spec = spec_text(row)
    if "Rockfish Rig" in parts and (
        any(p in parts for p in ["Texas Rig", "Free Rig", "Carolina Rig"])
        or has(desc, "ロックフィッシュ", "ボトム", "フロート", "キャロライナ")
    ):
        if "Carolina Rig" in parts or has(desc, "遠距離", "フルキャスト", "30g"):
            return (
                "轻型海水 / 远投 Light Game / Rockfish",
                "偏重 Jighead / Carolina 兼顾的强力轻海水竿",
                f"重量级浮漂、Carolina 和重 Jighead 的远投稳定性更好，遇到 Rockfish 或中型目标也有拉离结构的余量{f'，适配 {spec}' if spec else ''}",
            )
        return (
            "轻型海水 / Mebaring + Rockfish / 港湾礁岸",
            "Mebaring 与轻量 Rockfish 兼顾竿",
            f"小型 Jighead、Plug 与轻 Texas/Free Rig 都能保留触底和细线咬口反馈，结构边控鱼余量比纯精细竿更宽{f'，适配 {spec}' if spec else ''}",
        )
    if parts and parts[0] == "Mebaring Jighead":
        if any(p in parts for p in ["Micro Jigging", "Offshore Jigging", "Minnow"]) or has(desc, "尺", "ロックエリア", "磯ぎわ", "テトラ"):
            return (
                "轻型海水 / 尺级 Mebaring / 近距结构",
                "近距强力 Mebaring 多用竿",
                f"短距精确打点、偏重钓组操作和障碍边控鱼更直接，必要时可切小型 Minnow 或 Micro Jigging 搜索{f'，适配 {spec}' if spec else ''}",
            )
        return (
            "轻型海水 / Mebaring / 港湾礁岸",
            "小型 Jighead 与 Plug 的 Mebaring 竿",
            f"夜间慢速小饵、近距离结构边和尺级目标控鱼更稳定，细线咬口反馈更清楚{f'，适配 {spec}' if spec else ''}",
        )
    if "Ajing Jighead" in parts:
        return (
            "轻型海水 / Ajing / 港湾夜钓",
            "1g 级 Jighead 高感竿",
            f"细 PE 或氟碳线下潮感、吸入式小口和轻量 Jighead 存在感更容易读出来{f'，适配 {spec}' if spec else ''}",
        )
    if "Mebaring Jighead" in parts:
        return (
            "轻型海水 / Mebaring / 港湾礁岸",
            "小型 Jighead 与 Plug 的 Mebaring 竿",
            f"夜间慢速小饵、近距离结构边和尺级目标控鱼更稳定，细线咬口反馈更清楚{f'，适配 {spec}' if spec else ''}",
        )
    if "Rockfish Rig" in parts:
        return (
            "轻型海水 / Rockfish / 岩礁根鱼",
            "贴底避障的根鱼轻量竿",
            f"Jighead、Texas 或 Free Rig 贴底穿越障碍时，触底、挂底前兆和中鱼后拉离结构更直接{f'，适配 {spec}' if spec else ''}",
        )
    return (
        "轻型海水 / 港湾精细 / 小饵搜索",
        "Light Game Jighead 精细竿",
        f"轻量 Jighead、小型 Metal Jig 和 Small Plug 的出线、潮感和细微咬口更清楚{f'，适配 {spec}' if spec else ''}",
    )


def build_trout(parts, row):
    desc = lower(row.get("Description"))
    if has(desc, "渓流", "源流", "山"):
        return (
            "鳟鱼 / 溪流 / 短距精准抛投",
            "山溪 Minnow 精准操作竿",
            "狭窄标点低弹道送饵、落点修正和 Twitch 操作更直接，适合快节奏搜索",
        )
    if any(p in parts for p in ["Topwater Plug", "Minnow", "Vibration"]):
        return (
            "鳟鱼 / 管钓 / 操作型硬饵",
            "Area Trout 反应饵操作竿",
            "Topwater、Minnow 和 Vibration 的抽停、dart、lift 节奏更清楚，远距也能确实挂口",
        )
    return (
        "鳟鱼 / 管钓 / Spoon 慢速控层",
        "Area Trout 精细控层竿",
        "Spoon、Crank 和小型 Minnow 慢速控层更稳定，细线轻咬判断更清楚",
    )


def build_fields(row):
    parts = split_pairing(row.get("recommended_rig_pairing"))
    current_env = n(row.get("player_environment"))
    primary = parts[0] if parts else ""
    if primary in {"Casting Tairaba", "Tairaba"}:
        return build_tairaba(parts, row)
    if primary == "Eging" or any(p in parts for p in ["Deep Eging", "Tip-run Eging", "Slack-jerk Eging"]):
        return build_eging(parts, row)
    if ("海鲈" in current_env or "海鱸" in current_env) and not is_light_salt_env(row, parts):
        return build_seabass(parts, row)
    if is_light_salt_env(row, parts):
        return build_light_salt(parts, row)
    if any(p in parts for p in ["Seabass Plug", "Seabass Minnow", "Sinking Pencil", "Shore Plug"]):
        return build_seabass(parts, row)
    if any(p in parts for p in ["Slow Pitch Jig", "Long Fall Jig", "Offshore Jigging", "Light Jigging", "Micro Jigging"]):
        return build_jigging(parts, row)
    if "鳟" in current_env or "鱒" in current_env or any(p.startswith("Trout") or p in {"Topwater Plug", "Minnow", "Vibration"} for p in parts) and n(row.get("SKU")).startswith(("AATS", "AMSC", "AMSS")):
        return build_trout(parts, row)
    return build_bass(parts, row)


def snapshot_without_targets(ws, target_cols):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col not in target_cols
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    required = [
        "id", "SKU", "POWER", "LURE WEIGHT", "Line Wt N F", "PE Line Size",
        "Description", "recommended_rig_pairing", *TARGET_FIELDS,
    ]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    target_cols = {col[field] for field in TARGET_FIELDS}
    before = snapshot_without_targets(ws, target_cols)
    changes = []
    final_rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = {field: ws.cell(row_idx, col[field]).value for field in required}
        new_env, new_pos, new_sell = build_fields(row)
        detail_id = n(row["id"])
        sku = n(row["SKU"])
        row_change = {"row": row_idx, "id": detail_id, "sku": sku, "fields": {}}
        for field, new_value in zip(TARGET_FIELDS, [new_env, new_pos, new_sell]):
            old = n(ws.cell(row_idx, col[field]).value)
            if old != new_value:
                cell = ws.cell(row_idx, col[field])
                cell.value = new_value
                cell.fill = copy(YELLOW_FILL)
                row_change["fields"][field] = {"old": old, "new": new_value}
        if row_change["fields"]:
            changes.append(row_change)
        final_rows.append({
            "id": detail_id,
            "sku": sku,
            "recommended_rig_pairing": n(row["recommended_rig_pairing"]),
            "player_environment": new_env,
            "player_positioning": new_pos,
            "player_selling_points": new_sell,
        })

    after = snapshot_without_targets(ws, target_cols)
    if before != after:
        raise RuntimeError("unexpected changes outside player fields")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_file": str(XLSX_PATH),
        "changed_rows": len(changes),
        "fields": TARGET_FIELDS,
        "policy": "Only player_environment/player_positioning/player_selling_points are overwritten. Sources are sidecar evidence, recommended_rig_pairing, Description/spec boundaries, and conservative SKU/spec inference; evidence wording is not written into import fields.",
        "changes": changes,
        "final_rows": final_rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "changed_rows": len(changes),
        "report_file": str(REPORT_PATH),
        "samples": changes[:8],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
