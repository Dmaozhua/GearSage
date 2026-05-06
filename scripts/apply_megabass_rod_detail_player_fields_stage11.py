import json
import re
import subprocess
from collections import Counter
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook

import apply_megabass_rod_usage_hint_enrichment_stage10 as stage10


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('megabass_rod_import.xlsx')
REPORT_PATH = resolve_data_raw('megabass_rod_detail_player_fields_stage11_report.json')
SHADE_SCRIPT = ROOT / "scripts/shade_megabass_rod_detail_groups_stage2.py"
FIELDS = ["player_environment", "player_positioning", "player_selling_points"]


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def split_items(value):
    return [n(item) for item in n(value).split("/") if n(item)]


def has_any(items, names):
    return bool(set(items) & set(names))


def is_travel(model, sku):
    return bool(re.search(r"triza|valkyrie|huntsman|mountain stream|world expedition", lower(model))) or bool(re.search(r"-[234]p$|-[234][a-z]?$", lower(sku)))


def is_trout(model):
    return "greathunting" in lower(model) or "great hunting" in lower(model)


def is_valkyrie(model):
    return "valkyrie" in lower(model)


def is_triza(model):
    return "triza" in lower(model)


def env_from_context(row, model, items):
    desc = lower(row.get("Description"))
    sku = n(row.get("SKU"))
    travel = is_travel(model, sku)
    primary = items[0] if items else ""
    moving = {"Jerkbait", "Minnow", "Topwater Plug", "Pencil Bait", "Popper", "Bug Lure", "Deep Crankbait", "Crankbait", "Spinnerbait", "Chatterbait", "Vibration", "Shad", "Small Plug"}
    soft = {"Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Rubber Jig", "Football Jig", "Down Shot", "Neko Rig", "Jighead Rig", "BFS Jighead Rig", "Small Rubber Jig", "No Sinker", "Wacky Rig", "Nose Hook Rig"}

    if has_any(items, ["Lead Core Trolling", "Downrigger Trolling"]):
        if "lead core" in desc and "downrigger" in desc:
            return "湖泊鳟鱼拖钓 / Lead Core 与 Downrigger / 深水巡航"
        if "downrigger" in desc:
            return "湖泊鳟鱼拖钓 / Downrigger 深水线组 / 长时间巡航"
        return "湖泊鳟鱼拖钓 / Lead Core 线组 / 开阔水域搜索"

    if is_trout(model):
        if re.search(r"river&lake|river|lake|本流|湖|wide river|breakwater", lower(model) + " " + desc):
            return "本流与湖泊鳟鱼 / 远投搜索 / 大型 Trout"
        if re.search(r"mountain|huntsman|渓流|山溪|源流", lower(model) + " " + desc):
            return "山溪与小中型溪流 / 精准落点 / 移动钓行"
        if re.search(r"x-glass|glass|グラス", lower(model) + " " + desc):
            return "溪流鳟鱼 / 小型硬饵 / 近中距离控饵"
        return "溪流与本流鳟鱼 / Minnow 与 Spoon / 精细控线"

    if is_valkyrie(model):
        if has_any(items, ["Shore Plug", "Metal Jig"]):
            return "多鱼种远征 / 港湾岸投与河口 / 轻型金属饵搜索"
        if has_any(items, ["Big Bait", "Swimbait"]):
            return "多鱼种远征 / 开阔水域与船边结构 / 大型饵控鱼"
        if has_any(items, ["Small Plug", "Topwater Plug", "Minnow"]):
            return "多鱼种远征 / 小型 Plug 近距精准抛投 / 港湾与轻障碍"
        if has_any(items, ["Crankbait", "Spinnerbait", "Chatterbait", "Vibration"]):
            return "多鱼种远征 / 开放水域搜索 / 低弹性硬饵控线"
        return "多鱼种远征 / 淡水与轻型海水 / 便携主力场景"

    if primary in {"Jerkbait", "Minnow", "Topwater Plug", "Pencil Bait", "Popper", "Bug Lure"}:
        return f"{'旅行便携 / ' if travel else ''}浅场、草边与开阔水面 / 抽停硬饵 / 连续控线"
    if primary in {"Deep Crankbait", "Crankbait"}:
        if has_any(items, ["Spinnerbait", "Chatterbait", "Vibration", "Rubber Jig", "Punching"]):
            return f"{'旅行便携 / ' if travel else ''}开阔水域与障碍边 / 强搜索饵与底操补充 / 快速覆盖"
        return f"{'旅行便携 / ' if travel else ''}开阔水域与硬底结构 / Crankbait 搜索 / 稳定卷收"
    if primary in {"Spinnerbait", "Chatterbait", "Vibration", "Shad"}:
        if has_any(items, soft):
            return f"{'旅行便携 / ' if travel else ''}草边、浅滩与轻障碍 / 移动饵搜索与软饵补充 / 快速覆盖"
        return f"{'旅行便携 / ' if travel else ''}草边、浅滩与开放水域 / 移动饵搜索 / 快速覆盖"
    if primary == "Small Plug":
        if has_any(items, soft):
            return f"{'旅行便携 / ' if travel else ''}浅场与轻障碍 / 小型 Plug 搜索与轻量软饵补充 / 近中距离精准抛投"
        return f"{'旅行便携 / ' if travel else ''}岸投与开阔水域 / 小型硬饵搜索 / 远近距离切换"
    if has_any(items, ["Frog", "Punching"]):
        return f"{'旅行便携 / ' if travel else ''}草区与重障碍 / Frog 或 Punching / 近距强控鱼"
    if has_any(items, ["Big Bait", "Swimbait"]):
        if has_any(items, ["Deep Crankbait", "Spinnerbait", "Chatterbait"]):
            return f"{'旅行便携 / ' if travel else ''}开阔水域与深水结构 / 大饵和强搜索 / 高负荷搏鱼"
        return f"{'旅行便携 / ' if travel else ''}大饵与 Swimbait 场景 / 开放水域或障碍边 / 粗线控鱼"
    if has_any(items, ["Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Rubber Jig", "Football Jig"]):
        if has_any(items, ["Chatterbait", "Spinnerbait", "Crankbait", "Vibration"]):
            return f"{'旅行便携 / ' if travel else ''}浅中层障碍 / Contact Bait 与移动饵切换 / 快速搜索"
        return f"{'旅行便携 / ' if travel else ''}木桩、岩石与草边 / 底操软饵 / 障碍区读底"
    if has_any(items, ["Down Shot", "Neko Rig", "Jighead Rig", "BFS Jighead Rig", "Small Rubber Jig", "No Sinker", "Wacky Rig", "Nose Hook Rig"]):
        if row.get("TYPE") == "C":
            return f"{'旅行便携 / ' if travel else ''}轻障碍与近岸结构 / Bait Finesse / 小饵精准落点"
        return f"{'旅行便携 / ' if travel else ''}开放水域与轻障碍 / 轻线精细钓组 / 慢速诱食"
    if has_any(items, ["Small Plug", "Shore Plug", "Metal Jig"]):
        return f"{'旅行便携 / ' if travel else ''}岸投与开阔水域 / 小型硬饵搜索 / 远近距离切换"
    return f"{'旅行便携 / ' if travel else ''}淡水鲈鱼 / 常规结构区 / 按型号技法细分"


def positioning_from_context(row, model, items):
    sku = n(row.get("SKU"))
    travel = is_travel(model, sku)
    prefix = "便携" if travel else ""
    primary = items[0] if items else ""
    soft = {"Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Rubber Jig", "Football Jig", "Down Shot", "Neko Rig", "Jighead Rig", "BFS Jighead Rig", "Small Rubber Jig", "No Sinker", "Wacky Rig", "Nose Hook Rig", "Punching"}

    if has_any(items, ["Lead Core Trolling", "Downrigger Trolling"]):
        return "湖泊拖钓专项"
    if is_trout(model):
        if has_any(items, ["Lead Core Trolling", "Downrigger Trolling"]):
            return "湖泊拖钓专项"
        if re.search(r"river&lake|river|lake|本流|湖", lower(model) + " " + lower(row.get("Description"))):
            return f"{prefix}本流湖泊远投鳟鱼"
        if has_any(items, ["Jerkbait", "Minnow", "Trout Minnow"]):
            return f"{prefix}溪流 Minnow 精准操作"
        return f"{prefix}溪流轻饵精细控线"

    if is_valkyrie(model):
        if has_any(items, ["Shore Plug", "Metal Jig"]):
            return "远征岸投 Plug / Metal Jig"
        if has_any(items, ["Big Bait", "Swimbait"]):
            return "远征大饵强力控鱼"
        if has_any(items, ["Small Plug", "Topwater Plug", "Minnow"]):
            return "远征小型 Plug 泛用"
        return "远征多鱼种硬饵泛用"

    if primary in {"Topwater Plug", "Pencil Bait", "Popper", "Bug Lure"}:
        return f"{prefix}水面系与轻硬饵"
    if primary in {"Jerkbait", "Minnow"}:
        return f"{prefix}Jerkbait / Minnow 控饵"
    if primary in {"Deep Crankbait"} and has_any(items, soft):
        return f"{prefix}强搜索与重障碍混合"
    if primary in {"Deep Crankbait"}:
        return f"{prefix}Deep Crankbait 卷收搜索"
    if primary in {"Spinnerbait", "Chatterbait", "Vibration"} and has_any(items, soft):
        return f"{prefix}Moving Bait 与软饵补充"
    if primary in {"Spinnerbait", "Chatterbait", "Vibration"}:
        return f"{prefix}移动饵快速搜索"
    if primary == "Small Plug" and has_any(items, soft):
        return f"{prefix}小型 Plug 与 Bait Finesse 兼用"
    if primary == "Small Plug":
        return f"{prefix}小型 Plug 搜索"
    if has_any(items, ["Frog", "Punching"]):
        return f"{prefix}重障碍 Frog / Punching"
    if has_any(items, ["Big Bait", "Swimbait"]):
        if has_any(items, ["Deep Crankbait", "Spinnerbait", "Chatterbait", "Rubber Jig"]):
            return f"{prefix}大饵与强搜索混合"
        return f"{prefix}大饵 Swimbait 强力"
    if has_any(items, ["Chatterbait", "Spinnerbait", "Vibration"]) and has_any(items, ["Rubber Jig", "Texas Rig", "Light Texas Rig"]):
        return f"{prefix}Contact Bait 与 Moving Bait 混合"
    if has_any(items, ["Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Rubber Jig", "Football Jig"]):
        return f"{prefix}强力底操软饵"
    if has_any(items, ["BFS Jighead Rig", "Small Rubber Jig", "No Sinker"]) and row.get("TYPE") == "C":
        return f"{prefix}Bait Finesse 轻量枪柄"
    if has_any(items, ["Down Shot", "Neko Rig", "Jighead Rig", "Wacky Rig", "Nose Hook Rig"]):
        return f"{prefix}轻线精细软饵"
    if has_any(items, ["Crankbait", "Shad"]):
        return f"{prefix}Crank / Shad 搜索"
    if has_any(items, ["Small Plug"]):
        return f"{prefix}小型 Plug 搜索"
    return f"{prefix}{'直柄' if row.get('TYPE') == 'S' else '枪柄'}鲈钓泛用"


def selling_from_context(row, model, items):
    desc = lower(row.get("Description"))
    sku = n(row.get("SKU"))
    travel = is_travel(model, sku)
    primary = items[0] if items else ""
    soft = {"Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Rubber Jig", "Football Jig", "Down Shot", "Neko Rig", "Jighead Rig", "BFS Jighead Rig", "Small Rubber Jig", "No Sinker", "Wacky Rig", "Nose Hook Rig", "Punching"}
    extras = []
    if travel:
        extras.append("多节便携不牺牲核心技法")
    if re.search(r"pinpoint|accuracy|accurate|pitch|tight spots?|close quarters|精准", desc):
        extras.append("低弹道与精准落点")
    if re.search(r"long[- ]distance|distance|moon-shot|遠投|ロングディスタンス|wide|shore", desc):
        extras.append("长距离出线稳定")
    if re.search(r"fall|line tension|subtle bites?|light bites?|sensitivity|感度|喰わせ", desc):
        extras.append("线张力和短咬反馈清楚")
    if re.search(r"glass|グラス|low elasticity|低弾性", desc):
        extras.append("低弹性追随降低脱钩")
    if re.search(r"solid carbon|stinger tip|ソリッド|stinger", desc):
        extras.append("实心竿梢细信号")
    if re.search(r"torque|lifting|power|monster|heavy cover|mats?|vegetation|bush", desc):
        extras.append("中鱼后控鱼余量高")

    if has_any(items, ["Lead Core Trolling", "Downrigger Trolling"]):
        base = ["持续拖曳负荷稳定", "深水线组控线清楚", "适合湖泊大型鳟鱼巡航"]
    elif is_trout(model):
        if re.search(r"river&lake|lake|本流|湖|wide river", lower(model) + " " + desc):
            base = ["远投后线弧管理好", "Minnow / Spoon 漂流和回收更稳", "大型 Trout 搏鱼留有余量"]
        elif re.search(r"x-glass|glass|グラス", lower(model) + " " + desc):
            base = ["小型硬饵挂鱼更包容", "短距离抛投节奏自然", "溪流咬口不容易弹开"]
        else:
            base = ["小型 Minnow 连续 twitch 更顺", "落点控制直接", "溪流移动钓行负担低"]
    elif is_valkyrie(model):
        if has_any(items, ["Shore Plug", "Metal Jig"]):
            base = ["岸投远距离覆盖", "Plug 与金属饵切换明确", "跨淡水和轻型海水场景"]
        elif has_any(items, ["Big Bait", "Swimbait"]):
            base = ["100g 级大型饵抛投更稳", "短距离船边控饵直接", "大鱼冲刺时缓冲充足"]
        elif has_any(items, ["Small Plug", "Topwater Plug", "Minnow"]):
            base = ["3g 级小 Plug 也能送出", "近距精准抛投", "多鱼种轻量搜索宽容"]
        else:
            base = ["多鱼种路亚覆盖宽", "低弹性空白追随好", "远行携带和实钓兼顾"]
    elif primary in {"Topwater Plug", "Pencil Bait", "Popper", "Bug Lure"}:
        base = ["水面系走姿容易做出来", "轻硬饵抛投距离足", "停顿和小幅 twitch 更细"]
    elif primary in {"Jerkbait", "Minnow"}:
        base = ["连续抽停节奏稳定", "停顿时线弧好管理", "远近距离控饵都清楚"]
    elif primary in {"Deep Crankbait", "Crankbait", "Shad"} and has_any(items, soft):
        base = ["硬饵搜索和底操能衔接", "泳层维持稳定", "障碍边触碰后的控线更直接"]
    elif primary in {"Deep Crankbait", "Crankbait", "Shad"}:
        base = ["卷阻饵负荷承接自然", "泳层维持稳定", "长时间搜索手感更轻"]
    elif primary in {"Spinnerbait", "Chatterbait", "Vibration"} and has_any(items, soft):
        base = ["移动饵搜索和软饵补充能切换", "草边触碰后控线稳定", "搜索节奏不容易散"]
    elif primary in {"Spinnerbait", "Chatterbait", "Vibration"}:
        base = ["移动饵覆盖效率高", "草边触碰后控线稳定", "搜索节奏不容易散"]
    elif primary == "Small Plug" and has_any(items, soft):
        base = ["小型 Plug 抛投距离稳定", "轻量软饵可作为补充", "近中距离控饵不发散"]
    elif primary == "Small Plug":
        base = ["小型 Plug 抛投顺", "浅场搜索节奏轻快", "远近距离控饵清楚"]
    elif has_any(items, ["Frog", "Punching"]):
        base = ["粗线强制控鱼", "草洞和重障碍出鱼更有底气", "短距离精准打点"]
    elif has_any(items, ["Big Bait", "Swimbait"]):
        base = ["大饵抛投不发虚", "回收阻力和中鱼冲击承接好", "粗线高负荷搏鱼稳定"]
    elif has_any(items, ["Chatterbait", "Spinnerbait", "Vibration"]) and has_any(items, ["Rubber Jig", "Texas Rig", "Light Texas Rig"]):
        base = ["移动饵搜索和底操能切换", "浅障碍触碰反馈清楚", "挂底前后的控线更直接"]
    elif has_any(items, ["Texas Rig", "Light Texas Rig", "Heavy Texas Rig", "Free Rig", "Rubber Jig", "Football Jig"]):
        base = ["读底和触障反馈明确", "扬竿刺鱼直接", "障碍边慢速操作稳定"]
    elif has_any(items, ["BFS Jighead Rig", "Small Rubber Jig", "No Sinker"]):
        base = ["小饵低负荷抛投顺", "贴障碍轻量诱食更精准", "细线控线不拖泥带水"]
    elif has_any(items, ["Down Shot", "Neko Rig", "Jighead Rig", "Wacky Rig", "Nose Hook Rig"]):
        base = ["轻线控线细", "短咬和触底信号清楚", "慢速精细诱食稳定"]
    else:
        base = ["按型号技法分工清楚", "软硬饵切换宽容", "日常鲈钓适配面广"]

    merged = []
    for item in extras + base:
        if item not in merged:
            merged.append(item)
    return " / ".join(merged[:4])


def build_values(row, model):
    items = split_items(row.get("recommended_rig_pairing"))
    return {
        "player_environment": env_from_context(row, model, items),
        "player_positioning": positioning_from_context(row, model, items),
        "player_selling_points": selling_from_context(row, model, items),
    }


def main():
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_col = {n(cell.value): idx + 1 for idx, cell in enumerate(rod_ws[1])}
    detail_col = {n(cell.value): idx + 1 for idx, cell in enumerate(detail_ws[1])}
    required = ["id", "rod_id", "SKU", "TYPE", "POWER", "LURE WEIGHT", "ACTION", "Description", "recommended_rig_pairing"] + FIELDS
    missing = [field for field in required if field not in detail_col and field not in {"ACTION"}]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    model_by_rod_id = {
        n(rod_ws.cell(row_idx, rod_col["id"]).value): n(rod_ws.cell(row_idx, rod_col["model"]).value)
        for row_idx in range(2, rod_ws.max_row + 1)
    }

    changes = []
    protected_manual_cells = []
    coverage = Counter()
    unique_values = {field: Counter() for field in FIELDS}

    for row_idx in range(2, detail_ws.max_row + 1):
        row = {field: detail_ws.cell(row_idx, col_idx).value for field, col_idx in detail_col.items()}
        model = model_by_rod_id.get(n(row.get("rod_id")), "")
        next_values = build_values(row, model)
        row_change = {
            "row": row_idx,
            "id": n(row.get("id")),
            "rod_id": n(row.get("rod_id")),
            "sku": n(row.get("SKU")),
            "model": model,
            "fields": {},
        }
        for field in FIELDS:
            value = next_values[field]
            coverage[field] += bool(value)
            unique_values[field][value] += 1
            cell = detail_ws.cell(row_idx, detail_col[field])
            old = n(cell.value)
            if old != value:
                row_change["fields"][field] = {"old_value": old, "new_value": value}
                cell.value = value
        if row_change["fields"]:
            changes.append(row_change)

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    report = {
        "xlsx_file": str(XLSX_PATH),
        "fields": FIELDS,
        "changed_rows": len(changes),
        "changed_cells": sum(len(item["fields"]) for item in changes),
        "coverage": {field: f"{coverage[field]}/{detail_ws.max_row - 1}" for field in FIELDS},
        "unique_counts": {field: len(unique_values[field]) for field in FIELDS},
        "top_repeats": {
            field: [{"value": value, "count": count} for value, count in unique_values[field].most_common(12)]
            for field in FIELDS
        },
        "changes": changes,
        "protected_manual_cells": protected_manual_cells,
        "policy": {
            "source_priority": [
                "whitelist/player context when available",
                "confirmed recommended_rig_pairing",
                "Description and official specs as boundaries",
                "SKU/power/action/type/lure weight conservative inference",
            ],
            "write_scope": "Only rod_detail player_environment, player_positioning, player_selling_points were written.",
        },
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "changed_rows": report["changed_rows"],
        "changed_cells": report["changed_cells"],
        "coverage": report["coverage"],
        "unique_counts": report["unique_counts"],
        "report": str(REPORT_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
