from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')

HINT_BY_ID = {
    "DRD10043": "精細多用途：Jighead、Down Shot、Neko 與小型硬餌都能兼顧，細線出線、控線和小餌操作更穩。",
    "DRD10044": "輕量多用途：No Sinker、Down Shot、Light Texas 與小型 Crankbait 切換時，遠投、控線和揚竿刺魚更穩。",
    "DRD10054": "精細多用途：Jighead、Down Shot、Neko 與小型硬餌都能兼顧，細線出線、控線和小餌操作更穩。",
    "DRD10055": "輕量多用途：No Sinker、Down Shot、Light Texas 與小型 Crankbait 切換時，遠投、控線和揚竿刺魚更穩。",
    "DRD10285": "軟硬餌輕量泛用：Soft Plastic、Jighead、No Sinker 為主，也能兼顧小型 Crankbait、Spinnerbait，障礙邊控線更穩。",
    "DRD10286": "輕德州/萬用：Heavy Texas、Texas Rig 與 3/8-1/2oz Crankbait、Spinnerbait 都能兼顧，打點與搜索切換更自然。",
}


def n(value):
    return " ".join(str(value or "").split())


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["guide_use_hint"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        if detail_id not in HINT_BY_ID:
            continue
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        new = HINT_BY_ID[detail_id]
        if old != new:
            cell.value = new
            changed.append({"row": row, "id": detail_id, "old": old, "new": new})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-guide_use_hint changes")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed": len(changed),
        "changed_ids": [item["id"] for item in changed],
    })


if __name__ == "__main__":
    main()
