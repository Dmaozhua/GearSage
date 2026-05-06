#!/usr/bin/env python3
import hashlib
import json
import re
import ssl
import sys
import time
import urllib.request
from datetime import datetime, timezone
from html import unescape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from gear_data_paths import DATA_RAW_DIR


BASE_URL = "https://nories.com"
ENTRY_URL = f"{BASE_URL}/bass/category/rods/"
PRODUCT_URLS = [
    f"{BASE_URL}/bass/road-runner-voice-ltt/",
    f"{BASE_URL}/bass/road-runner-voice-hard-bait-special/",
    f"{BASE_URL}/bass/road-runner-structure-nxs/",
    f"{BASE_URL}/bass/road-runner-voice-jungle/",
]

OUTPUT_DIR = DATA_RAW_DIR
CACHE_DIR = OUTPUT_DIR / "nories_rods_cache"
NORMALIZED_PATH = OUTPUT_DIR / "nories_rod_normalized.json"
OUTPUT_FILE = OUTPUT_DIR / "nories_rod_import.xlsx"
WHITELIST_EVIDENCE_PATH = OUTPUT_DIR / "nories_rod_whitelist_player_evidence.json"
IMAGE_DIR = Path("/Users/tommy/Pictures/images/nories_rods")
OLD_IMAGE_DIR = Path("/Users/tommy/Pictures/images_old_copy/nories_rods")
STATIC_IMAGE_BASE = "https://static.gearsage.club/gearsage/Gearimg/images/nories_rods"
REFRESH_IMAGES = "--refresh-images" in sys.argv

BRAND_ID = 44
ROD_PREFIX = "NR"
ROD_DETAIL_PREFIX = "NRD"
FIT_STYLE_TAG_ORDER = ["bass", "溪流", "海鲈", "根钓", "岸投", "船钓", "旅行"]

ROD_HEADERS = [
    "id", "brand_id", "model", "model_cn", "model_year", "alias",
    "type_tips", "fit_style_tags", "images", "created_at", "updated_at",
    "series_positioning", "main_selling_points", "official_reference_price", "market_status",
    "Description", "player_positioning", "player_selling_points",
]

ROD_DETAIL_HEADERS = [
    "id", "rod_id", "TYPE", "SKU", "POWER", "TOTAL LENGTH", "Action", "PIECES",
    "CLOSELENGTH", "WEIGHT", "Tip Diameter", "LURE WEIGHT", "Line Wt N F",
    "PE Line Size", "Handle Length", "Reel Seat Position", "CONTENT CARBON",
    "Market Reference Price", "AdminCode", "Service Card", " Jig Weight",
    "Squid Jig Size", "Sinker Rating", "created_at", "updated_at",
    "LURE WEIGHT (oz)", "Sale Price", "Joint Type", "Code Name", "Fly Line",
    "Grip Type", "Reel Size",
    "guide_layout_type", "guide_use_hint", "recommended_rig_pairing",
    "hook_keeper_included", "sweet_spot_lure_weight_real",
    "official_environment", "player_environment", "player_positioning", "player_selling_points",
    "Description", "product_technical", "Extra Spec 1", "Extra Spec 2",
]

PLAYER_FIELD_OVERRIDES = {
    "LTT620PMH": {
        "player_positioning": "Versatile / Parabolic multi bait",
        "player_selling_points": "多用途拋投和控餌取向明確 / topwater、jig spinner、swimbait、free rig 等中量級路亞切換更自然",
        "guide_use_hint": "Bass 泛用：中量級線組出線穩定，硬餌、swimbait 和軟餌切換時更容易保持拋投節奏與控線角度。",
    },
    "LTT630MH": {
        "player_positioning": "Cover Texas / Frog",
        "player_selling_points": "Cover Texas、rubber jig 和 frog 取向明確 / 近距離準投、強補刺和障礙邊控魚更有餘量",
        "guide_use_hint": "Bass cover：太線和 frog / Texas 類高阻力釣組出線穩定，低彈道準投與近距離控魚更容易銜接。",
    },
    "STN580ML": {
        "player_positioning": "Tsuru-stroll / PE cover finesse",
        "player_selling_points": "PE 2 号級吊るし與 cover finesse 取向清楚 / 小型 swisher、surface bait 和障礙邊定點操作更好控線",
        "guide_use_hint": "Tsuru-stroll：PE 線出線和短距離 pitching 更穩，吊るし水面下誘魚與障礙邊補刺更容易維持角度。",
    },
    "STN680MS": {
        "player_positioning": "PE spin / Structure rig",
        "player_selling_points": "PE + leader 遠投結構攻略取向明確 / cover neko、down shot、jighead wacky，也能支撐 Carolina、Texas、heavy drop shot",
        "guide_use_hint": "PE spin：細 PE + leader 出線和 slack 管理更穩，遠投、深場 structure、Neko / down shot / Carolina 類釣組更容易讀底和補刺。",
    },
    "STN650M": {
        "player_positioning": "Neko / Heavy drop shot",
        "player_selling_points": "Neko、heavy drop shot 與輕 cover 精細操作取向明確 / 杭、沈木、障礙邊定點 shake 更容易讀底和補刺",
        "guide_use_hint": "Neko / drop shot：細線出線與 slack 處理更順，定點 shake、咬口判斷和短距離補刺更直接。",
    },
    "STN640MLS-Md": {
        "player_positioning": "Insect / Surface finesse",
        "player_selling_points": "蟲系、水面系與 PE twitch 取向清楚 / 低彈性竿尖讓高速 twitch 後的短咬口更容易留餌",
        "guide_use_hint": "Surface finesse：PE 出線順暢，細小 twitch、虫系停頓和水面咬口後的線路控制更穩。",
    },
    "680JMHS": {
        "player_positioning": "Power finesse / Heavy cover",
        "player_selling_points": "Power finesse 專用感更強 / PE + leader、smolaba、cover pitching 和 heavy cover 抽魚更有支撐",
        "guide_use_hint": "Power finesse：PE 出線和竿身回正更穩，cover 內 shake、短距離 pitching 和強補刺更容易連貫。",
    },
    "700JHS": {
        "player_positioning": "Long power finesse / Heavy cover",
        "player_selling_points": "長尺 power finesse / heavy cover 取向明確 / 遠距吊るし、高處補刺和 monster bass 控魚餘量更高",
        "guide_use_hint": "Long power finesse：PE 出線和長竿線路控制更穩，高處吊るし、遠距補刺和障礙邊控魚更有餘量。",
    },
}

WHITELIST_EVIDENCE_BY_SKU = {
    "LTT620PMH": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/search?rod=LTT620PMH",
        "confidence": "medium",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "matched_context": "検索結果にロードランナー ヴォイス LTT620PMH が複数件あり、トップウォーター、ジグスピナー、スイムベイト/ワーム系の組み合わせが確認できる。",
        "supported_fields": ["player_environment", "player_positioning", "player_selling_points", "guide_use_hint"],
    },
    "STN580ML": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/search?rod=STN580ML",
        "confidence": "medium",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "matched_context": "検索結果と公式説明の両方で PE MAX #2、ツルスト、吊るし、カバー周りの操作が一致する。",
        "supported_fields": ["player_positioning", "player_selling_points", "guide_use_hint"],
    },
    "STN680MS": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/collection?rod=%E3%83%AD%E3%83%BC%E3%83%89%E3%83%A9%E3%83%B3%E3%83%8A%E3%83%BC+%E3%82%B9%E3%83%88%E3%83%A9%E3%82%AF%E3%83%81%E3%83%A3%E3%83%BC+NXS+STN680MS",
        "confidence": "high",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "matched_context": "tackledb には STN680MS の伊藤巧使用例と AI 整理例があり、野良ネズミ、ISワスプ、スグリ60、アンクルミノーに加えて、PE+leader、Carolina、Texas、heavy downshot まで確認できる。公式説明の cover neko / down shot / jighead wacky と合わせ、単なる cover finesse より PE spin / structure rig が近い。",
        "supported_fields": ["player_positioning", "player_selling_points", "guide_use_hint"],
    },
    "HB6100ML-Gc": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/collection?rod=%E3%83%AD%E3%83%BC%E3%83%89%E3%83%A9%E3%83%B3%E3%83%8A%E3%83%BC+%E3%83%B4%E3%82%A9%E3%82%A4%E3%82%B9+HB6100ML",
        "confidence": "high",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "matched_context": "検索結果に HB6100ML-Gc / HB6100MLG / HB6100ML が混在し、ショットオーバー5、タダマキ112、コンプリートフラット68、デカピーナッツII SSR、コンバットクランク320、ラピッズブレードなど巻物寄りの使用例が多い。",
        "supported_fields": ["player_positioning", "player_selling_points", "guide_use_hint"],
    },
    "680JMHS": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/search?rod=680JMHS",
        "confidence": "medium",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "matched_context": "検索結果と公式説明の両方で power finesse / heavy cover / PE+leader の方向が一致する。",
        "supported_fields": ["player_positioning", "player_selling_points", "guide_use_hint"],
    },
    "700JHS": {
        "source_site": "tackledb.uosoku.com",
        "source_url": "https://tackledb.uosoku.com/search?rod=700JHS",
        "confidence": "medium",
        "matched_category": "バス釣り",
        "target_fish": ["ブラックバス"],
        "matched_context": "検索結果と公式説明の両方で heavy power finesse と長尺での cover 攻略方向が一致する。",
        "supported_fields": ["player_positioning", "player_selling_points", "guide_use_hint"],
    },
    "STN650LS": {
        "source_site": "rods.jp",
        "source_url": "https://www.rods.jp/brand-item.php?brand2_id=105",
        "confidence": "low",
        "matched_category": "バスルアー / ロッド仕様",
        "target_fish": ["ブラックバス"],
        "matched_context": "rods.jp は STN650LS の type、power、lure weight、price を確認できるが、玩家文案には直接転記しない。",
        "supported_fields": ["player_environment"],
    },
    "STN610LLS": {
        "source_site": "rods.jp",
        "source_url": "https://www.rods.jp/brand-item.php?brand2_id=105",
        "confidence": "low",
        "matched_category": "バスルアー / ロッド仕様",
        "target_fish": ["ブラックバス"],
        "matched_context": "rods.jp は STN610LLS の type、power、lure weight、price を確認できるが、玩家文案には直接転記しない。",
        "supported_fields": ["player_environment"],
    },
}

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
}

SSL_CONTEXT = ssl._create_unverified_context()


def ensure_dirs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)


def normalize_text(value):
    text = unescape(str(value or ""))
    text = text.replace("\u00a0", " ")
    text = text.replace("〜", "∼").replace("～", "∼").replace("~", "∼")
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r"\n\s*", "\n", text)
    return text.strip()


def strip_html(value):
    text = re.sub(r"(?i)<br\s*/?>", "\n", value or "")
    text = re.sub(r"(?is)<script.*?</script>|<style.*?</style>", " ", text)
    text = re.sub(r"(?is)<[^>]+>", " ", text)
    text = normalize_text(text)
    return re.sub(r"\s+", " ", text).strip()


def slug_from_url(url):
    return url.rstrip("/").split("/")[-1]


def sanitize_filename(value):
    base = normalize_text(value)
    base = re.sub(r"[^\x00-\x7F]+", "", base)
    base = re.sub(r'[\\/*?:"<>|]+', "_", base)
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
    base = re.sub(r"_+", "_", base).strip("_")
    return base or "nories_rod"


def fetch_url(url, cache_path):
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8")
    req = urllib.request.Request(url, headers=HTTP_HEADERS)
    with urllib.request.urlopen(req, timeout=45, context=SSL_CONTEXT) as response:
        raw = response.read()
        charset = response.headers.get_content_charset() or "utf-8"
    html = raw.decode(charset, errors="replace")
    cache_path.write_text(html, encoding="utf-8")
    time.sleep(0.4)
    return html


def download_image(url, file_stem):
    if not url:
        return ""
    ext = Path(url.split("?", 1)[0]).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
        ext = ".jpg"
    filename = f"{file_stem}{ext}"
    local_path = IMAGE_DIR / filename
    old_path = OLD_IMAGE_DIR / filename
    if REFRESH_IMAGES and local_path.exists():
        local_path.unlink()
    if local_path.exists():
        return f"{STATIC_IMAGE_BASE}/{filename}"
    if old_path.exists():
        local_path.write_bytes(old_path.read_bytes())
        return f"{STATIC_IMAGE_BASE}/{filename}"
    try:
        req = urllib.request.Request(url, headers=HTTP_HEADERS)
        with urllib.request.urlopen(req, timeout=45, context=SSL_CONTEXT) as response:
            local_path.write_bytes(response.read())
        time.sleep(0.25)
        return f"{STATIC_IMAGE_BASE}/{filename}"
    except Exception:
        return ""


def extract_article(html):
    match = re.search(r"(?is)<article\b[^>]*>(.*?)</article>", html)
    return match.group(1) if match else html


def first_text(pattern, html):
    match = re.search(pattern, html, flags=re.I | re.S)
    return strip_html(match.group(1)) if match else ""


def first_attr(pattern, html):
    match = re.search(pattern, html, flags=re.I | re.S)
    return unescape(match.group(1)).strip() if match else ""


def extract_master_description(article):
    h4_match = re.search(r"(?is)<h4[^>]*>.*?</h4>", article)
    if not h4_match:
        return ""
    stop = re.search(r"(?is)<table[^>]*>\s*<tbody>\s*<tr[^>]*>.*?<h3>", article[h4_match.end():])
    block = article[h4_match.end(): h4_match.end() + stop.start()] if stop else article[h4_match.end():]
    paragraphs = [strip_html(p) for p in re.findall(r"(?is)<p[^>]*>(.*?)</p>", block)]
    paragraphs = [p for p in paragraphs if p and p != "&nbsp;"]
    return "\n\n".join(paragraphs)


def collect_feature_text(article):
    chunks = []
    for block in re.findall(r"(?is)<dl[^>]*>.*?</dl>", article):
        text = strip_html(block)
        if text:
            chunks.append(text)
    return " ".join(dict.fromkeys(chunks))


def parse_spec_table(section_html):
    table_match = re.search(r"(?is)<table[^>]*>\s*<tbody>(.*?)</tbody>\s*</table>", section_html)
    if not table_match:
        return {}
    rows = {}
    for row in re.findall(r"(?is)<tr[^>]*>(.*?)</tr>", table_match.group(1)):
        key_match = re.search(r"(?is)<th[^>]*>(.*?)</th>", row)
        val_match = re.search(r"(?is)<td[^>]*>(.*?)</td>", row)
        if not key_match or not val_match:
            continue
        key = strip_html(key_match.group(1))
        val = strip_html(val_match.group(1))
        if key:
            rows[normalize_key(key)] = val
    return rows


def normalize_key(key):
    return re.sub(r"[^a-z0-9]+", "", normalize_text(key).lower())


def clean_lure_value(value):
    text = normalize_text(value)
    text = re.sub(r"oz\.?", "", text, flags=re.I)
    text = re.sub(r"(?<=\d)g\b", "", text, flags=re.I)
    return normalize_text(text)


def split_line_value(value):
    text = normalize_text(value).replace("PE #", "PE#").replace("PE #", "PE#")
    if not text:
        return "", ""
    pe_max_match = re.search(r"PE\s*MAX\s*#\s*[0-9.]+", text, flags=re.I)
    pe_match = pe_max_match or re.search(r"(?:PE\s*)?#\s*[0-9.]+(?:\s*[∼~]\s*[0-9.]+)?", text, flags=re.I)
    pe = ""
    line = text
    if pe_match:
        pe_raw = re.sub(r"\s+", "", pe_match.group(0))
        if pe_max_match:
            pe = re.sub(r"(?i)^PEMAX", "PE MAX", pe_raw)
        else:
            pe = pe_raw if pe_raw.upper().startswith("PE") else f"PE{pe_raw}"
        line = normalize_text((text[:pe_match.start()] + " " + text[pe_match.end():]).replace("/", " "))
    if "lb" not in line.lower() and re.fullmatch(r"(?:MAX\s*)?[0-9.]+(?:\s*[∼~]\s*[0-9.]+)?", line, flags=re.I):
        line = f"{line}lb."
    return normalize_text(line), normalize_text(pe)


def normalize_power(value, sku=""):
    text = normalize_text(value)
    upper = text.upper()
    replacements = [
        ("LIGHT-LIGHT", "LL"),
        ("LIGHT LIGHT", "LL"),
        ("MED-LIGHT", "ML"),
        ("MEDIUM-LIGHT", "ML"),
        ("MED LIGHT", "ML"),
        ("MED-HEAVY", "MH"),
        ("MEDIUM-HEAVY", "MH"),
        ("MED HEAVY", "MH"),
        ("EX-HEAVY", "XH"),
        ("EXTRA HEAVY", "XH"),
        ("HEAVY", "H"),
        ("MEDIUM", "M"),
        ("LIGHT", "L"),
    ]
    for raw, normalized in replacements:
        if raw in upper:
            return normalized
    sku_match = re.search(r"(XXH|XH|MH|ML|LL|UL|H|M|L)", sku.upper())
    return sku_match.group(1) if sku_match else text


def map_spec_to_fields(spec, sku):
    line = ""
    pe = ""
    for key, value in spec.items():
        if "line" in key:
            line, pe = split_line_value(value)
            break
    lure_g = ""
    lure_oz = ""
    for key, value in spec.items():
        if "lure" in key:
            if "g" in key and "oz" not in key:
                lure_g = clean_lure_value(value)
            else:
                lure_oz = clean_lure_value(value)
    total_length = spec.get("totallengthcmft", "") or spec.get("lengthcmft", "") or spec.get("lengthfttotallengthcm", "")
    if sku == "HB630M" and total_length == "19(6'3\")":
        total_length = "192(6'3\")"
    return {
        "SKU": sku,
        "POWER": normalize_power(spec.get("power", ""), sku),
        "TOTAL LENGTH": total_length,
        "Action": spec.get("taper", ""),
        "PIECES": spec.get("piece", ""),
        "CLOSELENGTH": spec.get("closedlengthcm", "") or spec.get("closedlength", ""),
        "WEIGHT": spec.get("weightg", ""),
        "Tip Diameter": spec.get("topdiamm", "") or spec.get("topdiamm", ""),
        "LURE WEIGHT": lure_g,
        "LURE WEIGHT (oz)": lure_oz,
        "Line Wt N F": line,
        "PE Line Size": pe,
        "Handle Length": spec.get("reargriplengthmm", ""),
        "Market Reference Price": spec.get("price", ""),
    }


def infer_positioning(series_model, sku, type_code, power, description):
    text = f"{series_model} {sku} {power} {description}".upper()
    if any(word in text for word in ["JUNGLE", "フリッピン", "PUNCHING", "パンチング"]):
        if type_code == "S":
            return "Power finesse / Heavy cover"
        return "Heavy cover / 打撃"
    if "FROG" in text:
        return "Frog / Cover"
    if "BIG" in text or "SWIM" in text or "EX-HEAVY" in text or power in {"XH", "XXH"}:
        return "Big bait / 強力"
    if any(word in text for word in ["CRANK", "HARD BAIT", "HARD"]) or any(
        word in text for word in ["SPINNER", "VIBRATION", "CHATTER", "JERK", "SHAD", "TOPWATER"]
    ):
        return "Hard bait / 巻物搜索"
    if any(
        word in text
        for word in [
            "FINESSE",
            "NEKO",
            "DROP SHOT",
            "JIGHEAD",
            "MID STROLL",
            "RAT INSECT",
            "ネコ",
            "ダウンショット",
            "ジグヘッド",
            "ワッキー",
            "ライトリグ",
            "ミドスト",
            "虫",
            "ムシ",
            "ツルスト",
        ]
    ):
        if any(word in text for word in ["COVER", "カバー", "吊るし", "ヘビーカバー"]):
            return "Cover finesse / 軟餌操作"
        return "Finesse / 軟餌操作"
    if any(
        word in text
        for word in [
            "TEXAS",
            "WORM",
            "JIG",
            "NO SINKER",
            "FREE RIG",
            "CAROLINA",
            "BOTTOM",
            "テキサス",
            "ワーム",
            "ラバージグ",
            "フットボール",
            "ノーシンカー",
            "フリーリグ",
            "キャロ",
            "ボトム",
            "ソフトベイト",
            "カバー撃ち",
        ]
    ):
        return "Soft bait / Structure"
    if any(word in text for word in ["COVER", "カバー", "障害物"]):
        return "Heavy cover / 打撃" if power in {"H", "XH", "XXH"} else "Cover / 打撃"
    if type_code == "S":
        return "Spinning finesse / Bass"
    return "Bass 泛用 / 打撃"


def infer_selling_points(positioning):
    if re.search(r"cover finesse", positioning, flags=re.I):
        return "Cover finesse 路線清楚 / 細線或中細線釣組在障礙邊、吊るし、定點 shake 時更容易控線與補刺"
    if "Hard bait" in positioning:
        return "硬餌路線清楚 / crank、shad、jerk、spinnerbait 等移動餌更容易穩定節奏與命中率"
    if "Heavy cover" in positioning or "Big bait" in positioning or "Frog" in positioning:
        return "強力線組、高負荷路亞與障礙區控魚餘量更高 / 適合 cover、frog、big bait 或 power game"
    if re.search(r"finesse|Spinning", positioning, flags=re.I):
        return "細線小餌與 PE 精細玩法適配 / 竿尖信號、短咬口與定點操作更容易判斷"
    if "Soft bait" in positioning:
        return "軟餌打撃與 structure 攻略取向明確 / Texas、jig、neko、free rig 等底層操作更好分工"
    return "Bass 日常泛用覆蓋較廣 / 軟餌打撃、硬餌搜索和岸釣常見場景都容易搭配"


def infer_guide_layout(series_model, type_code, text):
    source = normalize_text(text)
    series = normalize_text(series_model).upper()
    if "JUNGLE" in series and type_code == "C":
        return "Fuji ステンレスフレーム SiC + K 系導環：太線・PE 使用とテレスコ収納時の干渉軽減を重視"
    if "JUNGLE" in series and type_code == "S":
        return "Fuji チタンフレーム トルザイト + KR 系導環：power finesse の軽量化と PE 出線を重視"
    if "トルザイト" in source or "TORZITE" in source.upper():
        return "Fuji チタンフレーム トルザイト / KR 系導環：軽量化と細線 PE の糸抜けを重視した構成"
    if "SIC-S" in source.upper():
        return "Fuji SiC-S + K 系導環：PE 対応と出線安定を重視した NORIES オリジナル構成"
    if "SIC" in source.upper() or "SiC" in source:
        return "Fuji ステンレスフレーム SiC + K 系導環：太めのラインと PE 使用時の出線安定を重視"
    if "ガイド" in source:
        return "NORIES オリジナル導環セッティング：番手ごとの重量バランスとライン処理を重視"
    return ""


def infer_guide_hint(positioning):
    if "Hard bait" in positioning:
        return "Bass 巻物：出線が安定し、crank、shad、spinnerbait を一定速度で引く時のライン処理がしやすい。"
    if "Heavy cover" in positioning or "Big bait" in positioning or "Frog" in positioning:
        return "Bass 強力：太めのラインと高負荷ルアーの出線が安定し、障礙区の控魚と近距離打撃に余裕を持たせやすい。"
    if re.search(r"cover finesse", positioning, flags=re.I):
        return "Bass cover finesse：細線或 PE 的出線更穩，吊るし、短距離 shake 和障礙邊補刺時更容易維持線路角度。"
    if re.search(r"finesse|Spinning", positioning, flags=re.I):
        return "Bass 精細：細線小餌の出線が滑らかで、竿先信号、短いバイト、PE 精細操作を判断しやすい。"
    if "Soft bait" in positioning:
        return "Bass 軟餌：底物、Texas、jig のラインスラック処理が安定し、着底感と短い咬口を拾いやすい。"
    return "Bass 泛用：出線順暢、兼容多種線徑，軟餌、硬餌和搜索場景切換更自然。"


def merge_terms(terms):
    merged = []
    for term in terms:
        text = normalize_text(term)
        if text and text not in merged:
            merged.append(text)
    return " / ".join(merged)


def infer_product_technical(model, sku, type_code):
    model_upper = normalize_text(model).upper()
    terms = []
    if "VOICE LTT" in model_upper:
        terms.extend(
            [
                "富士工業製ステンレスフレームSICガイド",
                "Kガイド",
                "MNSTトップガイド",
                "富士工業製PTSリールシート",
                "富士工業製WBCバランサー",
            ]
        )
    elif "HARD BAIT SPECIAL" in model_upper:
        terms.append("バリアブルテーパー")
        if sku.endswith("-Gc"):
            terms.extend(["グラスコンポジット（-Gc）", "バキューム"])
        if type_code == "S":
            terms.extend(
                [
                    "富士工業製チタンフレームSiC KRコンセプトガイド",
                    "SGt（シャキットグラスティップ）",
                    "富士工業製VSSリールシート",
                ]
            )
        elif sku == "HB660L-Gc":
            terms.extend(
                [
                    "チタンフレームガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
        elif sku == "HB6100ML-Gc":
            terms.extend(
                [
                    "ステンレスフレームオールダブルフットガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
        elif sku in {"HB6100M-Gc", "HB730MH-Gc"}:
            terms.extend(
                [
                    "チタンフレームガイド",
                    "シングルフットティップガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
            if sku == "HB730MH-Gc":
                terms.append("グリップジョイント")
        else:
            terms.extend(
                [
                    "富士工業製GMステンレスフレームSiC-Sガイド",
                    "Kガイド",
                    "LGSTトップガイド",
                    "富士工業製PTSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
            if sku in {"HB710LL", "HB760L", "HB760M"}:
                terms.append("テレスコピックハンドル")
            if sku == "HB730MH-Gc":
                terms.append("グリップジョイント")
    elif "STRUCTURE NXS" in model_upper:
        terms.append("ストラクチャーNXSブランク")
        if type_code == "S":
            terms.extend(
                [
                    "富士工業製チタンフレームトルザイトガイド",
                    "KLガイド",
                    "KTガイド",
                    "富士工業製VSSシート",
                ]
            )
        else:
            terms.extend(
                [
                    "富士工業製チタンフレームトルザイトガイド",
                    "LRVガイド",
                    "KTガイド",
                    "富士工業製ECSリールシート",
                ]
            )
        if sku.endswith("-St"):
            terms.append("ショートカーボンソリッドティップ")
        if sku in {"STN720MH", "STN720H"}:
            terms.append("テレスコピック")
        if sku == "STN640MLS-Md":
            terms.append("モーメントディレイブランク")
        if sku == "STN511LLS":
            terms.append("アンサンドフィニッシュ")
    elif "VOICE JUNGLE" in model_upper:
        if type_code == "S":
            terms.extend(
                [
                    "富士工業製チタンフレームトルザイトリング",
                    "KRコンセプト",
                    "富士工業製VSSリールシート",
                ]
            )
        else:
            terms.extend(
                [
                    "テレスコピック",
                    "富士工業製ステンレスフレームSICガイド",
                    "Kガイド",
                    "MNSTトップガイド",
                    "テレスコピックストッパー",
                    "富士工業製TCSリールシート",
                    "富士工業製WBCバランサー",
                ]
            )
    return merge_terms(terms)


def apply_player_overrides(sku, fields):
    override = PLAYER_FIELD_OVERRIDES.get(sku)
    if override:
        fields.update(override)
    return fields


def parse_variants(article, model, feature_text):
    matches = list(re.finditer(r'(?is)<h3[^>]*>\s*<span\s+id="([^"]*)"[^>]*></span>(.*?)</h3>', article))
    variants = []
    for index, match in enumerate(matches):
        sku = normalize_text(match.group(1))
        if not sku:
            continue
        h3_html = match.group(2)
        h3_text = strip_html(h3_html)
        section = article[match.end(): matches[index + 1].start() if index + 1 < len(matches) else len(article)]
        small = first_text(r"(?is)<small[^>]*>(.*?)</small>", h3_html)
        type_code = "S" if "spinning" in f"{small} {h3_text}".lower() else "C"
        paragraphs = [strip_html(p) for p in re.findall(r"(?is)<p[^>]*>(.*?)</p>", section)]
        description = next((p for p in paragraphs if p and p != "&nbsp;" and "表示されている価格" not in p), "")
        spec = parse_spec_table(section)
        fields = map_spec_to_fields(spec, sku)
        power = fields["POWER"]
        positioning = infer_positioning(model, sku, type_code, power, f"{h3_text} {description}")
        guide_source = f"{feature_text} {description}"
        fields.update(
            {
                "TYPE": type_code,
                "official_environment": "Fresh Water / Bass",
                "player_environment": "淡水 / Bass",
                "player_positioning": positioning,
                "player_selling_points": infer_selling_points(positioning),
                "guide_layout_type": infer_guide_layout(model, type_code, guide_source),
                "guide_use_hint": infer_guide_hint(positioning),
                "product_technical": infer_product_technical(model, sku, type_code),
                "Description": description,
            }
        )
        apply_player_overrides(sku, fields)
        variants.append(
            {
                "sku": sku,
                "product_technical": fields.get("product_technical", ""),
                "title": h3_text,
                "model_type": small,
                "description": description,
                "specs": spec,
                "raw_spec": spec,
                "raw_specs": spec,
                "fields": fields,
            }
        )
    return variants


def parse_product(url, html):
    article = extract_article(html)
    model = first_text(r'(?is)<h2[^>]*class="[^"]*\bmainh2\b[^"]*"[^>]*>(.*?)</h2>', article)
    model_cn = first_text(r"(?is)<h4[^>]*>(.*?)</h4>", article)
    main_image_url = first_attr(r'(?is)<article\b[^>]*>.*?<img[^>]+src="([^"]+)"', html)
    description = extract_master_description(article)
    feature_text = collect_feature_text(article)
    variants = parse_variants(article, model, feature_text)
    price_values = []
    for variant in variants:
        price = variant["fields"].get("Market Reference Price", "")
        if price and price not in price_values:
            price_values.append(price)
    item = {
        "brand": "NORIES",
        "kind": "rod",
        "model": model,
        "model_cn": model_cn,
        "model_year": "",
        "alias": model,
        "source_url": url,
        "main_image_url": main_image_url,
        "description": description,
        "feature_text": feature_text,
        "official_reference_price": " / ".join(price_values),
        "variants": variants,
        "raw_data_hash": hashlib.sha256(html.encode("utf-8")).hexdigest(),
        "scraped_at": datetime.now(timezone.utc).isoformat(),
    }
    item["fit_style_tags"] = infer_master_fit_style_tags(item)
    return item


def row_from_headers(headers, values):
    return [values.get(header, "") for header in headers]


def parse_piece_count(value):
    match = re.search(r"\d+", str(value or ""))
    if not match:
        return None
    count = int(match.group(0))
    if 1 <= count <= 10:
        return count
    return None


def infer_master_fit_style_tags(item):
    tags = ["bass"]
    text = " ".join(
        [
            item.get("model", ""),
            item.get("model_cn", ""),
            item.get("alias", ""),
            item.get("description", ""),
        ]
    ).lower()
    has_travel_term = any(term in text for term in ["travel", "mobile", "多节", "多節", "振出"])
    has_three_plus = any(
        (parse_piece_count(variant.get("fields", {}).get("PIECES")) or 0) >= 3
        for variant in item.get("variants", [])
    )
    if has_travel_term or has_three_plus:
        tags.append("旅行")
    return ",".join(tag for tag in FIT_STYLE_TAG_ORDER if tag in tags)


def master_positioning(model):
    if "HARD BAIT" in model.upper():
        return "淡水 Bass / Hard bait"
    if "STRUCTURE" in model.upper():
        return "淡水 Bass / Soft bait structure"
    if "JUNGLE" in model.upper():
        return "淡水 Bass / Heavy cover"
    return "淡水 Bass / Versatile"


def master_selling_points(model):
    if "HARD BAIT" in model.upper():
        return "硬餌專用竿系 / cast accuracy、巻物節奏與咬口追従性是核心"
    if "STRUCTURE" in model.upper():
        return "軟餌與 structure 攻略竿系 / Texas、jig、neko、finesse 等番手分工清楚"
    if "JUNGLE" in model.upper():
        return "Heavy cover 與 power finesse 竿系 / 太線、障礙區控魚和強力打撃取向明確"
    return "Road Runner 泛用竿系 / 硬餌、軟餌與 power game 番手覆蓋完整"


def build_workbook(normalized):
    wb = Workbook()
    ws_rod = wb.active
    ws_rod.title = "rod"
    ws_detail = wb.create_sheet("rod_detail")
    ws_rod.append(ROD_HEADERS)
    ws_detail.append(ROD_DETAIL_HEADERS)

    detail_index = 10000
    detail_id_by_sku = {}
    for index, item in enumerate(normalized):
        rod_id = f"{ROD_PREFIX}{1000 + index}"
        image_url = download_image(item["main_image_url"], f"{rod_id}_{sanitize_filename(item['model'])}")
        ws_rod.append(
            row_from_headers(
                ROD_HEADERS,
                {
                    "id": rod_id,
                    "brand_id": BRAND_ID,
                    "model": item["model"],
                    "model_cn": item["model_cn"],
                    "model_year": item["model_year"],
                    "alias": item["alias"],
                    "type_tips": "BASS ROD",
                    "fit_style_tags": item.get("fit_style_tags", "bass"),
                    "images": image_url,
                    "series_positioning": "Fresh Water / Bass",
                    "main_selling_points": item["description"].replace("\n", " ")[:220],
                    "official_reference_price": item["official_reference_price"],
                    "market_status": "官网展示",
                    "Description": item["description"],
                    "player_positioning": master_positioning(item["model"]),
                    "player_selling_points": master_selling_points(item["model"]),
                },
            )
        )
        for variant in item["variants"]:
            detail_id = f"{ROD_DETAIL_PREFIX}{detail_index}"
            ws_detail.append(
                row_from_headers(
                    ROD_DETAIL_HEADERS,
                    {
                        "id": detail_id,
                        "rod_id": rod_id,
                        "AdminCode": detail_id,
                        **variant["fields"],
                    },
                )
            )
            detail_id_by_sku[variant["sku"]] = detail_id
            detail_index += 1

    for ws, headers in [(ws_rod, ROD_HEADERS), (ws_detail, ROD_DETAIL_HEADERS)]:
        ws.freeze_panes = "A2"
        for col_idx, header in enumerate(headers, start=1):
            width = 90 if header == "Description" else 48 if header == "product_technical" else 24 if header in {"player_selling_points", "guide_layout_type", "guide_use_hint"} else 16
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    shade_detail_groups(ws_detail)
    wb.save(OUTPUT_FILE)
    write_whitelist_evidence(normalized, detail_id_by_sku)


def write_whitelist_evidence(normalized, detail_id_by_sku):
    by_sku = {}
    rod_id_by_model = {item["model"]: f"{ROD_PREFIX}{1000 + idx}" for idx, item in enumerate(normalized)}
    for item in normalized:
        for variant in item["variants"]:
            by_sku[variant["sku"]] = {
                "rod_id": rod_id_by_model[item["model"]],
                "detail_id": detail_id_by_sku.get(variant["sku"], ""),
                "model": item["model"],
                "sku": variant["sku"],
            }

    evidence_rows = []
    for sku, evidence in WHITELIST_EVIDENCE_BY_SKU.items():
        base = by_sku.get(sku)
        if not base:
            continue
        evidence_rows.append(
            {
                "brand": "NORIES",
                **base,
                **evidence,
                "write_policy": "sidecar evidence only; official specs remain sourced from nories.com",
            }
        )

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "import_file": str(OUTPUT_FILE),
        "policy": {
            "source_boundary": "Nories official pages remain authoritative for official specs and descriptions.",
            "whitelist_role": "Whitelist sites support player-context fields only when they match exact SKU or clearly named series/model.",
            "non_write_fields": [
                "official_environment",
                "hook_keeper_included",
                "sweet_spot_lure_weight_real",
            ],
        },
        "source_summaries": [
            {
                "source_site": "tackledb.uosoku.com",
                "useful_for": [
                    "player_environment",
                    "player_positioning",
                    "player_selling_points",
                    "guide_use_hint",
                ],
                "limits": "Includes AI/virtual tackle records; use as supporting context, not official fact.",
            },
            {
                "source_site": "rods.jp",
                "useful_for": ["spec/category cross-check"],
                "limits": "Only sparse Nories coverage was found for this batch; not enough for broad player text.",
            },
            {
                "source_site": "rodsearch.com",
                "useful_for": [],
                "limits": "Search pages did not provide stable SKU-level Nories records for this batch.",
            },
        ],
        "evidence": evidence_rows,
    }
    WHITELIST_EVIDENCE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def shade_detail_groups(ws):
    fill_a = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
    fill_b = PatternFill(fill_type="solid", fgColor="FFE8F1FB")
    header = [cell.value for cell in ws[1]]
    rod_col = header.index("rod_id") + 1
    last_rod_id = None
    group = -1
    for row in range(2, ws.max_row + 1):
        rod_id = ws.cell(row=row, column=rod_col).value
        if rod_id != last_rod_id:
            group += 1
            last_rod_id = rod_id
        fill = fill_a if group % 2 == 0 else fill_b
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill


def main():
    ensure_dirs()
    fetch_url(ENTRY_URL, CACHE_DIR / "rod_category.html")
    normalized = []
    for url in PRODUCT_URLS:
        html = fetch_url(url, CACHE_DIR / f"{slug_from_url(url)}.html")
        normalized.append(parse_product(url, html))
    NORMALIZED_PATH.write_text(json.dumps(normalized, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    build_workbook(normalized)
    print(
        json.dumps(
            {
                "products": len(normalized),
                "variants": sum(len(item["variants"]) for item in normalized),
                "with_images": sum(1 for item in normalized if item.get("main_image_url")),
                "output": str(OUTPUT_FILE),
                "normalized": str(NORMALIZED_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
