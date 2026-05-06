from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel
import json
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('evergreen_rod_import.xlsx')
NORMALIZED_PATH = resolve_data_raw('evergreen_rod_normalized.json')
REPORT_PATH = resolve_data_raw('evergreen_rod_discontinued_notice_cleanup_report.json')
SHADE_SCRIPT = ROOT / "scripts/shade_evergreen_rod_detail_groups_stage2.py"

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")
DISCONTINUED_RE = re.compile(r"\s*※?\s*当製品の生産は終了いたしました。\s*")
SUCCESSOR_LINK_RE = re.compile(r"\s*当ロッドの後継機種・[^。]*?はコチラ→\s*")


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_model_prefix(value, model):
    text = n(value)
    model_text = n(model)
    if model_text and text.startswith(model_text):
        return n(text[len(model_text):])
    return text


def clean_display_text(value, model=""):
    original = n(value)
    if not original:
        return ""
    text = DISCONTINUED_RE.sub(" ", original)
    text = SUCCESSOR_LINK_RE.sub(" ", text)
    text = n(text)
    if text != original:
        text = strip_model_prefix(text, model)
    return text


def split_sentences(value):
    text = n(value)
    if not text:
        return []
    return [n(part) for part in re.findall(r"[^。.!！?？]+[。.!！?？]?", text) if n(part)]


def is_model_only(value, model):
    return bool(n(value) and n(model) and n(value) == n(model))


def summarize_description(description, model):
    text = strip_model_prefix(clean_display_text(description, model), model)
    for part in split_sentences(text):
        if len(part) >= 8 and not is_model_only(part, model):
            return part[:180]
    return text[:180]


def preview(value, limit=180):
    text = n(value)
    return text if len(text) <= limit else text[:limit] + "..."


def add_change(changes, sheet, row, source_id, field, old_value, new_value):
    changes.append({
        "sheet": sheet,
        "row": row,
        "id": source_id,
        "field": field,
        "old_preview": preview(old_value),
        "new_preview": preview(new_value),
    })


def clean_workbook():
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    rod_col = {n(cell.value): idx + 1 for idx, cell in enumerate(rod_ws[1])}
    detail_col = {n(cell.value): idx + 1 for idx, cell in enumerate(detail_ws[1])}

    for required in ["id", "model", "main_selling_points", "Description"]:
        if required not in rod_col:
            raise RuntimeError(f"rod missing column: {required}")
    for required in ["id", "rod_id", "Description"]:
        if required not in detail_col:
            raise RuntimeError(f"rod_detail missing column: {required}")

    model_by_rod_id = {}
    changes = []

    for row_idx in range(2, rod_ws.max_row + 1):
        rod_id = n(rod_ws.cell(row_idx, rod_col["id"]).value)
        model = n(rod_ws.cell(row_idx, rod_col["model"]).value)
        model_by_rod_id[rod_id] = model

        desc_cell = rod_ws.cell(row_idx, rod_col["Description"])
        old_desc = n(desc_cell.value)
        new_desc = clean_display_text(old_desc, model)
        if old_desc != new_desc:
            desc_cell.value = new_desc
            desc_cell.fill = copy(YELLOW_FILL)
            add_change(changes, "rod", row_idx, rod_id, "Description", old_desc, new_desc)

        main_cell = rod_ws.cell(row_idx, rod_col["main_selling_points"])
        old_main = n(main_cell.value)
        new_main = clean_display_text(old_main, model)
        if not new_main or is_model_only(new_main, model):
            new_main = summarize_description(new_desc, model)
        if old_main != new_main:
            main_cell.value = new_main
            main_cell.fill = copy(YELLOW_FILL)
            add_change(changes, "rod", row_idx, rod_id, "main_selling_points", old_main, new_main)

    for row_idx in range(2, detail_ws.max_row + 1):
        detail_id = n(detail_ws.cell(row_idx, detail_col["id"]).value)
        rod_id = n(detail_ws.cell(row_idx, detail_col["rod_id"]).value)
        model = model_by_rod_id.get(rod_id, "")
        desc_cell = detail_ws.cell(row_idx, detail_col["Description"])
        old_desc = n(desc_cell.value)
        new_desc = clean_display_text(old_desc, model)
        if old_desc != new_desc:
            desc_cell.value = new_desc
            desc_cell.fill = copy(YELLOW_FILL)
            add_change(changes, "rod_detail", row_idx, detail_id, "Description", old_desc, new_desc)

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)
    return changes


def clean_normalized():
    data = json.loads(NORMALIZED_PATH.read_text())
    changes = []

    for idx, item in enumerate(data):
        model = n(item.get("model"))
        old_desc = n(item.get("description"))
        new_desc = clean_display_text(old_desc, model)
        if old_desc != new_desc:
            item["description"] = new_desc
            add_change(changes, "normalized", idx, model, "description", old_desc, new_desc)

        for variant_idx, variant in enumerate(item.get("variants") or []):
            old_variant_desc = n(variant.get("description"))
            new_variant_desc = clean_display_text(old_variant_desc, model)
            if old_variant_desc != new_variant_desc:
                variant["description"] = new_variant_desc
                add_change(
                    changes,
                    "normalized.variants",
                    variant_idx,
                    n(variant.get("sku")) or model,
                    "description",
                    old_variant_desc,
                    new_variant_desc,
                )

    NORMALIZED_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n")
    return changes


def count_remaining_workbook():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=False)
    counts = {}
    for sheet_name, fields in {
        "rod": ["main_selling_points", "Description"],
        "rod_detail": ["Description"],
    }.items():
        ws = wb[sheet_name]
        col = {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
        for field in fields:
            count = 0
            for row_idx in range(2, ws.max_row + 1):
                value = n(ws.cell(row_idx, col[field]).value)
                if DISCONTINUED_RE.search(value):
                    count += 1
            counts[f"{sheet_name}.{field}"] = count
    return counts


def count_remaining_normalized():
    data = json.loads(NORMALIZED_PATH.read_text())
    count = 0
    for item in data:
        if DISCONTINUED_RE.search(n(item.get("description"))):
            count += 1
        for variant in item.get("variants") or []:
            if DISCONTINUED_RE.search(n(variant.get("description"))):
                count += 1
    return count


def main():
    workbook_changes = clean_workbook()
    normalized_changes = clean_normalized()
    report = {
        "xlsx_file": str(XLSX_PATH),
        "normalized_file": str(NORMALIZED_PATH),
        "rule": "Remove Evergreen discontinued product page notice from user-facing Description and main_selling_points. If main_selling_points becomes empty/model-only, refill it from the cleaned official description first sentence.",
        "workbook_changed_cells": len(workbook_changes),
        "normalized_changed_fields": len(normalized_changes),
        "remaining_workbook_notice_counts": count_remaining_workbook(),
        "remaining_normalized_notice_count": count_remaining_normalized(),
        "workbook_changes": workbook_changes,
        "normalized_changes": normalized_changes,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "workbook_changed_cells": len(workbook_changes),
        "normalized_changed_fields": len(normalized_changes),
        "remaining_workbook_notice_counts": report["remaining_workbook_notice_counts"],
        "remaining_normalized_notice_count": report["remaining_normalized_notice_count"],
        "report_file": str(REPORT_PATH),
        "samples": workbook_changes[:8],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
