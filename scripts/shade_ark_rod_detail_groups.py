from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/ark_rod_import.xlsx")
FILL_A = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
FILL_B = PatternFill(fill_type="solid", fgColor="FFE8F1FB")


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    try:
        rod_col = headers.index("rod_id") + 1
    except ValueError as exc:
        raise RuntimeError("rod_id column not found") from exc

    current_group = -1
    last_rod_id = None
    for row in range(2, ws.max_row + 1):
        rod_id = ws.cell(row=row, column=rod_col).value
        if rod_id != last_rod_id:
            current_group += 1
            last_rod_id = rod_id
        fill = FILL_A if current_group % 2 == 0 else FILL_B
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    wb.save(XLSX_PATH)
    print({"file": str(XLSX_PATH), "group_rows": max(ws.max_row - 1, 0), "fills": 2})


if __name__ == "__main__":
    main()
