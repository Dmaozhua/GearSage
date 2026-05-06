#!/usr/bin/env python3
"""Fill Raid Japan rod_detail.product_technical from official Raid pages only."""

from __future__ import annotations

import copy
import json
import re
import subprocess
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "raid_rod_normalized.json"
XLSX_PATH = DATA_DIR / "raid_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "raid_rod_product_technical_stage5_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_raid_rod_detail_groups.py"

FIELD = "product_technical"
INSERT_AFTER = "Description"
DELIMITER = " / "


KNOWN_MODEL_CODES = {
    "GX-65ML+C-ST",
    "GX-67MHC-ST",
    "GX-70M+C",
    "GX-70HC-ST",
    "GXT-70HC-ST",
    "GX-70H+C",
    "GX-71XHC-ST",
    "GX-72MH+C",
    "GX-59XLS-AS",
    "GX-59ULS-AS",
    "GX-61ULS-ST",
    "GX-61LS",
    "GX-64LS-ST",
    "GX-64LS-STMD",
    "GX-67MLS-PMD",
    "GA-62BF",
    "GA-64MLC",
    "GA-65PBF",
    "GA-67MHTC",
    "GA-610MC",
    "GA-610MHC",
    "GA-70MGC",
    "GA-72HC",
    "GA-74XHC",
    "GA-75XXHC",
    "GA-61ULS-ST",
    "GA-63LS",
    "GA-65MS",
    "GA-67L+S",
    "GA-611MLS-ST",
    "GA-70HS-ST",
    "GA-74ML+S",
}

SOURCE_REJECTION_REASON = (
    "Raid Japan rod pages expose technical wording only inside Description/spec blocks; "
    "current GearSage rules allow product_technical only from an independent official technology "
    "module, technology applicability note, or official technology page."
)


def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def model_code(sku: str) -> str:
    value = norm(sku)
    value = re.split(r"\s+【", value, maxsplit=1)[0]
    value = value.split()[0] if value.split() else value
    return value


def merge_terms(terms) -> str:
    merged = []
    for term in terms:
        value = norm(term)
        if value and value not in merged:
            merged.append(value)
    return DELIMITER.join(merged)


def strip_product_technical(data):
    stripped = copy.deepcopy(data)
    for item in stripped:
        item.pop(FIELD, None)
        fields = item.get("fields")
        if isinstance(fields, dict):
            fields.pop(FIELD, None)
        for variant in item.get("variants", []):
            variant.pop(FIELD, None)
            variant_fields = variant.get("fields")
            if isinstance(variant_fields, dict):
                variant_fields.pop(FIELD, None)
    return stripped


def sheet_values(ws, ignore_field):
    headers = [cell.value for cell in ws[1]]
    keep_cols = [idx for idx, header in enumerate(headers, start=1) if header != ignore_field]
    return [
        tuple(ws.cell(row=row, column=col).value for col in keep_cols)
        for row in range(1, ws.max_row + 1)
    ]


def copy_column_style(ws, source_col, target_col):
    source_letter = ws.cell(row=1, column=source_col).column_letter
    target_letter = ws.cell(row=1, column=target_col).column_letter
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row=row, column=source_col)
        target = ws.cell(row=row, column=target_col)
        target._style = copy.copy(source._style)
        if source.has_style:
            target.font = copy.copy(source.font)
            target.fill = copy.copy(source.fill)
            target.border = copy.copy(source.border)
            target.alignment = copy.copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy.copy(source.protection)


def delete_column_if_present(ws, field):
    headers = [cell.value for cell in ws[1]]
    removed = False
    for col in reversed([idx + 1 for idx, header in enumerate(headers) if header == field]):
        ws.delete_cols(col)
        removed = True
    return removed


def ensure_column_after(ws, field, after_field):
    headers = [cell.value for cell in ws[1]]
    matches = [idx + 1 for idx, header in enumerate(headers) if header == field]
    if len(matches) > 1:
        raise RuntimeError(f"duplicate {field} columns in {ws.title}: {matches}")
    if after_field not in headers:
        raise RuntimeError(f"{ws.title} missing anchor column: {after_field}")
    desired_col = headers.index(after_field) + 2
    if matches:
        current_col = matches[0]
        if current_col == desired_col:
            return current_col, False
        ws.delete_cols(current_col)
        headers = [cell.value for cell in ws[1]]
    after_col = headers.index(after_field) + 1
    insert_col = after_col + 1
    ws.insert_cols(insert_col)
    copy_column_style(ws, after_col, insert_col)
    ws.cell(row=1, column=insert_col).value = field
    ws.column_dimensions[ws.cell(row=1, column=insert_col).column_letter].width = 48
    return insert_col, True


def build_variant_values(data):
    values = []
    changed_variants = []
    missing_mapping = []
    removed_item_fields = []
    for item_index, item in enumerate(data):
        rod_id = f"RR{1000 + item_index}"
        if FIELD in item:
            removed_item_fields.append({"rod_id": rod_id, "before": item.get(FIELD)})
            item.pop(FIELD, None)
        item_fields = item.get("fields")
        if isinstance(item_fields, dict):
            item_fields.pop(FIELD, None)
        for variant in item.get("variants", []):
            variant_fields = variant.get("fields")
            sku = norm(variant.get("sku") or (variant_fields or {}).get("SKU"))
            code = model_code(sku)
            if code not in KNOWN_MODEL_CODES:
                missing_mapping.append({"rod_id": rod_id, "sku": sku, "model_code": code})
            value = ""
            before = norm(variant.get(FIELD) or (variant_fields or {}).get(FIELD))
            variant[FIELD] = value
            if isinstance(variant_fields, dict):
                variant_fields[FIELD] = value
            row = {"rod_id": rod_id, "sku": sku, "model_code": code, "value": value}
            values.append(row)
            if before != value:
                changed_variants.append({**row, "before": before})
    return {
        "values": values,
        "changed_variants": changed_variants,
        "missing_mapping": missing_mapping,
        "removed_item_fields": removed_item_fields,
    }


def update_normalized():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    before_guard = strip_product_technical(data)
    result = build_variant_values(data)
    if strip_product_technical(data) != before_guard:
        raise RuntimeError("normalized guard failed: fields other than product_technical changed")
    NORMALIZED_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {
        "items": len(data),
        "variants": len(result["values"]),
        "variant_product_technical_nonempty": sum(1 for row in result["values"] if row["value"]),
        **result,
    }


def update_workbook(variant_values):
    wb = load_workbook(XLSX_PATH)
    ws_rod = wb["rod"]
    ws_detail = wb["rod_detail"]

    rod_before = sheet_values(ws_rod, FIELD)
    detail_before = sheet_values(ws_detail, FIELD)

    rod_column_removed = delete_column_if_present(ws_rod, FIELD)
    detail_col, detail_column_inserted_or_moved = ensure_column_after(ws_detail, FIELD, INSERT_AFTER)

    headers = [cell.value for cell in ws_detail[1]]
    rod_id_col = headers.index("rod_id") + 1
    sku_col = headers.index("SKU") + 1

    values_by_key = {
        (row["rod_id"], model_code(row["sku"])): row["value"]
        for row in variant_values
    }
    workbook_changes = []
    unmatched_rows = []
    for row_idx in range(2, ws_detail.max_row + 1):
        rod_id = norm(ws_detail.cell(row=row_idx, column=rod_id_col).value)
        code = model_code(ws_detail.cell(row=row_idx, column=sku_col).value)
        key = (rod_id, code)
        if key not in values_by_key:
            unmatched_rows.append({"row": row_idx, "rod_id": rod_id, "model_code": code})
        value = values_by_key.get(key, "")
        cell = ws_detail.cell(row=row_idx, column=detail_col)
        before = norm(cell.value)
        cell.value = value
        if before != value:
            workbook_changes.append(
                {"row": row_idx, "rod_id": rod_id, "model_code": code, "before": before, "after": value}
            )

    if sheet_values(ws_rod, FIELD) != rod_before:
        raise RuntimeError("Unexpected rod sheet changes outside product_technical column removal")
    if sheet_values(ws_detail, FIELD) != detail_before:
        raise RuntimeError("Unexpected rod_detail changes outside product_technical")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    return {
        "rod_column_removed": rod_column_removed,
        "rod_detail_column_inserted_or_moved": detail_column_inserted_or_moved,
        "rod_detail_rows": ws_detail.max_row - 1,
        "rod_detail_product_technical_nonempty": sum(1 for row in variant_values if row["value"]),
        "workbook_changes": workbook_changes,
        "unmatched_rows": unmatched_rows,
    }


def validate_delimiter(values):
    bad = []
    for row in values:
        value = row["value"]
        if " / " not in value and "/" in value:
            bad.append(row)
    return bad


def main():
    normalized_report = update_normalized()
    xlsx_report = update_workbook(normalized_report["values"])
    delimiter_issues = validate_delimiter(normalized_report["values"])
    if normalized_report["missing_mapping"] or xlsx_report["unmatched_rows"] or delimiter_issues:
        raise RuntimeError("validation failed; see stdout")
    report = {
        "field": FIELD,
        "scope": "Raid Japan rod_detail SKU-level product_technical only",
        "source_policy": (
            "Official independent technology modules, technology applicability notes, or official technology pages only; "
            "Description/spec blocks/whitelist/player inference are not used."
        ),
        "delimiter": DELIMITER,
        "column_position": "rod_detail after Description; rod master must not contain product_technical",
        "blank_policy": SOURCE_REJECTION_REASON,
        "normalized": normalized_report,
        "xlsx": xlsx_report,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
