#!/usr/bin/env python3
"""Fill Jackall rod_detail.product_technical from official Jackall tech modules only."""

from __future__ import annotations

import json
import re
import subprocess
from copy import deepcopy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "jackall_rod_normalized.json"
XLSX_PATH = DATA_DIR / "jackall_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "jackall_rod_product_technical_stage5_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_jackall_rod_detail_groups.py"

FIELD = "product_technical"
ROD_PREFIX = "JR"
ROD_DETAIL_PREFIX = "JRD"
DELIMITER = " / "


def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def merge_terms(terms) -> str:
    seen = set()
    cleaned = []
    for term in terms:
        value = norm(term)
        if not value or value in seen:
            continue
        seen.add(value)
        cleaned.append(value)
    return DELIMITER.join(cleaned)


def base_sku(sku: str) -> str:
    return re.sub(r"/2$", "", norm(sku))


def infer_product_technical(item: dict, variant: dict) -> str:
    category = item.get("category_key")
    fields = variant.get("fields") or {}
    sku = norm(fields.get("SKU") or variant.get("sku"))
    base = base_sku(sku)
    rod_type = norm(fields.get("TYPE"))
    terms = []

    if category == "revoltage-rod":
        terms.append("富士工業社製SiCガイドリング")
        source_text = norm(f"{item.get('model', '')} {item.get('listing_model', '')} {item.get('url', '')}")
        is_revoltage_2pcs = bool(re.search(r"2pcs|ツーピース", source_text, re.I))
        if is_revoltage_2pcs:
            t1100_skus = {
                "RVⅡ-C69L+BF",
                "RVⅡ-C66M+",
                "RVⅡ-C67MH+",
                "RVⅡ-C68MH",
                "RVⅡ-C610M",
                "RVⅡ-C71H-ST",
                "RVⅡ-C73H",
                "RVⅡ-C711H",
                "RVⅡ-S65UL",
                "RVⅡ-S65L",
                "RVⅡ-S67ML",
                "RVⅡ-S68MH+",
                "RVⅡ-S69UL+",
                "RVⅡ-S78ML+",
            }
            m40x_skus = {"RVⅡ-C64ML-ST", "RVⅡ-S68MH+", "RVⅡ-C68MH", "RVⅡ-C71H-ST", "RVⅡ-C711H", "RVⅡ-S69UL+", "RVⅡ-S61L-ST"}
            grip_joint_skus = set()
            solid_30t_skus = {"RVⅡ-S61L-ST", "RVⅡ-C64ML-ST", "RVⅡ-C71H-ST"}
        else:
            t1100_skus = {
                "RVⅡ-C67L-FM",
                "RVⅡ-C69L+BF",
                "RVⅡ-C610M",
                "RVⅡ-C66M+",
                "RVⅡ-C68MH",
                "RVⅡ-C67MH+",
                "RVⅡ-C71H-ST",
                "RVⅡ-C73H",
                "RVⅡ-C711H",
                "RVⅡ-S510SUL-ST",
                "RVⅡ-S60SUL",
                "RVⅡ-S65UL",
                "RVⅡ-S68MH+",
                "RVⅡ-S69UL+",
                "RVⅡ-S78ML+",
            }
            m40x_skus = {
                "RVⅡ-C64ML-ST",
                "RVⅡ-C68MH",
                "RVⅡ-C71H-ST",
                "RVⅡ-C711H",
                "RVⅡ-S510SUL-ST",
                "RVⅡ-S61L-ST",
                "RVⅡ-S68MH+",
                "RVⅡ-S69UL+",
            }
            grip_joint_skus = {
                "RVⅡ-C67MH+",
                "RVⅡ-C68MH",
                "RVⅡ-C69L+BF",
                "RVⅡ-C610M",
                "RVⅡ-C71H-ST",
                "RVⅡ-C73H",
                "RVⅡ-C711H",
                "RVⅡ-S68MH+",
                "RVⅡ-S69UL+",
                "RVⅡ-S78ML+",
            }
            solid_30t_skus = {"RVⅡ-S510SUL-ST", "RVⅡ-S61L-ST", "RVⅡ-C64ML-ST", "RVⅡ-C71H-ST"}
        if base in t1100_skus:
            terms.append("トレカ®T1100G")
        if base in m40x_skus:
            terms.append("トレカ®M40X")
        if base in {"RVⅡ-C62L-GC", "RVⅡ-C66ML-GC"}:
            terms.extend(["UD Glass", "グラスコンポジットブランク"])
        if base in solid_30t_skus:
            terms.append("30tカーボンソリッドティップ")
        if not is_revoltage_2pcs and base == "RVⅡ-S61UL-ST":
            terms.append("ロングソリッドティップ")
        if not is_revoltage_2pcs and base == "RVⅡ-S56XUL/L-ST":
            terms.append("ショートソリッドティップ")
        if not is_revoltage_2pcs and base == "RVⅡ-C64M+PBF":
            terms.append("スパイラルガイドセッティング")
        if base in grip_joint_skus:
            terms.append("グリップジョイント構造")
        if rod_type == "S":
            terms.extend(["Kガイド", "LYガイド"])
            if (not is_revoltage_2pcs and base == "RVⅡ-S510SUL-ST") or base == "RVⅡ-S68MH+":
                terms.append("全ガイドKガイド")
        if base == "RVⅡ-S78ML+":
            terms.append("ダブルロックシステム")
        if not is_revoltage_2pcs and base == "RVⅡ-S69UL+":
            terms.append("大径トップ〜第4ガイド")

    elif category == "bpm":
        if sku.startswith("BP-"):
            terms.extend(["Fujiガイド", "1&ハーフ設計"])
            if rod_type == "C":
                terms.append("Fuji ECSリールシート")
            if rod_type == "S":
                terms.append("Fuji TVSリールシート")
            if re.search(r"BP-S70L-ST|BP-C70M\+\s*ST", sku):
                terms.append("カーボンソリッドティップ")
        else:
            terms.extend(["富士工業製アルコナイトガイド", "SiCトップガイド"])
            if sku in {"B1-C66MLG", "B1-C70MG", "B2-C66MLG"}:
                terms.append("グラスコンポジットブランク")
            if sku in {"B1-C72MH", "B1-C70H", "B1-C73XHSB", "B1-S73M", "B1-C70MG"}:
                terms.append("グリップジョイント構造")

    elif category == "nazzy-choice":
        if "SG" in sku:
            terms.extend(["スパイラルガイド", "蓄光スレッド", "グラスコンポジットブランクス", "ソフトティップ"])
        else:
            terms.extend(
                [
                    "Fuji製オールダブルフットガイド",
                    "ステンレスフレームSiCトップガイド",
                    "ステンレスフレームアルコナイトガイド",
                ]
            )
            if "CACAOBLACK" in sku:
                terms.extend(["グラスコンポジットティップ", "並継構造"])
                terms.append("PMNST6トップガイド")
            else:
                terms.extend(["グラスコンポジット素材", "並継構造"])

    return merge_terms(terms)


def strip_product_technical_from_normalized(data):
    cleaned = deepcopy(data)
    for item in cleaned:
        item.pop(FIELD, None)
        for variant in item.get("variants", []):
            variant.pop(FIELD, None)
            variant.get("fields", {}).pop(FIELD, None)
    return cleaned


def normalize_workbook_values(workbook, ignore_field: str):
    values = {}
    for sheet_name in ("rod", "rod_detail"):
        ws = workbook[sheet_name]
        headers = [cell.value for cell in ws[1]]
        ignore_indexes = {idx + 1 for idx, header in enumerate(headers) if header == ignore_field}
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(tuple(value for idx, value in enumerate(row, start=1) if idx not in ignore_indexes))
        values[sheet_name] = rows
    return values


def ensure_detail_field(ws):
    headers = [cell.value for cell in ws[1]]
    matches = [idx + 1 for idx, header in enumerate(headers) if header == FIELD]
    if len(matches) > 1:
      raise RuntimeError(f"duplicate {FIELD} columns in rod_detail: {matches}")
    if matches:
        return matches[0]
    if "Description" not in headers:
        raise RuntimeError("Description column not found in rod_detail")
    insert_at = headers.index("Description") + 2
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = FIELD
    return insert_at


def remove_rod_level_field(ws):
    headers = [cell.value for cell in ws[1]]
    for idx in reversed([i + 1 for i, header in enumerate(headers) if header == FIELD]):
        ws.delete_cols(idx)


def main():
    normalized = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    normalized_before_guard = strip_product_technical_from_normalized(normalized)

    technical_by_key = {}
    changed_variants = []
    for item_index, item in enumerate(normalized):
        rod_id = f"{ROD_PREFIX}{1000 + item_index}"
        item.pop(FIELD, None)
        for variant in item.get("variants", []):
            fields = variant.setdefault("fields", {})
            sku = norm(fields.get("SKU") or variant.get("sku"))
            value = infer_product_technical(item, variant)
            old_value = norm(fields.get(FIELD) or variant.get(FIELD))
            fields[FIELD] = value
            variant[FIELD] = value
            technical_by_key[(rod_id, sku)] = value
            if old_value != value:
                changed_variants.append({"rod_id": rod_id, "sku": sku, "old": old_value, "new": value})

    if strip_product_technical_from_normalized(normalized) != normalized_before_guard:
        raise RuntimeError("normalized guard failed: fields other than product_technical changed")

    wb = load_workbook(XLSX_PATH)
    workbook_before_guard = normalize_workbook_values(wb, FIELD)

    remove_rod_level_field(wb["rod"])
    detail_ws = wb["rod_detail"]
    tech_col = ensure_detail_field(detail_ws)
    headers = [cell.value for cell in detail_ws[1]]
    rod_id_col = headers.index("rod_id") + 1
    sku_col = headers.index("SKU") + 1

    changed_cells = []
    for row_idx in range(2, detail_ws.max_row + 1):
        key = (norm(detail_ws.cell(row=row_idx, column=rod_id_col).value), norm(detail_ws.cell(row=row_idx, column=sku_col).value))
        value = technical_by_key.get(key, "")
        cell = detail_ws.cell(row=row_idx, column=tech_col)
        old_value = norm(cell.value)
        cell.value = value
        if old_value != value:
            changed_cells.append({"row": row_idx, "rod_id": key[0], "sku": key[1], "old": old_value, "new": value})

    workbook_after_guard = normalize_workbook_values(wb, FIELD)
    if workbook_after_guard != workbook_before_guard:
        raise RuntimeError("workbook guard failed: fields other than product_technical changed")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)
    NORMALIZED_PATH.write_text(json.dumps(normalized, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    nonempty = sum(1 for value in technical_by_key.values() if value)
    report = {
        "field": FIELD,
        "source_policy": "official Jackall tech modules, SKU descriptions, lineup/spec lists with explicit model-technology mappings, or official tech pages only",
        "xlsx": str(XLSX_PATH),
        "normalized": str(NORMALIZED_PATH),
        "variants": len(technical_by_key),
        "nonempty": nonempty,
        "blank": len(technical_by_key) - nonempty,
        "changed_xlsx_cells": len(changed_cells),
        "changed_normalized_variants": len(changed_variants),
        "sample_changed_cells": changed_cells[:20],
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
