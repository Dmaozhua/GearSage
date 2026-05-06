#!/usr/bin/env python3
import copy
import json
import re
import subprocess
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "abu_rods_normalized.json"
XLSX_PATH = DATA_DIR / "abu_rod_import.xlsx"
EVIDENCE_PATH = DATA_DIR / "abu_rod_product_technical_official_evidence.json"
REPORT_PATH = DATA_DIR / "abu_rod_product_technical_stage7_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_abu_rod_detail_groups.py"

FIELD = "product_technical"
DETAIL_ANCHOR = "Description"


def n(value):
    return str(value or "").strip()


def headers(ws):
    return [cell.value for cell in ws[1]]


def header_col(ws, field):
    current = headers(ws)
    return current.index(field) + 1 if field in current else None


def copy_column_style(ws, source_col, target_col):
    ws.column_dimensions[ws.cell(row=1, column=target_col).column_letter].width = (
        ws.column_dimensions[ws.cell(row=1, column=source_col).column_letter].width
    )
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row=row, column=source_col)
        target = ws.cell(row=row, column=target_col)
        target._style = copy.copy(source._style)
        if source.has_style:
            target.font = copy.copy(source.font)
            target.fill = copy.copy(source.fill)
            target.border = copy.copy(source.border)
            target.alignment = copy.copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy.copy(source.protection)


def ensure_detail_column(ws):
    current = header_col(ws, FIELD)
    anchor = header_col(ws, DETAIL_ANCHOR)
    if not anchor:
        raise RuntimeError(f"{ws.title} missing {DETAIL_ANCHOR}")
    if current:
        if current != anchor + 1:
            raise RuntimeError(f"{ws.title}.{FIELD} must be immediately after {DETAIL_ANCHOR}")
        return current, False
    insert_at = anchor + 1
    ws.insert_cols(insert_at)
    copy_column_style(ws, anchor, insert_at)
    ws.cell(row=1, column=insert_at).value = FIELD
    return insert_at, True


def remove_rod_column_if_present(ws):
    col = header_col(ws, FIELD)
    if not col:
        return False
    ws.delete_cols(col)
    return True


def sheet_values_without_field(ws):
    sheet_headers = headers(ws)
    keep_cols = [idx for idx, header in enumerate(sheet_headers, start=1) if header != FIELD]
    values = []
    for row in range(1, ws.max_row + 1):
        values.append(tuple(ws.cell(row=row, column=col).value for col in keep_cols))
    return values


def normalized_without_field(data):
    def clean(value):
        if isinstance(value, dict):
            return {k: clean(v) for k, v in value.items() if k != FIELD}
        if isinstance(value, list):
            return [clean(v) for v in value]
        return value

    return clean(data)


def add_term(terms, evidence_terms, term, evidence):
    if term not in terms:
        terms.append(term)
        evidence_terms.append({"term": term, "evidence": evidence})


def infer_product_technical(item):
    features = [n(feature) for feature in item.get("features", []) if n(feature)]
    text = " ".join([n(item.get("description")), *features])
    terms = []
    evidence_terms = []

    powerlux_matches = re.findall(r"Powerlux[®™]?\s*(1000|500|200|100)\b", text, flags=re.I)
    for number in sorted(set(powerlux_matches), key=lambda value: int(value), reverse=True):
        source = next((feature for feature in features if re.search(rf"Powerlux[®™]?\s*{number}\b", feature, re.I)), n(item.get("description")))
        add_term(terms, evidence_terms, f"Powerlux® {number}", source)

    if re.search(r"\bROCS[™®]?\b|Robotically Optimized Casting System", text, flags=re.I):
        source = next((feature for feature in features if re.search(r"\bROCS[™®]?\b|Robotically Optimized Casting System", feature, re.I)), n(item.get("description")))
        add_term(terms, evidence_terms, "ROCS™", source)

    if re.search(r"\bIntraCarbon[™®]?\b", text, flags=re.I):
        source = next((feature for feature in features if re.search(r"\bIntraCarbon[™®]?\b", feature, re.I)), n(item.get("description")))
        add_term(terms, evidence_terms, "IntraCarbon™", source)

    if re.search(r"\bCCRS[™®]?\b|Carbon Constructed Reel Seat", text, flags=re.I):
        source = next((feature for feature in features if re.search(r"\bCCRS[™®]?\b|Carbon Constructed Reel Seat", feature, re.I)), n(item.get("description")))
        add_term(terms, evidence_terms, "CCRS™", source)

    return " / ".join(terms), evidence_terms


def infer_variant_product_technical(item, variant):
    terms_value, evidence_terms = infer_product_technical(item)
    terms = [term for term in terms_value.split(" / ") if term]
    specs = variant.get("specs") or {}
    pieces = n(specs.get("Number of Pieces") or specs.get("Pieces"))
    model = n(item.get("model"))
    features = [n(feature) for feature in item.get("features", []) if n(feature)]

    if (
        model == "Beast™ Casting Rod"
        and pieces == "2"
        and any(re.search(r"2 piece ferrule locking mechanism.*select models", feature, re.I) for feature in features)
    ):
        add_term(
            terms,
            evidence_terms,
            "2 piece ferrule locking mechanism",
            next(
                feature
                for feature in features
                if re.search(r"2 piece ferrule locking mechanism.*select models", feature, re.I)
            ),
        )

    return " / ".join(terms), evidence_terms


def validate_value(value, where):
    text = n(value)
    if not text:
        return []
    errors = []
    if "/" in text and " / " not in text:
        errors.append(f"{where}: slash separator must be ` / `")
    forbidden = ["白名单", "玩家", "钓组", "饵型", "来源", "官网确认", "General Lure", "Light Rig"]
    for fragment in forbidden:
        if fragment in text:
            errors.append(f"{where}: forbidden fragment {fragment}")
    if len(text) > 140:
        errors.append(f"{where}: value too long for technical term list")
    return errors


def update_normalized():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    before = normalized_without_field(data)
    changes = []
    evidence = []
    item_field_removed = []

    for item in data:
        model = n(item.get("model"))
        if FIELD in item:
            item_field_removed.append({"model": model, "before": n(item.get(FIELD))})
            item.pop(FIELD, None)
        for variant in item.get("variants", []):
            sku = n(variant.get("sku"))
            value, evidence_terms = infer_variant_product_technical(item, variant)
            before_value = n(variant.get(FIELD))
            if before_value != value:
                variant[FIELD] = value
                changes.append({"model": model, "sku": sku, "before": before_value, "after": value})
            if value:
                evidence.append(
                    {
                        "model": model,
                        "sku": sku,
                        "value": value,
                        "source_url": n(item.get("source_url")),
                        "source_type": "official_product_page_features",
                        "terms": evidence_terms,
                    }
                )

    if normalized_without_field(data) != before:
        raise RuntimeError("normalized guard failed: fields outside product_technical changed")

    NORMALIZED_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    EVIDENCE_PATH.write_text(json.dumps(evidence, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {
        "items": len(data),
        "variants": sum(len(item.get("variants", [])) for item in data),
        "item_product_technical_removed": item_field_removed,
        "variant_product_technical_nonempty": sum(
            1 for item in data for variant in item.get("variants", []) if n(variant.get(FIELD))
        ),
        "changes": changes,
        "evidence_count": len(evidence),
    }


def variant_values_by_sku():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    result = {}
    for item in data:
        for variant in item.get("variants", []):
            sku = n(variant.get("sku"))
            if sku:
                result[sku] = n(variant.get(FIELD))
    return result


def update_workbook():
    sku_values = variant_values_by_sku()
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]

    rod_before = sheet_values_without_field(rod_ws)
    detail_before = sheet_values_without_field(detail_ws)

    rod_removed = remove_rod_column_if_present(rod_ws)
    detail_col, detail_inserted = ensure_detail_column(detail_ws)

    detail_headers = headers(detail_ws)
    sku_col = detail_headers.index("SKU") + 1
    changes = []
    errors = []
    for row in range(2, detail_ws.max_row + 1):
        sku = n(detail_ws.cell(row=row, column=sku_col).value)
        value = sku_values.get(sku, "")
        before_value = n(detail_ws.cell(row=row, column=detail_col).value)
        if before_value != value:
            detail_ws.cell(row=row, column=detail_col).value = value
            changes.append({"row": row, "sku": sku, "before": before_value, "after": value})
        errors.extend(validate_value(value, f"rod_detail row {row} sku {sku}"))

    if sheet_values_without_field(rod_ws) != rod_before:
        raise RuntimeError("Unexpected rod sheet changes outside product_technical column removal")
    if sheet_values_without_field(detail_ws) != detail_before:
        raise RuntimeError("Unexpected rod_detail changes outside product_technical")
    if errors:
        raise RuntimeError(f"product_technical validation failed: {errors[:5]}")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    return {
        "rod_rows": rod_ws.max_row - 1,
        "rod_product_technical_removed": rod_removed,
        "rod_detail_rows": detail_ws.max_row - 1,
        "rod_detail_product_technical_inserted": detail_inserted,
        "rod_detail_product_technical_nonempty": sum(1 for value in sku_values.values() if value),
        "workbook_changes": changes,
    }


def main():
    normalized = update_normalized()
    xlsx = update_workbook()
    report = {
        "field": FIELD,
        "scope": "Abu Garcia rod_detail SKU-level product_technical only",
        "source_policy": "official Abu Garcia product page description/features only; no whitelist/player/spec inference",
        "separator": " / ",
        "column_position": "rod_detail after Description; rod master must not contain product_technical",
        "normalized": normalized,
        "xlsx": xlsx,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "variant_nonempty": normalized["variant_product_technical_nonempty"],
                "detail_nonempty": xlsx["rod_detail_product_technical_nonempty"],
                "workbook_changes": len(xlsx["workbook_changes"]),
                "evidence": str(EVIDENCE_PATH),
                "report": str(REPORT_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
