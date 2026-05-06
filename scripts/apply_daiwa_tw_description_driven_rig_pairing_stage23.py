from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')

PAIRING_BY_ID = {
    "DRD10038": "Jighead Rig / Texas Rig / Punching / Spoon / Heavy Carolina Rig / Spinnerbait",
    "DRD10058": "Shad / Minnow / Metal Vibration / Small Rubber Jig / Neko Rig / Jighead Rig / Wacky Jighead Rig / No Sinker / Down Shot / Jerkbait / Topwater Plug / Crankbait",
    "DRD10061": "Jerkbait / Spinnerbait / Topwater Plug / Neko Rig / Small Rubber Jig / No Sinker / Down Shot",
    "DRD10062": "Crankbait / Spinnerbait / Buzzbait / Vibration / Jerkbait / Topwater Plug / Popper / Neko Rig / Small Rubber Jig / No Sinker / Down Shot",
    "DRD10063": "Free Rig / Texas Rig / Leaderless Down Shot / Rubber Jig / Heavy Down Shot / No Sinker",
    "DRD10065": "Free Rig / Light Texas / Leaderless Down Shot / Heavy Down Shot / Rubber Jig / No Sinker",
    "DRD10066": "Texas Rig / Free Rig / Leaderless Down Shot / No Sinker / Rubber Jig / Swim Jig / Swimbait / Spinnerbait",
    "DRD10067": "Spinnerbait / Buzzbait / Chatterbait / Swim Jig / Vibration / Swimbait / Big Bait / Rubber Jig / Texas Rig / Leaderless Down Shot / No Sinker / Topwater Plug / Frog",
    "DRD10068": "Rubber Jig / Texas Rig / Leaderless Down Shot / Free Rig / Heavy Carolina Rig / Swim Jig / No Sinker / Swimbait / Spinnerbait / Frog",
}


def n(value):
    return " ".join(str(value or "").split())


def snapshot_without_target(ws, target_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col != target_col
        })
    return rows


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    target_col = col["recommended_rig_pairing"]
    before = snapshot_without_target(ws, target_col)

    changed = []
    for row in range(2, ws.max_row + 1):
        detail_id = n(ws.cell(row, col["id"]).value)
        if detail_id not in PAIRING_BY_ID:
            continue
        cell = ws.cell(row, target_col)
        old = n(cell.value)
        new = PAIRING_BY_ID[detail_id]
        if old != new:
            cell.value = new
            changed.append({"row": row, "id": detail_id, "old": old, "new": new})

    after = snapshot_without_target(ws, target_col)
    if before != after:
        raise RuntimeError("unexpected non-recommended_rig_pairing changes")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "changed": len(changed),
        "changed_ids": [item["id"] for item in changed],
    })


if __name__ == "__main__":
    main()
