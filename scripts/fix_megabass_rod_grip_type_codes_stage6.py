import json
import re
from collections import defaultdict
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('megabass_rod_import.xlsx')
REPORT_PATH = resolve_data_raw('megabass_rod_whitelist_backfill_report.json')

YELLOW = "FFFFFF00"
FILL_A = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
FILL_B = PatternFill(fill_type="solid", fgColor="FFE8F1FB")


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def readable_grip(value):
    text = n(value)
    if re.search(r"\bEVA\b", text, re.I):
        return "EVA casting grip"
    if re.search(r"\bSplit\b", text, re.I):
        return "Split casting grip"
    if re.search(r"\bFull\b", text, re.I):
        return "Full casting grip"
    return ""


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    col = {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
    grip_col = col["Grip Type"]
    rod_col = col["rod_id"]

    current_group = -1
    last_rod_id = None
    changes = []

    for row_idx in range(2, ws.max_row + 1):
        rod_id = n(ws.cell(row_idx, rod_col).value)
        if rod_id != last_rod_id:
            current_group += 1
            last_rod_id = rod_id

        cell = ws.cell(row_idx, grip_col)
        value = n(cell.value)
        if not value:
            continue
        rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else ""
        if rgb != YELLOW:
            continue
        if not re.match(r"^(?:Cast|Spin)\s+[A-Z](?:\s+-\s+(?:EVA|Split|Full))?$", value, re.I):
            continue

        next_value = readable_grip(value)
        if next_value:
            cell.value = next_value
            cell.fill = PatternFill(fill_type="solid", fgColor=YELLOW)
        else:
            cell.value = None
            cell.fill = copy(FILL_A if current_group % 2 == 0 else FILL_B)

        changes.append({
            "row": row_idx,
            "id": n(ws.cell(row_idx, col["id"]).value),
            "field": "Grip Type",
            "old_value": value,
            "new_value": next_value,
        })

    wb.save(XLSX_PATH)

    if REPORT_PATH.exists():
        report = json.loads(REPORT_PATH.read_text())
    else:
        report = {}

    report["grip_type_code_cleanup"] = {
        "changed_cells": len(changes),
        "cleared_as_opaque_tw_handle_code": sum(1 for item in changes if not item["new_value"]),
        "kept_as_readable_grip_type": sum(1 for item in changes if item["new_value"]),
        "rule": "Tackle Warehouse Cast/Spin A/B/C codes are retailer handle-template labels, not readable grip types. Only EVA/Split/Full suffixes were retained as readable Grip Type values.",
        "changes": changes,
    }

    by_field = defaultdict(int)
    for item in report.get("changes", []):
        if item.get("field") == "Grip Type":
            continue
        by_field[item.get("field")] += 1
    by_field["Grip Type"] = sum(1 for item in changes if item["new_value"])
    report["by_field_after_grip_cleanup"] = dict(sorted(by_field.items()))
    report["changed_cells_after_grip_cleanup"] = sum(by_field.values())

    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps(report["grip_type_code_cleanup"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
