#!/usr/bin/env python3
import json
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = ROOT / "GearSage-client/pkgGear/data_raw"
XLSX_PATH = DATA_DIR / "nories_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "nories_rod_recommended_rig_pairing_stage2_report.json"

FIELD = "recommended_rig_pairing"
AFTER_FIELD = "guide_use_hint"


PAIRINGS = {
    "LTT620PMH": {
        "pairing": "Large Topwater / Bulky Neko Rig / Metal Jig / Football Jig / Vibration / Chatterbait",
        "source": "official_description+whitelist_check",
        "notes": "Official description names large topwater, bulky Neko, metal jig, football jig, vibration, and chatter; tackledb also supports topwater/jig-spinner/softbait use.",
    },
    "LTT630M": {
        "pairing": "Small Crankbait / Wake Minnow Prop Topwater / Light Texas / No Sinker Paddle-tail Worm",
        "source": "official_description",
        "notes": "Official description prioritizes small crank and Laydown Minnow Wake 110 Prop before light Texas and no-sinker shad-tail/paddle-tail worm.",
    },
    "LTT630MH": {
        "pairing": "Cover Texas / Rubber Jig / Frog",
        "source": "official_description",
        "notes": "Official title and description directly name Cover Texas Jig & Frog.",
    },
    "LTT650M": {
        "pairing": "Spinnerbait / Crankbait / Big Minnow Jerkbait / Heavy Down Shot / Bait Neko / Light Metal Jig / Chatterbait",
        "source": "official_description",
        "notes": "Official description names reeling baits first, then heavy down shot and bait Neko around light cover.",
    },
    "LTT650MH": {
        "pairing": "Spinnerbait / Vibration / Mid-size Crankbait / Jerkbait / Texas Rig / Heavy Down Shot",
        "source": "official_description",
        "notes": "Official description emphasizes 3/8oz+ spinnerbait, vibration, mid crank, and jerkbait, then Texas/heavy down shot.",
    },
    "LTT660H": {
        "pairing": "Slow-roll Spinnerbait / Texas Rig / Guarded Rubber Jig / Dairakka Spoon",
        "source": "official_description",
        "notes": "Official description names slow-roll spinnerbait, 6/0 Texas, and guarded rubber jig; no Big Bait evidence.",
    },
    "LTT680MH": {
        "pairing": "Chatterbait / Swim Jig / Buzzbait / Long Cast Texas / Cover Texas",
        "source": "official_description",
        "notes": "Official description leads with chatter, swim jig, and buzzbait, then long-cast/cover Texas.",
    },
    "LTT680H": {
        "pairing": "Large Body Crankbait / Texas Rig / Cherry Rig / No Sinker Flip Gill / Guarded Rubber Jig",
        "source": "official_description",
        "notes": "Official description names large body plugs, Texas with Escape Twin, Cherry Rig, no-sinker Flip Gill, and guarded rubber jig.",
    },
    "LTT690PH": {
        "pairing": "Rubber Jig / Big Crankbait / Large Spinnerbait / Heavy Texas / Chatterbait",
        "source": "official_description",
        "notes": "Official description says slow-down jig through big crank and large spinnerbait; secondary soft-rig range is conservative from heavy versatile positioning.",
    },
    "LTT6100H": {
        "pairing": "Slow-roll Spinnerbait / Swimbait / Big Spoon / 6/0 Texas Rig / Guarded Jig / Heavy Carolina",
        "source": "official_description",
        "notes": "Official description names Crystal S slow rolling, large swimbait, big spoon, 6/0 Texas, guarded jig, and heavy Carolina.",
    },
    "HB660L-Gc": {
        "pairing": "Light Crankbait / Chatterbait / Buzzbait / Spinnerbait / Topwater Plug / Jerkbait",
        "source": "official_description",
        "notes": "Official hard-bait description lists these moving bait and topwater uses.",
    },
    "HB6100ML-Gc": {
        "pairing": "Small Crankbait / 3m Diver Crankbait / 10g Shad / Light Plug",
        "source": "official_description+whitelist_check",
        "notes": "Official Elite Spec description and tackledb examples both support crank/shad/light plug use.",
    },
    "HB6100M-Gc": {
        "pairing": "Spinnerbait / Chatterbait / Mid-size Crankbait",
        "source": "official_description",
        "notes": "Official title BLADE BAIT VACUUM and description directly name spinnerbait, bladed jig/chatterbait, and mid crank.",
    },
    "HB730MH-Gc": {
        "pairing": "Heavy Spinnerbait / Chatterbait / Mid-to-Magnum Crankbait",
        "source": "official_description",
        "notes": "Official description names heavier spinnerbait, bladed jig/chatterbait, and mid-to-magnum crankbait.",
    },
    "HB560L": {
        "pairing": "Jerkbait / Topwater Plug / Crankbait / Spinnerbait",
        "source": "official_description",
        "notes": "Official JERK & ACCURACY description leads with jerkbait/topwater operation, then crank and spinnerbait.",
    },
    "HB511LL": {
        "pairing": "Shad Plug / Small Crankbait / Sub-7g Plug",
        "source": "official_description",
        "notes": "Official description names shad plug, small crank, and sub-7g air-resistant plugs.",
    },
    "HB600L": {
        "pairing": "Short-distance Rod-work Plug / Bank-cover Light Plug / Crankbait / Topwater Plug",
        "source": "official_description",
        "notes": "Official description highlights rod-work plug handling from crankbait to topwater.",
    },
    "HB600M": {
        "pairing": "3/8-1/2oz Spinnerbait / Crankbait",
        "source": "official_description",
        "notes": "Official description names 3/8-1/2oz spinnerbait and crankbait.",
    },
    "HB630LL": {
        "pairing": "Small Crankbait / Shad / Topwater Plug / Sub-7g Plug",
        "source": "official_description",
        "notes": "Official description names small crank, shad, topwater, and sub-7g plug use.",
    },
    "HB630L": {
        "pairing": "Topwater Plug / Crankbait / Spinnerbait / Buzzbait / Vibration / Chatterbait",
        "source": "official_description",
        "notes": "Official description calls it the first hardbait-special rod and lists these moving baits.",
    },
    "HB630M": {
        "pairing": "1/2oz+ Spinnerbait / Cover Crankbait / Vibration",
        "source": "official_description",
        "notes": "Official description prioritizes 1/2oz+ spinnerbaiting, cover cranking, and vibration.",
    },
    "HB640ML": {
        "pairing": "Jerkbait / Shallow Crankbait / 3m Diver Crankbait / Topwater Plug / Buzzbait / Chatterbait / Spinnerbait / Vibration",
        "source": "official_description",
        "notes": "Official description emphasizes jerkbait suitability and then broad hard moving bait coverage.",
    },
    "HB660H": {
        "pairing": "Slow-roll Spinnerbait / 4m Diver Crankbait",
        "source": "official_description",
        "notes": "Official description names 3/4oz spinnerbait slow rolling and 4m diver crankbait.",
    },
    "HB680L": {
        "pairing": "Small Crankbait / Flat-side Shallow Crankbait / Shad / Long-bill Minnow / 3-4m Deep Diver Crankbait",
        "source": "official_description",
        "notes": "Official description names small crank, flat-side shallow crank, shad, long-bill minnow, and 3-4m deep diver.",
    },
    "HB680M": {
        "pairing": "Medium Diver Crankbait / Spinnerbait / Tail Spinner / Pencil Bait / Prop Bait / Buzzbait",
        "source": "official_description",
        "notes": "Official description names medium diver crank, spinnerbait, tail spinner, pencil, prop, and buzzbait.",
    },
    "HB680XH": {
        "pairing": "180-class Big Bait / 50g Magnum Crankbait / Heavy Treble Plug",
        "source": "official_description",
        "notes": "Official description explicitly names 180-class big bait and 50g-class magnum crankbait.",
    },
    "HB710LL": {
        "pairing": "Small Crankbait / Medium Crankbait / Long Cranking",
        "source": "official_description",
        "notes": "Official MASTER OF CRANKING description is specifically small-to-medium crankbait.",
    },
    "HB760L": {
        "pairing": "Crankbait / Deep Diver Crankbait / Spinnerbait / Vibration / Swim Jig",
        "source": "official_description",
        "notes": "Official description says shallow runner through large deep diver, high-speed spinnerbait, vibration, and swim jig.",
    },
    "HB760M": {
        "pairing": "Large Deep Crankbait / Heavy Spinnerbait / Swimming Worm / Swim Jig",
        "source": "official_description",
        "notes": "Official description names large deep crank, heavy spinnerbait, and swimming worm/jig compatibility.",
    },
    "HB640LS-SGt": {
        "pairing": "High-speed Shad Plug / Shad Plug",
        "source": "official_description",
        "notes": "Official title and description define this as a long-cast high-speed shad spinning model.",
    },
    "HB660MLS-SGt": {
        "pairing": "Metal Bait / Metal Vibration / Blade Bait",
        "source": "official_description",
        "notes": "Official SPINNING METAL description supports metal-type reaction baits; blade bait is a conservative related term.",
    },
    "STN680MS": {
        "pairing": "Cover Neko / Down Shot / Jighead Wacky / Fall Bait No Sinker / No Sinker Jerk / Carolina Rig / Texas Rig / Heavy Down Shot",
        "source": "official_description+whitelist_check",
        "notes": "Official description gives Cover Neko, Down Shot, Jighead Wacky, Fall Bait, and No Sinker Jerk; tackledb confirms Carolina/Texas/heavy downshot as secondary.",
    },
    "STN6100MLS": {
        "pairing": "Mid Strolling / Neko Rig / Jighead Wacky / Bug Cover / Offset Neko / No Sinker Fall Bait",
        "source": "official_description",
        "notes": "Official description lists mid-strolling, Neko, jighead wacky, bug-cover PE use, offset Neko, and no-sinker fall bait.",
    },
    "STN580ML": {
        "pairing": "Tsuru-stroll / Hanging Cover Softbait / Small Swisher / Small Popper",
        "source": "official_description+whitelist_check",
        "notes": "Official TSURU STROLL description and previous whitelist check support hanging cover use, with small swisher/popper as secondary.",
    },
    "STN650M": {
        "pairing": "Neko Rig / Heavy Drop Shot",
        "source": "official_description",
        "notes": "Official title and description directly name Neko and heavy drop shot.",
    },
    "STN670H": {
        "pairing": "Heavy Texas / Football Jig",
        "source": "official_description",
        "notes": "Official BOTTOM SENSITIVE description directly names heavy Texas and football jig.",
    },
    "STN680MH": {
        "pairing": "Texas Rig / Vegetation Texas / Block-hole Texas",
        "source": "official_description",
        "notes": "Official TEXAS VERSATILE description is centered on Texas rig variations.",
    },
    "STN6100M": {
        "pairing": "No Sinker Fall Bait / Soft Jerk / Neko Rig / Heavy Down Shot / Light Texas / Swimming Softbait",
        "source": "official_description",
        "notes": "Official description names no-sinker fall bait, soft jerk, Neko, heavy downshot, light Texas, and swimming action.",
    },
    "STN6100MH": {
        "pairing": "Long Cast Texas / Cover Texas / Heavy Worm Rig",
        "source": "official_title_description",
        "notes": "Official title says LONG CAST TEXAS; description confirms long line, cover contact, and worming versatile use.",
    },
    "STN700H": {
        "pairing": "Reaction Carolina / Heavy Carolina / Carolina Worm",
        "source": "official_description",
        "notes": "Official title and description are explicitly reaction Carolina focused.",
    },
    "STN720MH": {
        "pairing": "High-density No Sinker / Free Rig / No Sinker Flip",
        "source": "official_description",
        "notes": "Official title and description name no-sinker flip and free rig.",
    },
    "STN720H": {
        "pairing": "Bulky Texas / Flip Gill Neko Rig / Guarded Rubber Jig",
        "source": "official_description",
        "notes": "Official title and description name bulky Texas, Flip Gill Neko, and guarded rubber jig.",
    },
    "STN660M-St": {
        "pairing": "Reaction Down Shot / Neko Rig / Small Rubber Jig / Paddle-tail Worm",
        "source": "official_description",
        "notes": "Official bait-finesse description names reaction downshot, Neko, small rubber jig, and shad tail.",
    },
    "STN670MH-St": {
        "pairing": "Cover Neko / Cover Small Rubber Jig / Mid-column Shake",
        "source": "official_description",
        "notes": "Official COVER BAIT FINESSE description names cover Neko, cover bait/small rubber jig, and mid-column shake.",
    },
    "STN6100H-St": {
        "pairing": "Cover Texas / Escape Twin 5g Texas / Football Jig",
        "source": "official_description",
        "notes": "Official SENSITIVE COVER WORMING description names cover Texas, Escape Twin 5g Texas, and football.",
    },
    "STN511LLS": {
        "pairing": "Ultra-light Down Shot / Neko Rig / No Sinker",
        "source": "official_description",
        "notes": "Official description names ultra-light sinker downshot, Neko, and no-sinker.",
    },
    "STN610LLS": {
        "pairing": "No Sinker / Down Shot / Split Shot",
        "source": "official_description+rodsjp_category_check",
        "notes": "Official description names no-sinker, downshot, and split shot; rods.jp only supports bass/category/spec context.",
    },
    "STN620LS": {
        "pairing": "Jighead Rig / Split Shot / Down Shot / Neko Rig / No Sinker",
        "source": "official_description",
        "notes": "Official description lists these finesse rigs; no generic Light Rig label is used.",
    },
    "STN640LLS": {
        "pairing": "Mid Strolling Jighead / Mid-strolling Down Shot",
        "source": "official_description",
        "notes": "Official MID STROLLING description names jighead and downshot for mid-column swimming.",
    },
    "STN640MLS-Md": {
        "pairing": "Insect Surface Bait / Rat Insect / Surface Twitch Bait",
        "source": "official_description",
        "notes": "Official RAT INSECT description supports bug/surface twitch techniques.",
    },
    "STN650LS": {
        "pairing": "Neko Rig / Jighead Wacky / No Sinker Boil Shot / High-speed Twitch Finesse",
        "source": "official_description+rodsjp_category_check",
        "notes": "Official description names Neko, jighead wacky, no-sinker boil shot, and high-speed twitch; rods.jp is category/spec only.",
    },
    "760JMH": {
        "pairing": "Flipping Rubber Jig / Texas Rig / High-density No Sinker",
        "source": "official_description",
        "notes": "Official Jungle Stick Light description names flipping cover, rubber jig, Texas, and high-density no-sinker.",
    },
    "760JH": {
        "pairing": "Punching Rig / Heavy Flipping / 1-2oz Sinker Texas",
        "source": "official_description",
        "notes": "Official Jungle Stick description names flipping and 1-2oz sinker punching with heavy PE/line.",
    },
    "680JMHS": {
        "pairing": "Small Rubber Jig Power Finesse / Heavy-cover PE Finesse / Hanging Cover Shake",
        "source": "official_description+whitelist_check",
        "notes": "Official Jungle Spin description puts small rubber jig power finesse in heavy cover first.",
    },
    "700JHS": {
        "pairing": "Long-cast Power Finesse / Hanging Small Rubber Jig / Heavy-cover PE Finesse",
        "source": "official_description+whitelist_check",
        "notes": "Official Jungle Spin Heavy description extends 680JMHS for long cast/high-position hanging power finesse.",
    },
}


GUIDE_HINT_OVERRIDES = {
    "LTT620PMH": "Parabolic versatile：短尺準投後可從 large topwater 切換到 bulky Neko、metal jig / football jig；調性重點是讓重餌不彈口，同時保留 cover 旁控線與補刺角度。",
    "LTT630M": "Light all-round：小型 crank 與 Laydown Minnow Wake 110 Prop topwater 先處理低彈道準投，也能切換 light Texas / no-sinker paddle-tail worm；竿尖要讓硬餌吸入口，butt 則保留 1/0-3/0 hook 的補刺餘量。",
    "LTT630MH": "Cover Texas / Frog：Texas 和 rubber jig 需要連續 pitching 不跳 rig，frog 則重視 side-hand 準投和出水後第一拍控線；太線近距補刺要乾脆。",
    "LTT650M": "Light multi-purpose：spinnerbait、crank、big minnow jerkbait 先做 open-water 搜索，heavy down shot / bait Neko 再處理 light cover；軟硬切換時線路穩定比單純遠投更重要。",
    "LTT650MH": "Mid reeling & worming：3/8oz+ spinnerbait、vibration、mid crank 是主線，也能切換到 Texas / heavy down shot；較長 tip 幫助巻物追従，belly torque 保留單鉤補刺。",
    "LTT660H": "Power strategy：slow-roll spinnerbait 穿 weed、Texas / guarded rubber jig 打硬 vegetation，Dairakka 類 spoon 需要整支竿帶動 lifting；軟硬切換時以太線控線和強補刺為核心。",
    "LTT680MH": "Moving-to-worming：chatterbait、swim jig、buzzbait 先做 slack-line reeling，再切換到 long cast Texas / cover Texas；長線下要能維持泳姿、讀 bottom 並完成長行程補刺。",
    "LTT680H": "Heavy technical versatile：large body crankbait 可慢巻搜索，並能切換到 Texas、Cherry Rig、No Sinker Flip Gill 與 guarded rubber jig 等 heavy softbait；重點是 16lb 級線組下的出線、停頓和強補刺。",
    "LTT690PH": "Parabolic heavy：rubber jig、big crank 和 large spinnerbait 都可用，整支竿的深彎曲讓重餌拋投、吸入口和 landed rate 更穩；軟硬餌切換時靠 blank 承接負荷。",
    "LTT6100H": "Heavy cover versatile：slow-roll spinnerbait、swimbait、big spoon 做 3-5m 強搜索，並能切換到 6/0 Texas、guarded jig、heavy Carolina 處理 cover 和 clear deep；太線長距控線要穩。",
    "HB660L-Gc": "Vacuum versatile：light crank 到 1/2oz chatterbait 都靠 glass composite 追従短咬；buzzbait、spinnerbait、topwater、jerkbait 切換時，重點是低阻尼回彈和準投平衡。",
    "HB6100ML-Gc": "Crank bait vacuum：小型 crank、3m diver crankbait 和 10g shad 是主場，light plug 只作延伸；柔軟 blank 讓 lure 保持原生泳姿，cover 下滑入時仍能黏住短咬。",
    "HB6100M-Gc": "Blade bait vacuum：spinnerbait、chatterbait、mid crank 這類高阻力巻物需要穩定牽引；glass tip 吸入口、fast belly 和 butt 負責單鉤上顎補刺。",
    "HB730MH-Gc": "Power reserve vacuum：heavy spinnerbait、chatterbait、mid-to-magnum crank 需要長距離覆蓋水域；長竿節奏重點是不用 full cast 也能出距離，遠處咬口仍能壓進 hook。",
    "HB560L": "Jerk & accuracy：jerkbait 和 topwater 需要不同角度短抽與準投，crankbait / spinnerbait 則用來穿 cover 邊；短竿優勢是低彈道入點和 rod-work 控姿。",
    "HB511LL": "Hardbait finesse：shad plug、small crank 和 sub-7g plug 要靠竿身載重低彈道送入 cover；柔軟 tip 提升輕 plug 投放，butt torque 負責大魚牽制。",
    "HB600L": "Back-hand accuracy light：bank、reed、dock 周邊的短距 rod-work plug 是主場，light plug 可從 crankbait 切到 topwater；每投距離和角度都不同時，重點是穩定出線、低彈道入點與控姿。",
    "HB600M": "Back-hand accuracy mid：3/8-1/2oz spinnerbait 和 crankbait 面向短距 cover 撞擊；竿身要在 backhand cast 時容易壓載，進入 cover 後保留抗阻和黏魚。",
    "HB630LL": "Side-hand light plug：small crank、shad、topwater 和 sub-7g plug 用較太線送進 cover 周邊；細緻 tip 讓小 plug 有水感，低彈性追従短咬。",
    "HB630L": "Hardbait first rod：topwater、crankbait、spinnerbait、buzzbait、vibration、chatterbait 都可覆蓋；重點是 360 度 casting / backhand 下的出線穩定和 moving bait 乘り。",
    "HB630M": "Side-hand mid power：1/2oz+ spinnerbait、cover crankbait、vibration 面向近距 cover 和 shallow flat；太線進 cover 時需要較強竿身支撐抗阻和補刺。",
    "HB640ML": "Technical long cast：jerkbait 放首位，shallow crank 到 3m diver、topwater、buzzbait、chatterbait、spinnerbait、vibration 都能延伸；tip 強度決定 jerk 節奏和遠投後控線。",
    "HB660H": "Over-head drive：3/4oz spinnerbait slow roll 和 4m diver crankbait 需要長時間穩定 trace；強 power 支撐深層阻力，仍保留 hardbait 咬口追従。",
    "HB680L": "Extra energy drive light：small crank、flat-side shallow crank、shad、long-bill minnow 到 3-4m diver 都重視長投後換線角；適合 weed / rock / cover 多點一次攻完。",
    "HB680M": "Extra energy drive mid：medium diver crankbait、spinnerbait、tail spinner、pencil / prop bait、buzzbait 都偏遠投搜索；長竿重點是泳層維持、 ripping retrieve 和水面下控速。",
    "HB680XH": "Technical power cast：180 class big bait 和 50g magnum crankbait 是明確主軸，不能只按 MH/H 泛寫 big bait；低彈道安靜入水後，XH power 負責太軸 treble 補刺，較短長度保留大型餌操作自由度。",
    "HB710LL": "Master of cranking：small-to-medium crankbait 和 long cranking 是唯一主軸；tele butt 支撐穩定拋投與深咬口，tip 反發要能依 lip 阻力自然變化。",
    "HB760L": "Away distance light：crankbait、deep diver、spinnerbait、vibration 先做遠投搜索，swim jig 可切換到中層單鉤 moving bait；長線下重點是泳層、碰障礙和反應咬口承接。",
    "HB760M": "Away distance mid：large deep crank 和 heavy spinnerbait 是主軸，也能切換到 swimming worm / swim jig；深 weed 與 break edge 長投後，要維持 trace line 和確實補刺。",
    "HB640LS-SGt": "High-speed shad spin：shad plug 低速到超高速都要直線飛行並穩定泳姿；SGt tip 讓高速 reaction bite 不易彈掉，spinning 設定服務長投和細線控速。",
    "HB660MLS-SGt": "Spinning metal：metal bait、metal vibration、blade bait 用快速反應誘發低活性魚；SGt glass tip 讓 fast retrieve 的 miss bite 也能吸入，PE deep game 時降低脫鉤。",
    "STN680MS": "PE spin structure：Cover Neko、Down Shot、Jighead Wacky 先遠投打 deep structure，Fall Bait No Sinker / No Sinker Jerk 再做大鉤自然下沉；Carolina / Texas / heavy downshot 只作白名單補強副線。",
    "STN6100MLS": "Versatile power spin：fluoro 下可做 mid strolling、Neko、Jighead Wacky；PE+leader 下切 bug cover、offset Neko、No Sinker Fall Bait，長距 finesse 重點是線弧和 torque。",
    "STN580ML": "Tsuru-stroll：PE 2 号級線組和 hanging cover softbait 是主軸，short length 追求 1cm 級 pitching；也能切換到 small swisher / small popper 做障礙邊表層精準投放。",
    "STN650M": "Neko & heavy drop shot：5in Neko 和 1/4oz drop shot 面向杭、沈木、light cover；parabolic taper 讓水阻大的 rig 能吃進去，同時保留 tubular tip 的掛障礙控制。",
    "STN670H": "Bottom sensitive：Heavy Texas 打 cover，Football Jig 沿 break 操作；短強 tip 負責躲 laydown / rock 卡點，blank 感度用來讀枝條、岩縫和喰わせ timing。",
    "STN680MH": "Texas versatile：1/4oz Texas + Escape Twin 是核心，vegetation、floating cover、消波 block hole 都要安靜進入並讀 fall / shake；補刺靠整支竿順彎保護線。",
    "STN6100M": "Versatile mid softbait：No Sinker Fall Bait、Soft Jerk、Neko、Heavy Down Shot、Light Texas 都是標準 softbait；長度服務 cover long pitch 和 flat 連續 shake swimming。",
    "STN6100MH": "Long cast Texas：強長 tip 在長線和 cover cushion 下仍能細控 rig，Heavy Worm Rig 不是泛用 jig 竿；重點是遠距 Texas 的讀感、延遲咬口和曲竿浮魚。",
    "STN700H": "Reaction Carolina：Heavy Carolina / Carolina Worm 用來打 flat bottom、structure 和 break edge pin spot；tip 要能帶動 heavy sinker，又能感知 leader 後方 worm 咬口。",
    "STN720MH": "No Sinker flip & free rig：High-density No Sinker 打 vegetation，Free Rig 打 block，輕 rig + 太線需要長距 pitching、細緻 rig control 和低活性淺掛後的順彎承接。",
    "STN720H": "Bulky Texas & Gill bait：Bulky Texas、Flip Gill Neko、guarded rubber jig 都面向 heavy cover；20lb 級線組下要保持單手操作輕快、cover 越し補刺和控魚 torque。",
    "STN660M-St": "Bait finesse solid tip：Reaction Down Shot、Neko、Small Rubber Jig、Paddle-tail Worm 面向高壓魚；solid tip 要吸收極小咬口，同時避開小石縫與 cover 纏線。",
    "STN670MH-St": "Cover bait finesse：Cover Neko 和 cover small rubber jig 是主軸，中層 shake 需要 solid tip 保持 range；MH butt 在 cover 奧補刺後要快速把魚帶出。",
    "STN6100H-St": "Sensitive cover worming：Cover Texas / Escape Twin 5g Texas 是主場，Football Jig 作 fall bite 補強；短 solid tip 讓魚含餌不吐，H power 負責 heavy rig 補刺。",
    "STN511LLS": "Near-the-target finesse：Ultra-light Down Shot、Neko、No Sinker 用 0.2 號 PE 或 2lb fluoro 精細操作；重點是極近目標的底感、短咬口和深水變化感知。",
    "STN610LLS": "Finesse versatile：No Sinker、Down Shot、Split Shot 需要毫米級操作和 sudden bite 追従；短尺 finesse 強調快速貫穿 barb，同時不讓魚取得主導權。",
    "STN620LS": "High-speed finesse：Jighead、Split Shot、Down Shot、Neko、No Sinker 都能攻擊性操作；soft tip 讓魚吃入，extra-fast belly 負責瞬間掛魚，不是等待型 light rig。",
    "STN640LLS": "Mid strolling：Mid Strolling Jighead 和 mid-strolling Down Shot 都靠 slack line 讓 rig 在中層自然漂游；柔軟 tip/belly 防止跳動，強 butt 保留突然大魚補刺。",
    "STN640MLS-Md": "Rat insect surface：虫系 surface bait、Rat Insect 和高速 twitch bait 是主軸；低彈性 carbon tip 配 PE 能在細小 twitch 後追従水面短咬，ML butt 負責 landed rate。",
    "STN650LS": "Neko & jighead wacky：Neko Rig、Jighead Wacky 和 No Sinker Boil Shot 都要長投後精準 twitch / hook set；強 tip 處理水阻大 rig，L butt 保留遠距補刺。",
    "760JMH": "Jungle stick light：Flipping Rubber Jig、Texas、High-density No Sinker 用 20lb line 打 weed clump、芦 / 蒲內側；重點是 cover 奧入點、torque 拔魚和不過硬的食わせ餘量。",
    "760JH": "Jungle stick punching：Punching Rig、Heavy Flipping、1-2oz Sinker Texas 面向 heavy cover；厚 blank 靠 torque 拔魚，太 PE / 25lb 線下仍要能感覺重量並穩定 flip。",
    "680JMHS": "Jungle spin power finesse：small rubber jig power finesse 打 heavy cover 深處，短距 pitching、空中 cover 穿入與細 shake 是核心；PE 出線、太軸 hook 補刺和強 butt 拔魚要連貫。",
    "700JHS": "Long jungle spin：long-cast power finesse 和 hanging small rubber jig 面向遠距、高處吊るし與 monster bass；比 680JMHS 更重視長線 hook stroke、full-power hookset 和高位控線。",
}

FORBIDDEN_GENERIC = ["General Lure", "Light Rig", "Hardbait", "Soft Bait"]

HARD_TERMS = [
    "Crankbait", "Shad", "Minnow", "Spinnerbait", "Chatterbait", "Vibration", "Buzzbait",
    "Topwater", "Jerkbait", "Plug", "Big Bait", "Swimbait", "Spoon", "Metal", "Popper",
    "Swisher", "Pencil", "Prop", "Tail Spinner", "Bladed Jig", "Blade Bait",
]
SOFT_TERMS = [
    "Texas", "Neko", "Down Shot", "Drop Shot", "Jighead", "Wacky", "No Sinker", "Free Rig",
    "Carolina", "Rubber Jig", "Football Jig", "Small Rubber Jig", "Punching", "Flipping",
    "Power Finesse", "Fall Bait", "Split Shot", "Cherry Rig", "Worm", "Soft Jerk",
    "Tsuru-stroll", "Mid Strolling", "Bug", "Insect", "Swim Jig",
]


def normalize(value):
    return " ".join(str(value or "").split())


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.font = copy(source.font)
    target.fill = copy(source.fill)


def header_map(ws):
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def ensure_pairing_column(ws):
    col = header_map(ws)
    if FIELD in col:
        return False
    if AFTER_FIELD not in col:
        raise RuntimeError(f"missing column: {AFTER_FIELD}")
    insert_at = col[AFTER_FIELD] + 1
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = FIELD
    copy_cell_style(ws.cell(row=1, column=insert_at - 1), ws.cell(row=1, column=insert_at))
    for row in range(2, ws.max_row + 1):
        copy_cell_style(ws.cell(row=row, column=insert_at - 1), ws.cell(row=row, column=insert_at))
    ws.column_dimensions[ws.cell(row=1, column=insert_at).column_letter].width = 42
    return True


def shade_detail_groups(ws):
    fill_a = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
    fill_b = PatternFill(fill_type="solid", fgColor="FFE8F1FB")
    col = header_map(ws)
    rod_col = col["rod_id"]
    last_rod_id = None
    group = -1
    for row in range(2, ws.max_row + 1):
        rod_id = ws.cell(row=row, column=rod_col).value
        if rod_id != last_rod_id:
            group += 1
            last_rod_id = rod_id
        fill = fill_a if group % 2 == 0 else fill_b
        for column in range(1, ws.max_column + 1):
            ws.cell(row=row, column=column).fill = fill


def has_any(text, terms):
    return any(term.lower() in text.lower() for term in terms)


def first_item(pairing):
    return normalize(pairing).split(" / ", 1)[0]


def classify_pairing(pairing):
    first = first_item(pairing)
    has_hard = has_any(pairing, HARD_TERMS)
    has_soft = has_any(pairing, SOFT_TERMS)
    first_hard = has_any(first, HARD_TERMS)
    first_soft = has_any(first, SOFT_TERMS)
    return {
        "has_hard": has_hard,
        "has_soft": has_soft,
        "first_hard": first_hard,
        "first_soft": first_soft,
    }


def detect_conflicts(sku, desc, pairing, guide_hint):
    issues = []
    for token in FORBIDDEN_GENERIC:
        if token.lower() in pairing.lower():
            issues.append(f"forbidden_generic:{token}")
    cls = classify_pairing(pairing)
    hint = normalize(guide_hint)
    if cls["first_hard"] and ("軟餌" in hint or "底物" in hint) and not any(word in hint for word in ["泛用", "切換", "切换", "兼顧", "兼顾"]):
        issues.append("first_hard_hint_soft_only")
    if cls["first_soft"] and ("巻物" in hint or "硬餌" in hint) and not any(word in hint for word in ["泛用", "切換", "切换", "兼顧", "兼顾"]):
        issues.append("first_soft_hint_hard_only")
    if cls["has_hard"] and cls["has_soft"] and not any(
        word in hint for word in ["泛用", "切換", "切换", "兼顧", "兼顾", "軟硬", "moving-to-worming"]
    ):
        issues.append("mixed_pairing_hint_not_switching")

    desc_lower = normalize(desc).lower()
    explicit_checks = [
        ("テキサス", "Texas"),
        ("ラバージグ", "Rubber Jig"),
        ("フットボール", "Football Jig"),
        ("ネコ", "Neko"),
        ("ダウンショット", "Down Shot"),
        ("ドロップショット", "Drop Shot"),
        ("ジグヘッド", "Jighead"),
        ("ノーシンカー", "No Sinker"),
        ("フリーリグ", "Free Rig"),
        ("キャロ", "Carolina"),
        ("フロッグ", "Frog"),
        ("クランク", "Crankbait"),
        ("シャッド", "Shad"),
        ("ミノー", "Minnow"),
        ("スピナーベイト", "Spinnerbait"),
        ("チャター", "Chatterbait"),
        ("バイブレーション", "Vibration"),
        ("バズベイト", "Buzzbait"),
        ("トップウォーター", "Topwater"),
        ("ジャークベイト", "Jerkbait"),
        ("スイムベイト", "Swimbait"),
        ("ビッグベイト", "Big Bait"),
        ("パンチング", "Punching"),
    ]
    for raw, expected in explicit_checks:
        if raw.lower() in desc_lower and expected.lower() not in pairing.lower():
            if raw == "シャッド" and "シャッドテール" in desc and "Paddle-tail Worm".lower() in pairing.lower():
                continue
            if raw == "ジグヘッド" and "Jighead Wacky".lower() in pairing.lower():
                continue
            issues.append(f"description_term_not_carried:{expected}")
    return issues


def apply_updates(ws):
    col = header_map(ws)
    required = ["id", "rod_id", "SKU", "Description", "guide_use_hint", FIELD]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")

    changes = []
    rows = []
    seen_skus = set()
    for row in range(2, ws.max_row + 1):
        sku = normalize(ws.cell(row=row, column=col["SKU"]).value)
        seen_skus.add(sku)
        item = PAIRINGS.get(sku)
        if not item:
            rows.append({"row": row, "sku": sku, "issues": ["missing_pairing_mapping"]})
            continue
        old_pairing = normalize(ws.cell(row=row, column=col[FIELD]).value)
        new_pairing = item["pairing"]
        if old_pairing != new_pairing:
            ws.cell(row=row, column=col[FIELD]).value = new_pairing
            changes.append(
                {
                    "row": row,
                    "id": ws.cell(row=row, column=col["id"]).value,
                    "sku": sku,
                    "field": FIELD,
                    "old": old_pairing,
                    "new": new_pairing,
                    "source": item["source"],
                }
            )

        old_hint = normalize(ws.cell(row=row, column=col["guide_use_hint"]).value)
        new_hint = GUIDE_HINT_OVERRIDES.get(sku)
        if new_hint and old_hint != new_hint:
            ws.cell(row=row, column=col["guide_use_hint"]).value = new_hint
            changes.append(
                {
                    "row": row,
                    "id": ws.cell(row=row, column=col["id"]).value,
                    "sku": sku,
                    "field": "guide_use_hint",
                    "old": old_hint,
                    "new": new_hint,
                    "source": "consistency_fix_with_official_pairing",
                }
            )

        guide_hint = normalize(ws.cell(row=row, column=col["guide_use_hint"]).value)
        desc = normalize(ws.cell(row=row, column=col["Description"]).value)
        rows.append(
            {
                "row": row,
                "id": ws.cell(row=row, column=col["id"]).value,
                "rod_id": ws.cell(row=row, column=col["rod_id"]).value,
                "sku": sku,
                "recommended_rig_pairing": new_pairing,
                "guide_use_hint": guide_hint,
                "source": item["source"],
                "notes": item["notes"],
                "issues": detect_conflicts(sku, desc, new_pairing, guide_hint),
            }
        )
    missing_skus = sorted(set(PAIRINGS) - seen_skus)
    if missing_skus:
        raise RuntimeError(f"pairing mappings not matched in workbook: {missing_skus}")
    return changes, rows


def write_report(inserted, changes, rows):
    source_counts = {}
    for row in rows:
        source_counts[row["source"]] = source_counts.get(row["source"], 0) + 1
    issue_rows = [row for row in rows if row["issues"]]
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "import_file": str(XLSX_PATH),
        "field": FIELD,
        "inserted_column": inserted,
        "coverage": {
            "rows": len(rows),
            "filled": sum(1 for row in rows if row["recommended_rig_pairing"]),
        },
        "source_counts": source_counts,
        "write_scope": [
            "rod_detail.recommended_rig_pairing",
            "rod_detail.guide_use_hint",
            "No Description/spec/player fields were edited in this stage.",
        ],
        "forbidden_generic_terms": FORBIDDEN_GENERIC,
        "changes": changes,
        "issue_rows": issue_rows,
        "rows": rows,
    }
    REPORT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return payload


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    inserted = ensure_pairing_column(ws)
    changes, rows = apply_updates(ws)
    shade_detail_groups(ws)
    wb.save(XLSX_PATH)
    report = write_report(inserted, changes, rows)
    print(
        json.dumps(
            {
                "inserted_column": inserted,
                "rows": report["coverage"]["rows"],
                "filled": report["coverage"]["filled"],
                "changes": len(changes),
                "issue_rows": len(report["issue_rows"]),
                "report": str(REPORT_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    if report["issue_rows"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
