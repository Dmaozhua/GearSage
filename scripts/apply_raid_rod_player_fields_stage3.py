import json
import subprocess
from collections import Counter
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = resolve_data_raw('raid_rod_import.xlsx')
REPORT_PATH = resolve_data_raw('raid_rod_player_fields_stage3_report.json')
SHADE_SCRIPT = ROOT / "scripts/shade_raid_rod_detail_groups.py"

TARGET_FIELDS = ["player_environment", "player_positioning", "player_selling_points"]

PLAYER_FIELDS_BY_SKU = {
    "GX-65ML+C-ST": {
        "player_environment": "淡水 Bass / 岸钓近距 cover、轻量 worming",
        "player_positioning": "近距轻量 cover 打撃枪柄",
        "player_selling_points": "短尺轻量但不软脚，Light Texas、Free Rig、Neko 和 backslide no sinker 能低弹道落进 cover；竿尖反馈直接，近距离刺鱼和控鱼更利落。",
    },
    "GX-67MHC-ST": {
        "player_environment": "淡水 Bass / 硬 cover、近距离吊るし与轻中量软饵",
        "player_positioning": "近距 cover finesse 强控枪柄",
        "player_selling_points": "solid tip 负责精密投放和轻咬读取，强 belly/butt 负责挂鱼后快速拉离障碍；适合 Heavy Down Shot、Free Rig、Light Texas 和小橡胶铅头钩。",
    },
    "GX-70M+C": {
        "player_environment": "淡水 Bass / 开放水域到 light cover 岸钓泛用",
        "player_positioning": "12-14lb 中轻量软硬饵高精度泛用",
        "player_selling_points": "以 vibration、spinnerbait、shad tail worm、no sinker 和 light cover rig 切换为核心，抛投距离变化时仍能保持落点精度和线弧控制。",
    },
    "GX-70HC-ST": {
        "player_environment": "淡水 Bass / heavy cover bait finesse",
        "player_positioning": "重 cover 轻量钓组特化枪柄",
        "player_selling_points": "H power 配 solid tip，能把 Light Texas、Free Rig、Neko、Small Rubber Jig 精准打进 cover，吃口后不拖泥带水地把鱼带出来。",
    },
    "GXT-70HC-ST": {
        "player_environment": "淡水 Bass / cover worming 强化、轻量 moving bait 辅助",
        "player_positioning": "轻量化 heavy cover worming 强化版",
        "player_selling_points": "比 THE MAXX 下限更宽，轻量 worming 操作更轻快，同时 butt power 更足；主要服务 cover 软饵打撃，必要时可补少量 vibration 等 moving bait。",
    },
    "GX-70H+C": {
        "player_environment": "淡水 Bass / 大饵横向 power game 与 cover 纵向打撃",
        "player_positioning": "大饵 + heavy rig 双向 power 枪柄",
        "player_selling_points": "Big Bait、Swimbait 和 Alabama Rig 横向搜索有余量，Rubber Jig / Texas Rig 进 cover 时也能保持控底和强制控鱼，是横纵 power game 的切换型。",
    },
    "GX-71XHC-ST": {
        "player_environment": "淡水 Bass / 粗线重饵精细操作、强 cover",
        "player_positioning": "XH 重饵精细操作 power finesse",
        "player_selling_points": "不是只靠硬度硬拉的 XH，30t solid tip 让 Alabama Rig、Swimbait、Rubber Jig 这类有重量和体积的搭配仍能细腻操控。",
    },
    "GX-72MH+C": {
        "player_environment": "淡水 Bass / 岸钓远投、moving bait 与 bottom worming",
        "player_positioning": "MH+ 岸钓软硬饵强泛用",
        "player_selling_points": "Swim Jig、wire bait、vibration、crank、topwater 和 rubber jig、sinker rig、backslide worm 都能覆盖；长握把和 slack line 控制适合岸钓远投切换。",
    },
    "GX-59XLS-AS": {
        "player_environment": "淡水 Bass / 高压场轻量定点精细",
        "player_positioning": "一点シェイク慢诱直柄精细",
        "player_selling_points": "更适合 Down Shot、Neko、Jighead、Small Rubber Jig 的定点慢诱，轻线下能把细小抖动和短促吃口读得更清楚。",
    },
    "GX-59ULS-AS": {
        "player_environment": "淡水 Bass / cover 周边快节奏轻量 finesse",
        "player_positioning": "cover 投放型轻量直柄精细",
        "player_selling_points": "同样是轻量 finesse，但更偏把小钓组快速送到 cover 周边，连续打点、轻线控线和即时回收线弧比慢诱更重要。",
    },
    "GX-61ULS-ST": {
        "player_environment": "淡水 Bass / 高压场 light rig 与 2.5-5g worming",
        "player_positioning": "UL light rig 宽容直柄",
        "player_selling_points": "不只服务超轻小饵，也能承接 2.5g、3.5g 甚至 5g sinker worming；轻量、反馈和搏鱼缓冲之间平衡好。",
    },
    "GX-61LS": {
        "player_environment": "淡水 Bass / 轻量精细到小硬饵快速搜索",
        "player_positioning": "直柄 finesse utility",
        "player_selling_points": "Neko、Down Shot、Jighead 可细腻操作，也能用 Shad、小型 topwater 做反应搜索；适合需要一支短尺直柄快速判断鱼情的场景。",
    },
    "GX-64LS-ST": {
        "player_environment": "淡水 Bass / FC 与 PE 双线精细、近中距 light plugging",
        "player_positioning": "全域 finesse 直柄",
        "player_selling_points": "Down Shot、Jighead、No Sinker、Neko、Hover Strolling、Small Rubber Jig 到 Shad、Minnow、小 topwater 都能覆盖；FC4-5lb 和 PE0.6+leader 都有使用空间。",
    },
    "GX-64LS-STMD": {
        "player_environment": "淡水 Bass / PE mid-strolling、hobast、水面微抖",
        "player_positioning": "PE 专用 mid-strolling 直柄",
        "player_selling_points": "长 regular solid tip 让 PE 也能持续高频 shake，适合 Mid Strolling、Hover Strolling 和水面ピクピク这类需要稳定节奏的精细中层玩法。",
    },
    "GX-67MLS-PMD": {
        "player_environment": "淡水 Bass / power mid-strolling 中层攻略",
        "player_positioning": "重 jighead 中层 roll 操作直柄",
        "player_selling_points": "面向 5-6.5inch 体积感软饵和偏重 jighead 的中层游动，能维持节奏、线弧和泳层；也可兼顾小 Shad / Minnow 的反应操作。",
    },
    "GA-62BF": {
        "player_environment": "淡水 Bass / bait finesse 轻量软饵与 5g plug",
        "player_positioning": "real bait finesse 短尺枪柄",
        "player_selling_points": "1/16oz 级 Small Rubber Jig、Neko、Down Shot 和 5g light plug 都能用枪柄完成，底感直接，适合短距离高精度轻量操作。",
    },
    "GA-64MLC": {
        "player_environment": "淡水 Bass / jerkbait 与小中型 hard plug",
        "player_positioning": "短尺 jerkbait / hard plug 专用",
        "player_selling_points": "80-120mm jerkbait 是主轴，抽停节奏清楚；Crankbait、Vibration、Popper、Pencil、Spinnerbait 也能稳定维持泳层和动作。",
    },
    "GA-65PBF": {
        "player_environment": "淡水 Bass / power bait finesse、轻量 plug cover",
        "player_positioning": "BFS 与标准枪柄之间的 power finesse",
        "player_selling_points": "Straight worm Neko 与 Small Rubber Jig 可快速打点，也能承接 10g 以下 plug；适合普通 BFS 嫌弱、标准 bait tackle 又嫌钝的空档。",
    },
    "GA-67MHTC": {
        "player_environment": "淡水 Bass / 草区 Frog、topwater 与重阻 moving bait",
        "player_positioning": "Frog + 操作系 topwater power 枪柄",
        "player_selling_points": "PE Frog 是核心，也能操作 DODGE、G.i、chatter、spinnerbait 和重阻 crank；强而会弯，适合短空间抛投和连续 dog walk。",
    },
    "GA-610MC": {
        "player_environment": "淡水 Bass / 7-14g moving bait 与同级软饵泛用",
        "player_positioning": "中轻量 moving bait 泛用主轴",
        "player_selling_points": "Minnow、Vibration、Crankbait、Spinnerbait 是主菜，No Sinker、Neko 等同重量 worming 可补位；抛投手感和准确落点是关键价值。",
    },
    "GA-610MHC": {
        "player_environment": "淡水 Bass / 10-20g 中量级 moving / soft bait",
        "player_positioning": "M-H 中量级强泛用",
        "player_selling_points": "比 Joker 更强，适合 Chatterbait、Spinnerbait、Vibration，也能切 Rubber Jig、Texas、Swimbait；casting 和 hookset 两端都更有底气。",
    },
    "GA-70MGC": {
        "player_environment": "淡水 Bass / 广域 hard plug 巻物搜索",
        "player_positioning": "glass composite hard-plug 专项",
        "player_selling_points": "Crankbait、Vibration、Spinnerbait、Buzzbait、Chatterbait 和 crawler bait 都能稳定巻き；中型 Big Bait / Swimbait 是补充项，玻纤复合的水阻反馈和咬口包容度更突出。",
    },
    "GA-72HC": {
        "player_environment": "淡水 Bass / cover worming 与 open water heavy versatile",
        "player_positioning": "heavy bottom + 重阻 moving 泛用",
        "player_selling_points": "Rubber Jig、Texas、Heavy Carolina 和高比重 worming 可远投控底，也能切 Spinnerbait、Chatterbait、Deep Crank；不是单纯大饵竿。",
    },
    "GA-74XHC": {
        "player_environment": "淡水 Bass / 2oz big bait 与 cover shift",
        "player_positioning": "2oz 大饵精细操作兼 cover",
        "player_selling_points": "Big Bait、Big Plug、Swimbait、S-shaped bait 和 crawler bait 是主线，同时可快速切 Rubber Jig / Texas 进 cover；重饵中仍强调准确抛投和 slack 操作。",
    },
    "GA-75XXHC": {
        "player_environment": "淡水 Bass / 5oz+ big bait 长时间强攻",
        "player_positioning": "超大饵专用强力长竿",
        "player_selling_points": "面向 5oz+ Big Bait / Swimbait，也能处理 dead-slow big bait 和 Swimming Jig；重点是长时间大饵抛投不累、负载下仍保留操控。",
    },
    "GA-61ULS-ST": {
        "player_environment": "淡水 Bass / 高压场喰わせ light rig",
        "player_positioning": "极细腻喰わせ直柄",
        "player_selling_points": "Down Shot、Neko、Jighead、No Sinker 这类轻量钓组更容易做出细微动作；solid tip 和柔软握持感适合骗咬和短促吃口判断。",
    },
    "GA-63LS": {
        "player_environment": "淡水 Bass / light game 探索、少竿出行",
        "player_positioning": "light rig / light plug 守备型直柄",
        "player_selling_points": "Small Rubber Jig、Neko、Jighead Wacky、Jighead 到 7g 级 Shad、Minnow、Topwater 都能覆盖，适合陌生水域或少带竿时找鱼。",
    },
    "GA-65MS": {
        "player_environment": "淡水 Bass / PE 表层 worming、light power finesse",
        "player_positioning": "PE light power finesse / light plugging",
        "player_selling_points": "PE0.8-1.2 表层 worming、Minnow、Shad、Vibration、小 Topwater 和軽吊るし都能做；结节通过和远投后的控线是主要价值。",
    },
    "GA-67L+S": {
        "player_environment": "淡水 Bass / 岸钓单直柄泛用、近远兼顾",
        "player_positioning": "finesse rig / plug 万能副手",
        "player_selling_points": "一支直柄覆盖 Neko、Down Shot、Jighead 到 Shad、Minnow、小 Topwater；FC4-5lb 偏操作，PE0.8 偏远投，适合岸钓少竿配置。",
    },
    "GA-611MLS-ST": {
        "player_environment": "淡水 Bass / 大水面远投 light plug / worming",
        "player_positioning": "长距离直柄操作搜索",
        "player_selling_points": "6'11 solid tip 让 Shad、Minnow、Metal Vibration 和 worming 在远投后仍能被细腻操控；适合大水面、河川和远端反应咬口。",
    },
    "GA-70HS-ST": {
        "player_environment": "淡水 Bass / 重 cover 吊るし power finesse",
        "player_positioning": "PE 吊るし专用 monster spin",
        "player_selling_points": "PE1.5-2.0 + 粗 leader 操作 Egu-Dama Type Level，在枝、沉木、水下 cover 中挂线诱鱼并强制拔鱼；这是明确的吊るし power finesse。",
    },
    "GA-74ML+S": {
        "player_environment": "淡水 Bass / 岸投超远投 light rig / plug",
        "player_positioning": "长尺 shore spin 远投泛用",
        "player_selling_points": "No Sinker、Neko、JHW、mid-strolling 到 Shad、Minnow、Vibration 都以远投后可控为前提；line mending、trace/range control 和小 plug 反馈是核心。",
    },
}


def n(value):
    return str(value or "").strip()


def headers(ws):
    return {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}


def has_forbidden_environment(value):
    forbidden = ["海水", "港湾", "船钓", "SLSJ", "Eging", "Rockfish", "Ajing", "Mebaring"]
    return any(term.lower() in n(value).lower() for term in forbidden)


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    col = headers(ws)
    missing_cols = [field for field in ["SKU", "id", "recommended_rig_pairing", *TARGET_FIELDS] if field not in col]
    if missing_cols:
        raise RuntimeError(f"missing columns: {missing_cols}")

    changed = []
    missing_mapping = []
    protected_rows = []
    before_uniques = {field: set() for field in TARGET_FIELDS}

    for row_idx in range(2, ws.max_row + 1):
        sku = n(ws.cell(row_idx, col["SKU"]).value)
        row_id = n(ws.cell(row_idx, col["id"]).value)
        for field in TARGET_FIELDS:
            before_uniques[field].add(n(ws.cell(row_idx, col[field]).value))

        next_values = PLAYER_FIELDS_BY_SKU.get(sku)
        if not next_values:
            missing_mapping.append({"row": row_idx, "id": row_id, "sku": sku})
            continue

        for field in TARGET_FIELDS:
            cell = ws.cell(row_idx, col[field])
            old = n(cell.value)
            new = next_values[field]
            if old != new:
                changed.append({"row": row_idx, "id": row_id, "sku": sku, "field": field, "old": old, "new": new})
                cell.value = new

    if missing_mapping:
        raise RuntimeError(f"missing player mappings: {missing_mapping}")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["rod_detail"]
    col = headers(ws)
    coverage = {}
    unique_counts = {}
    forbidden_environment = []
    conflict_rows = []

    for field in TARGET_FIELDS:
        vals = [n(ws.cell(row_idx, col[field]).value) for row_idx in range(2, ws.max_row + 1)]
        coverage[field] = f"{sum(bool(v) for v in vals)}/{ws.max_row - 1}"
        unique_counts[field] = len(set(vals))

    for row_idx in range(2, ws.max_row + 1):
        row_id = n(ws.cell(row_idx, col["id"]).value)
        sku = n(ws.cell(row_idx, col["SKU"]).value)
        env = n(ws.cell(row_idx, col["player_environment"]).value)
        pos = n(ws.cell(row_idx, col["player_positioning"]).value)
        selling = n(ws.cell(row_idx, col["player_selling_points"]).value)
        rig = n(ws.cell(row_idx, col["recommended_rig_pairing"]).value)
        if has_forbidden_environment(env):
            forbidden_environment.append({"id": row_id, "sku": sku, "player_environment": env})
        if "Big Bait" in rig and "大饵" not in pos and "big bait" not in selling.lower() and "大饵" not in selling:
            conflict_rows.append({"id": row_id, "sku": sku, "reason": "big_bait_pairing_without_player_big_bait_context"})
        if "Frog" in rig and "Frog" not in pos and "frog" not in selling.lower() and "草区" not in env:
            conflict_rows.append({"id": row_id, "sku": sku, "reason": "frog_pairing_without_player_frog_context"})
        if any(term in rig for term in ["Mid Strolling", "Hover Strolling"]) and not any(
            term in pos + selling for term in ["mid-strolling", "Mid", "中层", "水面微抖", "hobast", "Hover"]
        ):
            conflict_rows.append({"id": row_id, "sku": sku, "reason": "strolling_pairing_without_player_context"})

    report = {
        "xlsx_file": str(XLSX_PATH),
        "fields": TARGET_FIELDS,
        "managed_rows": ws.max_row - 1,
        "managed_target_cells": (ws.max_row - 1) * len(TARGET_FIELDS),
        "changed_cells": len(changed),
        "changed_rows": len({item["id"] for item in changed}),
        "protected_rows": protected_rows,
        "coverage": coverage,
        "unique_counts": unique_counts,
        "before_unique_counts": {field: len(values) for field, values in before_uniques.items()},
        "forbidden_environment": forbidden_environment,
        "conflict_rows": conflict_rows,
        "source_policy": {
            "primary": "whitelist/player usage context and long-term angling vocabulary where available",
            "secondary": "confirmed recommended_rig_pairing",
            "boundary": "official Description/spec/SKU/power/action/type prevent player-field overreach",
        },
        "changes": changed,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if any(value != f"{ws.max_row - 1}/{ws.max_row - 1}" for value in coverage.values()):
        raise SystemExit(1)
    if forbidden_environment or conflict_rows:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
