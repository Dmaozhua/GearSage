from copy import copy
from html.parser import HTMLParser
from pathlib import Path
import re
from urllib.request import Request, urlopen

from openpyxl import load_workbook


XLSX_PATH = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw/daiwa_rod_import.xlsx")
OFFICIAL_URL = "https://www.daiwa.com/tw/product/e0yd31e"
TARGET_ROD_ID = "DR1019"
TARGET_FIELDS = ["Description", "recommended_rig_pairing", "guide_use_hint"]


RIG_AND_HINT_BY_CODE = {
    "C63MH-FR": (
        "Frog / Jointed Swimbait / Topwater Plug / Jerkbait / Spinnerbait / Buzzbait / Chatterbait / Swim Jig / Swimbait / No Sinker / Rubber Jig / Texas Rig / Leaderless Down Shot / Free Rig",
        "Frog/重障礙：Frog、Jointed Swimbait 與 Topwater 打點為主，兼顧 Spinnerbait、Chatterbait 和軟餌，PE 粗線低彈道拋投與障礙邊控魚更穩。",
    ),
    "C64L-BF": (
        "Neko Rig / No Sinker / Down Shot / Small Rubber Jig / Small Plug",
        "BF 精細輕餌：Neko、No Sinker、Down Shot 和小型軟膠鉤低彈道拋投更順，輕餌控線與細微咬口判斷更直接。",
    ),
    "C65MH-FR": (
        "Frog / Rubber Jig / Texas Rig / Leaderless Down Shot / Free Rig / No Sinker / Spinnerbait / Swim Jig / Jointed Swimbait / Topwater Plug / Soft Plastic / Swimbait",
        "Frog/重障礙：Frog 為主，兼顧 Rubber Jig、Texas、Free Rig 和 Swim Jig，草洞、覆蓋物邊的落點控制與拉離障礙更穩。",
    ),
    "C66ML-LM": (
        "Shallow Crankbait / Medium Crankbait / Small Crankbait / Flat Side Crankbait / Shad / Small Topwater Plug",
        "LM 卷阻硬餌：Crankbait、Flat Side、Shad 和小型 Topwater 低彈道拋投更容易，平收時鬆線與碰障回饋更穩。",
    ),
    "C66ML+": (
        "Topwater Plug / Jerkbait / Spinnerbait / Buzzbait / Crankbait / Vibration / Neko Rig / Small Rubber Jig / No Sinker / Down Shot",
        "硬餌泛用：Topwater、Jerkbait、Spinnerbait、Crankbait 等操作型與平收硬餌為主，也能兼顧 Neko、No Sinker 等輕量軟餌。",
    ),
    "C66M-ST": (
        "Down Shot / Neko Rig / Small Rubber Jig / Free Rig / Texas Rig / Leaderless Down Shot / No Sinker / Metal Vibration",
        "高感底操：Mega Top R 實心竿尖放大 Down Shot、Neko、Free Rig 和 Texas 的觸底變化，也能兼顧 Metal Vibration 的障礙反應。",
    ),
    "C67L/ML+-BF": (
        "Neko Rig / Down Shot / Small Rubber Jig / No Sinker / Free Rig / Leaderless Down Shot",
        "大湖 BFS 精細：Neko、Down Shot、小型軟膠鉤與 No Sinker 為主，水草、淺層結構和障礙邊的控線與脫草更穩。",
    ),
    "SC C68H-ST-SB": (
        "Swimbait / Big Bait / Alabama Rig / Rubber Jig / Texas Rig / Leaderless Down Shot / No Sinker / Spinnerbait",
        "SB 大餌：Swimbait、Big Bait 和 Alabama Rig 拋投與回收負荷更穩，實心竿尖也能兼顧 Rubber Jig、Texas 等軟餌打點。",
    ),
    "SC C69M+-ST": (
        "Neko Rig / No Sinker / Down Shot / Texas Rig / Leaderless Down Shot / Free Rig / Rubber Jig / Spinnerbait / Vibration / Crankbait / Soft Plastic",
        "SC 全能底操：Neko、No Sinker、Down Shot、Texas 和 Free Rig 為主，實心竿尖也能兼顧 Spinnerbait、Vibration、Crankbait 的鬆線回收與觸感。",
    ),
    "SC C69MH": (
        "Rubber Jig / Texas Rig / Leaderless Down Shot / No Sinker / Spinnerbait / Soft Plastic / Swimbait / Big Bait / Chatterbait / Frog",
        "SC 重型全能：Rubber Jig、Texas、No Sinker 與 Spinnerbait 為主，可擴展到 Swimbait、Big Bait、Chatterbait 和 Frog。",
    ),
    "C610M": (
        "Spinnerbait / Buzzbait / Chatterbait / Vibration / Swimbait / Prop Bait / Topwater Plug / Jerkbait / Craw / Neko Rig / No Sinker / Down Shot / Texas Rig / Leaderless Down Shot / Free Rig / Rubber Jig",
        "Wire bait 泛用：Spinnerbait、Buzzbait、Chatterbait、Vibration 等移動餌為主，兼顧 Neko、No Sinker、Texas 和 Free Rig 的軟餌切換。",
    ),
    "C610MH": (
        "Spinnerbait / Buzzbait / Chatterbait / Swim Jig / Vibration / Swimbait / Big Bait / Soft Plastic / Topwater Plug / Frog / Rubber Jig / Texas Rig / Leaderless Down Shot / No Sinker",
        "強力移動餌：Spinnerbait、Buzzbait、Chatterbait、Swim Jig 與 Swimbait 為主，兼顧 Big Bait、Frog 和中重型軟餌。",
    ),
    "C70H": (
        "Heavy Carolina Rig / Texas Rig / Rubber Jig / Leaderless Down Shot / Free Rig / No Sinker / Swim Jig / Spinnerbait / Swimbait / Big Bait / Soft Plastic",
        "重底操長竿：Heavy Carolina、Texas、Rubber Jig、Free Rig 為主，遠投探底、水草邊脫草和中魚後控魚更有餘量。",
    ),
    "C73ML+": (
        "Shallow Crankbait / Medium Crankbait / Deep Crankbait / Shad / Vibration / Spinnerbait / Buzzbait / Chatterbait",
        "遠投卷阻硬餌：Crankbait、Shad、Vibration 和 Wire bait 為主，長竿遠投後的鬆線回收與泳層控制更穩。",
    ),
    "C74H": (
        "Rubber Jig / Texas Rig / Leaderless Down Shot / Free Rig / Heavy Carolina Rig / No Sinker / Swim Jig / Spinnerbait / Swimbait / Big Bait / Soft Plastic",
        "大湖強力泛用：Rubber Jig、Texas、Heavy Carolina 和 Free Rig 為主，長竿可做遠距探底、縱向控線和水草邊強制控魚。",
    ),
    "C74MH": (
        "Deep Crankbait / Spinnerbait / Buzzbait / Chatterbait / Swim Jig / Vibration / Big Crankbait / Swimbait / Big Bait / Soft Plastic / Topwater Plug / Frog / Rubber Jig / Texas Rig / Leaderless Down Shot / No Sinker",
        "重型卷阻硬餌：Deep Crankbait、Spinnerbait、Chatterbait、Swim Jig 與大型硬餌為主，也能承接重型軟餌和 Frog。",
    ),
    "S61L": (
        "Small Rubber Jig / Neko Rig / Down Shot / Jighead Rig / Wacky Jighead Rig / No Sinker / Small Plug",
        "終極精細紡車：小型軟膠鉤、Neko、Down Shot、Jighead 和 No Sinker 為主，細線輕餌精準拋投與微弱咬口判斷更清楚。",
    ),
    "S64UL": (
        "Hover Strolling / Mid Strolling / Bottom Strolling / Down Shot / Neko Rig / No Sinker / Small Rubber Jig / Jighead Rig / Wacky Jighead Rig / Small Plug",
        "Strolling 精細：Hover、Mid、Bottom Strolling 為主，搭配 Down Shot、Neko、No Sinker 時水層節奏與鬆線變化更好掌握。",
    ),
    "SC S64L-ST": (
        "Small Rubber Jig / No Sinker / Down Shot / Neko Rig / Jighead Rig / Wacky Jighead Rig / Stickbait / I-shaped Plug / Minnow / Prop Bait / Bug Lure / Small Metal Vibration",
        "SC 精細多用途：Small Rubber Jig、No Sinker、Down Shot、Neko 等輕釣組為主，也能兼顧 I-shaped、Minnow、Bug 和小型金屬振動。",
    ),
    "S64L/ML+": (
        "Neko Rig / Down Shot / No Sinker / Jighead Rig / Wacky Jighead Rig / Small Rubber Jig / I-shaped Plug / Prop Bait / Bug Lure / Small Frog",
        "進取型精細：Neko、Down Shot、No Sinker、Jighead 和小型軟膠鉤為主，水草邊、灌木叢和反應式底操更容易控線。",
    ),
    "S68UL-ST": (
        "Neko Rig / Small Rubber Jig / Wacky Jighead Rig / Jighead Rig / No Sinker / Down Shot / Light Carolina Rig / Split Shot / Mid Strolling / Bottom Strolling / Small Plug",
        "UL 實心竿尖精細：Neko、Small Rubber Jig、Jighead、No Sinker 和輕量 Carolina 為主，小餌遠投、鬆線與中底層細操作更穩。",
    ),
    "S69ML-ST": (
        "Neko Rig / Down Shot / No Sinker / Jighead Rig / Wacky Jighead Rig / Light Carolina Rig / Split Shot / Small Rubber Jig / Mid Strolling / Shad / Prop Bait / Bug Lure / Metal Vibration",
        "長竿高感精細：Neko、Down Shot、Jighead、Light Carolina 和 Mid Strolling 為主，兼顧 Shad、Bug、Metal Vibration 的節奏變化。",
    ),
}


class TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_tr = False
        self.in_td = False
        self.current_row = []
        self.current_text = []
        self.rows = []

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self.in_tr = True
            self.current_row = []
        elif self.in_tr and tag == "td":
            self.in_td = True
            self.current_text = []
        elif self.in_td and tag == "br":
            self.current_text.append(" ")

    def handle_endtag(self, tag):
        if tag == "td" and self.in_td:
            self.current_row.append(clean_text("".join(self.current_text)))
            self.in_td = False
        elif tag == "tr" and self.in_tr:
            if self.current_row:
                self.rows.append(self.current_row)
            self.in_tr = False

    def handle_data(self, data):
        if self.in_td:
            self.current_text.append(data)


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_code(value):
    text = clean_text(value).upper()
    text = re.sub(r"^BLX\s+", "", text)
    text = text.replace("•", "-").replace("・", "-").replace("－", "-")
    text = text.replace("和", "-").replace("、", "-")
    text = re.sub(r"\s+", "", text)
    text = text.replace("SC", "SC")
    return text


def code_from_title(title):
    known = sorted(RIG_AND_HINT_BY_CODE, key=len, reverse=True)
    norm_title = normalize_code(title)
    for code in known:
        if norm_title.startswith(normalize_code(code)):
            return code
    return ""


def clean_official_description(code, title, desc):
    text = clean_text(f"{title} {desc}")
    if code == "C64L-BF" and "ㄆㄨ" in text:
        text = re.sub(r"竿尖採用「ㄆ.*?一旦中魚", "竿尖採用「へ」字形錐度設計。 一旦中魚", text)
    text = text.replace("ワーミング", "軟餌")
    return text


def fetch_official_descriptions():
    req = Request(OFFICIAL_URL, headers={"User-Agent": "Mozilla/5.0"})
    html = urlopen(req, timeout=20).read().decode("utf-8", "ignore")
    parser = TableParser()
    parser.feed(html)
    descriptions = {}
    for row in parser.rows:
        if len(row) < 2:
            continue
        code = code_from_title(row[0])
        if not code or code in descriptions:
            continue
        descriptions[code] = clean_official_description(code, row[0], row[1])
    missing = sorted(set(RIG_AND_HINT_BY_CODE) - set(descriptions))
    if missing:
        raise RuntimeError(f"missing official descriptions: {missing}")
    return descriptions


def snapshot_without_targets(ws, target_cols):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            header: ws.cell(row, col).value
            for col, header in enumerate(headers, start=1)
            if col not in target_cols
        })
    return rows


def copy_fills(src_row, dst_row):
    for src, dst in zip(src_row, dst_row):
        dst.fill = copy(src.fill)


def main():
    descriptions = fetch_official_descriptions()
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    ws = wb["rod_detail"]
    rod_headers = [cell.value for cell in rod_ws[1]]
    headers = [cell.value for cell in ws[1]]
    rod_col = {name: i + 1 for i, name in enumerate(rod_headers)}
    col = {name: i + 1 for i, name in enumerate(headers)}
    for field in ["id", "rod_id", "SKU", *TARGET_FIELDS]:
        if field not in col:
            raise RuntimeError(f"missing rod_detail column: {field}")

    target_cols = {col[field] for field in TARGET_FIELDS}
    before_detail = snapshot_without_targets(ws, target_cols)
    before_rod = [[rod_ws.cell(row, c).value for c in range(1, rod_ws.max_column + 1)] for row in range(1, rod_ws.max_row + 1)]

    changed = []
    matched = []
    for row in range(2, ws.max_row + 1):
        if clean_text(ws.cell(row, col["rod_id"]).value) != TARGET_ROD_ID:
            continue
        sku = clean_text(ws.cell(row, col["SKU"]).value)
        code = ""
        norm_sku = normalize_code(sku)
        for item_code in sorted(RIG_AND_HINT_BY_CODE, key=len, reverse=True):
            if normalize_code(item_code) in norm_sku:
                code = item_code
                break
        if not code:
            raise RuntimeError(f"unmatched DR1019 SKU: {sku}")
        matched.append(code)
        new_values = {
            "Description": descriptions[code],
            "recommended_rig_pairing": RIG_AND_HINT_BY_CODE[code][0],
            "guide_use_hint": RIG_AND_HINT_BY_CODE[code][1],
        }
        row_changes = {}
        for field, new_value in new_values.items():
            cell = ws.cell(row, col[field])
            old = clean_text(cell.value)
            if old != new_value:
                cell.value = new_value
                row_changes[field] = {"old": old, "new": new_value}
        if row_changes:
            changed.append({"id": clean_text(ws.cell(row, col["id"]).value), "sku": sku, "code": code, "fields": sorted(row_changes)})

    if sorted(matched) != sorted(RIG_AND_HINT_BY_CODE):
        missing = sorted(set(RIG_AND_HINT_BY_CODE) - set(matched))
        extra = sorted(set(matched) - set(RIG_AND_HINT_BY_CODE))
        raise RuntimeError(f"DR1019 matched set mismatch: missing={missing}, extra={extra}")

    after_detail = snapshot_without_targets(ws, target_cols)
    if before_detail != after_detail:
        raise RuntimeError("unexpected non-target field changes in rod_detail")
    after_rod = [[rod_ws.cell(row, c).value for c in range(1, rod_ws.max_column + 1)] for row in range(1, rod_ws.max_row + 1)]
    if before_rod != after_rod:
        raise RuntimeError("unexpected changes in rod sheet")

    wb.save(XLSX_PATH)
    print({
        "file": str(XLSX_PATH),
        "official_url": OFFICIAL_URL,
        "target_rod_id": TARGET_ROD_ID,
        "matched_rows": len(matched),
        "changed_rows": len(changed),
        "changed_cells": sum(len(item["fields"]) for item in changed),
        "changed": changed,
    })


if __name__ == "__main__":
    main()
