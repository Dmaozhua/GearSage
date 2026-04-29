import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
XLSX_PATH = ROOT / "GearSage-client/pkgGear/data_raw/raid_rod_import.xlsx"
REPORT_PATH = ROOT / "GearSage-client/pkgGear/data_raw/raid_rod_recommended_rig_pairing_check.json"

FIELD = "recommended_rig_pairing"
OVER_GENERIC = {
    "General Lure",
    "Light Rig",
    "Hardbait",
    "Hard Bait",
    "Soft Bait",
    "General Hardbait",
    "General Soft Bait",
}

DESC_REQUIREMENTS = [
    (r"テキサス|texas", "Texas Rig"),
    (r"フリーリグ|free rig", "Free Rig"),
    (r"ネコリグ|neko", "Neko Rig"),
    (r"ダウンショット|drop shot|down shot", "Down Shot"),
    (r"ノーシンカー|no sinker|weightless", "No Sinker"),
    (r"ラバージグ|rubber jig", "Rubber Jig"),
    (r"スモラバ|small rubber jig", "Small Rubber Jig"),
    (r"ジグヘッドワッキー|jighead wacky", "Jighead Wacky"),
    (r"ミドスト|mid strolling", "Mid Strolling Jighead"),
    (r"ホバスト|hover", "Hover Strolling"),
    (r"シャッド|shad", "Shad"),
    (r"ミノー|minnow", "Minnow"),
    (r"バイブ|vibration", "Vibration"),
    (r"クランク|crank", "Crankbait"),
    (r"スピナ|spinner", "Spinnerbait"),
    (r"チャター|chatter", "Chatterbait"),
    (r"フロッグ|frog", "Frog"),
    (r"ビッグベイト|big bait", "Big Bait"),
    (r"スイムベイト|swimbait|swim bait", "Swimbait"),
    (r"バマスト|alabama", "Alabama Rig"),
    (r"トップウォーター|topwater", "Topwater Plug"),
    (r"ペンシル|pencil", "Pencil Bait"),
    (r"ポッパー|popper", "Popper"),
    (r"虫系|虫", "Bug Lure"),
]

SOFT = {
    "Light Texas Rig",
    "Heavy Texas Rig",
    "Texas Rig",
    "Free Rig",
    "Neko Rig",
    "Heavy Neko Rig",
    "Down Shot",
    "Leaderless Down Shot",
    "Heavy Down Shot",
    "No Sinker",
    "Backslide No Sinker",
    "Small Rubber Jig",
    "Rubber Jig",
    "Finesse Cover Jig",
    "Jighead Rig",
    "Jighead Wacky",
    "Mid Strolling Jighead",
    "Power Mid Strolling Jighead",
    "Hover Strolling",
    "Surface Twitch Rig",
    "Hanging Rubber Jig",
    "Power Finesse Jig",
    "Cover Suspended Rig",
    "Surface Worming",
    "Shad Tail Worm",
    "Sinker Rig",
}

HARD = {
    "Small Plug",
    "Small Topwater Plug",
    "Topwater Plug",
    "Crawler Bait",
    "Big Plug",
    "S-shaped Bait",
    "Jerkbait",
    "Minnow",
    "Shad",
    "Crankbait",
    "Deep Crankbait",
    "Vibration",
    "Metal Vibration",
    "Spinnerbait",
    "Buzzbait",
    "Chatterbait",
    "Swim Jig",
    "Swimming Jig",
    "Swimbait",
    "Big Bait",
    "Frog",
    "Pencil Bait",
    "Popper",
    "I-shaped Plug",
    "Bug Lure",
    "Alabama Rig",
}


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def split_pairing(value):
    return [n(part) for part in n(value).split("/") if n(part)]


def has_equivalent(label, items):
    equivalents = {
        "Texas Rig": {"Texas Rig", "Light Texas Rig", "Heavy Texas Rig"},
        "Down Shot": {"Down Shot", "Leaderless Down Shot", "Heavy Down Shot"},
        "Neko Rig": {"Neko Rig", "Heavy Neko Rig"},
        "Shad": {"Shad", "Shad Tail Worm"},
        "Vibration": {"Vibration", "Metal Vibration"},
        "Rubber Jig": {"Rubber Jig", "Small Rubber Jig", "Finesse Cover Jig", "Hanging Rubber Jig", "Power Finesse Jig"},
        "Topwater Plug": {"Topwater Plug", "Small Topwater Plug", "Crawler Bait"},
        "Jighead Rig": {"Jighead Rig", "Jighead Wacky", "Mid Strolling Jighead", "Power Mid Strolling Jighead"},
        "No Sinker": {"No Sinker", "Backslide No Sinker"},
        "Big Bait": {"Big Bait", "Big Plug"},
    }
    return bool(set(items) & equivalents.get(label, {label}))


def row_dict(ws, headers, row_idx):
    return {field: ws.cell(row_idx, col_idx).value for field, col_idx in headers.items()}


def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["rod_detail"]
    headers = {n(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
    required = ["id", "SKU", "Description", "guide_use_hint", FIELD]
    missing_headers = [field for field in required if field not in headers]
    if missing_headers:
        raise RuntimeError(f"missing headers: {missing_headers}")

    rows = []
    over_generic = []
    desc_misses = []
    guide_conflicts = []
    coverage = 0
    source_families = Counter()

    for row_idx in range(2, ws.max_row + 1):
        row = row_dict(ws, headers, row_idx)
        pairing = n(row.get(FIELD))
        items = split_pairing(pairing)
        desc = n(row.get("Description"))
        guide = n(row.get("guide_use_hint"))
        if pairing:
            coverage += 1
        if any(term in items or term == pairing for term in OVER_GENERIC):
            over_generic.append({"id": n(row.get("id")), "sku": n(row.get("SKU")), "value": pairing})

        for pattern, label in DESC_REQUIREMENTS:
            if re.search(pattern, desc, re.I) and not has_equivalent(label, items):
                desc_misses.append({"id": n(row.get("id")), "sku": n(row.get("SKU")), "missing": label})

        first = items[0] if items else ""
        has_soft = any(item in SOFT for item in items)
        has_hard = any(item in HARD for item in items)
        if first in SOFT:
            source_families["soft_primary"] += 1
            if re.search(r"硬餌搜索|硬饵搜索|巻物", guide) and not re.search(r"軟餌|软饵|切|兼容|泛用|cover|障礙|障碍", guide):
                guide_conflicts.append({"id": n(row.get("id")), "sku": n(row.get("SKU")), "reason": "soft_primary_hard_hint"})
        elif first in HARD:
            source_families["hard_primary"] += 1
            if re.search(r"軟餌底操|软饵底操", guide):
                guide_conflicts.append({"id": n(row.get("id")), "sku": n(row.get("SKU")), "reason": "hard_primary_soft_hint"})
        else:
            source_families["other_primary"] += 1

        if has_soft and has_hard and not re.search(r"切|兼容|泛用|両立|兼顧|切り替え|切換|切换|cover|障礙|障碍|moving bait|worming|plug", guide, re.I):
            guide_conflicts.append({"id": n(row.get("id")), "sku": n(row.get("SKU")), "reason": "mixed_pairing_needs_switch_logic"})

        rows.append({"id": n(row.get("id")), "sku": n(row.get("SKU")), FIELD: pairing})

    report = {
        "xlsx_file": str(XLSX_PATH),
        "field": FIELD,
        "rows": ws.max_row - 1,
        "coverage": f"{coverage}/{ws.max_row - 1}",
        "over_generic_count": len(over_generic),
        "description_miss_count": len(desc_misses),
        "guide_conflict_count": len(guide_conflicts),
        "primary_family_counts": dict(source_families),
        "over_generic": over_generic,
        "description_misses": desc_misses,
        "guide_conflicts": guide_conflicts,
        "row_values": rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if over_generic or desc_misses or guide_conflicts or coverage != ws.max_row - 1:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
