#!/usr/bin/env python3
"""Fill Olympic rod_detail.product_technical from official Olympic pages only."""

from __future__ import annotations

import json
import re
import subprocess
from copy import deepcopy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "olympic_rod_normalized.json"
XLSX_PATH = DATA_DIR / "olympic_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "olympic_rod_product_technical_stage5_report.json"
SHADE_SCRIPT = ROOT / "scripts/shade_olympic_rod_detail_groups.py"

FIELD = "product_technical"
DELIMITER = " / "


def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def merge_terms(terms) -> str:
    seen = set()
    result = []
    for term in terms:
        value = norm(term)
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return DELIMITER.join(result)


def infer_product_technical(item: dict, variant: dict) -> tuple[str, str]:
    model = norm(item.get("model"))
    fields = variant.get("fields") or {}
    sku = norm(fields.get("SKU") or variant.get("sku"))
    rod_type = norm(fields.get("TYPE"))
    description = norm(fields.get("Description") or variant.get("description"))
    terms = []
    evidence = ""

    if model == "20 VIGORE":
        terms.extend(["トレカ®T1100G", "ナノアロイ®", "G-MAPS製法", "クワトログラファイトクロスXX"])
        evidence = "official product Description, BLANK MATERIAL & TECHNOLOGY, FEATURES, material icons"
        if sku == "20GVIGC-77XH":
            terms.extend(["スーパークワトログラファイトクロス", "ステンレスフレームSiC-Sリング", "オールダブルフットKガイド", "グリップ脱着式", "ECS17リールシート"])
        elif sku == "20GVIGC-71H":
            terms.extend(["チタンフレームトルザイトリングガイド", "KWガイド", "シングルフットガイド", "ECS16リールシート"])
        elif sku in {"20GVIGC-75M", "20GVIGC-76MH"}:
            terms.extend(["チタンフレームトルザイトリングガイド", "LRVガイド", "KWガイド", "シングルフットガイド", "グリップ脱着式"])
            terms.append("ECS16リールシート" if sku == "20GVIGC-75M" else "ECS17リールシート")
        elif sku == "GVIGS-6102ML":
            terms.extend(["チタンフレームトルザイトリングガイド", "スピゴットフェルール（印籠継）", "オリジナルカーボンリールシート OP-02"])
        elif sku in {"20GVIGS-610ML", "20GVIGS-742M"}:
            terms.extend(["チタンフレームトルザイトリングガイド", "VSSリールシート"])
            if sku == "20GVIGS-742M":
                terms.append("スピゴットフェルール（印籠継）")

    elif model == "21 VELOCE UX":
        terms.extend(["グラファイトクロスLV", "ステンレスフレームSiC-SリングKガイド"])
        evidence = "official BLANK MATERIAL & TECHNOLOGY, FEATURES"
        if sku == "21GVELUC-74X":
            terms.extend(["オールダブルフット仕様", "グリップ脱着式", "TCSリールシート"])
        elif rod_type == "C":
            terms.append("ECSリールシート")
        elif rod_type == "S":
            terms.append("VSSリールシート")
        if sku == "21GVELUS-742M":
            terms.append("フェルール（逆並継）")

    elif model == "18 Super Bellezza":
        terms.extend(
            [
                "トレカ®T1100G",
                "ナノアロイ®",
                "スーパークワトログラファイトクロスLV",
                "G-MAPS製法",
                "チタンフレームトルザイトリングガイド",
                "Fuji TORZITE",
                "オリジナルシート",
                "スピゴットフェルール（印籠継）",
            ]
        )
        evidence = "official product Description, FEATURES, MATERIAL & TECHNOLOGY"

    elif model == "26 Bellezza PROTOTYPE":
        terms.append("トレカ®T1100G")
        if sku not in {"26GBELPS-572XUL-S", "26GBELPS-582SUL-T"}:
            terms.append("トレカ®M40X")
        terms.extend(
            [
                "ナノアロイ®",
                "スーパークワトログラファイトクロスLV",
                "G-MAPS製法",
                "O.S.S",
                "チタンフレームトルザイトリングガイド",
                "チタンフレームSiCトップガイド",
                "スピゴットフェルール（印籠継）",
                "オリジナルシート",
                "エステルライン対応ガイドセッティング",
            ]
        )
        if "ソリッドティップ" in description:
            terms.append("ソリッドティップ")
        evidence = "official product Description, BLANK MATERIAL, FEATURES, material icons"

    elif model == "24 BELLEZZA UX":
        terms.extend(
            [
                "グラファイトクロスLV",
                "ステンレスフレームSiC-Sリングガイド",
                "スリップオーバーフェルール（逆並継）",
                "オリジナルカーボンリールシート OP-01",
            ]
        )
        if sku == "25GBELUS-612SUL-S":
            terms.append("ソリッドティップ")
        evidence = "official product Description, BLANK MATERIAL, FEATURES"

    return merge_terms(terms), evidence


def strip_product_technical(data):
    cleaned = deepcopy(data)
    for item in cleaned:
        item.pop(FIELD, None)
        for variant in item.get("variants", []):
            variant.pop(FIELD, None)
            fields = variant.get("fields")
            if isinstance(fields, dict):
                fields.pop(FIELD, None)
    return cleaned


def workbook_values(wb, ignore_field):
    result = {}
    for sheet_name in ("rod", "rod_detail"):
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        ignored = {idx + 1 for idx, header in enumerate(headers) if header == ignore_field}
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(tuple(value for idx, value in enumerate(row, start=1) if idx not in ignored))
        result[sheet_name] = rows
    return result


def remove_master_field(ws):
    headers = [cell.value for cell in ws[1]]
    removed = False
    for col in reversed([idx + 1 for idx, header in enumerate(headers) if header == FIELD]):
        ws.delete_cols(col)
        removed = True
    return removed


def ensure_detail_field(ws):
    headers = [cell.value for cell in ws[1]]
    matches = [idx + 1 for idx, header in enumerate(headers) if header == FIELD]
    if len(matches) > 1:
        raise RuntimeError(f"duplicate {FIELD} columns in rod_detail: {matches}")
    if matches:
        return matches[0], False
    if "Description" not in headers:
        raise RuntimeError("rod_detail missing Description column")
    insert_at = headers.index("Description") + 2
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = FIELD
    return insert_at, True


def validate_values(rows):
    errors = []
    for row in rows:
        value = row["value"]
        if not value:
            errors.append({"sku": row["sku"], "type": "blank"})
        if "/" in value and DELIMITER not in value:
            errors.append({"sku": row["sku"], "type": "bad_delimiter", "value": value})
        for banned in ("白名单", "玩家", "钓组", "饵型", "来源"):
            if banned in value:
                errors.append({"sku": row["sku"], "type": "forbidden_fragment", "fragment": banned})
    return errors


def main():
    normalized = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    normalized_guard = strip_product_technical(normalized)

    rows = []
    value_by_key = {}
    for item_index, item in enumerate(normalized):
        rod_id = f"OR{1000 + item_index}"
        item.pop(FIELD, None)
        for variant in item.get("variants", []):
            fields = variant.setdefault("fields", {})
            sku = norm(fields.get("SKU") or variant.get("sku"))
            value, evidence = infer_product_technical(item, variant)
            variant[FIELD] = value
            fields[FIELD] = value
            value_by_key[(rod_id, sku)] = value
            rows.append({"rod_id": rod_id, "model": item.get("model"), "sku": sku, "value": value, "evidence": evidence})

    if strip_product_technical(normalized) != normalized_guard:
        raise RuntimeError("normalized guard failed: fields other than product_technical changed")

    errors = validate_values(rows)
    if errors:
        raise RuntimeError(json.dumps({"validation_errors": errors}, ensure_ascii=False, indent=2))

    wb = load_workbook(XLSX_PATH)
    before_guard = workbook_values(wb, FIELD)
    rod_removed = remove_master_field(wb["rod"])
    detail_ws = wb["rod_detail"]
    tech_col, inserted = ensure_detail_field(detail_ws)
    headers = [cell.value for cell in detail_ws[1]]
    rod_id_col = headers.index("rod_id") + 1
    sku_col = headers.index("SKU") + 1

    changed_cells = []
    for row_idx in range(2, detail_ws.max_row + 1):
        key = (norm(detail_ws.cell(row=row_idx, column=rod_id_col).value), norm(detail_ws.cell(row=row_idx, column=sku_col).value))
        value = value_by_key.get(key, "")
        cell = detail_ws.cell(row=row_idx, column=tech_col)
        old = norm(cell.value)
        cell.value = value
        if old != value:
            changed_cells.append({"row": row_idx, "rod_id": key[0], "sku": key[1], "old": old, "new": value})

    if workbook_values(wb, FIELD) != before_guard:
        raise RuntimeError("workbook guard failed: fields other than product_technical changed")

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)
    NORMALIZED_PATH.write_text(json.dumps(normalized, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    report = {
        "field": FIELD,
        "scope": "rod_detail SKU level only; rod master must not contain product_technical",
        "source_policy": "official Olympic product pages and official technology/material page links only; whitelist/player/spec inference not used",
        "delimiter": DELIMITER,
        "xlsx": str(XLSX_PATH),
        "normalized": str(NORMALIZED_PATH),
        "variants": len(rows),
        "nonempty": sum(1 for row in rows if row["value"]),
        "rod_product_technical_removed": rod_removed,
        "rod_detail_column_inserted": inserted,
        "changed_xlsx_cells": len(changed_cells),
        "changed_cells": changed_cells,
        "values": rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
