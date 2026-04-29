#!/usr/bin/env python3
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = ROOT / "GearSage-client/pkgGear/data_raw"
XLSX_PATH = DATA_DIR / "nories_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "nories_rod_player_fields_stage3_report.json"

FIELDS = ["player_environment", "player_positioning", "player_selling_points"]


PLAYER_FIELDS = {
    "LTT620PMH": {
        "player_environment": "淡水 Bass / cover 邊與深水結構多餌切換",
        "player_positioning": "Parabolic 多餌泛用 / topwater 到 bottom jig",
        "player_selling_points": "短尺準投和 parabolic 曲線兼顧，大型 topwater、bulky Neko、metal / football jig 與 vibration / chatter 切換時不容易彈口。",
    },
    "LTT630M": {
        "player_environment": "淡水 Bass / cover 邊小硬餌與輕量軟餌切換",
        "player_positioning": "Light cover all-round / 小 crank 與 paddle-tail no sinker 切換",
        "player_selling_points": "小 crank、wake minnow prop topwater 可低彈道送入 cover 邊，也保留 light Texas、no-sinker paddle-tail worm 的補刺餘量。",
    },
    "LTT630MH": {
        "player_environment": "淡水 Bass / cover pitching 與 frog 近距攻障礙",
        "player_positioning": "Cover Texas / Rubber Jig / Frog 專向",
        "player_selling_points": "硬 tip 讓 Texas、rubber jig 不易跳離障礙，side-hand frog 準投和近距強補刺都有足夠支撐。",
    },
    "LTT650M": {
        "player_environment": "淡水 Bass / open water 搜索與 light cover",
        "player_positioning": "Light versatile / 巻物搜索到 bait Neko 切換",
        "player_selling_points": "spinnerbait、crank、big minnow jerkbait 可先找魚，再切換到 heavy down shot、bait Neko 處理杭邊和 light cover。",
    },
    "LTT650MH": {
        "player_environment": "淡水 Bass / 中量巻物與 worming 混合場景",
        "player_positioning": "Mid reeling & worming / 中量軟硬餌泛用",
        "player_selling_points": "3/8oz 以上 spinnerbait、vibration、中型 crank 是主線，Texas 和 heavy down shot 可補底操，軟硬餌切換寬容度高。",
    },
    "LTT660H": {
        "player_environment": "淡水 Bass / vegetation 邊 slow roll 與 power softbait",
        "player_positioning": "Power strategy / spinnerbait 與 heavy rig 雙線",
        "player_selling_points": "slow-roll spinnerbait 可切 weed，6/0 Texas、guarded rubber jig 和 Dairakka spoon 需要的整體 power 也足夠。",
    },
    "LTT680MH": {
        "player_environment": "淡水 Bass / 長線 moving bait 與遠投 Texas",
        "player_positioning": "Moving-to-worming versatile / chatterbait 到 long Texas",
        "player_selling_points": "chatterbait、swim jig、buzzbait 可用 slack-line reeling 搜索，長距 Texas 讀底和長行程補刺也不鬆散。",
    },
    "LTT680H": {
        "player_environment": "淡水 Bass / heavy plug 與 heavy softbait 技術場",
        "player_positioning": "Heavy technical versatile / 大型 crank 與重系 softbait",
        "player_selling_points": "大型 body crank 可慢巻，Texas、Cherry Rig、No Sinker Flip Gill 和 guarded rubber jig 都能用 16lb 級線組細控。",
    },
    "LTT690PH": {
        "player_environment": "淡水 Bass / 重餌遠投與軟硬餌綜合 power game",
        "player_positioning": "Parabolic heavy versatile / rubber jig 到 big crank",
        "player_selling_points": "深彎曲調性讓 rubber jig、big crank、large spinnerbait 和 heavy Texas 都能穩定拋投、吸口和落魚率。",
    },
    "LTT6100H": {
        "player_environment": "淡水 Bass / heavy cover 與 3-5m slow rolling",
        "player_positioning": "Heavy cover versatile / slow roll 到 heavy Carolina",
        "player_selling_points": "3/4oz spinnerbait slow rolling、swimbait、big spoon 可做強搜索，也能轉 6/0 Texas、guarded jig、heavy Carolina。",
    },
    "HB660L-Gc": {
        "player_environment": "淡水 Bass / 中近距 moving bait 泛用搜索",
        "player_positioning": "Glass composite moving bait / 輕 crank 到 chatterbait",
        "player_selling_points": "glass composite 追従短咬，light crank、chatterbait、buzzbait、spinnerbait、topwater 和 jerkbait 切換自然。",
    },
    "HB6100ML-Gc": {
        "player_environment": "淡水 Bass / shallow 到 3m crank 與 shad 搜索",
        "player_positioning": "Crank & shad reeling / 小 crank、3m diver、10g shad",
        "player_selling_points": "柔軟 blank 保留 small crank、3m diver 和 10g shad 的原生泳姿，light plug 只是延伸，cover 下滑入後遇短咬也更容易黏住。",
    },
    "HB6100M-Gc": {
        "player_environment": "淡水 Bass / 高阻力 chatterbait、spinnerbait 與中型 crank",
        "player_positioning": "High-resistance moving bait / spinnerbait 與 chatterbait",
        "player_selling_points": "高阻力 spinnerbait、chatterbait 和 mid crank 可穩定牽引，glass tip 吸口、butt 則負責單鉤上顎補刺。",
    },
    "HB730MH-Gc": {
        "player_environment": "淡水 Bass / 遠距 heavy spinnerbait 與 magnum crank",
        "player_positioning": "Power reserve vacuum / 長距強巻物",
        "player_selling_points": "7'3 長度讓 heavy spinnerbait、chatterbait、mid-to-magnum crank 不必全力揮也能出距離，遠處咬口仍能壓 hook。",
    },
    "HB560L": {
        "player_environment": "淡水 Bass / overhang 與 cover 邊短尺準投",
        "player_positioning": "Jerk & accuracy / jerkbait 與 topwater 操作",
        "player_selling_points": "短竿可用不同角度低彈道送餌，jerkbait、topwater 的 rod-work 清楚，也能讓 crank、spinnerbait 貼 cover 通過。",
    },
    "HB511LL": {
        "player_environment": "淡水 Bass / cover 周邊 small plug finesse",
        "player_positioning": "Small plug finesse / shad 與小 crank 精準投放",
        "player_selling_points": "sub-7g plug 能載重低彈道送進 cover，柔軟 tip 提升輕 plug 投放感，butt torque 不會遇大魚就失控。",
    },
    "HB600L": {
        "player_environment": "淡水 Bass / bank、reed、dock 短距 plug 操作",
        "player_positioning": "Short-distance rod-work plug / bank cover 準投",
        "player_selling_points": "bank、reed、dock 周邊的短距 rod-work plug 是主場，light plug 可從 crank 切到 topwater，角度變化投放時仍能保持出線和控姿。",
    },
    "HB600M": {
        "player_environment": "淡水 Bass / reed、芦邊中量硬餌短打",
        "player_positioning": "Back-hand accuracy mid / spinnerbait 與 crank cover shot",
        "player_selling_points": "3/8-1/2oz spinnerbait、crankbait 進 cover 時有足夠抗阻，backhand cast 時竿身容易壓載、出手不拖。",
    },
    "HB630LL": {
        "player_environment": "淡水 Bass / cover 邊輕量 plug 與 bait finesse",
        "player_positioning": "Side-hand light plug / small crank、shad、topwater",
        "player_selling_points": "小 crank、shad、topwater 可用較太線送入 cover 邊，細緻 tip 有水感，短咬追従比一般硬竿更好。",
    },
    "HB630L": {
        "player_environment": "淡水 Bass / 日常 moving bait 全段搜索",
        "player_positioning": "Moving bait first rod / topwater 到 chatterbait",
        "player_selling_points": "topwater、crank、spinnerbait、buzzbait、vibration、chatterbait 都能覆蓋，適合作為表層到巻物的第一支泛用竿。",
    },
    "HB630M": {
        "player_environment": "淡水 Bass / cover cranking 與中量 spinnerbait",
        "player_positioning": "Side-hand mid power / cover crank 與 spinnerbait",
        "player_selling_points": "1/2oz+ spinnerbait、cover crank、vibration 可用太線強攻 cover，竿身支撐比輕量 HB 更充足。",
    },
    "HB640ML": {
        "player_environment": "淡水 Bass / 遠投 jerkbait 與全段 moving bait",
        "player_positioning": "Technical long cast / jerkbait 高適性 moving bait",
        "player_selling_points": "jerkbait 節奏清楚，shallow crank 到 3m diver、topwater、buzzbait、chatterbait、spinnerbait、vibration 都能延伸。",
    },
    "HB660H": {
        "player_environment": "淡水 Bass / invisible structure 深層巻物",
        "player_positioning": "Over-head drive / slow-roll spinnerbait 與 4m crank",
        "player_selling_points": "3/4oz spinnerbait slow roll 和 4m diver crank 可長時間 trace，power 支撐深層阻力但不犧牲 treble moving bait 追従。",
    },
    "HB680L": {
        "player_environment": "淡水 Bass / weed、岩盤、cover 多點長投 crank",
        "player_positioning": "Extra energy drive light / 小 crank 到 deep diver",
        "player_selling_points": "small crank、flat-side shallow crank、shad、long-bill minnow 到 3-4m diver 都能長投後換線角攻完整段 cover。",
    },
    "HB680M": {
        "player_environment": "淡水 Bass / 遠距中量 moving bait 搜索",
        "player_positioning": "Extra energy drive mid / crank、spinnerbait 與 surface plug",
        "player_selling_points": "medium diver、spinnerbait、tail spinner、pencil、prop、buzzbait 都可遠投搜索，長竿有助於泳層維持和 ripping retrieve。",
    },
    "HB680XH": {
        "player_environment": "淡水 Bass / 大型餌低彈道 power cast",
        "player_positioning": "Heavy treble power cast / 180 級 big bait 與 50g magnum crank",
        "player_selling_points": "180 class big bait、50g magnum crank 可安靜低彈道入水，XH power 能打太軸 treble，也保留大型 treble plug 的操作自由度。",
    },
    "HB710LL": {
        "player_environment": "淡水 Bass / 長距 small-to-medium cranking",
        "player_positioning": "Master of cranking / 小中型 crank 專門",
        "player_selling_points": "小到中型 crank 長時間巻收穩定，tele butt 支撐拋投和深咬口，tip 會隨 lip 阻力自然變化。",
    },
    "HB760L": {
        "player_environment": "淡水 Bass / 大水面遠投中層 moving bait",
        "player_positioning": "Away distance light / crank 與中層 moving bait 切換",
        "player_selling_points": "crank、deep diver、spinnerbait、vibration 可遠投搜索，並可切換到 swim jig 作中層單鉤補充，長線下泳層穩。",
    },
    "HB760M": {
        "player_environment": "淡水 Bass / deep weed 與 break 遠投強巻物",
        "player_positioning": "Away distance mid / deep crank 到 swimming bait 切換",
        "player_selling_points": "large deep crank、heavy spinnerbait 可穩定 trace deep weed，並可切換到 swimming worm / swim jig 作強補刺單鉤延伸。",
    },
    "HB640LS-SGt": {
        "player_environment": "淡水 Bass / spinning high-speed shad 長投",
        "player_positioning": "High-speed shad spin / shad plug 專門",
        "player_selling_points": "shad plug 從低速到超高速都能保持直線和泳姿，SGt tip 提升高速 reaction bite 的吸入口和 landed rate。",
    },
    "HB660MLS-SGt": {
        "player_environment": "淡水 Bass / spinning metal reaction 與 deep PE",
        "player_positioning": "Spinning metal / metal bait reaction",
        "player_selling_points": "metal bait、metal vibration、blade bait 可快速誘發低活性魚，SGt glass tip 降低 fast retrieve miss bite 和 PE deep game 脫鉤。",
    },
    "STN680MS": {
        "player_environment": "淡水 Bass / deep structure PE 遠投與 cover finesse",
        "player_positioning": "PE spin structure / 遠投 cover Neko 與 down shot",
        "player_selling_points": "PE+leader 遠投時可用 Cover Neko、Down Shot、Jighead Wacky 打 deep structure，副線也能承接 Carolina、Texas、heavy downshot。",
    },
    "STN6100MLS": {
        "player_environment": "淡水 Bass / power spin 遠投 finesse 與 bug cover",
        "player_positioning": "Versatile power spin / mid stroll 到 PE cover finesse",
        "player_selling_points": "fluoro 可做 mid strolling、Neko、Jighead Wacky，PE+leader 可切 bug cover、offset Neko、No Sinker Fall Bait。",
    },
    "STN580ML": {
        "player_environment": "淡水 Bass / 空中 cover 吊るし與障礙邊表層",
        "player_positioning": "Tsuru-stroll / PE cover pitching 與表層切換",
        "player_selling_points": "PE 2 号級吊るし cover softbait 是主軸，短尺 pitching 精度高，也能切換到 small swisher / small popper 補表層精準投放。",
    },
    "STN650M": {
        "player_environment": "淡水 Bass / 杭、沈木與 light cover 精細底操",
        "player_positioning": "Neko & heavy drop shot / bait finesse bottom",
        "player_selling_points": "5in Neko 和 1/4oz drop shot 在杭、沈木旁有水阻也能操作，parabolic taper 讓咬口吃得進去。",
    },
    "STN670H": {
        "player_environment": "淡水 Bass / laydown、rock、break 底部精準操作",
        "player_positioning": "Bottom sensitive / heavy Texas 與 football jig",
        "player_selling_points": "Heavy Texas 打 cover、Football Jig 沿 break 操作，短強 tip 避卡，blank 感度可讀枝條和岩縫變化。",
    },
    "STN680MH": {
        "player_environment": "淡水 Bass / vegetation、floating cover、消波 block Texas",
        "player_positioning": "Texas versatile / 1/4oz Texas 專精泛用",
        "player_selling_points": "1/4oz Texas + Escape Twin 可安靜進 vegetation 和 block hole，fall / shake 讀感清楚，曲竿保護線結。",
    },
    "STN6100M": {
        "player_environment": "淡水 Bass / cover long pitch 與 flat softbait swimming",
        "player_positioning": "Versatile mid softbait / no sinker 到 light Texas",
        "player_selling_points": "No Sinker Fall Bait、Soft Jerk、Neko、Heavy Down Shot、Light Texas 都能用，長度讓 cover pitch 和 flat 連續 shake 更自然。",
    },
    "STN6100MH": {
        "player_environment": "淡水 Bass / 長線 cover Texas 與 heavy worming",
        "player_positioning": "Long cast Texas / 遠距 worming 專向",
        "player_selling_points": "強長 tip 在長線和 cover cushion 下仍能控 rig，遠距 Texas 的讀感、延遲咬口和曲竿浮魚是重點。",
    },
    "STN700H": {
        "player_environment": "淡水 Bass / flat、structure、break edge reaction Carolina",
        "player_positioning": "Reaction Carolina / heavy Carolina 專門",
        "player_selling_points": "heavy sinker 能被清楚帶動，leader 後方 worm 接觸 pin spot 和吸入口也能感知，適合深場精準 Carolina。",
    },
    "STN720MH": {
        "player_environment": "淡水 Bass / vegetation no sinker 與 block free rig",
        "player_positioning": "No Sinker Flip & Free Rig / 輕 rig 太線攻障礙",
        "player_selling_points": "High-density No Sinker 和 Free Rig 可長距 pitching，輕 rig + 太線下仍能細控，低活性淺掛後承接穩。",
    },
    "STN720H": {
        "player_environment": "淡水 Bass / heavy cover bulky Texas 與 gill bait",
        "player_positioning": "Bulky Texas & Gill bait / heavy cover softbait",
        "player_selling_points": "Bulky Texas、Flip Gill Neko、guarded rubber jig 面向 20lb 級 heavy cover，單手操作輕快但補刺和控魚 torque 充足。",
    },
    "STN660M-St": {
        "player_environment": "淡水 Bass / 高壓場 bait finesse 與小石縫 cover",
        "player_positioning": "Bait finesse solid tip / reaction down shot 到 small rubber jig",
        "player_selling_points": "solid tip 能吸收極小咬口，Reaction Down Shot、Neko、Small Rubber Jig、Paddle-tail Worm 都能細控且避開小石縫。",
    },
    "STN670MH-St": {
        "player_environment": "淡水 Bass / cover 奧中層 shake 與 bait finesse",
        "player_positioning": "Cover bait finesse / cover Neko 與中層 shake",
        "player_selling_points": "Cover Neko、cover small rubber jig 和 mid-column shake 能維持 range，MH butt 在 cover 奧補刺後能快速帶魚。",
    },
    "STN6100H-St": {
        "player_environment": "淡水 Bass / heavy cover worming 與 fall bite",
        "player_positioning": "Sensitive cover worming / cover Texas 與 football",
        "player_selling_points": "Cover Texas / Escape Twin 5g Texas 可讓魚含餌不易吐，Football Jig 可捕捉 fall bite，H power 負責 heavy rig 補刺。",
    },
    "STN511LLS": {
        "player_environment": "淡水 Bass / 深水與近目標 ultra-light finesse",
        "player_positioning": "Near-the-target finesse / 極細線 down shot、Neko",
        "player_selling_points": "0.2 號 PE 或 2lb fluoro 下可精細操作 Ultra-light Down Shot、Neko、No Sinker，底感和短咬口判斷清楚。",
    },
    "STN610LLS": {
        "player_environment": "淡水 Bass / 高壓近距 finesse 與短尺控魚",
        "player_positioning": "Finesse versatile / no sinker、down shot、split shot",
        "player_selling_points": "No Sinker、Down Shot、Split Shot 能毫米級操作，短尺讓突然咬口快速貫穿 barb，也能保持控魚主導權。",
    },
    "STN620LS": {
        "player_environment": "淡水 Bass / bank 與 boat high-speed finesse",
        "player_positioning": "High-speed finesse / jighead 到 no sinker 攻擊型精細",
        "player_selling_points": "Jighead、Split Shot、Down Shot、Neko、No Sinker 都能主動操作，soft tip 吃口、extra-fast belly 瞬間掛魚。",
    },
    "STN640LLS": {
        "player_environment": "淡水 Bass / 中層 mid strolling 與 slack line 控餌",
        "player_positioning": "Mid strolling / jighead 與 down shot 中層游動",
        "player_selling_points": "Mid Strolling Jighead 和 Down Shot 依靠 slack line 自然漂游，柔軟 tip 防跳動，強 butt 保留大魚補刺。",
    },
    "STN640MLS-Md": {
        "player_environment": "淡水 Bass / PE 表層虫系 twitch 與水面短咬",
        "player_positioning": "Insect surface twitch / Rat Insect 與 PE 表層細操",
        "player_selling_points": "Insect surface bait、Rat Insect 和 surface twitch bait 可高速 twitch 後停頓，低彈性 tip 配 PE 更能黏住水面短咬。",
    },
    "STN650LS": {
        "player_environment": "淡水 Bass / 長投 Neko、jighead wacky 與 boil shot",
        "player_positioning": "Long light finesse / Neko、Jighead Wacky、No Sinker Boil Shot",
        "player_selling_points": "Neko、Jighead Wacky 和 No Sinker Boil Shot 可長距精準 twitch，強 tip 處理水阻大的輕量 rig，L butt 保留遠距補刺。",
    },
    "760JMH": {
        "player_environment": "淡水 Bass / weed clump、芦、蒲內側 flipping",
        "player_positioning": "Jungle stick light / flipping jig、Texas、high-density no sinker",
        "player_selling_points": "20lb line 打 cover 奧入點，rubber jig、Texas、high-density no sinker 都能用 torque 拔魚，硬中帶食わせ餘量。",
    },
    "760JH": {
        "player_environment": "淡水 Bass / heavy cover punching 與強 flipping",
        "player_positioning": "Jungle stick punching / 1-2oz sinker heavy cover",
        "player_selling_points": "Punching Rig、Heavy Flipping、1-2oz Sinker Texas 可用太 PE / 25lb 線強攻，厚 blank 靠 torque 拔魚而不是只靠硬度。",
    },
    "680JMHS": {
        "player_environment": "淡水 Bass / heavy cover 內側 power finesse",
        "player_positioning": "Heavy-cover PE finesse / smolaba cover pitching",
        "player_selling_points": "Small Rubber Jig Power Finesse 可打 heavy cover 深處，PE 出線、短距 pitching、hanging cover shake 和太軸 hook 補刺銜接緊密。",
    },
    "700JHS": {
        "player_environment": "淡水 Bass / 房總型遠距與高處吊るし power finesse",
        "player_positioning": "Long-cast heavy-cover PE finesse / hanging smolaba",
        "player_selling_points": "Long-cast Power Finesse 和 Hanging Small Rubber Jig 面向遠距、高處吊るし與 monster bass，長線 hook stroke、full-power hookset 和高位控線更有餘量。",
    },
}


HARD_TERMS = [
    "Crankbait", "Shad", "Minnow", "Spinnerbait", "Chatterbait", "Vibration", "Buzzbait",
    "Topwater", "Jerkbait", "Plug", "Big Bait", "Swimbait", "Spoon", "Metal", "Popper",
    "Swisher", "Pencil", "Prop", "Tail Spinner", "Bladed Jig", "Blade Bait",
]
SOFT_TERMS = [
    "Texas", "Neko", "Down Shot", "Drop Shot", "Jighead", "Wacky", "No Sinker", "Free Rig",
    "Carolina", "Rubber Jig", "Football Jig", "Small Rubber Jig", "Punching", "Flipping",
    "Power Finesse", "Fall Bait", "Split Shot", "Cherry Rig", "Worm", "Soft Jerk",
    "Tsuru-stroll", "Mid Strolling", "Bug", "Insect", "Swim Jig",
]


def normalize(value):
    return " ".join(str(value or "").split())


def header_map(ws):
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def shade_detail_groups(ws):
    fill_a = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
    fill_b = PatternFill(fill_type="solid", fgColor="FFE8F1FB")
    col = header_map(ws)
    rod_col = col["rod_id"]
    last_rod_id = None
    group = -1
    for row in range(2, ws.max_row + 1):
        rod_id = ws.cell(row=row, column=rod_col).value
        if rod_id != last_rod_id:
            group += 1
            last_rod_id = rod_id
        fill = fill_a if group % 2 == 0 else fill_b
        for column in range(1, ws.max_column + 1):
            ws.cell(row=row, column=column).fill = fill


def has_any(text, terms):
    text = normalize(text).lower()
    return any(term.lower() in text for term in terms)


def detect_issues(row_values):
    sku = row_values["sku"]
    env = row_values["player_environment"]
    pos = row_values["player_positioning"]
    sell = row_values["player_selling_points"]
    pair = row_values["recommended_rig_pairing"]
    desc = row_values["Description"]
    issues = []

    for field, value in [("player_environment", env), ("player_positioning", pos), ("player_selling_points", sell)]:
        if not normalize(value):
            issues.append(f"blank:{field}")
        if any(token in normalize(value).lower() for token in ["官网", "白名单", "tackledb", "rods.jp", "source"]):
            issues.append(f"source_leak:{field}")

    if "淡水 Bass" not in env:
        issues.append("environment_not_bass")
    if any(word in env for word in ["海水", "木蝦", "船釣", "SLSJ", "rockfish"]):
        issues.append("wrong_environment_domain")

    pair_hard = has_any(pair, HARD_TERMS)
    pair_soft = has_any(pair, SOFT_TERMS)
    player_text = f"{pos} {sell}"
    if pair_hard and not pair_soft and any(word in player_text for word in ["底操", "worming", "softbait 專向", "軟餌專向"]):
        issues.append("hard_pairing_soft_only_player")
    if pair_soft and not pair_hard and any(word in player_text for word in ["硬餌專向", "巻物專向", "hardbait 專向"]):
        issues.append("soft_pairing_hard_only_player")
    if pair_hard and pair_soft and not any(word in player_text for word in ["切換", "兼用", "泛用", "雙線", "mixed", "多餌", "versatile"]):
        issues.append("mixed_pairing_player_not_mixed")

    explicit_checks = [
        ("テキサス", "Texas"),
        ("ラバージグ", "Rubber Jig"),
        ("フットボール", "Football Jig"),
        ("ネコ", "Neko"),
        ("ダウンショット", "Down Shot"),
        ("ジグヘッド", "Jighead"),
        ("ノーシンカー", "No Sinker"),
        ("フリーリグ", "Free Rig"),
        ("キャロ", "Carolina"),
        ("フロッグ", "Frog"),
        ("クランク", "Crankbait"),
        ("スピナーベイト", "Spinnerbait"),
        ("チャター", "Chatterbait"),
        ("バイブレーション", "Vibration"),
        ("バズベイト", "Buzzbait"),
        ("スイムベイト", "Swimbait"),
        ("ビッグベイト", "Big Bait"),
        ("パンチング", "Punching"),
    ]
    for raw, expected in explicit_checks:
        if raw in desc and expected.lower() not in pair.lower():
            if raw == "ジグヘッド" and "jighead wacky" in pair.lower():
                continue
            issues.append(f"description_pairing_gap:{expected}")

    return issues


def apply_updates(ws):
    col = header_map(ws)
    required = ["id", "rod_id", "SKU", "Description", "recommended_rig_pairing", *FIELDS]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")

    changes = []
    rows = []
    seen = set()
    for row in range(2, ws.max_row + 1):
        sku = normalize(ws.cell(row=row, column=col["SKU"]).value)
        seen.add(sku)
        target = PLAYER_FIELDS.get(sku)
        if not target:
            rows.append({"row": row, "sku": sku, "issues": ["missing_player_mapping"]})
            continue
        for field in FIELDS:
            old = normalize(ws.cell(row=row, column=col[field]).value)
            new = target[field]
            if old != new:
                ws.cell(row=row, column=col[field]).value = new
                changes.append(
                    {
                        "row": row,
                        "id": ws.cell(row=row, column=col["id"]).value,
                        "sku": sku,
                        "field": field,
                        "old": old,
                        "new": new,
                    }
                )
        row_values = {
            "row": row,
            "id": ws.cell(row=row, column=col["id"]).value,
            "rod_id": ws.cell(row=row, column=col["rod_id"]).value,
            "sku": sku,
            "recommended_rig_pairing": normalize(ws.cell(row=row, column=col["recommended_rig_pairing"]).value),
            "Description": normalize(ws.cell(row=row, column=col["Description"]).value),
            **target,
        }
        row_values["issues"] = detect_issues(row_values)
        rows.append(row_values)

    missing_skus = sorted(set(PLAYER_FIELDS) - seen)
    if missing_skus:
        raise RuntimeError(f"player mappings not matched in workbook: {missing_skus}")
    return changes, rows


def write_report(changes, rows):
    issue_rows = [row for row in rows if row["issues"]]
    unique_counts = {
        field: len({normalize(row[field]) for row in rows if normalize(row[field])})
        for field in FIELDS
    }
    coverage = {
        field: sum(1 for row in rows if normalize(row[field]))
        for field in FIELDS
    }
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "import_file": str(XLSX_PATH),
        "write_scope": FIELDS,
        "protected_fields": [
            "Description",
            "recommended_rig_pairing",
            "guide_use_hint",
            "official/spec columns",
        ],
        "coverage": coverage,
        "unique_counts": unique_counts,
        "changes": changes,
        "issue_rows": issue_rows,
        "rows": rows,
    }
    REPORT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return payload


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    before_header = [cell.value for cell in ws[1]]
    changes, rows = apply_updates(ws)
    after_header = [cell.value for cell in ws[1]]
    if before_header != after_header:
        raise RuntimeError("headers changed unexpectedly")
    shade_detail_groups(ws)
    wb.save(XLSX_PATH)
    report = write_report(changes, rows)
    print(
        json.dumps(
            {
                "rows": len(rows),
                "changes": len(changes),
                "coverage": report["coverage"],
                "unique_counts": report["unique_counts"],
                "issue_rows": len(report["issue_rows"]),
                "report": str(REPORT_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    if report["issue_rows"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
