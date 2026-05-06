import json
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "dstyle_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "dstyle_rod_player_fields_stage3_report.json"

FIELDS = ["player_environment", "player_positioning", "player_selling_points"]


PLAYER_FIELDS = {
    "DHRS-GD-64L": {
        "player_environment": "淡水 Bass / 岸钓与船钓 / 近中距离精细到小型硬饵搜索",
        "player_positioning": "轻量精细泛用 spinning",
        "player_selling_points": "Neko、Down Shot 和 No Sinker 的控线手感比较自然，小型 Plug 搜索也能兼顾，适合只带一支 spinning 应对高压场。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRC-GD-68M-LM": {
        "player_environment": "淡水 Bass / 开放水域・船钓 / 中层游动与轻移动饵",
        "player_positioning": "Bait mid-strolling 兼轻移动饵",
        "player_selling_points": "用枪柄系统做 Mid Strolling 时更好控线，Swimming Football Jig 也能维持泳姿；需要时可切 Spinnerbait、Crankbait 做弱搜索。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-511UL": {
        "player_environment": "淡水 Bass / 高压野池・近岸标点 / 短距超精细",
        "player_positioning": "近距离超精细食わせ",
        "player_selling_points": "短竿在狭小标点更容易压低弹道，Neko、Down Shot、Small Rubber Jig 的轻抖和短咬反馈更直接。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-60L": {
        "player_environment": "淡水 Bass / 高压湖库・野池 / 精细软饵到小型 Plug",
        "player_positioning": "精细食わせ泛用 spinning",
        "player_selling_points": "比 XUL/UL 更有余量，细线软饵仍然好操作，遇到鱼追但不咬时可换小型 Plug 做慢速搜索。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-63L": {
        "player_environment": "淡水 Bass / 岸钓中距离 / 轻量软饵・表层・Shad 切换",
        "player_positioning": "轻量多用途 spinning",
        "player_selling_points": "Neko、Down Shot、No Sinker 是主轴，表层 Plug、Shad 和虫系可补搜索；适合一支竿覆盖高压轻量路线。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-63UL-FS": {
        "player_environment": "淡水 Bass / 开放水域・缓坡 / 中层 finesse swim",
        "player_positioning": "轻量中层游动专向",
        "player_selling_points": "Jighead、Swimming Neko 和 Swimming Down Shot 的线弧更容易维持，适合持续抖动和轻量软饵水平泳姿控制。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-66ML": {
        "player_environment": "淡水 Bass / cover 边缘・码头・草洞 / PE power finesse",
        "player_positioning": "PE power finesse 多用途",
        "player_selling_points": "Small Rubber Jig 吊挂、Elastomer dog-walk 和 Power Mid Strolling 都能处理，价值在细操作后仍有把鱼带出 cover 的支撑。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-66L/ML-FS": {
        "player_environment": "淡水 Bass / 开放水域・大型湖库 / 4-5 寸 worm 中层游动",
        "player_positioning": "中型软饵 Mid Strolling",
        "player_selling_points": "4-5 寸 worm 做 Mid Strolling 时不容易被饵重拖垮，远处线弧和 roll 感更好维持，兼顾 PE power finesse。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-68M": {
        "player_environment": "淡水 Bass / 大型湖库・远投 cover / 强化 spinning finesse",
        "player_positioning": "远投型 power finesse spinning",
        "player_selling_points": "Cover Neko、Small Rubber Jig 和 6 寸以上 worm 的远距离操作更有底气，适合细线体系下想兼顾距离和控鱼的人。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRC-66L/ML": {
        "player_environment": "淡水 Bass / cover 边缘・浅场 / BFS 软饵精细",
        "player_positioning": "软饵向 bait finesse",
        "player_selling_points": "No Sinker、Neko、Down Shot 和 Small Rubber Jig 的落点控制好，适合用枪柄在障碍边缘做低弹道精细投放。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRC-69L+": {
        "player_environment": "淡水 Bass / 轻 cover・芦苇边・浮物 / cover finesse",
        "player_positioning": "轻 cover 精细软饵",
        "player_selling_points": "Cover Neko 和 Small Rubber Jig 是最自然的搭配，竿长让贴 cover 投放和中鱼后导鱼更有优势。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRC-68M": {
        "player_environment": "淡水 Bass / 岸钓・船钓 / 软饵底操到卷物搜索",
        "player_positioning": "软硬饵切换型 bait versatile",
        "player_selling_points": "No Sinker、Texas、Down Shot、Neko、Football 可以打点，Crankbait 和 Spinnerbait 可搜索，适合一支竿覆盖日常主力 bait 场景。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRC-610MH": {
        "player_environment": "淡水 Bass / 湖库・草边・硬底 / 中量软饵与移动饵",
        "player_positioning": "中量级强泛用 bait",
        "player_selling_points": "Texas、Football、Neko 和 Spinnerbait 都在合理范围，Mid-size Big Bait 只是扩展用途；更适合中重钓组的连续切换。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRC-70H": {
        "player_environment": "淡水 Bass / 重 cover・深水硬底・大型饵搜索",
        "player_positioning": "强力底操与重系搜索",
        "player_selling_points": "Heavy Texas、Rubber Jig、Carolina、Free Rig 能打重 cover，Big Spoon、Big Bait 和 Heavy Spinnerbait 可做强搜索，适合需要控鱼余量的场景。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-E-62UL-S": {
        "player_environment": "淡水 Bass / 高压开放水・清水 / PE finesse shooting",
        "player_positioning": "极细线轻量投射 finesse",
        "player_selling_points": "No Sinker、Down Shot、Jighead 和 Hover Strolling 的轻量投射更稳定，长 solid tip 对短咬和细微负重变化更友好。",
        "source": "recommended_pairing_plus_description",
    },
    "DHRS-E-510XUL-S": {
        "player_environment": "淡水 Bass / 极高压近距・ sight fishing / XUL finesse",
        "player_positioning": "短距极轻食わせ专向",
        "player_selling_points": "XUL solid 更适合轻 No Sinker、Down Shot、Neko 与 Hover Strolling，优势是轻饵存在感和短咬判断。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-SS-63ML": {
        "player_environment": "淡水 Bass / 大型湖库・风浪边 / Neko 与 power finesse",
        "player_positioning": "湖库 Neko power finesse",
        "player_selling_points": "Neko Rig 放在核心，Down Shot、Small Rubber Jig 和 No Sinker 可跟进；ML power 让远处中鱼后更容易保持主动。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-SS-68L": {
        "player_environment": "淡水 Bass / 大型湖库・远投开阔水 / long-cast mid strolling",
        "player_positioning": "远投中层游动 spinning",
        "player_selling_points": "6'8 长度适合把 Jighead Mid Strolling 和 Long Cast Down Shot 送远，PE 或较粗 fluorocarbon 下仍能维持线弧。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-SS-64MH": {
        "player_environment": "淡水 Bass / 浅草边・开口区 / 高比重 worm jerk",
        "player_positioning": "高比重无铅 jerking 专向",
        "player_selling_points": "短尺 MH 让 High-density No Sinker 和 Jerk Worm 更好做连续抽停，Texas、Free Rig 可作为同重量软饵替代。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-SS-69MH+": {
        "player_environment": "淡水 Bass / 琵琶湖型大型湖・草区边缘 / power versatile",
        "player_positioning": "强力软硬饵泛用",
        "player_selling_points": "Texas、Free Rig 和 No Sinker 可打点，Spinnerbait、Chatterbait、Swimbait 和 Big Bait 可搜索，适合大水面重线组切换。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-612UL+-S": {
        "player_environment": "淡水 Bass / 高压岸边・小型湖库 / 2pc super finesse",
        "player_positioning": "便携超精细 spinning",
        "player_selling_points": "No Sinker、Down Shot、Neko、Jighead 和 Hover Strolling 都能细致操作，2pc 适合移动频繁但仍要保持轻饵手感的玩家。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-632UL/L-S": {
        "player_environment": "淡水 Bass / 高压开放水・轻 cover 边 / all-round finesse",
        "player_positioning": "食わせ与控鱼兼顾 finesse",
        "player_selling_points": "UL tip 负责轻咬承接，L butt 负责控鱼；No Sinker、Down Shot、Small Rubber Jig、Hover 和 Elastomer 都能覆盖。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-642UL+-MIDSP": {
        "player_environment": "淡水 Bass / 开放水・缓坡・中层鱼 / 2pc mid strolling",
        "player_positioning": "便携 mid/bottom strolling 专向",
        "player_selling_points": "Mid Strolling Jighead 和 Bottom Strolling 的节奏更明确，Down Shot 也能按中层或贴底游动思路使用。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-662L": {
        "player_environment": "淡水 Bass / 岸钓远近切换 / finesse 到小型 Minnow・Shad",
        "player_positioning": "轻量软饵兼小硬饵",
        "player_selling_points": "Neko、Down Shot、No Sinker 可以食わせ，Small Minnow 和 Shad 能补搜索，适合岸边只带一支 spinning 时拉开距离。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-662M": {
        "player_environment": "淡水 Bass / 浓 cover・虫系・大型湖 light rig / PE power finesse",
        "player_positioning": "cover power finesse spinning",
        "player_selling_points": "Cover Neko、Guarded Small Rubber Jig 和 Bug Lure 适合打 cover，Long Cast Down Shot 可作为大型湖远投精细补充。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-6102ML-S": {
        "player_environment": "淡水 Bass / 岸钓远投・开阔水 / light rig 与表层搜索",
        "player_positioning": "远投轻量泛用 spinning",
        "player_selling_points": "Long Cast Down Shot 和 Neko 能保持距离下的操作感，I-shaped、Bug、Surface Plug 用来处理表层或清水弱波动搜索。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-662ML-BF": {
        "player_environment": "淡水 Bass / 岸边障碍・小型河川・野池 / multi BFS",
        "player_positioning": "小硬饵与精细软饵 BFS",
        "player_selling_points": "Small Plug、Small Spinnerbait 能搜索，Small Rubber Jig、Neko、Down Shot 可打点；适合近中距离精准抛投和手返し。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-672MH-S": {
        "player_environment": "淡水 Bass / tough cover・沉木・草边 / cover soft rig",
        "player_positioning": "cover 底操与 heavy finesse",
        "player_selling_points": "Cover Neko、Heavy Down Shot、Texas 和 Rubber Jig 都围绕 cover 展开，定位保守清晰，不把 MH 误扩成大型饵竿。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-6102M": {
        "player_environment": "淡水 Bass / 岸钓・船钓通用 / worm rig 到 Chatterbait",
        "player_positioning": "软硬饵日常主力 bait",
        "player_selling_points": "Texas、Free Rig、No Sinker 负责打点，Crankbait、Spinnerbait、Chatterbait 负责搜索，适合一支竿快速切换路线。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-6102MH": {
        "player_environment": "淡水 Bass / 中重 cover・硬底・浅草 / middle-heavy versatile",
        "player_positioning": "中重软饵与卷物泛用",
        "player_selling_points": "Texas、Free Rig、Rubber Jig 可读底和穿 cover，Spinnerbait、Chatterbait、Crankbait 能做中量搜索；不把用途夸大到纯大饵。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-6102XH": {
        "player_environment": "淡水 Bass / 大型湖・开阔水・大型鱼标点 / big bait 操作",
        "player_positioning": "操作系大型饵专向",
        "player_selling_points": "Big Bait、Swimbait、Crawler 和 Glide Bait 的抽停与承接是核心，适合需要明确操控大型饵而不是单纯重底操的玩家。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-702H": {
        "player_environment": "淡水 Bass / 重 cover・草区・大型饵补充搜索 / heavy versatile",
        "player_positioning": "强力 cover 与大型饵补充",
        "player_selling_points": "Heavy Texas、Rubber Jig 和 Frog 是主要场景，Big Bait、Swimbait 属副线搜索；适合重线组和控鱼优先的玩家。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-60XUL-S": {
        "player_environment": "淡水 Bass / 极高压近岸・清水 / ultimate finesse",
        "player_positioning": "短尺 XUL 食わせ finesse",
        "player_selling_points": "Down Shot、Neko、Hover Strolling 和轻 No Sinker 更容易做细节动作，适合轻咬口多、鱼不愿追饵的场景。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-67UL-S": {
        "player_environment": "淡水 Bass / 开阔水远投・缓坡 / light Carolina 与弱波动搜索",
        "player_positioning": "长尺 UL 远投 finesse",
        "player_selling_points": "Light Carolina 和长 leader Down Shot 适合远处慢拖，No Sinker 与 I-shaped Plug 用来处理清水弱波动鱼。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-60UL-S": {
        "player_environment": "淡水 Bass / 高压野池・近岸 cover 边 / super finesse",
        "player_positioning": "短距轻量食わせ",
        "player_selling_points": "Down Shot、Neko、No Sinker 和 Small Rubber Jig 都是高压场常用组合，短尺让贴边落点和轻抖更好控制。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-61UL+-S": {
        "player_environment": "淡水 Bass / 小型湖库・清水高压 / 轻量 finesse",
        "player_positioning": "UL+ 细线精细操作",
        "player_selling_points": "Down Shot、Neko、Jighead、No Sinker 的轻量操作更稳，适合需要比 XUL 多一点支撑但仍重视细节反馈的玩家。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-61L": {
        "player_environment": "淡水 Bass / 水库・清水岸钓 / Neko 与细线 light finesse",
        "player_positioning": "Neko 主轴轻量泛用",
        "player_selling_points": "Neko Rig 实战适配度高，Down Shot、No Sinker、Small Rubber Jig 和 Small Plug 能补齐食わせ到轻搜索的路线。",
        "source": "official_plus_whitelist",
    },
    "DBTS-63UL-MIDSP": {
        "player_environment": "淡水 Bass / 开放水・中层悬浮鱼 / mid strolling",
        "player_positioning": "轻量 Mid Strolling 专向",
        "player_selling_points": "Mid Strolling Jighead 和 Hover Strolling 是主轴，Down Shot、No Sinker 可按同样线弧和轻抖逻辑补充。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-65L+": {
        "player_environment": "淡水 Bass / 岸钓远投・开阔水 / finesse 与 Small Minnow",
        "player_positioning": "远投轻量软硬饵泛用",
        "player_selling_points": "Neko、Down Shot、No Sinker 用于食わせ，Small Minnow 用于远处搜索；L+ power 让轻量路线多一点距离和控鱼余量。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-66ML-S-MIDSP": {
        "player_environment": "淡水 Bass / 池・野池・开阔水 / mid strolling 与 No Sinker",
        "player_positioning": "ML solid 中层游动",
        "player_selling_points": "Mid Strolling Jighead 是主线，No Sinker 和 Hover Strolling 能处理更弱波动的中层鱼，小 Plug 只作为轻搜索补充。",
        "source": "official_plus_whitelist",
    },
    "DBTS-66M": {
        "player_environment": "淡水 Bass / cover 边缘・远投点 / spinning power finesse",
        "player_positioning": "M power finesse spinning",
        "player_selling_points": "Power Finesse、Cover Neko 和 Small Rubber Jig 的控鱼力比轻量 spinning 更充足，适合细线体系下处理更厚的 cover。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-68H-S-PF": {
        "player_environment": "淡水 Bass / 高压重 cover・草洞 / heavy power finesse",
        "player_positioning": "H solid power finesse",
        "player_selling_points": "Cover Neko、Small Rubber Jig 和高比重 No Sinker 进 cover 后有足够竿身支撑，适合 PE 强行控鱼。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTS-610L-S": {
        "player_environment": "淡水 Bass / 岸钓远投・开放水・表层 / long shooting finesse",
        "player_positioning": "远投 Neko 与表层轻搜索",
        "player_selling_points": "Long Cast Neko、I-shaped Plug、Surface Plug 和 Small Rubber Jig 都围绕距离展开，适合想把轻量路线送到更远标点的玩家。",
        "source": "official_plus_whitelist",
    },
    "DBTC-64ML-FM": {
        "player_environment": "淡水 Bass / 浅场・硬物边・护岸 / fast moving plug",
        "player_positioning": "Full-glass 卷物硬饵",
        "player_selling_points": "Crankbait、Shad、Minnow、Topwater 连续收线时更容易维持泳姿，Full-glass 调性对碰硬物和短咬承接更宽容。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-65ML-FM": {
        "player_environment": "淡水 Bass / 小型河川・野池・清水浅场 / light plugging",
        "player_positioning": "轻硬饵 plug 专向",
        "player_selling_points": "Minnow、Shad、轻 Crankbait 和 Topwater 的控速、停顿、twitch 更自然，适合用低弹性手感处理轻硬饵。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-65M+-PF": {
        "player_environment": "淡水 Bass / light cover・沉木边・草洞 / bait power finesse",
        "player_positioning": "轻 cover bait power finesse",
        "player_selling_points": "PE 时 Small Rubber Jig 和 Bait Neko 是主轴，换 fluorocarbon 后可用 Heavy Down Shot、Free Rig、Texas、Football 打底。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-67ML-BF": {
        "player_environment": "淡水 Bass / 野池・护岸・轻 cover / 攻击型 BFS",
        "player_positioning": "进攻型 bait finesse",
        "player_selling_points": "Small Plug 与 Small Spinnerbait 可快速搜索，Small Rubber Jig、Neko、Down Shot 可补精细打点，适合手返し很高的近中距离作钓。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-68M": {
        "player_environment": "淡水 Bass / 湖泊・船钓・开阔水 / 中量硬饵搜索",
        "player_positioning": "硬饵搜索兼 bait versatile",
        "player_selling_points": "Crankbait、Shad、Jerkbait 是更强项，Spinnerbait、Chatterbait 可扩展搜索，Texas 作为软饵补充而不是主轴。",
        "source": "official_plus_whitelist",
    },
    "DBTC-610MH": {
        "player_environment": "淡水 Bass / 野池・草边・浅 cover / Chatterbait 与中量 bait",
        "player_positioning": "Chatterbait 主轴中量泛用",
        "player_selling_points": "Chatterbait 和 Spinnerbait 的卷动稳定性更突出，Texas、Rubber Jig 可在同一套线组下补打点，适合移动饵和底操来回切。",
        "source": "official_plus_whitelist",
    },
    "DBTC-70M+-FM": {
        "player_environment": "淡水 Bass / 开阔水・硬物边・风口 / fast moving search",
        "player_positioning": "Glass-composite 强搜索卷物",
        "player_selling_points": "Crankbait、Big Minnow、Topwater、Spinnerbait、Buzzbait、Chatterbait 和 Shad Tail 都围绕连续搜索，适合覆盖水面和承接追咬。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-70H-S": {
        "player_environment": "淡水 Bass / heavy cover・沉木・厚草 / heavy cover finesse",
        "player_positioning": "重 cover 软饵控制",
        "player_selling_points": "Cover Neko、Texas、Rubber Jig 和 Heavy Down Shot 都服务于重 cover 内的精细进攻，重点是进得去、挂住后拉得出来。",
        "source": "recommended_pairing_plus_description",
    },
    "DBTC-71MH-PF": {
        "player_environment": "淡水 Bass / cover 池塘・草洞・浮物 / bait power finesse 与 Frog",
        "player_positioning": "Bait power finesse 兼 Frog cover",
        "player_selling_points": "Small Rubber Jig、Bait Neko 是主轴，Frog 可处理草面或浮物，Heavy Down Shot、Free Rig、Texas、Football 让同一支竿能切到底层强攻。",
        "source": "official_plus_whitelist",
    },
    "DBTC-73H": {
        "player_environment": "淡水 Bass / 大型湖・草区・重 cover / big bait 与 heavy cover",
        "player_positioning": "大型移动饵与 heavy cover 双线",
        "player_selling_points": "Big Bait、Crawler、Swimbait、Swim Jig 负责大范围搜索，Heavy Texas 和 Flipping 可处理重 cover，是偏强力路线的一支。",
        "source": "recommended_pairing_plus_description",
    },
}


def normalize(value):
    return " ".join(str(value or "").split())


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    required = ["id", "rod_id", "SKU", "Description", "recommended_rig_pairing", *FIELDS]
    missing = [name for name in required if name not in col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")

    rows = []
    missing_skus = []
    source_counts = {}
    for row in range(2, ws.max_row + 1):
        sku = normalize(ws.cell(row=row, column=col["SKU"]).value)
        detail_id = normalize(ws.cell(row=row, column=col["id"]).value)
        rod_id = normalize(ws.cell(row=row, column=col["rod_id"]).value)
        item = PLAYER_FIELDS.get(sku)
        if not item:
            missing_skus.append({"xlsx_row": row, "id": detail_id, "rod_id": rod_id, "SKU": sku})
            continue

        before = {field: normalize(ws.cell(row=row, column=col[field]).value) for field in FIELDS}
        for field in FIELDS:
            ws.cell(row=row, column=col[field]).value = item[field]
        source_counts[item["source"]] = source_counts.get(item["source"], 0) + 1
        rows.append({
            "xlsx_row": row,
            "id": detail_id,
            "rod_id": rod_id,
            "SKU": sku,
            "recommended_rig_pairing": normalize(ws.cell(row=row, column=col["recommended_rig_pairing"]).value),
            "source": item["source"],
            "before": before,
            "after": {field: item[field] for field in FIELDS},
            "changed_fields": [field for field in FIELDS if before[field] != item[field]],
        })

    if missing_skus:
        raise RuntimeError(f"missing player field mappings: {missing_skus}")

    wb.save(XLSX_PATH)

    report = {
        "schema": "dstyle_rod_player_fields_stage3",
        "source_xlsx": str(XLSX_PATH),
        "fields": FIELDS,
        "total_detail_rows": ws.max_row - 1,
        "updated_rows": len(rows),
        "source_counts": source_counts,
        "rows": rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "xlsx": str(XLSX_PATH),
        "report": str(REPORT_PATH),
        "updated_rows": len(rows),
        "source_counts": source_counts,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
