import json
import re
import sys
import time
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from scrapling.fetchers import Fetcher
except ModuleNotFoundError:
    Fetcher = None


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
NORMALIZED_PATH = DATA_DIR / "shimano_rod_normalized.json"
XLSX_PATH = DATA_DIR / "shimano_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "shimano_rod_product_technical_stage25_report.json"
EVIDENCE_PATH = DATA_DIR / "shimano_rod_product_technical_official_evidence.json"
FIELD = "product_technical"


TECH_ALIASES = {
    "X导环": "X 导环",
    "X GUIDE": "X 导环",
    "FULL CARBON MONOCOQUE": "全碳纤维一体成型握把",
    "FULLCARBONMONOCOQUE": "全碳纤维一体成型握把",
    "CARBON MONOCOQUE": "碳纤维一体成型握把",
    "CARBONMONOCOQUE": "碳纤维一体成型握把",
    "HI-POWERX": "HI-POWER X",
    "HI POWER X": "HI-POWER X",
    "HI-POWERXSOLID": "HI-POWER X SOLID",
    "HI POWER X SOLID": "HI-POWER X SOLID",
    "EXCITE TOP": "EXCITETOP",
    "TAFTEC INFINITY": "TAFTEC∞",
    "TAFTEC∞": "TAFTEC∞",
    "SOFTUBETOP": "SOFTUBE TOP",
    "SOFTUBE TOP": "SOFTUBE TOP",
    "NANOPITCH": "NANO PITCH",
    "NANO PITCH": "NANO PITCH",
    "ULTIMATE BLANKS DESIGN": "UBD",
    "MULTI PIECE ULTIMATE BLANKS DESIGN": "MULTI PIECE ULTIMATE BLANKS DESIGN",
    "PERFECTION CI4+": "PERFECTION CI4+ 渔轮座",
    "PERFCTION CI4+": "PERFECTION CI4+ 渔轮座",
    "G-CLOTH PROTECTOR": "G-CLOTH PROTECTOR",
    "CARBON SHELL GRIP": "CARBONSHELL GRIP",
    "CARBONSHELL GRIP": "CARBONSHELL GRIP",
    "MUSCLE CARBON": "MUSCLE CARBON",
    "MUSCLECARBON": "MUSCLE CARBON",
}

IGNORED_ALT_RE = re.compile(r"_RD$|^Image:?", re.I)
SKU_BOUNDARY = r"A-Z0-9/+.\-"


def canonical_tech_name(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    text = re.sub(r"^Image:\s*", "", text, flags=re.I)
    if not text or IGNORED_ALT_RE.search(text):
        return ""
    compact_upper = text.upper().replace("　", " ")
    compact_upper = re.sub(r"\s+", " ", compact_upper).strip()
    no_space_upper = compact_upper.replace(" ", "")
    if compact_upper in TECH_ALIASES:
        return TECH_ALIASES[compact_upper]
    if no_space_upper in TECH_ALIASES:
        return TECH_ALIASES[no_space_upper]
    return text


def merge_terms(values):
    merged = []
    for value in values:
        for part in str(value or "").split("/"):
            term = part.strip()
            if term and term not in merged:
                merged.append(term)
    return " / ".join(merged)


def normalize_code(text):
    value = str(text or "").upper()
    value = value.replace("＋", "+").replace("－", "-").replace("／", "/")
    value = re.sub(r"\s+", "", value)
    return value


def strip_model_prefix(sku, model):
    value = str(sku or "").strip()
    prefix = str(model or "").strip()
    if not value or not prefix or not value.startswith(prefix):
        return value
    rest = value[len(prefix):]
    if not rest or (not rest[0].isspace() and rest[0] not in "-_/（(［["):
        return value
    return re.sub(r"^[-_/（(［[\s]+", "", rest).strip() or value


def infer_rod_type(variant):
    raw_specs = variant.get("raw_specs") or {}
    code = str(raw_specs.get("型号") or raw_specs.get("货号") or variant.get("variant_name") or "").upper()
    if (
        re.search(r"(^|\s)2\d{2}", code)
        or re.search(r"(^|\s)S\d{2}", code)
        or "SJR" in code
        or re.search(r"\d+S(?:\b|[-/])", code)
        or re.search(r"^TW\d{2}", code)
    ):
        return "S"
    if (
        re.search(r"(^|\s)1\d{2}", code)
        or re.search(r"^[BC]\d{2}", code)
        or "BFS" in code
        or "MBR" in code
        or re.search(r"\d+C(?:\b|[-/])", code)
    ):
        return "C"
    return ""


def sku_in_text(sku, text):
    sku_norm = re.escape(normalize_code(sku))
    text_norm = normalize_code(text)
    return bool(re.search(rf"(?<![{SKU_BOUNDARY}]){sku_norm}(?![{SKU_BOUNDARY}])", text_norm))


def note_mentions_tech(note, tech):
    label = str(note or "").lstrip("※").split("：", 1)[0].split(":", 1)[0]
    label = re.sub(r"（.*?）|\(.*?\)", "", label).strip()
    label_canonical = canonical_tech_name(label)
    return label_canonical == tech


def parse_note_applicability(note, tech, variants):
    if not note_mentions_tech(note, tech):
        return None

    sku_matches = [row["sku"] for row in variants if sku_in_text(row["sku"], note)]
    note_text = str(note)
    if "仅直柄" in note_text or "直柄规格" in note_text and not sku_matches:
        return {row["sku"] for row in variants if row["type"] == "S"}
    if "仅枪柄" in note_text or "枪柄规格" in note_text and not sku_matches:
        return {row["sku"] for row in variants if row["type"] == "C"}
    if "除" in note_text and "之外" in note_text and ("直柄" in note_text or "纺车" in note_text):
        excluded = set(sku_matches)
        return {row["sku"] for row in variants if row["type"] == "S" and row["sku"] not in excluded}
    if "除" in note_text and "之外" in note_text and ("枪柄" in note_text or "两轴" in note_text):
        excluded = set(sku_matches)
        return {row["sku"] for row in variants if row["type"] == "C" and row["sku"] not in excluded}
    if "除外" in note_text:
        excluded = set(sku_matches)
        return {row["sku"] for row in variants if row["sku"] not in excluded}
    if sku_matches:
        return set(sku_matches)
    return None


def fetch_html(url, retries=2):
    if Fetcher is None:
        return "", None, "scrapling is unavailable; use cached evidence or install the crawler dependency"
    last_error = None
    for attempt in range(retries + 1):
        try:
            page = Fetcher.get(url, timeout=40)
            html = str(page.html_content or "")
            if page.status == 200 and "Access Denied" not in html and len(html) > 1000:
                return html, page.status, None
            last_error = f"status={page.status}, len={len(html)}"
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
        time.sleep(0.8 + attempt * 0.7)
    return "", None, last_error


def extract_official_tech(url):
    html, status, error = fetch_html(url)
    evidence = {
        "url": url,
        "status": status,
        "error": error,
        "tech_terms": [],
        "notes": [],
        "has_official_technology_section": False,
    }
    if not html:
        return evidence

    soup = BeautifulSoup(html, "html.parser")
    section = soup.select_one("section.tech-listing") or soup.select_one(".technologylisting")
    if not section:
        return evidence

    thumb = section.select_one(".tech-listing-thumbnail")
    imgs = thumb.select("img[alt]") if thumb else section.select(".tech-listing-thumbnail__item img[alt]")
    terms = []
    for img in imgs:
        term = canonical_tech_name(img.get("alt"))
        if term and term not in terms:
            terms.append(term)

    text = section.get_text("\n", strip=True)
    notes = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("※") and stripped not in notes:
            notes.append(stripped)

    evidence["has_official_technology_section"] = bool(terms)
    evidence["tech_terms"] = terms
    evidence["notes"] = notes
    return evidence


def values_for_item(item, item_idx, evidence):
    variants = []
    for variant_idx, variant in enumerate(item.get("variants") or []):
        variants.append({
            "variant_index": variant_idx,
            "sku": strip_model_prefix(variant.get("variant_name"), item.get("model_name")),
            "type": infer_rod_type(variant),
        })

    values = {row["sku"]: [] for row in variants}
    tech_terms = evidence.get("tech_terms") or []
    notes = evidence.get("notes") or []
    note_matches = {}

    for tech in tech_terms:
        scoped = None
        matched_notes = []
        for note in notes:
            parsed = parse_note_applicability(note, tech, variants)
            if parsed is not None:
                scoped = parsed if scoped is None else scoped & parsed
                matched_notes.append(note)
        target_skus = scoped if scoped is not None else {row["sku"] for row in variants}
        note_matches[tech] = matched_notes
        for sku in target_skus:
            if sku in values:
                values[sku].append(tech)

    return [
        {
            "rod_id": f"SR{1000 + item_idx}",
            "model": item.get("model_name") or "",
            "variant_index": row["variant_index"],
            "sku": row["sku"],
            "value": merge_terms(values[row["sku"]]),
        }
        for row in variants
    ], note_matches


def headers(ws):
    return [cell.value for cell in ws[1]]


def ensure_header(ws, field, after=None):
    current = headers(ws)
    if field in current:
        return current.index(field) + 1
    idx = current.index(after) + 2 if after and after in current else len(current) + 1
    ws.insert_cols(idx)
    ws.cell(row=1, column=idx, value=field)
    return idx


def delete_header_if_exists(ws, field):
    current = headers(ws)
    if field not in current:
        return False
    ws.delete_cols(current.index(field) + 1)
    return True


def row_dicts(ws, ignored_fields=None):
    ignored = set(ignored_fields or [])
    current = headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            header: row[idx] if idx < len(row) else None
            for idx, header in enumerate(current)
            if header not in ignored
        })
    return rows


def restore_rod_detail_fills(ws):
    current_headers = headers(ws)
    rod_id_idx = current_headers.index("rod_id") + 1
    fill_a = PatternFill(fill_type="solid", fgColor="FFF8F3C8")
    fill_b = PatternFill(fill_type="solid", fgColor="FFE8F1FB")
    current_rod_id = None
    group_index = -1
    for row_idx in range(2, ws.max_row + 1):
        rod_id = ws.cell(row=row_idx, column=rod_id_idx).value
        if rod_id != current_rod_id:
            current_rod_id = rod_id
            group_index += 1
        fill = fill_a if group_index % 2 == 0 else fill_b
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def main():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    cached_evidence = {}
    if EVIDENCE_PATH.exists() and "--refetch" not in sys.argv:
        cached_evidence = json.loads(EVIDENCE_PATH.read_text(encoding="utf-8"))
    evidence_by_rod = {}
    variant_values = []
    fetch_failures = []
    no_section = []

    for item_idx, item in enumerate(data):
        rod_id = f"SR{1000 + item_idx}"
        url = item.get("url") or ""
        evidence = cached_evidence.get(rod_id)
        if evidence:
            cached_terms = []
            for term in evidence.get("tech_terms") or []:
                canonical = canonical_tech_name(term)
                if canonical and canonical not in cached_terms:
                    cached_terms.append(canonical)
            evidence = {
                "url": evidence.get("url") or url,
                "status": evidence.get("status"),
                "error": evidence.get("error"),
                "tech_terms": cached_terms,
                "notes": evidence.get("notes") or [],
                "has_official_technology_section": bool(cached_terms),
            }
        else:
            evidence = extract_official_tech(url) if url else {
            "url": url,
            "status": None,
            "error": "missing url",
            "tech_terms": [],
            "notes": [],
            "has_official_technology_section": False,
            }
        rows, note_matches = values_for_item(item, item_idx, evidence)
        evidence["note_matches"] = note_matches
        evidence["nonempty_variant_count"] = sum(1 for row in rows if row["value"])
        evidence["variant_count"] = len(rows)
        evidence_by_rod[rod_id] = evidence
        variant_values.extend(rows)
        if evidence.get("error"):
            fetch_failures.append(rod_id)
        if not evidence.get("has_official_technology_section"):
            no_section.append(rod_id)
        print(f"{rod_id} {item.get('model_name')}: {len(evidence.get('tech_terms') or [])} techs, {evidence['nonempty_variant_count']}/{len(rows)} variants")
        sys.stdout.flush()

    # normalized: product_technical only exists on variants.
    value_iter = iter(variant_values)
    item_field_removed = 0
    variant_nonempty = 0
    for item in data:
        if FIELD in item:
            item.pop(FIELD, None)
            item_field_removed += 1
        for variant in item.get("variants") or []:
            row = next(value_iter)
            variant[FIELD] = row["value"]
            if row["value"]:
                variant_nonempty += 1

    NORMALIZED_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]
    before_rod = row_dicts(rod_ws, {FIELD})
    before_detail = row_dicts(detail_ws, {FIELD})
    rod_had_field = delete_header_if_exists(rod_ws, FIELD)
    tech_col = ensure_header(detail_ws, FIELD, after="Description")

    if detail_ws.max_row - 1 != len(variant_values):
        raise RuntimeError(f"rod_detail rows ({detail_ws.max_row - 1}) != normalized variants ({len(variant_values)})")

    detail_nonempty = 0
    for row_idx, row in enumerate(variant_values, start=2):
        detail_ws.cell(row=row_idx, column=tech_col, value=row["value"])
        if row["value"]:
            detail_nonempty += 1

    if before_rod != row_dicts(rod_ws, {FIELD}):
        raise RuntimeError("Unexpected rod sheet changes outside product_technical column removal")
    if before_detail != row_dicts(detail_ws, {FIELD}):
        raise RuntimeError("Unexpected rod_detail sheet changes outside product_technical")

    restore_rod_detail_fills(detail_ws)
    wb.save(XLSX_PATH)

    report = {
        "scope": "shimano rod product_technical rebuilt from official product technology section only",
        "source_rule": "Only .tech-listing / technologylisting official product technology icons and official applicability notes are used; Description and whitelist are not used.",
        "normalized_items": len(data),
        "normalized_variants": len(variant_values),
        "item_product_technical_fields_removed": item_field_removed,
        "variant_product_technical_nonempty": variant_nonempty,
        "rod_sheet_product_technical_removed": rod_had_field,
        "rod_detail_product_technical_nonempty": detail_nonempty,
        "fetch_failures": fetch_failures,
        "no_official_technology_section": no_section,
        "sample_evidence": {key: evidence_by_rod[key] for key in sorted(evidence_by_rod)[:10]},
    }
    EVIDENCE_PATH.write_text(json.dumps(evidence_by_rod, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
