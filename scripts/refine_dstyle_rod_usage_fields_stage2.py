import json
from copy import copy
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "dstyle_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "dstyle_rod_usage_fields_stage2_report.json"

PAIRING_FIELD = "recommended_rig_pairing"
HINT_FIELD = "guide_use_hint"
PAIRING_AFTER_FIELD = "guide_use_hint"


USAGE_BY_SKU = {
    "DHRS-GD-64L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "hint": "RIKU 是轻量泛用 spinning：先以 Neko、Down Shot、No Sinker 做精细操作，需要搜索时再切到小型 Plug；温和 Fast 调性兼顾远投、控线和中鱼后的牵制。",
        "source": "official_description",
    },
    "DHRC-GD-68M-LM": {
        "pairing": "Mid Strolling Jighead / Swimming Football Jig / Hover Strolling / Spinnerbait / Crankbait",
        "hint": "THE SUTO 以 bait mid-strolling 为主轴：Jighead 或 Football Jig 自然游动放首位，低弹性碳布让短咬更容易黏住；硬饵和 wire bait 只是副线用途。",
        "source": "official_description",
    },
    "DHRS-511UL": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "hint": "短竿超精细定位，适合高压近距离用 Neko、Down Shot、No Sinker 和 Small Rubber Jig 做精准落点与细抖；小型 Plug 属补充搜索。",
        "source": "official_description",
    },
    "DHRS-60L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "hint": "Finesse Shake 2 仍以食わせ系 Fast taper 为核心，Neko、Down Shot、No Sinker 和 Small Rubber Jig 是主场；需要拉开距离时可兼顾小型 Plug。",
        "source": "official_description",
    },
    "DHRS-63L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Surface Plug / Shad / Bug Lure",
        "hint": "Finesse Versatile 覆盖 spinning light-rig 全段：底层 Neko/Down Shot、无铅软饵、表层 Plug、Shad 和虫系都能切换，重点是轻量多用途而不是强障碍。",
        "source": "official_description",
    },
    "DHRS-63UL-FS": {
        "pairing": "Mid Strolling Jighead / Swimming Neko Rig / Swimming Down Shot / Hover Strolling",
        "hint": "Finesse Swim 不是普通底操竿，重点是用 Jighead、Neko 或 Down Shot 做中层游动；竿尖和导环设定服务于线弧、抖动节奏和泳姿稳定。",
        "source": "official_description",
    },
    "DHRS-66ML": {
        "pairing": "Power Finesse / Hanging Small Rubber Jig / Elastomer Dog Walk / Power Mid Strolling / Heavy Down Shot",
        "hint": "MAX Finesse 面向 PE cover power finesse：吊挂 Small Rubber Jig、Elastomer dog-walk、Power Mid Strolling 和稍重 Neko/Down Shot 都可用，重点是细致操作后把鱼带出 cover。",
        "source": "official_description",
    },
    "DHRS-66L/ML-FS": {
        "pairing": "Mid Strolling Jighead / Hover Strolling / 4-5in Worm Swim / Power Finesse",
        "hint": "Power Finesse Swim 是 4-5 寸中型 worm 的 mid-strolling 专用向，导环给线松弛和 blank 震动传递留空间；Power Finesse 只是同一套 PE 系统下的延伸。",
        "source": "official_description",
    },
    "DHRS-68M": {
        "pairing": "Power Finesse / Cover Neko / Small Rubber Jig / 6in+ Mid Strolling / Heavy No Sinker",
        "hint": "Power Finesse Shake 以 PE spinning 远投和 cover 拖鱼为核心，Cover Neko、Small Rubber Jig、6 寸以上 worm 的 Mid Strolling 和较重无铅软饵都比普通轻量底操更匹配。",
        "source": "official_description",
    },
    "DHRC-66L/ML": {
        "pairing": "No Sinker / Neko Rig / Down Shot / Small Rubber Jig",
        "hint": "Bait Finesse Versatile 适合在 cover 与 open water 间精准投放 No Sinker、Neko、Down Shot 和 Small Rubber Jig；优势是短中距离落点、读底和控线。",
        "source": "official_description",
    },
    "DHRC-69L+": {
        "pairing": "Cover Neko / Small Rubber Jig / No Sinker / Light Texas",
        "hint": "Cover Finesse 的重点是把轻量软饵送进 cover：Cover Neko 和 Small Rubber Jig 放首位，No Sinker 与 Light Texas 用于更隐蔽或更抗挂的切换。",
        "source": "official_description",
    },
    "DHRC-68M": {
        "pairing": "No Sinker / Texas Rig / Down Shot / Neko Rig / Football Jig / Crankbait / Spinnerbait",
        "hint": "FX 是 bait 万用核心竿，描述明确覆盖从撃ち到巻き：软饵端以 No Sinker、Texas、Down Shot、Neko、Football 为主，搜索端再接 Crankbait 和 Spinnerbait。",
        "source": "official_description",
    },
    "DHRC-610MH": {
        "pairing": "Texas Rig / No Sinker / Down Shot / Neko Rig / Football Jig / Spinnerbait / Mid-size Big Bait",
        "hint": "FXR 是中量级 super versatile，先承担 Texas、No Sinker、Down Shot、Neko 和 Football 这类 worm/jig 场景，再扩展到 wire bait 与中型 Big Bait。",
        "source": "official_description",
    },
    "DHRC-70H": {
        "pairing": "Heavy Texas / Rubber Jig / Carolina Rig / Free Rig / Big Spoon / Big Bait / Heavy Spinnerbait",
        "hint": "Power Fishing Versatile 明确面向重 cover 和重系钓组：Heavy Texas、Rubber Jig、Carolina、Free Rig 是主线，Big Spoon、Big Bait 和 Heavy Spinnerbait 属强力搜索补充。",
        "source": "official_description",
    },
    "DHRS-E-62UL-S": {
        "pairing": "No Sinker / Down Shot / Neko Rig / Jighead Rig / Small Rubber Jig / Hover Strolling",
        "hint": "The Finesse Shooting 用长 solid tip 和细 PE 做轻量投射与操作，No Sinker、Down Shot、Neko、Jighead、Small Rubber Jig 和 Hover Strolling 都是官网明确适用。",
        "source": "official_description",
    },
    "DHRS-E-510XUL-S": {
        "pairing": "No Sinker / Down Shot / Neko Rig / Jighead Rig / Hover Strolling",
        "hint": "Finesse Shake KIWAMI 是 XUL solid 的极限食わせ竿，适合短咬明显的 No Sinker、Down Shot、Neko、Jighead 和 Hover Strolling，重点是轻量操控和瞬间挂鱼。",
        "source": "official_description",
    },
    "DBTS-SS-63ML": {
        "pairing": "Neko Rig / Down Shot / Small Rubber Jig / No Sinker / Power Finesse",
        "hint": "Lightning 为 big lake finesse pattern 设计，Neko 是核心，Down Shot、Small Rubber Jig 和 No Sinker 承接 light-rig 全般；ML power 用来补足远处中鱼后的牵制。",
        "source": "official_description",
    },
    "DBTS-SS-68L": {
        "pairing": "Mid Strolling Jighead / Long Cast Down Shot / No Sinker / Small Rubber Jig",
        "hint": "BM1 的主轴是 long-cast mid-strolling，6'8 长度和较粗 fluorocarbon/PE 导环设定服务于远投控线；Down Shot、No Sinker 只是 light-rig 侧向扩展。",
        "source": "official_description",
    },
    "DBTC-SS-64MH": {
        "pairing": "High-density No Sinker / Jerk Worm / Texas Rig / Free Rig",
        "hint": "Dagger 明确为高比重 worm no-sinker jerking 设计，6'4 长度便于连续 jerk 和控姿；Texas、Free Rig 只是同等重量软饵的补充用法。",
        "source": "official_description",
    },
    "DBTC-SS-69MH+": {
        "pairing": "Texas Rig / Free Rig / No Sinker / Spinnerbait / Chatterbait / Swimbait / Big Bait",
        "hint": "Saber versatile SP 是琵琶湖向 power versatile，先覆盖 worm 系 rig，再扩展到 wire bait、Chatterbait、Swimbait 和 Big Bait；不是单纯 heavy bottom 竿。",
        "source": "official_description",
    },
    "DBTS-612UL+-S": {
        "pairing": "No Sinker / Down Shot / Neko Rig / Jighead Rig / Small Rubber Jig / Hover Strolling",
        "hint": "2-piece 612UL+-S 是更轻量的 super finesse，solid tip 让 No Sinker、Down Shot、Neko、Jighead、Small Rubber Jig 与 Hover Strolling 的轻咬更容易吃住。",
        "source": "official_description",
    },
    "DBTS-632UL/L-S": {
        "pairing": "No Sinker / Down Shot / Neko Rig / Jighead Rig / Small Rubber Jig / Hover Strolling / Elastomer Softbait",
        "hint": "632UL/L-S 用 UL tip 搭配 L butt，官网明确覆盖 No Sinker、Down Shot、Neko、Jighead、Small Rubber Jig、Hover Strolling 与 Elastomer；适合食わせ和控鱼并重。",
        "source": "official_description",
    },
    "DBTS-642UL+-MIDSP": {
        "pairing": "Mid Strolling Jighead / Bottom Strolling / Down Shot / Jighead Rig",
        "hint": "642UL+-MIDSP 是 mid-strolling 专门项，Jighead 与 Down Shot 的 shake 线弧是核心；Bottom Strolling 是同一调性下的底层游动延伸。",
        "source": "official_description",
    },
    "DBTS-662L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Minnow / Shad",
        "hint": "662L 是 finesse 到小型 plug 的多用途 spinning，食わせ端以 Neko、Down Shot、No Sinker 为主；远投搜索时更适合 Small Minnow 与 Shad，而不是重 cover。",
        "source": "official_description",
    },
    "DBTS-662M": {
        "pairing": "Cover Neko / Guarded Small Rubber Jig / Bug Lure / Power Finesse / Long Cast Down Shot",
        "hint": "662M 是 PE power finesse spinning，重点是 Guarded Small Rubber Jig、Cover Neko 和虫系打浓 cover；big lake 远投 Down Shot/Neko 可用，但不是第一定位。",
        "source": "official_description",
    },
    "DBTS-6102ML-S": {
        "pairing": "Long Cast Down Shot / Neko Rig / I-shaped Plug / Bug Lure / Surface Plug",
        "hint": "6102ML-S 是 long-shooting special，ML solid tip 和 PE 导环适合把 Down Shot/Neko、I-shaped、虫系和表层 Plug 送远；核心价值是距离下仍保留操作感。",
        "source": "official_description",
    },
    "DBTC-662ML-BF": {
        "pairing": "Bait Finesse Small Plug / Small Spinnerbait / Small Rubber Jig / Neko Rig / Down Shot",
        "hint": "662ML-BF 是 multi bait finesse，小型 Plug 与小型 wire bait 可以搜索，Small Rubber Jig、Neko、Down Shot 用于精细投放；优势是近中距离准确和手返し。",
        "source": "official_description",
    },
    "DBTC-672MH-S": {
        "pairing": "Cover Neko / Heavy Down Shot / Texas Rig / Rubber Jig",
        "hint": "672MH-S 描述虽短，但明确是 tough-cover rod；推荐只保守落在 Cover Neko、Heavy Down Shot、Texas 与 Rubber Jig，不外扩到 Big Bait。",
        "source": "official_short_description",
    },
    "DBTC-6102M": {
        "pairing": "Texas Rig / Free Rig / No Sinker / Crankbait / Spinnerbait / Chatterbait",
        "hint": "6102M 是岸钓到船钓都能带的一支 super versatile，worm 系 rig 负责撃ち，Crankbait、Spinnerbait、Chatterbait 负责巻き；适合软硬饵切换。",
        "source": "official_description",
    },
    "DBTC-6102MH": {
        "pairing": "Texas Rig / Free Rig / Rubber Jig / Spinnerbait / Chatterbait / Crankbait",
        "hint": "6102MH 只有中量级 versatile 描述，按 MH 规格保守放在 Texas、Free Rig、Rubber Jig 与中量移动饵；不因 MH 自动扩写到大型饵。",
        "source": "official_short_description",
    },
    "DBTC-6102XH": {
        "pairing": "Big Bait / Swimbait / Crawler Bait / Glide Bait",
        "hint": "6102XH 的描述明确是操作系 Big Bait 竿，优先服务 Big Bait、Swimbait、Crawler 和 Glide 系的操控与承接，不再套用普通 Texas/Jig 泛用逻辑。",
        "source": "official_short_description",
    },
    "DBTC-702H": {
        "pairing": "Heavy Texas / Rubber Jig / Frog / Big Bait / Swimbait",
        "hint": "702H 仅有 heavy super versatile 描述，按 H power 保守覆盖 Heavy Texas、Rubber Jig、Frog 与中大型移动饵；具体大型饵排序低于底操和 cover 用途。",
        "source": "official_short_description",
    },
    "DBTS-60XUL-S": {
        "pairing": "Down Shot / Neko Rig / Hover Strolling / No Sinker",
        "hint": "60XUL-S 是 10th ultimate finesse 的短尺 XUL solid，适合 Down Shot、Neko、Hover Strolling 和轻 No Sinker；强调极轻饵控制和短咬承接。",
        "source": "official_spec_image",
    },
    "DBTS-67UL-S": {
        "pairing": "Light Carolina Rig / Long Leader Down Shot / No Sinker / I-shaped Plug",
        "hint": "67UL-S 是 10th 长尺 finesse，Light Carolina 与长 leader Down Shot 放首位，No Sinker 和 I-shaped Plug 用于需要距离和弱波动搜索的场景。",
        "source": "official_spec_image",
    },
    "DBTS-60UL-S": {
        "pairing": "Down Shot / Neko Rig / No Sinker / Small Rubber Jig",
        "hint": "60UL-S 是 super finesse，适合高压短距用 Down Shot、Neko、No Sinker 和 Small Rubber Jig 抢一口；重点是轻量操控和不放过短咬。",
        "source": "official_short_description",
    },
    "DBTS-61UL+-S": {
        "pairing": "Down Shot / Neko Rig / Jighead Rig / No Sinker",
        "hint": "61UL+-S 是进化型 super finesse，UL+ solid 更适合轻量 Down Shot、Neko、Jighead 与 No Sinker；比普通 L power 更重视细节反馈。",
        "source": "official_short_description",
    },
    "DBTS-61L": {
        "pairing": "Neko Rig / Down Shot / No Sinker / Small Rubber Jig / Small Plug",
        "hint": "61L 是 super all-round spinning，但白名单实钓明确支持 Neko；因此以 Neko、Down Shot、No Sinker 和 Small Rubber Jig 为主，小型 Plug 作为轻量搜索补充。",
        "source": "official_plus_whitelist",
    },
    "DBTS-63UL-MIDSP": {
        "pairing": "Mid Strolling Jighead / Hover Strolling / Down Shot / No Sinker",
        "hint": "63UL-MIDSP 明确是 mid-strolling rod，Jighead/Hover Strolling 优先，Down Shot 和 No Sinker 只作为同样需要线弧与轻抖的补充。",
        "source": "official_short_description",
    },
    "DBTS-65L+": {
        "pairing": "Neko Rig / Down Shot / Small Minnow / No Sinker",
        "hint": "65L+ 是兼具 long distance 和细致度的 light versatile，Neko、Down Shot、No Sinker 负责食わせ，Small Minnow 负责远处搜索。",
        "source": "official_short_description",
    },
    "DBTS-66ML-S-MIDSP": {
        "pairing": "Mid Strolling Jighead / No Sinker / Hover Strolling / Small Plug",
        "hint": "66ML-S-MIDSP 以 mid-strolling 为主，白名单 No Sinker 实例可补强中层软饵操作；Small Plug 是远投轻搜索，不应盖过 mid-strolling 主轴。",
        "source": "official_plus_whitelist",
    },
    "DBTS-66M": {
        "pairing": "Power Finesse / Cover Neko / Small Rubber Jig / No Sinker",
        "hint": "66M 是 power finesse spin，重点是 PE 下的 Cover Neko、Small Rubber Jig 和偏重 No Sinker，把轻量精细推进到 cover 与远投场景。",
        "source": "official_short_description",
    },
    "DBTS-68H-S-PF": {
        "pairing": "Power Finesse / Cover Neko / Small Rubber Jig / High-density No Sinker",
        "hint": "68H-S-PF 明确面向高压 cover power finesse，H power solid 负责把鱼从 cover 中带出；Cover Neko、Small Rubber Jig 和高比重 No Sinker 是主要用途。",
        "source": "official_short_description",
    },
    "DBTS-610L-S": {
        "pairing": "Long Cast Neko Rig / I-shaped Plug / Surface Plug / Small Rubber Jig",
        "hint": "610L-S 是 long shooting special，白名单和描述都支持远投 light finesse 与表层/I-shaped plug；Neko 远投和 Small Rubber Jig 放在精细端。",
        "source": "official_plus_whitelist",
    },
    "DBTC-64ML-FM": {
        "pairing": "Crankbait / Shad / Minnow / Topwater Plug",
        "hint": "64ML-FM 是 full-glass fast-moving 专用，Crankbait、Shad、Minnow 和 Topwater Plug 的连收、碰障碍与短咬吸收是核心，不应写成软饵底操。",
        "source": "official_short_description",
    },
    "DBTC-65ML-FM": {
        "pairing": "Minnow / Shad / Crankbait / Topwater Plug",
        "hint": "65ML-FM 是低弹性碳布 light-plugging special，适合 Minnow、Shad、轻量 Crankbait 和 Topwater 的控速与停顿；不是普通软硬泛用。",
        "source": "official_short_description",
    },
    "DBTC-65M+-PF": {
        "pairing": "Bait Power Finesse / Small Rubber Jig / Bait Neko / Heavy Down Shot / Free Rig / Texas Rig / Football Jig",
        "hint": "65M+-PF 是 light-cover bait power finesse，PE 时以 Small Rubber Jig、Bait Neko 为主，换 fluorocarbon 后可转 Heavy Down Shot、Free Rig、Texas 和 Football。",
        "source": "official_short_description",
    },
    "DBTC-67ML-BF": {
        "pairing": "Bait Finesse Small Plug / Small Rubber Jig / Neko Rig / Down Shot / Small Spinnerbait",
        "hint": "67ML-BF 是进攻型 bait finesse，既能投 Small Plug 和小型 Spinnerbait 搜索，也能用 Small Rubber Jig、Neko、Down Shot 做精准软饵操作。",
        "source": "official_short_description",
    },
    "DBTC-68M": {
        "pairing": "Crankbait / Shad / Jerkbait / Spinnerbait / Chatterbait / Texas Rig",
        "hint": "68M 官方为 super versatile，白名单实钓强化 Crankbait/Shad/Jerkbait 搜索端；Spinnerbait/Chatterbait 次之，Texas 作为软饵补充。",
        "source": "official_plus_whitelist",
    },
    "DBTC-610MH": {
        "pairing": "Chatterbait / Spinnerbait / Texas Rig / Rubber Jig / Crankbait",
        "hint": "610MH 官方是 middle versatile，白名单明确 bladed jig 实例，因此 Chatterbait 放首位；Spinnerbait 与 Crankbait 负责巻き，Texas/Rubber Jig 负责撃ち。",
        "source": "official_plus_whitelist",
    },
    "DBTC-70M+-FM": {
        "pairing": "Crankbait / Big Minnow / Topwater Plug / Spinnerbait / Buzzbait / Chatterbait / Shad Tail",
        "hint": "70M+-FM 是 glass-composite fast-moving 专用，官网明确列出 Crankbait、Big Minnow、中大型 Topwater、Spinnerbait、Buzzbait、Blade Jig 和 Shad Tail，全部属于巻物/移动饵体系。",
        "source": "official_description",
    },
    "DBTC-70H-S": {
        "pairing": "Cover Neko / Texas Rig / Rubber Jig / Heavy Down Shot",
        "hint": "70H-S 以 heavy cover finesse 为主题，Cover Neko 放首位，Texas、Rubber Jig 和 Heavy Down Shot 用于更重 cover 或更强穿透；不扩写到大型饵。",
        "source": "official_short_description",
    },
    "DBTC-71MH-PF": {
        "pairing": "Bait Power Finesse / Small Rubber Jig / Bait Neko / Frog / Heavy Down Shot / Free Rig / Texas Rig / Football Jig",
        "hint": "71MH-PF 官方主轴是 bait power finesse：PE 时 Small Rubber Jig 与 Bait Neko 优先；白名单 Frog 可作为 cover 补强，fluorocarbon 时再切 Heavy Down Shot、Free Rig、Texas、Football。",
        "source": "official_plus_whitelist",
    },
    "DBTC-73H": {
        "pairing": "Big Bait / Crawler Bait / Swimbait / Swim Jig / Heavy Texas / Flipping",
        "hint": "73H 是 heavy versatile，适合 Big Bait、Crawler、Swimbait 和 Swim Jig 这类强搜索，也能承担 Heavy Texas 与 Flipping；排序体现大型移动饵和 heavy cover 双线用途。",
        "source": "official_short_description",
    },
}


FORBIDDEN_GENERIC = ["General Lure", "Light Rig", "Hardbait", "Soft Bait"]


def normalize(value):
    return " ".join(str(value or "").split())


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)


def ensure_pairing_field(ws):
    headers = [cell.value for cell in ws[1]]
    if PAIRING_FIELD in headers:
        return False
    if PAIRING_AFTER_FIELD not in headers:
        raise RuntimeError(f"missing anchor column: {PAIRING_AFTER_FIELD}")
    insert_at = headers.index(PAIRING_AFTER_FIELD) + 2
    ws.insert_cols(insert_at)
    for row in range(1, ws.max_row + 1):
        copy_cell_style(ws.cell(row=row, column=insert_at - 1), ws.cell(row=row, column=insert_at))
    ws.cell(row=1, column=insert_at).value = PAIRING_FIELD
    return True


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    inserted = ensure_pairing_field(ws)
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    required = ["id", "rod_id", "SKU", "Description", HINT_FIELD, PAIRING_FIELD]
    missing = [name for name in required if name not in col]
    if missing:
        raise RuntimeError(f"missing rod_detail columns: {missing}")

    rows = []
    missing_skus = []
    source_counts = {}
    forbidden_residuals = []
    for row in range(2, ws.max_row + 1):
        sku = normalize(ws.cell(row=row, column=col["SKU"]).value)
        item = USAGE_BY_SKU.get(sku)
        detail_id = normalize(ws.cell(row=row, column=col["id"]).value)
        rod_id = normalize(ws.cell(row=row, column=col["rod_id"]).value)
        if not item:
            missing_skus.append({"xlsx_row": row, "id": detail_id, "rod_id": rod_id, "SKU": sku})
            continue

        old_hint = normalize(ws.cell(row=row, column=col[HINT_FIELD]).value)
        old_pairing = normalize(ws.cell(row=row, column=col[PAIRING_FIELD]).value)
        ws.cell(row=row, column=col[HINT_FIELD]).value = item["hint"]
        ws.cell(row=row, column=col[PAIRING_FIELD]).value = item["pairing"]
        source_counts[item["source"]] = source_counts.get(item["source"], 0) + 1
        for field_name, value in [(HINT_FIELD, item["hint"]), (PAIRING_FIELD, item["pairing"])]:
            hits = [term for term in FORBIDDEN_GENERIC if term.lower() in value.lower()]
            if hits:
                forbidden_residuals.append({
                    "xlsx_row": row,
                    "id": detail_id,
                    "SKU": sku,
                    "field": field_name,
                    "terms": hits,
                    "value": value,
                })
        rows.append({
            "xlsx_row": row,
            "id": detail_id,
            "rod_id": rod_id,
            "SKU": sku,
            "source": item["source"],
            "old_guide_use_hint": old_hint,
            "new_guide_use_hint": item["hint"],
            "old_recommended_rig_pairing": old_pairing,
            "new_recommended_rig_pairing": item["pairing"],
            "guide_use_hint_changed": old_hint != item["hint"],
            "recommended_rig_pairing_changed": old_pairing != item["pairing"],
        })

    if missing_skus:
        raise RuntimeError(f"missing usage mappings: {missing_skus}")
    if forbidden_residuals:
        raise RuntimeError(f"forbidden generic terms remain: {forbidden_residuals}")

    wb.save(XLSX_PATH)

    report = {
        "schema": "dstyle_rod_usage_fields_stage2",
        "source_xlsx": str(XLSX_PATH),
        "inserted_recommended_rig_pairing_column": inserted,
        "total_detail_rows": ws.max_row - 1,
        "updated_rows": len(rows),
        "guide_use_hint_changed_count": sum(1 for row in rows if row["guide_use_hint_changed"]),
        "recommended_rig_pairing_changed_count": sum(1 for row in rows if row["recommended_rig_pairing_changed"]),
        "source_counts": source_counts,
        "forbidden_generic_terms": FORBIDDEN_GENERIC,
        "forbidden_residuals": forbidden_residuals,
        "rows": rows,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "xlsx": str(XLSX_PATH),
        "report": str(REPORT_PATH),
        "updated_rows": report["updated_rows"],
        "guide_use_hint_changed_count": report["guide_use_hint_changed_count"],
        "recommended_rig_pairing_changed_count": report["recommended_rig_pairing_changed_count"],
        "source_counts": source_counts,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
