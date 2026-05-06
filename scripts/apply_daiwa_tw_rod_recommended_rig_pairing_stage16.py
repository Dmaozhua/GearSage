import json
import re
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from openpyxl import load_workbook


XLSX_PATH = resolve_data_raw('daiwa_rod_import.xlsx')
SNAPSHOT_PATH = Path("/tmp/daiwa_rod_before_rig_whitelist_pass.json")
FIELD = "recommended_rig_pairing"


def n(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def lower(value):
    return n(value).lower()


def has_any(text, *patterns):
    return any(re.search(pattern, text, re.I) for pattern in patterns)


def join(items):
    out = []
    for item in items:
        if item and item not in out:
            out.append(item)
    return " / ".join(out[:6])


def lure_oz_max(value):
    text = lower(value).replace("〜", "~").replace("～", "~")
    if not text:
        return None
    fractions = re.findall(r"(\d+)\s*/\s*(\d+)", text)
    nums = []
    for a, b in fractions:
        den = int(b)
        if den:
            nums.append(int(a) / den)
    decimals = re.findall(r"(?<!/)\b\d+(?:\.\d+)?\b(?!\s*/)", text)
    nums.extend(float(x) for x in decimals)
    return max(nums) if nums else None


def bass_pairing(row, model):
    sku = n(row["SKU"])
    text = lower(f"{model} {sku} {row.get('Description')} {row.get('player_positioning')} {row.get('player_selling_points')}")
    sku_text = lower(sku)
    rod_type = n(row.get("TYPE"))
    power = lower(row.get("POWER"))
    action = lower(row.get("Action"))
    max_oz = lure_oz_max(row.get("LURE WEIGHT (oz)")) or None

    if has_any(text, "frog", r"\bfr\b"):
        return join(["Frog", "Punching", "Heavy Texas"])
    if has_any(text, "big bait", "swim bait", "swimbait", "大型餌", "大餌", "sb"):
        return join(["Swimbait", "Big Bait", "A-rig", "Wakebait"])

    if "tatula" in lower(model):
        if has_any(sku_text, "661mlrb", "662mlrb"):
            return join(["Soft Plastic", "Topwater Plug", "Crankbait", "Small Hardbait"])
        if has_any(sku_text, "641lfs", "642lfs", "644lfs"):
            return join(["Jighead Rig", "Down Shot", "Neko Rig", "Small Hardbait", "I-shaped Plug"])
        if has_any(sku_text, "681mlfs", "682mlfs"):
            return join(["No Sinker", "Down Shot", "Light Texas", "Small Crankbait", "Soft Plastic"])
        if has_any(sku_text, "6011ulxs", "6012ulxs", "631ulfs", "632ulfs"):
            return join(["Down Shot", "Neko Rig", "No Sinker", "Small Rubber Jig", "Jighead Rig"])

    if rod_type == "S":
        if power in {"xul", "ul"} or (max_oz is not None and max_oz <= 0.15):
            return join(["Jighead Rig", "No Sinker", "Wacky Rig", "Bug Lure", "I-shaped Plug"])
        if power == "l" or "solid" in text or "-st" in sku_text or "sv-st" in sku_text:
            return join(["Jighead Rig", "Down Shot", "Neko Rig", "Wacky Rig", "I-shaped Plug"])
        if power in {"ml", "ml+"} or (max_oz is not None and max_oz <= 0.5):
            return join(["No Sinker", "Down Shot", "Light Texas", "Small Crankbait", "Soft Plastic"])
        return join(["No Sinker", "Jighead Rig", "Minnow", "Small Crankbait"])

    if has_any(sku_text, "c66ml-g", "c66ml-lm"):
        return join(["Crankbait", "Shad", "Chatterbait", "Minnow", "Soft Plastic"])
    if power in {"ul", "l"} or "bf" in sku_text:
        return join(["Neko Rig", "Down Shot", "No Sinker", "Small Rubber Jig", "Bug Lure"])
    if has_any(sku_text, "st") or action in {"xf", "xf-s"}:
        return join(["Texas Rig", "Free Rig", "Rubber Jig", "Down Shot", "Soft Plastic"])
    if power in {"ml", "ml+"} or (max_oz is not None and max_oz <= 0.55):
        return join(["Crankbait", "Shad", "Minnow", "Down Shot", "No Sinker"])
    if power in {"m", "m/ML".lower()} or (max_oz is not None and max_oz <= 0.8):
        return join(["Texas Rig", "Free Rig", "Rubber Jig", "Spinnerbait", "Chatterbait"])
    if power in {"m+", "mh", "h", "xh", "xxh"} or (max_oz is not None and max_oz >= 0.85):
        return join(["Texas Rig", "Rubber Jig", "Swim Jig", "Spinnerbait", "Frog"])
    return join(["Texas Rig", "Crankbait", "Spinnerbait", "Down Shot"])


def salt_pairing(row, model):
    sku = n(row["SKU"])
    text = lower(f"{model} {sku} {row.get('Description')} {row.get('player_positioning')} {row.get('player_selling_points')}")
    model_text = lower(model)

    if has_any(model_text, "morethan", "lateo"):
        return join(["Minnow", "Sinking Pencil", "Vibration", "Seabass Plug"])
    if has_any(model_text, "over there", "overthere"):
        if has_any(text, "heavy plug", "重餌", "plug"):
            return join(["Heavy Plug", "Sinking Pencil", "Metal Jig", "Surf Plug"])
        return join(["Metal Jig", "Minnow", "Sinking Pencil", "Surf Plug"])
    if has_any(model_text, "dragger"):
        if has_any(text, "slsj", "小型"):
            return join(["SLSJ", "Small Metal Jig", "Minnow", "Sinking Pencil"])
        return join(["Metal Jig", "Minnow", "Sinking Pencil", "Surf Plug"])
    if has_any(model_text, "crossbeat") and has_any(text, "wind", "木蝦"):
        return join(["Eging", "WIND", "Small Metal Jig"])
    if has_any(model_text, "crossbeat") and has_any(text, "金屬", "jig", "青背", "比目魚"):
        return join(["Small Metal Jig", "Minnow", "Surf Plug"])
    if has_any(model_text, "emeraldas", "eging"):
        if has_any(text, "tip"):
            return join(["Tip-run Eging", "Eging"])
        if has_any(text, "mlm", "mmh", "lml", "sink", "仮面"):
            return join(["Eging", "Sinker Rig"])
        return join(["Eging", "Shallow Eging", "Sinker Rig"])
    if has_any(model_text, "lowresponse"):
        return join(["Slow Pitch Jig", "Metal Jig"])
    if has_any(model_text, "saltiga j"):
        return join(["Offshore Jigging", "One Pitch Jig", "Metal Jig"])
    if has_any(model_text, "outrage xv sj"):
        return join(["Slow Jigging", "Metal Jig"])
    if has_any(model_text, "outrage xv lj"):
        return join(["Light Jigging", "Metal Jig"])
    if has_any(model_text, "outrage xv j"):
        return join(["Offshore Jigging", "Metal Jig"])
    if has_any(model_text, "outrage sj"):
        return join(["Slow Pitch Jig", "Metal Jig"])
    if has_any(model_text, "鏡牙"):
        return join(["Tachiuo Jigging", "Metal Jig"])
    if has_any(model_text, "outrage br") and has_any(text, "plug"):
        return join(["Diving Pencil", "Popper", "Stickbait", "Offshore Plug"])
    if has_any(model_text, "outrage br") and has_any(text, "鐵板", "jig"):
        return join(["Offshore Jigging", "Metal Jig"])
    if has_any(model_text, "saltiga c"):
        return join(["Diving Pencil", "Popper", "Stickbait", "Offshore Plug"])
    if has_any(model_text, "hardrock"):
        return join(["Texas Rig", "Jighead Rig", "Free Rig", "Metal Jig", "Rockfish Rig"])
    return ""


def refine_pairing(row, model):
    current = n(row.get(FIELD))
    text = lower(f"{model} {row.get('SKU')} {row.get('Description')} {row.get('player_positioning')} {row.get('player_selling_points')}")
    model_text = lower(model)

    if has_any(model_text, "tatula", "steez", "black label", "heartland", "wilderness", "vertice", "crossfire"):
        refined = bass_pairing(row, model)
    else:
        refined = salt_pairing(row, model)

    if not refined:
        return current

    too_generic = current in {
        "Light Rig",
        "Light Rig / Small Hardbait",
        "Texas Rig / Crankbait / Spinnerbait / Down Shot",
        "Metal Jig",
        "Plug",
        "Eging",
        "Big Bait",
    }
    if too_generic:
        return refined

    if "Soft Plastic" not in current and has_any(text, "軟餌", "軟蟲", "worm", "ワーム"):
        return join(current.split(" / ") + ["Soft Plastic"])
    return current


def snapshot_workbook(ws, field_col):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        data = {}
        for col, header in enumerate(headers, start=1):
            if col == field_col:
                continue
            data[header] = ws.cell(row, col).value
        rows.append(data)
    return {"headers": [h for i, h in enumerate(headers, start=1) if i != field_col], "rows": rows}


def main():
    wb = load_workbook(XLSX_PATH)
    rod_ws = wb["rod"]
    detail_ws = wb["rod_detail"]

    rod_headers = [cell.value for cell in rod_ws[1]]
    rod_col = {name: i + 1 for i, name in enumerate(rod_headers)}
    models = {
        n(rod_ws.cell(row, rod_col["id"]).value): n(rod_ws.cell(row, rod_col["model"]).value)
        for row in range(2, rod_ws.max_row + 1)
    }

    headers = [cell.value for cell in detail_ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    if FIELD not in col:
        raise RuntimeError(f"missing {FIELD}")

    if not SNAPSHOT_PATH.exists():
        SNAPSHOT_PATH.write_text(json.dumps(snapshot_workbook(detail_ws, col[FIELD]), ensure_ascii=False, indent=2))

    changed = []
    required = ["id", "rod_id", "TYPE", "SKU", "POWER", "Action", "LURE WEIGHT", "LURE WEIGHT (oz)", "Description", "player_positioning", "player_selling_points", FIELD]
    missing = [name for name in required if name not in col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    for row_idx in range(2, detail_ws.max_row + 1):
        row = {name: detail_ws.cell(row_idx, col[name]).value for name in required}
        model = models.get(n(row["rod_id"]), "")
        old = n(row[FIELD])
        new = refine_pairing(row, model)
        if new and new != old:
            detail_ws.cell(row_idx, col[FIELD]).value = new
            changed.append({
                "row": row_idx,
                "id": n(row["id"]),
                "rod_id": n(row["rod_id"]),
                "sku": n(row["SKU"]),
                "old": old,
                "new": new,
            })

    wb.save(XLSX_PATH)
    print(json.dumps({"changed": len(changed), "samples": changed[:30]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
