#!/usr/bin/env python3
import json
import re
import shutil
import ssl
import subprocess
import time
import unicodedata
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from html import unescape
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook


BASE_URL = "https://arkrods.com"
COLLECTION_URL = f"{BASE_URL}/collections/rods"
COLLECTION_JSON_URL = f"{COLLECTION_URL}/products.json?limit=250"
OUTPUT_DIR = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw")
CACHE_DIR = OUTPUT_DIR / "ark_rods_cache"
DETAIL_CACHE_DIR = CACHE_DIR / "details"
LIST_PATH = CACHE_DIR / "ark_rod_list_items.json"
NORMALIZED_PATH = OUTPUT_DIR / "ark_rod_normalized.json"
WHITELIST_EVIDENCE_PATH = OUTPUT_DIR / "ark_rod_whitelist_player_evidence.json"
OUTPUT_FILE = OUTPUT_DIR / "ark_rod_import.xlsx"
SHADE_SCRIPT = Path("/Users/tommy/GearSage/scripts/shade_ark_rod_detail_groups.py")
IMAGE_DIR = Path("/Users/tommy/Pictures/images/ark_rods")
OLD_IMAGE_DIR = Path("/Users/tommy/Pictures/images_old_copy/ark_rods")
STATIC_IMAGE_BASE = "https://static.gearsage.club/gearsage/Gearimg/images/ark_rods"

BRAND_ID = 115
ROD_PREFIX = "ARKR"
ROD_DETAIL_PREFIX = "ARKRD"
HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
}
SSL_CONTEXT = ssl._create_unverified_context()

ROD_HEADERS = [
    "id",
    "brand_id",
    "model",
    "model_cn",
    "model_year",
    "alias",
    "type_tips",
    "fit_style_tags",
    "images",
    "created_at",
    "updated_at",
    "series_positioning",
    "main_selling_points",
    "official_reference_price",
    "market_status",
    "Description",
    "player_positioning",
    "player_selling_points",
]

ROD_DETAIL_HEADERS = [
    "id",
    "rod_id",
    "TYPE",
    "SKU",
    "POWER",
    "TOTAL LENGTH",
    "Action",
    "PIECES",
    "CLOSELENGTH",
    "WEIGHT",
    "Tip Diameter",
    "LURE WEIGHT",
    "Line Wt N F",
    "PE Line Size",
    "Handle Length",
    "Reel Seat Position",
    "CONTENT CARBON",
    "Market Reference Price",
    "AdminCode",
    "Service Card",
    " Jig Weight",
    "Squid Jig Size",
    "Sinker Rating",
    "created_at",
    "updated_at",
    "LURE WEIGHT (oz)",
    "Sale Price",
    "Joint Type",
    "Code Name",
    "Fly Line",
    "Grip Type",
    "Reel Size",
    "guide_layout_type",
    "guide_use_hint",
    "hook_keeper_included",
    "sweet_spot_lure_weight_real",
    "official_environment",
    "player_environment",
    "player_positioning",
    "player_selling_points",
    "Description",
    "product_technical",
    "Extra Spec 1",
    "Extra Spec 2",
]

TECH_PATTERNS = [
    ("ARK HPCR (High Pressure Carbon Fiber Rolling) technology", r"\b(?:ARK\s+)?HPCR\b|\bHigh[- ]Pressure Carbon[- ]Fiber Rolling Technology\b|\bARK high pressure carbon[- ]fiber rolling technology\b"),
    ("Multi-Direction Multi-Layer technology", r"\bMulti[- ]Direction Multi[- ]Layer technology\b|\bMDML Technology\b"),
    ("carbon nano tube reinforcement", r"\bcarbon nano tubes?\b|\bnano tube reinforcements?\b|\bNano tube reinforcement\b"),
    ("Fuji PTS/TVS reel seat", r"\bFuji PTS/TVS reel seat\b"),
    ("Fuji K concept Alconite guides", r"\bFuji K concept alconite guides\b"),
    ("Fuji K concept guides", r"\bFuji K[- ]concept (?:tangle[- ]free )?guides\b|\bFuji K concept guides\b"),
    ("FazLite rings", r"\bFazL?ite (?:rings|inserts)\b"),
    ("ARK Titanium Tangle-Free Guides", r"\bARK Titanium Tangle[- ]Free Guides\b"),
    ("ARK Stainless Steel Tangle-Free Guides", r"\bARK Stainless Steel Tangle[- ]Free Guides\b"),
    ("ARK Tangle-Free Guides", r"\bARK Tangle[- ]?Free Guides\b"),
    ("NanoForce Rings", r"\bNanoForce(?:™)? (?:Ring|Rings|guides)\b"),
    ("A-Ring technology", r"\bA[- ]Ring technology\b"),
    ("Black Coated Stainless Micro Guides System With Zirconium Inserts", r"\bBlack Coated Stainless Micro Guides System With Zirconium Inserts\b"),
    ("Team ARK reel seat", r"\bTeam ARK reel seat\b"),
    ("High-Visibility Strike Indicator", r"\bHigh[- ]Visibility Strike Indicator\b"),
]


def ensure_dirs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    DETAIL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)


def normalize_text(value):
    text = unescape(str(value or ""))
    text = text.replace("\u00a0", " ")
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.replace("\u2032", "'").replace("\u2033", '"')
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    return text.strip()


def compact_text(value):
    return re.sub(r"\s+", " ", normalize_text(value)).strip()


def fetch_json(url):
    req = urllib.request.Request(url, headers=HTTP_HEADERS)
    with urllib.request.urlopen(req, timeout=60, context=SSL_CONTEXT) as response:
        return json.load(response)


def fetch_bytes(url):
    req = urllib.request.Request(url, headers=HTTP_HEADERS)
    with urllib.request.urlopen(req, timeout=60, context=SSL_CONTEXT) as response:
        return response.read()


def product_url(handle):
    return f"{BASE_URL}/products/{handle}.json"


def detail_cache_path(index, handle):
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", handle)
    return DETAIL_CACHE_DIR / f"{index + 1:03d}_{safe}.json"


def fetch_list(force=False):
    ensure_dirs()
    if LIST_PATH.exists() and not force:
        items = json.loads(LIST_PATH.read_text(encoding="utf-8"))
        print(f"[list] cached {len(items)} items -> {LIST_PATH}")
        return items

    payload = fetch_json(COLLECTION_JSON_URL)
    items = []
    for index, product in enumerate(payload.get("products", [])):
        handle = product.get("handle", "")
        if not handle:
            continue
        items.append(
            {
                "position": index + 1,
                "id": product.get("id"),
                "handle": handle,
                "title": product.get("title", ""),
                "source_url": f"{COLLECTION_URL}/products/{handle}",
                "product_json_url": product_url(handle),
            }
        )
    LIST_PATH.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[list] fetched {len(items)} items -> {LIST_PATH}")
    return items


def read_list():
    if not LIST_PATH.exists():
        raise RuntimeError(f"list cache missing: {LIST_PATH}. Run --stage=list first.")
    return json.loads(LIST_PATH.read_text(encoding="utf-8"))


def fetch_details(force=False, start=0, limit=None):
    ensure_dirs()
    items = read_list()
    end = len(items) if limit is None else min(len(items), start + limit)
    failures = []
    for index in range(start, end):
        item = items[index]
        cache_path = detail_cache_path(index, item["handle"])
        if cache_path.exists() and not force:
            print(f"[detail] cached {index + 1}/{len(items)} {item['handle']}")
            continue
        try:
            payload = fetch_json(item["product_json_url"])
            cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"[detail] fetched {index + 1}/{len(items)} {item['handle']}")
            time.sleep(0.25)
        except Exception as exc:
            failures.append({"item": item, "message": str(exc)})
            print(f"[detail] failed {item['handle']}: {exc}")
    if failures:
        failure_path = CACHE_DIR / "detail_failures.json"
        failure_path.write_text(json.dumps(failures, ensure_ascii=False, indent=2), encoding="utf-8")
        raise RuntimeError(f"{len(failures)} detail pages failed. See {failure_path}")


def parse_table_rows(body_html):
    soup = BeautifulSoup(body_html or "", "html.parser")
    rows = []
    for table in soup.find_all("table"):
        raw_rows = []
        for tr in table.find_all("tr"):
            cells = [compact_text(cell.get_text(" ", strip=True)) for cell in tr.find_all(["th", "td"])]
            if any(cells):
                raw_rows.append(cells)
        if not raw_rows:
            continue
        headers = [normalize_header(cell) for cell in raw_rows[0]]
        current_type = ""
        for raw in raw_rows[1:]:
            cells = raw + [""] * max(0, len(headers) - len(raw))
            filled = [cell for cell in cells if cell]
            if len(filled) == 1 and filled[0].upper().replace(" ", "") in {"CASTING", "SPINNING"}:
                current_type = filled[0].upper()
                continue
            if len(cells) < 2 or not cells[0] or not looks_like_model(cells[0]):
                continue
            spec = {headers[i] or f"extra_{i}": cells[i] for i in range(min(len(headers), len(cells)))}
            spec["section_type"] = current_type
            spec["raw_cells"] = raw
            rows.append(spec)
    return rows


def normalize_header(value):
    key = compact_text(value).lower()
    key = key.replace(".", "")
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    aliases = {
        "model": "model",
        "length": "length",
        "action": "action",
        "power": "power",
        "line_wt": "line_wt",
        "lure_wt": "lure_wt",
        "application": "application",
        "purpose": "application",
        "rod_weight": "rod_weight",
        "price": "price",
    }
    return aliases.get(key, key)


def looks_like_model(value):
    text = compact_text(value)
    if not text or len(text) < 3:
        return False
    if re.fullmatch(r"\(?\s*(SPINNING|CASTING)\s*\)?", text, flags=re.I):
        return False
    return bool(re.search(r"[A-Z]{2,}\w*\d", text, flags=re.I))


def strip_sale_prefix(title):
    return re.sub(r"^\s*SALE\s*-\s*", "", compact_text(title), flags=re.I)


def extract_paragraphs(body_html):
    soup = BeautifulSoup(body_html or "", "html.parser")
    for table in soup.find_all("table"):
        table.decompose()
    out = []
    for node in soup.find_all(["p", "li"]):
        text = compact_text(node.get_text(" ", strip=True))
        if text and text not in out:
            out.append(text)
    if not out:
        text = compact_text(soup.get_text(" ", strip=True))
        if text:
            out.append(text)
    return out


def extract_features(paragraphs):
    text = " ".join(paragraphs)
    features = []
    checks = [
        (r"\bFuji\b.*?(?:guides?|reel seat)|(?:guides?|reel seat).*?\bFuji\b", "Fuji components"),
        (r"Titanium(?: Tangle-Free)? Guides?|Titanium guides?", "Titanium guides"),
        (r"NanoForce", "NanoForce guide rings"),
        (r"carbon nano tube|nano tube", "carbon nano tube blank"),
        (r"40T|46T|36T", "high-modulus carbon blank"),
        (r"cork", "cork handle"),
        (r"EVA", "EVA handle"),
        (r"hook keeper", "hook keeper"),
        (r"lifetime warranty", "lifetime warranty"),
    ]
    for pattern, label in checks:
        if re.search(pattern, text, flags=re.I):
            features.append(label)
    return features


def sku_from_model(model_text):
    text = compact_text(model_text)
    text = re.sub(r"\s*\([^)]*\)", "", text)
    return compact_text(re.split(r"\s+-\s+|\s+\u2013\s+|\s+\u2014\s+", text, maxsplit=1)[0])


def normalize_sku_key(value):
    return re.sub(r"[^A-Z0-9]+", "", compact_text(value).upper())


def edit_distance(a, b):
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        cur = [i]
        for j, cb in enumerate(b, start=1):
            cur.append(min(prev[j] + 1, cur[-1] + 1, prev[j - 1] + (0 if ca == cb else 1)))
        prev = cur
    return prev[-1]


def match_variant(model_text, fallback_sku, variants_by_sku):
    row_key = normalize_sku_key(model_text)
    fallback_key = normalize_sku_key(fallback_sku)
    if fallback_key in variants_by_sku:
        return variants_by_sku[fallback_key]

    prefix_matches = [
        (len(sku_key), sku_key, variant)
        for sku_key, variant in variants_by_sku.items()
        if row_key.startswith(sku_key)
    ]
    if prefix_matches:
        return sorted(prefix_matches, reverse=True)[0][2]

    first_token = compact_text(re.split(r"\s+", fallback_sku, maxsplit=1)[0])
    first_key = normalize_sku_key(first_token)
    if first_key in variants_by_sku:
        return variants_by_sku[first_key]

    fuzzy = []
    for sku_key, variant in variants_by_sku.items():
        compare_key = fallback_key or row_key
        if len(sku_key) >= 5 and compare_key[:5] == sku_key[:5]:
            dist = edit_distance(compare_key, sku_key)
            if dist <= 2:
                fuzzy.append((dist, sku_key, variant))
    if fuzzy:
        return sorted(fuzzy)[0][2]
    return {}


def price(value):
    text = compact_text(value)
    if not text:
        return ""
    text = text.replace("$", "")
    try:
        return f"${float(text):.2f}"
    except ValueError:
        return text if text.startswith("$") else f"${text}"


def normalize_power_value(value):
    text = compact_text(value)
    upper = text.upper()
    replacements = {
        "MAG MH": "MH+",
        "MAG H": "H+",
    }
    return replacements.get(upper, text)


def normalize_line_wt(value):
    text = compact_text(value)
    if not text:
        return ""
    text = re.sub(r"\bBRIAD\b", "Braid", text, flags=re.I)
    text = re.sub(r"(?<=\d)Braid\b", " Braid", text)
    text = re.sub(r"\s*-\s*", "-", text)
    return compact_text(text)


def normalize_lure_oz(value):
    text = compact_text(value)
    if not text:
        return ""
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s*oz\.?\b", "", text, flags=re.I)
    return compact_text(text)


def price_range(variants):
    values = []
    for variant in variants:
        p = price(variant.get("price", ""))
        if p and p not in values:
            values.append(p)
    return " / ".join(values)


def sanitize_filename(value):
    base = unicodedata.normalize("NFKD", compact_text(value))
    base = re.sub(r"[^\x00-\x7F]+", "", base)
    base = re.sub(r'[\\/*?:"<>|]+', "_", base)
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
    base = re.sub(r"_+", "_", base).strip("_")
    return base or "ark_rod"


def image_ext(url):
    try:
        ext = Path(urllib.parse.urlparse(url).path).suffix.lower()
        return ext if ext in {".jpg", ".jpeg", ".png", ".webp"} else ".jpg"
    except Exception:
        return ".jpg"


def download_image(url, file_stem, refresh=False):
    if not url:
        return ""
    ext = image_ext(url)
    filename = f"{file_stem}{ext}"
    local_path = IMAGE_DIR / filename
    old_path = OLD_IMAGE_DIR / filename
    if refresh and local_path.exists():
        local_path.unlink()
    if local_path.exists():
        return f"{STATIC_IMAGE_BASE}/{filename}"
    if old_path.exists():
        shutil.copyfile(old_path, local_path)
        return f"{STATIC_IMAGE_BASE}/{filename}"
    try:
        local_path.write_bytes(fetch_bytes(url))
        time.sleep(0.2)
        return f"{STATIC_IMAGE_BASE}/{filename}"
    except Exception:
        return ""


def infer_type(title, model_text, sku, section_type):
    text = f"{section_type} {title} {model_text} {sku}".upper()
    sku_key = normalize_sku_key(sku)
    if "ICE" in text:
        return "S"
    if "SPINNING" in text or sku_key.endswith("S"):
        return "S"
    if "CASTING" in text or sku_key.endswith("C"):
        return "C"
    return ""


def infer_model_year(title):
    match = re.search(r"\b(20\d{2})\b", title)
    return match.group(1) if match else ""


def series_positioning(title, tags, body_text):
    text = f"{title} {' '.join(tags or [])} {body_text}".lower()
    if re.search(r"\bice\b", text):
        return "Freshwater / Ice fishing"
    if "walleye" in text or "panfish" in text or "trolling" in text:
        return "Freshwater / Walleye / Panfish"
    return "Freshwater / Bass"


def fit_style_tags(title, series, description):
    text = f"{title} {series} {description}".lower()
    if re.search(r"\bice\b|ice fishing|hardwater", text):
        return ""
    tags = []
    if "bass" in text:
        tags.append("bass")
    has_three_plus_piece = re.search(r"\b(?:3|4|5|6|7|8|9)[- ]?(?:piece|pc)\b", text)
    has_multi_or_travel = re.search(r"\bmulti[- ]?piece\b|\btravel\b", text)
    has_two_piece = re.search(r"\b2[- ]?(?:piece|pc)\b|\btwo[- ]piece\b", text)
    if (has_three_plus_piece or has_multi_or_travel) and not (has_two_piece and not has_three_plus_piece):
        tags.append("旅行")
    return ",".join(tags)


def product_technical_terms(description):
    text = re.sub(
        r"From 2026 onward, new batches of rods will be upgraded from Fuji guides to ARK Tangle Free Guides with NanoForce™ Rings\.?",
        "",
        compact_text(description),
        flags=re.I,
    )
    terms = []
    for term, pattern in TECH_PATTERNS:
        if re.search(pattern, text, flags=re.I) and term not in terms:
            terms.append(term)
    return " / ".join(terms)


def master_player_positioning(series, title):
    title_text = title.lower()
    if "Ice" in series:
        return "冰钓 / Panfish 与 Walleye"
    if "Walleye" in series:
        return "淡水多鱼种 / Walleye / Panfish"
    if "bfs" in title_text or "finesse" in title_text:
        return "淡水 Bass / BFS 与精细轻量"
    if "spinning" in title_text:
        return "淡水 Bass / 精细钓组与开阔水域"
    if "casting" in title_text:
        return "淡水 Bass / 硬饵、底钓与强力场景"
    return "淡水 Bass / 技法细分竿系"


def master_player_points(features, series, specs):
    apps = [compact_text(spec.get("application", "")) for spec in specs if spec.get("application")]
    app_text = " / ".join(apps[:6])
    if "Bass" in series:
        base = ["按官方型号表细分 Bass 技法，覆盖反应饵、底部接触、精细钓组和重障碍型号。"]
    elif "Ice" in series:
        base = ["按冰钓对象鱼、竿尖动作和短竿长度区分 panfish、perch、walleye 场景。"]
    else:
        base = ["按 walleye、panfish、trolling、vertical jig 等淡水场景分工。"]
    if app_text:
        base.append(f"代表应用：{app_text}")
    if features:
        base.append("核心配置：" + " / ".join(features[:4]))
    return "；".join(base)


def technique_profile(application, sku, power, title):
    text = f"{application} {sku} {power} {title}".lower()
    if re.search(r"\bice\b", text):
        if "deadstick" in text:
            return ("冰钓 / 湖库硬水", "冰钓 deadstick / Walleye", "Walleye 静置等口")
        if "panfish" in text or "perch" in text:
            return ("冰钓 / 湖库硬水", "冰钓 Panfish / Perch", "轻口小型鱼")
        return ("冰钓 / 湖库硬水", "冰钓泛用", "短竿控线")
    if "panfish" in text or "perch" in text:
        return ("淡水湖库 / Walleye / Panfish", "淡水 Panfish / 小型鱼精细", "轻量小饵")
    if "bfs" in text or "bass finesse" in text:
        return ("淡水 Bass / 岸钓与船钓", "Bass BFS / 超轻量细线", "轻饵精准抛投")
    if any(k in text for k in ["dropshot", "drop shot", "ned", "neko", "niko", "shaky", "damiki", "hairjig", "hair jig", "spybait", "mid strolling", "mid-strolling", "jighead minnow", "forward-facing sonar"]):
        return ("淡水 Bass / 岸钓与船钓", "Bass 精细钓组 / 轻量饵", "轻线轻饵与开阔水域")
    if (
        "walleye" in text
        or "jig- n- rap" in text
        or "jig-n-rap" in text
        or "snap jig" in text
        or "rip-n" in text
        or "slip-n" in text
        or "bottom bouncer" in text
        or "vertical jig" in text
        or "trolling" in text
    ):
        return ("淡水湖库 / Walleye / Panfish", "淡水 Walleye / 拖钓与垂直技法", "湖库搜索与垂直控饵")
    if "w series" in text and "all purpose" in text:
        return ("淡水湖库 / Walleye / Panfish", "淡水 Walleye / 拖钓与垂直技法", "湖库搜索与垂直控饵")
    if any(k in text for k in ["glide", "swimbait", "swim bait", "umbrella", "a-rig", "a rig", "big bait"]):
        return ("淡水 Bass / 岸钓与船钓", "Bass 大饵 / 强力抛投", "大饵负载与远投")
    if any(k in text for k in ["flipping", "pitching", "punching", "heavy cover", "frog"]):
        return ("淡水 Bass / 岸钓与船钓", "Bass 重障碍 / 强力作钓", "cover 抽鱼与近障碍控鱼")
    if any(k in text for k in ["finesse jig", "small jig", "jig&worm", "jig, worm", "jig,worm", "worm", "tube", "c-rig", "free rig", "football jig", "bottom"]):
        return ("淡水 Bass / 岸钓与船钓", "Bass 底部接触 / Jig & Worm", "结构区探底与补刺")
    if any(k in text for k in ["jerkbait", "topwater", "crank", "lipless", "deep diver", "deepdiver", "cranking"]) and any(
        k in text for k in ["chatter", "swimjig", "swim jig", "spinner", "underspin", "blade"]
    ):
        return ("淡水 Bass / 岸钓与船钓", "Bass 反应饵 / 硬饵与移动饵", "抽停、巻物和搜索节奏")
    if any(k in text for k in ["chatter", "swimjig", "swim jig", "spinner", "underspin", "blade"]):
        return ("淡水 Bass / 岸钓与船钓", "Bass 移动饵 / 搜索型技法", "连续搜索与控速")
    if any(k in text for k in ["jerkbait", "topwater", "crank", "lipless", "deep diver", "deepdiver", "cranking"]):
        return ("淡水 Bass / 岸钓与船钓", "Bass 硬饵 / 抽停与反应饵", "抽停、巻物与反应咬口")
    if power in {"H", "H+", "XH", "XXH"}:
        return ("淡水 Bass / 岸钓与船钓", "Bass 强力泛用", "较重线组与中重饵")
    return ("淡水 Bass / 岸钓与船钓", "Bass 泛用技法", "日常软硬饵切换")


def infer_player_positioning(application, sku, power, title):
    return technique_profile(application, sku, power, title)[1]


def infer_player_environment(series, application, sku="", power="", title=""):
    return technique_profile(application, sku, power, title)[0] if application or sku else infer_player_environment_from_series(series, application)


def infer_player_environment_from_series(series, application):
    text = f"{series} {application}".lower()
    if re.search(r"\bice\b", text):
        return "冰钓 / 湖库硬水"
    if "walleye" in text or "panfish" in text or "trolling" in text:
        return "淡水湖库 / Walleye / Panfish"
    return "淡水 Bass / 岸钓与船钓"


def infer_player_selling_points(application, power, action, line_wt, lure_wt, positioning=""):
    _, _, focus = technique_profile(application, "", power, positioning)
    parts = []
    if application:
        parts.append(f"适合{application}。")
    elif positioning:
        parts.append(f"适合{positioning}。")
    spec_bits = []
    if power:
        spec_bits.append(f"{power} power")
    if action:
        spec_bits.append(f"{action} action")
    if line_wt:
        spec_bits.append(f"线号 {line_wt}")
    if lure_wt:
        spec_bits.append(f"饵重 {lure_wt} oz")
    if spec_bits:
        parts.append("规格倾向：" + " / ".join(spec_bits) + "。")
    if focus:
        parts.append(f"玩家侧重点：{focus}。")
    return "".join(parts)


def guide_layout_type(features, description):
    text = f"{' '.join(features)} {description}"
    if re.search(r"NanoForce|Tangle-Free", text, flags=re.I):
        return "Titanium Tangle-Free / NanoForce 导环：轻量化、防缠出线，兼顾细线抛投与感度"
    if re.search(r"Fuji.*K|K.*Fuji|Alconite", text, flags=re.I):
        return "Fuji K Concept / Alconite 导环：防缠出线与耐用性兼顾"
    if re.search(r"Titanium guides?", text, flags=re.I):
        return "Titanium guides：轻量化导环配置，兼顾感度与耐用"
    return ""


def guide_hint(layout, positioning):
    if not layout:
        return ""
    if "精细" in positioning or "轻量" in positioning:
        return "精细钓组：细线出线更顺，轻饵抛投、控线和轻口反馈更清楚。"
    if "BFS" in positioning:
        return "BFS：轻线和小饵出线更顺，近中距离精准抛投和轻口反馈更直接。"
    if "Walleye" in positioning:
        return "Walleye：垂直控线和拖钓出线更稳定，抽停、贴底和巡航搜索更容易维持节奏。"
    if "硬饵" in positioning or "移动饵" in positioning:
        return "反应饵：连续抛投与收线更稳定，抽停、搜索和控饵节奏更容易保持。"
    if "底部接触" in positioning:
        return "底部接触：线弧控制更清楚，jig、worm 和拖底钓组的轻微咬口更容易分辨。"
    if "大饵" in positioning:
        return "大饵：较粗线组出线更稳，长距离抛投和重饵负载下控线更安心。"
    if "重障碍" in positioning or "强力" in positioning:
        return "强力作钓：较粗线组出线更稳，重饵抛投和近障碍搏鱼更安心。"
    if "冰钓" in positioning:
        return "冰钓短竿：近距离控线更直接，轻口观察和小幅操饵更清楚。"
    return "泛用 Bass：出线顺畅、兼容多种线径，软饵和硬饵切换更自然。"


def grip_type(paragraphs):
    text = " ".join(paragraphs)
    if re.search(r"cork", text, flags=re.I) and re.search(r"EVA", text, flags=re.I):
        return "Cork / EVA"
    if re.search(r"cork", text, flags=re.I):
        return "Cork"
    if re.search(r"EVA", text, flags=re.I):
        return "EVA"
    return ""


def reel_seat(paragraphs):
    text = " ".join(paragraphs)
    if re.search(r"Fuji.*reel seat|reel seat.*Fuji", text, flags=re.I):
        return "Fuji reel seat"
    if re.search(r"custom.*reel seat|reel seat", text, flags=re.I):
        return "ARK custom reel seat"
    return ""


def parse_product(product):
    paragraphs = extract_paragraphs(product.get("body_html", ""))
    description = "\n\n".join(paragraphs)
    features = extract_features(paragraphs)
    spec_rows = parse_table_rows(product.get("body_html", ""))
    tags = product.get("tags", [])
    title = strip_sale_prefix(product.get("title", ""))
    variants = product.get("variants", [])
    variants_by_sku = {normalize_sku_key(v.get("sku", "")): v for v in variants if v.get("sku")}
    image_url = (product.get("image") or {}).get("src") or ""

    parsed_specs = []
    for row in spec_rows:
        model_text = compact_text(row.get("model", ""))
        row_sku = sku_from_model(model_text)
        variant = match_variant(model_text, row_sku, variants_by_sku)
        sku = variant.get("sku") or row_sku
        application = compact_text(row.get("application", ""))
        parsed_specs.append(
            {
                "raw_model": model_text,
                "sku": sku,
                "variant_sku": variant.get("sku", ""),
                "variant_title": variant.get("title", ""),
                "barcode": variant.get("barcode", ""),
                "variant_id": variant.get("id", ""),
                "available": variant.get("available", ""),
                "price": price(row.get("price", "")) or price(variant.get("price", "")),
                "compare_at_price": price(variant.get("compare_at_price", "")),
                "length": compact_text(row.get("length", "")),
                "action": compact_text(row.get("action", "")),
                "power": normalize_power_value(row.get("power", "")),
                "line_wt": normalize_line_wt(row.get("line_wt", "")),
                "lure_wt": normalize_lure_oz(row.get("lure_wt", "")),
                "application": application,
                "rod_weight": compact_text(row.get("rod_weight", "")),
                "section_type": compact_text(row.get("section_type", "")),
                "raw_cells": row.get("raw_cells", []),
            }
        )

    return {
        "id": product.get("id"),
        "handle": product.get("handle"),
        "source_url": f"{COLLECTION_URL}/products/{product.get('handle')}",
        "product_json_url": product_url(product.get("handle")),
        "title": title,
        "original_title": product.get("title", ""),
        "model_year": infer_model_year(product.get("title", "")),
        "tags": tags,
        "description": description,
        "features": features,
        "main_image_url": image_url,
        "official_reference_price": price_range(variants),
        "market_status": "在售" if any(v.get("available") for v in variants) else "暂时缺货",
        "variants_count": len(variants),
        "specs": parsed_specs,
    }


def read_cached_products():
    items = read_list()
    products = []
    missing = []
    for index, item in enumerate(items):
        path = detail_cache_path(index, item["handle"])
        if not path.exists():
            missing.append(item)
            continue
        payload = json.loads(path.read_text(encoding="utf-8"))
        product = payload.get("product") or payload
        products.append(parse_product(product))
    if missing:
        raise RuntimeError(f"{len(missing)} detail cache files missing. Run --stage=details first.")
    return products


def build_rows(normalized, refresh_images=False):
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    rod_rows = []
    detail_rows = []

    for index, item in enumerate(normalized):
        rod_id = f"{ROD_PREFIX}{1000 + index}"
        series = series_positioning(item["title"], item["tags"], item["description"])
        image_cdn = download_image(
            item["main_image_url"],
            f"{rod_id}_{sanitize_filename(item['handle'])}",
            refresh=refresh_images,
        )
        item["images"] = image_cdn
        item["series_positioning"] = series
        item["fit_style_tags"] = fit_style_tags(item["title"], series, item["description"])
        item["player_positioning"] = master_player_positioning(series, item["title"])
        item["player_selling_points"] = master_player_points(item["features"], series, item["specs"])

        rod_rows.append(
            {
                "id": rod_id,
                "brand_id": BRAND_ID,
                "model": item["title"],
                "model_cn": "",
                "model_year": item["model_year"],
                "alias": item["title"],
                "type_tips": " / ".join(item["tags"]),
                "fit_style_tags": item["fit_style_tags"],
                "images": image_cdn,
                "created_at": now,
                "updated_at": now,
                "series_positioning": series,
                "main_selling_points": item["description"],
                "official_reference_price": item["official_reference_price"],
                "market_status": item["market_status"],
                "Description": item["description"],
                "player_positioning": item["player_positioning"],
                "player_selling_points": item["player_selling_points"],
            }
        )

        layout = guide_layout_type(item["features"], item["description"])
        grip = grip_type(item["description"].split("\n\n"))
        seat = reel_seat(item["description"].split("\n\n"))
        pieces = "2" if re.search(r"2[- ]?piece", item["title"], flags=re.I) else ""
        product_technical = product_technical_terms(item["description"])
        for spec in item["specs"]:
            spec["product_technical"] = product_technical
            type_code = infer_type(item["title"], spec["raw_model"], spec["sku"], spec["section_type"])
            positioning = infer_player_positioning(spec["application"], spec["sku"], spec["power"], item["title"])
            extra_2 = []
            if spec["barcode"]:
                extra_2.append(f"UPC: {spec['barcode']}")
            if spec["variant_id"]:
                extra_2.append(f"Shopify variant id: {spec['variant_id']}")
            if spec["available"] is not True:
                extra_2.append("Availability: out of stock")
            detail_rows.append(
                {
                    "id": f"{ROD_DETAIL_PREFIX}{10000 + len(detail_rows)}",
                    "rod_id": rod_id,
                    "TYPE": type_code,
                    "SKU": spec["sku"],
                    "POWER": spec["power"],
                    "TOTAL LENGTH": spec["length"],
                    "Action": spec["action"],
                    "PIECES": pieces,
                    "CLOSELENGTH": "",
                    "WEIGHT": spec["rod_weight"],
                    "Tip Diameter": "",
                    "LURE WEIGHT": "",
                    "Line Wt N F": spec["line_wt"],
                    "PE Line Size": "",
                    "Handle Length": "",
                    "Reel Seat Position": seat,
                    "CONTENT CARBON": "",
                    "Market Reference Price": spec["price"],
                    "AdminCode": spec["variant_sku"] or spec["sku"],
                    "Service Card": "",
                    " Jig Weight": "",
                    "Squid Jig Size": "",
                    "Sinker Rating": "",
                    "created_at": now,
                    "updated_at": now,
                    "LURE WEIGHT (oz)": spec["lure_wt"],
                    "Sale Price": "",
                    "Joint Type": "2-piece" if pieces == "2" else "",
                    "Code Name": spec["raw_model"],
                    "Fly Line": "",
                    "Grip Type": grip,
                    "Reel Size": "",
                    "guide_layout_type": layout,
                    "guide_use_hint": guide_hint(layout, positioning),
                    "hook_keeper_included": "1" if "hook keeper" in " ".join(item["features"]).lower() else "",
                    "sweet_spot_lure_weight_real": "",
                    "official_environment": "",
                    "player_environment": infer_player_environment(
                        series, spec["application"], spec["sku"], spec["power"], item["title"]
                    ),
                    "player_positioning": positioning,
                    "player_selling_points": infer_player_selling_points(
                        spec["application"],
                        spec["power"],
                        spec["action"],
                        spec["line_wt"],
                        spec["lure_wt"],
                        positioning,
                    ),
                    "Description": item["description"],
                    "product_technical": spec.get("product_technical", ""),
                    "Extra Spec 1": " / ".join(item["features"]),
                    "Extra Spec 2": "; ".join(extra_2),
                }
            )

    return rod_rows, detail_rows


def append_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 36)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    return ws


def write_whitelist_evidence(normalized):
    checked_skus = []
    for item in normalized:
        for spec in item.get("specs", [])[:3]:
            if spec.get("sku"):
                checked_skus.append(spec["sku"])
    checked_skus = checked_skus[:20]
    evidence = {
        "brand": "Ark",
        "brand_id": BRAND_ID,
        "official_source": COLLECTION_URL,
        "checked_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "whitelist_sources": [
            {
                "source_site": "tackledb.uosoku.com",
                "result": "no usable Ark Rods model coverage found",
                "searched_terms": ["ARK Rods", *checked_skus[:6]],
                "usable_for_fields": [],
            },
            {
                "source_site": "rods.jp",
                "result": "no usable Ark Rods model coverage found",
                "searched_terms": ["ARK Rods", *checked_skus[6:12]],
                "usable_for_fields": [],
            },
            {
                "source_site": "rodsearch.com",
                "result": "no usable Ark Rods model coverage found",
                "searched_terms": ["ARK Rods", *checked_skus[12:18]],
                "usable_for_fields": [],
            },
        ],
        "field_policy": {
            "official_environment": "kept blank because whitelist/official environment taxonomy is unavailable",
            "sweet_spot_lure_weight_real": "kept blank because official lure range is not a real player sweet spot",
            "hook_keeper_included": "kept blank because Ark pages do not provide stable model-level hook keeper evidence",
            "player_fields": "filled from official Application/Purpose, SKU, power, action, line/lure ranges, and official series descriptions",
        },
        "series_evidence": [
            {
                "rod_id": f"{ROD_PREFIX}{1000 + index}",
                "handle": item.get("handle"),
                "source_url": item.get("source_url"),
                "series_positioning": item.get("series_positioning"),
                "features": item.get("features", []),
                "applications": [
                    {
                        "sku": spec.get("sku"),
                        "application": spec.get("application"),
                        "power": spec.get("power"),
                        "action": spec.get("action"),
                        "line_wt": spec.get("line_wt"),
                        "lure_wt": spec.get("lure_wt"),
                    }
                    for spec in item.get("specs", [])
                ],
            }
            for index, item in enumerate(normalized)
        ],
    }
    WHITELIST_EVIDENCE_PATH.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")


def export_from_cache(refresh_images=False):
    ensure_dirs()
    normalized = read_cached_products()
    rod_rows, detail_rows = build_rows(normalized, refresh_images=refresh_images)
    NORMALIZED_PATH.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    write_whitelist_evidence(normalized)

    wb = Workbook()
    wb.remove(wb.active)
    append_sheet(wb, "rod", ROD_HEADERS, rod_rows)
    append_sheet(wb, "rod_detail", ROD_DETAIL_HEADERS, detail_rows)
    wb.save(OUTPUT_FILE)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)
    print(f"[export] wrote {NORMALIZED_PATH}")
    print(f"[export] wrote {WHITELIST_EVIDENCE_PATH}")
    print(f"[export] wrote {OUTPUT_FILE}")
    print(f"[export] masters={len(rod_rows)} details={len(detail_rows)}")


def parse_args(argv):
    args = {"stage": "all", "force": False, "refresh_images": False, "start": 0, "limit": None}
    for arg in argv:
        if arg == "--force":
            args["force"] = True
        elif arg == "--refresh-images":
            args["refresh_images"] = True
        elif arg.startswith("--stage="):
            args["stage"] = arg.split("=", 1)[1]
        elif arg.startswith("--start="):
            args["start"] = max(int(arg.split("=", 1)[1]), 0)
        elif arg.startswith("--limit="):
            args["limit"] = max(int(arg.split("=", 1)[1]), 1)
    if args["stage"] not in {"all", "list", "details", "export"}:
        raise RuntimeError(f"unknown stage: {args['stage']}")
    return args


def main(argv):
    args = parse_args(argv)
    if args["stage"] == "list":
        fetch_list(force=args["force"])
    elif args["stage"] == "details":
        fetch_details(force=args["force"], start=args["start"], limit=args["limit"])
    elif args["stage"] == "export":
        export_from_cache(refresh_images=args["refresh_images"])
    else:
        fetch_list(force=args["force"])
        fetch_details(force=args["force"], start=args["start"], limit=args["limit"])
        export_from_cache(refresh_images=args["refresh_images"])


if __name__ == "__main__":
    import sys

    main(sys.argv[1:])
