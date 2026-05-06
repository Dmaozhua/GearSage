from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel
import json
import re

from openpyxl import load_workbook


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
REPORT_PATH = DATA_DIR / "evergreen_rod_player_fields_stage10_qa_report.json"
FIELDS = ["player_environment", "player_positioning", "player_selling_points"]

LIGHT_SALT = {"Ajing Jighead", "Mebaring Jighead", "Light Game Jighead", "Rockfish Rig"}
EGING = {"Eging", "Deep Eging", "Tip-run Eging", "Slack-jerk Eging"}
JIGGING = {"Slow Pitch Jig", "Long Fall Jig", "Offshore Jigging", "Light Jigging", "Micro Jigging"}
TAIRABA = {"Casting Tairaba", "Tairaba"}
BASS = {
    "Small Rubber Jig", "Jighead Rig", "Down Shot", "Neko Rig", "No Sinker", "Wacky Rig",
    "Power Finesse", "Hover Strolling", "Mid Strolling Jighead", "Bottom Strolling",
    "Texas Rig", "Light Texas", "Heavy Texas", "Free Rig", "Leaderless Down Shot",
    "Heavy Down Shot", "Rubber Jig", "Guarded Rubber Jig", "Football Jig", "Carolina Rig",
    "Heavy Carolina Rig", "High-density No Sinker", "Chatterbait", "Spinnerbait", "Buzzbait",
    "Swim Jig", "Vibration", "Metal Vibration", "Crankbait", "Deep Crankbait",
    "Shallow Crankbait", "Shad", "Minnow", "Jerkbait", "Topwater Plug", "Pencil Bait",
    "Popper", "Crawler Bait", "Spoon", "Small Plug", "Big Bait", "Swimbait", "Alabama Rig",
    "Frog", "Punching",
}

SOURCE_TERMS = ["白名单", "官网", "官方", "来源", "证据", "Plus Fishing", "tackledb", "retail", "外部"]
LOW_VALUE_TERMS = ["性能优秀", "适合多场景", "多场景使用", "泛泛", "通用", "万能"]


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def split_pairing(value):
    return [part.strip() for part in clean(value).split(" / ") if part.strip()]


def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    required = ["id", "SKU", "recommended_rig_pairing", "Description", "guide_use_hint", *FIELDS]
    missing = [field for field in required if field not in col]
    if missing:
        raise RuntimeError(f"missing columns: {missing}")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = {field: clean(ws.cell(row_idx, col[field]).value) for field in required}
        row["row"] = row_idx
        row["parts"] = split_pairing(row["recommended_rig_pairing"])
        rows.append(row)

    coverage = {}
    top_duplicates = {}
    for field in FIELDS:
        values = [row[field] for row in rows]
        coverage[field] = {
            "filled": sum(1 for value in values if value),
            "total": len(values),
            "unique": len(set(values)),
        }
        top_duplicates[field] = Counter(values).most_common(12)

    source_hits = []
    low_value_hits = []
    conflict_flags = []
    for row in rows:
        all_player_text = " ".join(row[field] for field in FIELDS)
        for term in SOURCE_TERMS:
            if term and term in all_player_text:
                source_hits.append({"row": row["row"], "id": row["id"], "sku": row["SKU"], "term": term})
        for term in LOW_VALUE_TERMS:
            if term and term in all_player_text:
                low_value_hits.append({"row": row["row"], "id": row["id"], "sku": row["SKU"], "term": term})

        parts = row["parts"]
        primary = parts[0] if parts else ""
        env = row["player_environment"]
        pos = row["player_positioning"]
        sell = row["player_selling_points"]
        guide = row["guide_use_hint"]
        text = f"{env} {pos} {sell}"
        sku = row["SKU"]
        sea_context = (
            sku.startswith(("ZAG", "SPR", "PSS", "NEOS"))
            or any(p in row["recommended_rig_pairing"] for p in ["Seabass", "Light Game", "Ajing", "Mebaring", "Rockfish"])
            or "海鱸" in guide
            or "輕海水" in guide
            or "Mebaring" in guide
        )
        flags = []
        if primary in LIGHT_SALT and ("船钓" in env or "近海船钓" in env or "Bass" in env):
            flags.append("light_salt_primary_environment_mismatch")
        if primary in EGING and "木虾" not in env and "烏賊" not in env:
            flags.append("eging_primary_environment_mismatch")
        if primary in TAIRABA and "Tairaba" not in env:
            flags.append("tairaba_primary_environment_mismatch")
        if primary in TAIRABA and ("Bass" in guide or "木蝦" in guide):
            flags.append("tairaba_guide_conflict")
        if primary in LIGHT_SALT and ("船釣鐵板" in guide or "木蝦泛用" in guide):
            flags.append("light_salt_guide_conflict")
        if primary in JIGGING and "船钓" not in env and "近海船钓" not in env and "木虾" not in env:
            flags.append("jigging_primary_environment_mismatch")
        if primary in BASS and not sea_context and ("船钓" in env or "轻型海水" in env or "海鲈" in env):
            flags.append("bass_primary_environment_mismatch")
        if (
            "Big Bait" not in row["recommended_rig_pairing"]
            and "Swimbait" not in row["recommended_rig_pairing"]
            and ("大饵远投" in env or "Big Bait" in text or "Swimbait" in text)
        ):
            flags.append("big_bait_without_pairing_evidence")
        if primary in LIGHT_SALT and "Rockfish Rig" in row["recommended_rig_pairing"] and "Rockfish" in row["recommended_rig_pairing"] and "Rockfish" not in text:
            flags.append("rockfish_pairing_not_reflected")
        if flags:
            conflict_flags.append({
                "row": row["row"],
                "id": row["id"],
                "sku": row["SKU"],
                "recommended_rig_pairing": row["recommended_rig_pairing"],
                "flags": flags,
                "player_environment": env,
                "player_positioning": pos,
                "player_selling_points": sell,
                "guide_use_hint": guide,
            })

    fill_samples = {}
    for cell_ref in ["A2", f"A{ws.max_row}", f"{ws.cell(1, col[FIELDS[0]]).coordinate[:-1]}2"]:
        fill_samples[cell_ref] = ws[cell_ref].fill.fgColor.rgb
    target_fill_samples = {
        f"{field}_row2": ws.cell(2, col[field]).fill.fgColor.rgb
        for field in FIELDS
    }
    target_fill_samples.update({
        f"{field}_last": ws.cell(ws.max_row, col[field]).fill.fgColor.rgb
        for field in FIELDS
    })

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_file": str(XLSX_PATH),
        "rows": len(rows),
        "coverage": coverage,
        "top_duplicates": top_duplicates,
        "source_term_hits": source_hits,
        "low_value_term_hits": low_value_hits,
        "conflict_flags": conflict_flags,
        "bottom_color_samples": {
            "row_group_fills": {"A2": ws["A2"].fill.fgColor.rgb, f"A{ws.max_row}": ws.cell(ws.max_row, 1).fill.fgColor.rgb},
            "target_field_fills": target_fill_samples,
        },
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    print(json.dumps({
        "rows": len(rows),
        "coverage": coverage,
        "source_term_hits": len(source_hits),
        "low_value_term_hits": len(low_value_hits),
        "conflict_flags": len(conflict_flags),
        "report_file": str(REPORT_PATH),
        "top_duplicates": top_duplicates,
        "bottom_color_samples": report["bottom_color_samples"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
