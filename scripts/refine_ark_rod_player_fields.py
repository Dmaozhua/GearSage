#!/usr/bin/env python3
import json
import re
import subprocess
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


DATA_DIR = Path("/Users/tommy/GearSage/GearSage-client/pkgGear/data_raw")
XLSX_PATH = DATA_DIR / "ark_rod_import.xlsx"
NORMALIZED_PATH = DATA_DIR / "ark_rod_normalized.json"
REPORT_PATH = DATA_DIR / "ark_rod_player_fields_refine_report.json"
SHADE_SCRIPT = Path("/Users/tommy/GearSage/scripts/shade_ark_rod_detail_groups.py")

PLAYER_FIELDS = ["player_environment", "player_positioning", "player_selling_points"]


def norm(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def key(value):
    return re.sub(r"[^A-Z0-9]+", "", norm(value).upper())


def parse_max_oz(value):
    nums = []
    for raw in re.findall(r"\d+(?:/\d+)?(?:\.\d+)?", norm(value)):
        try:
            nums.append(float(raw.split("/", 1)[0]) / float(raw.split("/", 1)[1]) if "/" in raw else float(raw))
        except (ValueError, ZeroDivisionError):
            pass
    return max(nums) if nums else 0


def first_pairing(row):
    return norm(row.get("recommended_rig_pairing")).split(" / ")[0]


def pairing_parts(row):
    return [part.strip() for part in norm(row.get("recommended_rig_pairing")).split("/") if part.strip()]


def compact_pairing(row, n=3):
    return " / ".join(pairing_parts(row)[:n])


def source_basis(spec):
    if norm(spec.get("application")):
        return "official_application + recommended_rig_pairing + spec_boundary"
    return "recommended_rig_pairing + spec_boundary + conservative_player_inference"


def load_spec_index():
    data = json.loads(NORMALIZED_PATH.read_text(encoding="utf-8"))
    out = {}
    for index, item in enumerate(data):
        rod_id = f"ARKR{1000 + index}"
        for spec in item.get("specs", []):
            spec = dict(spec)
            spec["series_title"] = item.get("title", "")
            spec["handle"] = item.get("handle", "")
            out[(rod_id, key(spec.get("sku", "")))] = spec
    return out


def category(row, spec):
    first = first_pairing(row)
    parts = set(pairing_parts(row))
    sku = norm(row.get("SKU"))
    title = norm(spec.get("series_title", ""))
    if "Ice" in title or sku.startswith(("CTI", "CIR")):
        return "ice"
    if sku.startswith(("IVW", "LCW", "LCWT")) and first in {"Shaky Head", "Ned Rig", "Neko Rig", "Damiki Rig"}:
        return "freshwater_finesse"
    if sku.startswith(("IVW", "LCW", "LCWT")) or first in {
        "Jigging Rap", "Ripping Rap", "Snap Jig", "Walleye Jig", "Bottom Bouncer Rig",
        "Slip Rig", "Trolling Crankbait", "Vertical Jig", "Panfish Jig", "Perch Spoon",
    }:
        return "walleye"
    if first in {"BFS Jighead Rig", "Small Jighead Rig", "Small Rubber Jig"} and ("BFS" in title or sku.startswith("BF")):
        return "bfs"
    if first in {"Down Shot", "Neko Rig", "Ned Rig", "Damiki Rig", "Shaky Head", "Jighead Minnow", "Jighead Rig", "Hair Jig", "Spybait", "Mid Strolling"}:
        return "finesse"
    if first in {"Rubber Jig", "Texas Rig", "Light Texas Rig", "Free Rig", "Carolina Rig", "Tube Rig", "Football Jig", "Dock Skipping Jig"}:
        if parts & {"Frog", "Flipping Jig", "Punching Rig"}:
            return "bottom_cover_mix"
        return "bottom_contact"
    if first in {"Frog", "Flipping Jig", "Punching Rig"}:
        return "heavy_cover"
    if first in {"Swim Jig", "Chatterbait", "Spinnerbait", "Buzzbait", "Underspin", "Blade Bait"}:
        if parts & {"Football Jig", "Carolina Rig", "Offshore Jig", "Rubber Jig", "Texas Rig"}:
            return "moving_bottom_mix"
        return "moving_bait"
    if first in {"Jerkbait", "Crankbait", "Mini Crankbait", "Small Crankbait", "Squarebill Crankbait", "Shallow Crankbait", "Mid-depth Crankbait", "Deep Crankbait", "Lipless Crankbait", "Vibration", "Topwater Plug", "Small Topwater Plug"}:
        return "hardbait"
    if first in {"Glidebait", "Swimbait", "Alabama Rig", "Umbrella Rig"}:
        return "big_bait"
    return "bass_general"


def environment_for(cat, row, spec):
    parts = set(pairing_parts(row))
    if cat == "ice":
        if any(item in parts for item in ["Deadstick Minnow", "Walleye Jig", "Ice Spoon"]):
            return "冰钓湖库 / Walleye 静置与轻抽"
        return "冰钓湖库 / Panfish、Perch 轻口短竿"
    if cat == "walleye":
        if "Trolling Crankbait" in parts or "Bottom Bouncer Rig" in parts:
            return "淡水湖库 / Walleye 拖钓、漂流和贴底搜索"
        return "淡水湖库 / Walleye 垂直抽停与轻量 jig"
    if cat == "freshwater_finesse":
        return "淡水湖库 / Bass 与 Walleye 开阔水细线慢钓"
    if cat == "bfs":
        return "淡水 Bass / 岸边、码头、浅滩和轻障碍 BFS"
    if cat == "finesse":
        if first_pairing(row) in {"Hair Jig", "Spybait", "Mid Strolling", "Jighead Minnow"}:
            return "淡水 Bass / 开阔水域、饵鱼层和前视声呐精细搜索"
        return "淡水 Bass / 开阔水域、码头边和草边精细慢钓"
    if cat == "bottom_contact":
        if "Dock Skipping Jig" in parts:
            return "淡水 Bass / 码头、树荫和近岸结构跳投"
        if "Carolina Rig" in parts or "Football Jig" in parts:
            return "淡水 Bass / 离岸结构、硬底和深浅交界"
        return "淡水 Bass / 木桩、石区、草边和结构区底操"
    if cat == "bottom_cover_mix":
        return "淡水 Bass / cover 边缘、码头和草洞的底操加 Frog"
    if cat == "heavy_cover":
        if "Punching Rig" in parts:
            return "淡水 Bass / 厚草垫、草洞和重 cover 穿透"
        return "淡水 Bass / 草区 Frog、木障碍和近岸重 cover"
    if cat == "moving_bottom_mix":
        return "淡水 Bass / 草边、水草通道、离岸结构的搜索与慢探"
    if cat == "moving_bait":
        if "Chatterbait" in parts:
            return "淡水 Bass / 草边、浑水和浅水覆盖区移动搜索"
        return "淡水 Bass / 岸线、草边和浅中层移动饵搜索"
    if cat == "hardbait":
        if any(item in parts for item in ["Deep Crankbait", "Mid-depth Crankbait"]):
            return "淡水 Bass / 硬底、坎位和离岸结构 crank 搜索"
        if any(item in parts for item in ["Topwater Plug", "Small Topwater Plug"]):
            return "淡水 Bass / 浅水、清晨傍晚和草边水面反应"
        return "淡水 Bass / 开阔水、岸线和草边硬饵反应搜索"
    if cat == "big_bait":
        return "淡水 Bass / 开阔水、岬角、浅滩边和大鱼巡游线"
    return "淡水 Bass / 岸边、码头、开阔水和常规结构泛用场景"


def positioning_for(cat, row, spec):
    pair = compact_pairing(row, 2)
    power = norm(row.get("POWER"))
    action = norm(row.get("Action"))
    if cat == "ice":
        return f"冰钓短竿 / {pair} 控线"
    if cat == "walleye":
        return f"Walleye 湖库技法竿 / {pair}"
    if cat == "freshwater_finesse":
        return f"淡水湖库细线精细竿 / {pair}"
    if cat == "bfs":
        return f"BFS 轻量精准竿 / {pair}"
    if cat == "finesse":
        return f"细线精细操作竿 / {pair}"
    if cat == "bottom_contact":
        return f"底部接触专向竿 / {pair}"
    if cat == "bottom_cover_mix":
        return f"底操与轻重 cover 切换竿 / {pair}"
    if cat == "heavy_cover":
        return f"重 cover 强力竿 / {pair}"
    if cat == "moving_bottom_mix":
        return f"移动饵与结构慢探切换竿 / {pair}"
    if cat == "moving_bait":
        return f"移动饵搜索竿 / {pair}"
    if cat == "hardbait":
        return f"反应硬饵操作竿 / {pair}"
    if cat == "big_bait":
        return f"大饵负载竿 / {pair}"
    return f"{power}/{action} Bass 泛用竿 / {pair}".strip("/")


def selling_points_for(cat, row, spec):
    pair = compact_pairing(row, 3)
    power = norm(row.get("POWER"))
    action = norm(row.get("Action"))
    line = norm(row.get("Line Wt N F"))
    lure = norm(row.get("LURE WEIGHT (oz)"))
    spec_text = " / ".join(x for x in [power, action, f"{line}lb" if line else "", f"{lure}oz" if lure else ""] if x)
    prefix = f"{spec_text} 的设定" if spec_text else "这支竿的设定"
    if cat == "ice":
        return f"{prefix}利于短距离小幅操饵和看竿尖；{pair} 场景下轻口反馈更清楚，静置和轻抽节奏更好控。"
    if cat == "walleye":
        return f"{prefix}适合湖库垂直控线和拖钓节奏；{pair} 切换时贴底感、停顿和轻咬信号更容易分辨。"
    if cat == "freshwater_finesse":
        return f"{prefix}服务开阔水细线精细钓组；{pair} 作钓时线弧、停顿和轻口反馈更容易控制，Bass 与 Walleye 慢钓都比较稳。"
    if cat == "bfs":
        return f"{prefix}偏轻线小饵；{pair} 能低弹道送入小标点，近中距离控线和轻口反馈更直接。"
    if cat == "finesse":
        return f"{prefix}服务细线精细钓组；{pair} 作钓时抛投落点、线弧管理和轻口补刺更稳定。"
    if cat == "bottom_contact":
        return f"{prefix}突出读底和结构感；{pair} 贴底时能分辨硬底、草边和障碍，补刺力量也更充足。"
    if cat == "bottom_cover_mix":
        return f"{prefix}兼顾底操和 cover 边缘强度；{pair} 之间切换时，能保持读底清楚并保留近障碍控鱼余量。"
    if cat == "heavy_cover":
        return f"{prefix}强调强力补刺和控鱼；{pair} 面对厚草、木障碍或草洞时，能更快把鱼带离 cover。"
    if cat == "moving_bottom_mix":
        return f"{prefix}兼顾搜索和慢探；{pair} 可先用移动饵找鱼，再用结构钓组补点，节奏切换宽容度高。"
    if cat == "moving_bait":
        return f"{prefix}适合持续搜索；{pair} 操作中震动、泳姿和碰草反馈更明确，长时间抛投不容易失控。"
    if cat == "hardbait":
        return f"{prefix}利于硬饵节奏；{pair} 需要的抽停、巻物和咬口缓冲更平衡，连续抛投稳定性更好。"
    if cat == "big_bait":
        return f"{prefix}提供大饵负载和控鱼余量；{pair} 远投、控摆和搏大鱼时更稳，不容易被高阻力饵拖散节奏。"
    return f"{prefix}适合 {pair} 等常见钓组；优势在于软硬饵切换宽容、抛投稳定和补刺力量均衡。"


def main():
    spec_index = load_spec_index()
    wb = load_workbook(XLSX_PATH)
    ws = wb["rod_detail"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    changed = []
    protected_notes = []

    for row_idx in range(2, ws.max_row + 1):
        row = {name: ws.cell(row=row_idx, column=idx).value for name, idx in col.items()}
        spec = spec_index.get((norm(row.get("rod_id")), key(row.get("SKU"))))
        if not spec:
            protected_notes.append({"row": row_idx, "sku": row.get("SKU"), "reason": "no normalized SKU match"})
            continue
        cat = category(row, spec)
        updates = {
            "player_environment": environment_for(cat, row, spec),
            "player_positioning": positioning_for(cat, row, spec),
            "player_selling_points": selling_points_for(cat, row, spec),
        }
        for field, new_value in updates.items():
            old_value = norm(ws.cell(row=row_idx, column=col[field]).value)
            if old_value != new_value:
                ws.cell(row=row_idx, column=col[field]).value = new_value
                changed.append(
                    {
                        "row": row_idx,
                        "id": row.get("id"),
                        "sku": row.get("SKU"),
                        "field": field,
                        "old": old_value,
                        "new": new_value,
                        "category": cat,
                        "recommended_rig_pairing": row.get("recommended_rig_pairing"),
                    }
                )

    wb.save(XLSX_PATH)
    subprocess.run(["python3", str(SHADE_SCRIPT)], check=True)

    wb_check = load_workbook(XLSX_PATH, data_only=True)
    ws_check = wb_check["rod_detail"]
    headers = [cell.value for cell in ws_check[1]]
    c = {name: idx for idx, name in enumerate(headers)}
    rows = list(ws_check.iter_rows(min_row=2, values_only=True))
    coverage = {
        field: {
            "filled": sum(1 for row in rows if norm(row[c[field]])),
            "total": len(rows),
            "unique": len({norm(row[c[field]]) for row in rows if norm(row[c[field]])}),
        }
        for field in PLAYER_FIELDS
    }
    category_counts = Counter()
    source_counts = Counter()
    row_summaries = []
    conflict_rows = []
    for idx, row in enumerate(rows, start=2):
        spec = spec_index.get((norm(row[c["rod_id"]]), key(row[c["SKU"]])))
        rec = norm(row[c["recommended_rig_pairing"]])
        env = norm(row[c["player_environment"]])
        pos = norm(row[c["player_positioning"]])
        sell = norm(row[c["player_selling_points"]])
        first = rec.split(" / ")[0] if rec else ""
        if any(word in f"{env} {pos} {sell}" for word in ["海水", "船钓", "岸投", "木虾", "SLSJ", "rockfish"]):
            conflict_rows.append({"row": idx, "sku": row[c["SKU"]], "reason": "unexpected non-Ark environment term"})
        if first in {"Frog", "Punching Rig", "Flipping Jig"} and "cover" not in f"{env} {sell}" and "障碍" not in f"{env} {sell}":
            conflict_rows.append({"row": idx, "sku": row[c["SKU"]], "reason": "heavy cover pairing not reflected"})
        if first in {"Down Shot", "Neko Rig", "Ned Rig"} and "精细" not in f"{env} {pos} {sell}":
            conflict_rows.append({"row": idx, "sku": row[c["SKU"]], "reason": "finesse pairing not reflected"})
        category_counts[norm(row[c["player_positioning"]]).split(" / ")[0]] += 1
        if spec:
            source_counts[source_basis(spec)] += 1
        row_summaries.append(
            {
                "row": idx,
                "id": row[c["id"]],
                "rod_id": row[c["rod_id"]],
                "sku": row[c["SKU"]],
                "recommended_rig_pairing": rec,
                "player_environment": env,
                "player_positioning": pos,
                "player_selling_points": sell,
                "source_basis": source_basis(spec) if spec else "missing_normalized_spec",
            }
        )

    report = {
        "fields": PLAYER_FIELDS,
        "updated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "xlsx": str(XLSX_PATH),
        "normalized": str(NORMALIZED_PATH),
        "refined_rows": len(rows),
        "refined_row_range": "2-227",
        "changed_cells": len(changed),
        "coverage": coverage,
        "source_basis_counts": dict(source_counts),
        "positioning_prefix_counts": dict(category_counts),
        "conflict_rows": conflict_rows,
        "protected_notes": protected_notes,
        "changed_rows": changed,
        "row_summaries": row_summaries,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"changed_cells": len(changed), "coverage": coverage, "conflicts": len(conflict_rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
