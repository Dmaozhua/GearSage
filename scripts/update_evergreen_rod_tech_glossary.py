#!/usr/bin/env python3
import json
import re
import subprocess
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from gear_data_paths import DATA_RAW_DIR, EXCEL_DIR, resolve_data_raw, resolve_excel

from bs4 import BeautifulSoup
from openpyxl import load_workbook

import importlib.util


ROOT = Path("/Users/tommy/GearSage")
DATA_DIR = DATA_RAW_DIR
GLOSSARY_PATH = ROOT / "GearSage-client/pkgGear/utils/官网解释_text_simple_rod.js"
XLSX_PATH = DATA_DIR / "evergreen_rod_import.xlsx"
EVIDENCE_PATH = DATA_DIR / "evergreen_rod_product_technical_official_evidence.json"
CACHE_DIR = DATA_DIR / "evergreen_rod_product_technical_cache"
APPLY_SCRIPT = ROOT / "scripts/apply_evergreen_rod_product_technical.py"

EVERGREEN_GROUP = "evergreen_rod_import"
EVERGREEN_BRAND = "evergreen"

OFFICIAL_TEXT_OVERRIDES = {
    "トレカ®T1100G": {
        "url": "https://www.evergreen-fishing.com/news_html/toraycat1100g_nanoalloy.php",
        "text": "『トレカ®T1100G』＆『ナノアロイ®技術』 ナノレベル（10億分の1）で繊維構造を緻密にコントロールする焼成技術により高強度と高弾性率化を両立し、30トンを頂点に弾性率が高くなるほど強度が低下するというカーボン繊維の力学特性マップ（下図）を塗り替えた画期的素材『トレカ®T1100G』と、革新的テクノロジーをベースに引張強度と耐衝撃性を両立したマトリクス樹脂技術『ナノアロイ®技術』。",
    },
    "ナノアロイ®技術": {
        "url": "https://www.evergreen-fishing.com/news_html/toraycat1100g_nanoalloy.php",
        "text": "『トレカ®T1100G』＆『ナノアロイ®技術』 ナノレベル（10億分の1）で繊維構造を緻密にコントロールする焼成技術により高強度と高弾性率化を両立し、30トンを頂点に弾性率が高くなるほど強度が低下するというカーボン繊維の力学特性マップ（下図）を塗り替えた画期的素材『トレカ®T1100G』と、革新的テクノロジーをベースに引張強度と耐衝撃性を両立したマトリクス樹脂技術『ナノアロイ®技術』。",
    },
    "高強度・高弾性率炭素繊維": {
        "url": "https://www.evergreen-fishing.com/special/concept_grandcobra.php",
        "text": "カーボン製造の最大手メーカーである東レ株式会社から、これまで技術難度が高いとされた高強度と高弾性率化の両立を実現したカーボン（炭素）繊維が発表された。20年に一度の画期的素材と謳われる『トレカ®「T1100G」』である。",
    },
    "トレカ®M40X": {
        "url": "https://www.evergreen-fishing.com/news_html/toraycam40x.php",
        "text": "『トレカ®M40X』＆『ナノアロイ®技術』 技術難易度が高く、大きな課題とされる繊維強度と高弾性率の両立を極限まで追求し、高弾性率（40t）を保持したまま繊維強度を約30％向上させたトレカ最新カーボン素材。革新的なマトリクス樹脂技術『ナノアロイ®技術』と組み合わせることで、従来の40tカーボンプリプレグと比較して引張強度・圧縮強度・耐衝撃性が大幅に向上、理想的な剛性設計が可能となり製品の軽量化にも貢献。",
    },
    "ZIGZAGガイドシステム": {
        "url": "https://www.evergreen-fishing.com/news_html/delgesu_story.php",
        "text": "さらに、このロッドには象徴的な新構造が搭載されている。それは青木氏が考案した『ZIGZAGガイドシステム』である。その名のとおり、ティップセクションガイドを左右ジグザグに配置しており、搭載個数も非常に多い。ラインからのボトム変化やバスのバイト信号を同時多発的にブランクスに増幅伝達するギリギリの角度にセットしている。",
    },
    "カレイド・スーパークワトロクロス": {
        "url": "https://www.evergreen-fishing.com/special/concept_grandcobra.php",
        "text": "『トレカ®「T1100G」』の特性を打ち消すことなく、その優位性をより発揮させるために、全身をカレイドスーパークワトロクロスで補強しネジレを抑制。ネジレにくさはロッドのパワーロス減少につながり、ひいてはビッグベイトのキャスト精度やフッキングパワーの向上に貢献する。",
    },
    "オールダブルフットガイド": {
        "url": "https://www.evergreen-fishing.com/special/concept_inspirare.php",
        "text": "用途・目的に合わせ一部機種にオールダブルフットガイド採用。対超ビッグフィッシュやビッグベイトを扱う場合、あるいはオカッパリでのハードな使用等を考慮して、意図的にセミマイクロ・オールダブルガイドを採用した例外機種も一部存在する。",
    },
    "ブランクタッチ方式": {
        "url": "https://www.evergreen-fishing.com/special/concept_inspirare.php",
        "text": "スリムタイプのブランクタッチ式リールシートを採用したことも軽量化に大きく貢献しており、そのうえグリップが細身で握りやすいことから手首の自由度が増し、キャスト精度の向上に一役買っている。",
    },
}


def load_apply_module():
    spec = importlib.util.spec_from_file_location("evergreen_tech", APPLY_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def norm(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def js_string(value):
    return json.dumps(value, ensure_ascii=False)


def load_current_module():
    code = (
        "const m=require('./GearSage-client/pkgGear/utils/官网解释_text_simple_rod.js');"
        "process.stdout.write(JSON.stringify(m));"
    )
    result = subprocess.run(
        ["node", "-e", code],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout)


def read_evergreen_terms():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["rod_detail"]
    header = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}
    terms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[col["product_technical"]]
        if not value:
            continue
        for term in str(value).split(" / "):
            term = term.strip()
            if term and term not in terms:
                terms.append(term)
    return terms


def cache_path(url):
    import hashlib

    return CACHE_DIR / f"{hashlib.sha1(url.encode('utf-8')).hexdigest()}.html"


def source_blocks(url):
    path = cache_path(url)
    if not path.exists():
        return []
    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
    blocks = []
    title_area = soup.select_one(".titleArea")
    if title_area:
        for node in title_area.find_all(["em", "p", "dd", "li"], recursive=True):
            text = norm(node.get_text(" ", strip=True))
            if len(text) >= 12:
                blocks.append(text)
        text = norm(title_area.get_text(" ", strip=True))
        if len(text) >= 12:
            blocks.append(text)
    for block in soup.select("li.item-feature"):
        text = norm(block.get_text(" ", strip=True))
        if len(text) >= 12:
            blocks.append(text)
    box = soup.select_one(".productsBox")
    if box:
        text = norm(box.get_text(" ", strip=True))
        if len(text) >= 12:
            blocks.append(text)
    # Preserve order while removing duplicate text.
    seen = set()
    unique = []
    for text in blocks:
        if text not in seen:
            unique.append(text)
            seen.add(text)
    return unique


def concise_official_text(text):
    text = norm(text)
    text = re.sub(r"PAGE TOP.*$", "", text).strip()
    text = re.sub(r"※リールシートフード.*$", "", text).strip()
    if len(text) <= 360:
        return text
    pieces = re.split(r"(?<=[。！？])\s*|<br\s*/?>", text)
    picked = []
    size = 0
    for piece in pieces:
        piece = norm(piece)
        if not piece:
            continue
        if size + len(piece) > 360 and picked:
            break
        picked.append(piece)
        size += len(piece)
        if size >= 240:
            break
    return norm(" ".join(picked))[:420]


def build_official_evidence(terms):
    extractor = load_apply_module()
    evidence = json.loads(EVIDENCE_PATH.read_text(encoding="utf-8"))["items"]
    term_sources = defaultdict(list)
    term_texts = defaultdict(list)

    for item in evidence:
        url = item.get("source_url") or ""
        value = item.get("product_technical") or ""
        row_terms = [part.strip() for part in value.split(" / ") if part.strip()]
        for term in row_terms:
            if url and url not in term_sources[term]:
                term_sources[term].append(url)

        for block in source_blocks(url):
            extracted = extractor.extract_product_technical(block)
            block_terms = [part.strip() for part in extracted.split(" / ") if part.strip()]
            for term in block_terms:
                if term in terms:
                    text = concise_official_text(block)
                    if text and text not in term_texts[term]:
                        term_texts[term].append(text)

    result = {}
    for term in terms:
        texts = sorted(term_texts.get(term, []), key=lambda x: (len(x), x))
        sources = term_sources.get(term, [])
        if term in OFFICIAL_TEXT_OVERRIDES:
            override = OFFICIAL_TEXT_OVERRIDES[term]
            texts = [override["text"]] + [text for text in texts if text != override["text"]]
            if override["url"] not in sources:
                sources = [override["url"]] + sources
        result[term] = {
            "text": texts[0] if texts else "",
            "sources": sources[:5],
        }
    return result


SIMPLE_OVERRIDES = {
    "トレカ®M40X": "高弹高强度碳纤维，用在需要轻量、高感度和操作响应的型号。",
    "トレカ®T1100G": "高强度碳纤维，提升竿身强度、韧性和控鱼稳定性。",
    "ナノアロイ®技術": "配合高强度碳材的树脂技术，让轻量、强度和韧性更平衡。",
    "高強度・高弾性率炭素繊維": "高强度高弹性碳纤维，兼顾硬挺响应和受力强度。",
    "カレイド・スーパークワトロクロス": "Kaleido 系列的多轴补强，重点抑制扭转和竿身晃动。",
    "ZIGZAGガイドシステム": "ZIGZAG 导环系统通过特殊导环角度增强线接触和感度反馈。",
    "ナノアロイ®技術": "配合 T1100G 等高强度碳材的树脂技术，平衡轻量、强度和韧性。",
    "AAAシャンパンコルク": "高等级香槟软木握把，提升握持质感和重量控制。",
    "ブランクタッチ方式": "手部更接近竿坯，轻咬、底感和震动传递更直接。",
    "アンサンドフィニッシュ": "竿坯不打磨/少涂装，减少多余重量并保留直接反馈。",
    "完全ブランクスルー構造": "竿坯贯通到握把末端，力量和震动传递更连续。",
    "高品位スピゴットジョイント": "高品质插节接合，兼顾多节收纳、顺畅弯曲和力量传递。",
    "スレッドレス仕様": "减少导环绑线和涂装重量，让竿稍反应更轻快。",
    "エバーグリーンオリジナルアーバー": "Evergreen 自有轻量垫圈/填隙结构，提升握把处感度传递。",
    "EGスーパーバスシート・スリム": "Evergreen 低轮廓卷线器座，强调贴手握持和操作稳定性。",
}


def simple_text(term):
    if term in SIMPLE_OVERRIDES:
        return SIMPLE_OVERRIDES[term]
    if re.match(r"^\d+トンカーボン$", term):
        return f"{term}表示该型号竿坯使用的碳布等级，用于调整轻量、硬挺度和回弹。"
    if "高弾性" in term:
        return "高弹性碳素材让竿身反应更快，控饵和底感反馈更清晰。"
    if "中弾性" in term:
        return "中弹性碳素材让竿身保留韧性和负载感，适合更宽容的操作。"
    if "低弾性" in term:
        return "低弹性碳素材让竿身更柔顺，抛投和咬口追随更宽容。"
    if "低レジン" in term:
        return "低树脂碳素材减少多余重量，让竿身更清脆、更有张力。"
    if "高レジン" in term:
        return "高树脂碳素材带来更湿润的弯曲手感，降低过硬过弹的突兀感。"
    if "カーボン" in term and "ソリッド" in term:
        return "碳实心竿稍提升轻量钓组的咬口识别和精细操作能力。"
    if "チューブラー" in term:
        return "管状竿坯/竿稍保留更直接的回弹和操控响应。"
    if "グラス" in term:
        return "玻纤复合让竿身更柔顺，硬饵咬口更不容易弹开。"
    if "4軸" in term or "クロス" in term:
        return "多轴碳布补强用于抑制竿身扭转，提升抛投、抽动和控鱼稳定性。"
    if "トルザイト" in term and "ガイド" in term:
        return "TorZite 导环更轻、内径利用率高，有利于减重、出线和感度。"
    if "SiC" in term and "ガイド" in term:
        return "SiC 导环耐磨顺滑，适合长时间出线和较高负载使用。"
    if "Kガイド" in term:
        return "K 导环配置降低缠线风险，让 PE 或较细线出线更稳定。"
    if "LKWガイド" in term:
        return "LKW 导环强调轻量和强度平衡，适合承受较高负载的导环段。"
    if "LGトップ" in term or "KGトップ" in term:
        return "轻量顶导环降低竿稍负担，减少缠线并提升竿稍反应。"
    if "リールシート" in term or "シート" in term:
        return "卷线器座影响握持贴合、重量和手部感度传递。"
    if "カーボンパイプ" in term or "カーボンスリーブ" in term:
        return "碳管/碳套件用于减重和加强握把区域的整体感。"
    if "コルク" in term:
        return "软木握把兼顾轻量、触感和长时间握持舒适度。"
    if "ダブルフット" in term:
        return "双脚导环提升导环支撑强度，适合更高负载或强力控鱼。"
    if "シングルフット" in term:
        return "单脚导环减轻竿身负担，让竿稍和中前段动作更轻快。"
    return "该技术用于对应型号的竿坯、导环或握把配置，影响重量、感度、出线和控鱼稳定性。"


def source_obj(url):
    return {"group": EVERGREEN_GROUP, "url": url}


def merge_sources(existing, urls):
    merged = list(existing or [])
    seen = {(item.get("group"), item.get("url")) for item in merged if isinstance(item, dict)}
    for url in urls:
        item = source_obj(url)
        key = (item["group"], item["url"])
        if key not in seen:
            merged.append(item)
            seen.add(key)
    return merged


def merge_groups(existing):
    groups = list(existing or [])
    if EVERGREEN_GROUP not in groups:
        groups.append(EVERGREEN_GROUP)
    return groups


def render_value(value, indent=0):
    space = " " * indent
    if isinstance(value, dict):
        if not value:
            return "{}"
        lines = ["{"]
        items = list(value.items())
        for idx, (key, val) in enumerate(items):
            comma = "," if idx < len(items) - 1 else ""
            lines.append(f'{space}  {js_string(key)}: {render_value(val, indent + 2)}{comma}')
        lines.append(f"{space}}}")
        return "\n".join(lines)
    if isinstance(value, list):
        if not value:
            return "[]"
        if all(not isinstance(item, (dict, list)) for item in value):
            return "[\n" + ",\n".join(f"{space}  {js_string(item)}" for item in value) + f"\n{space}]"
        lines = ["["]
        for idx, item in enumerate(value):
            comma = "," if idx < len(value) - 1 else ""
            rendered = render_value(item, indent + 2)
            lines.append(f"{space}  {rendered}{comma}")
        lines.append(f"{space}]")
        return "\n".join(lines)
    return js_string(value)


def render_js(meta, tech_glossary, terms_by_brand):
    lines = []
    lines.append(f"const meta = {render_value(meta, 0)};\n")
    lines.append('const JACKALL_GROUP = "jackall_rod_import";')
    lines.append('const JACKALL_BRAND = "jackall";')
    lines.append('const SHIMANO_GROUP = "shimano_rod_import";')
    lines.append('const ARK_GROUP = "ark_rod_import";')
    lines.append('const NORIES_GROUP = "nories_rod_import";')
    lines.append('const DSTYLE_GROUP = "dstyle_rod_import";')
    lines.append('const EVERGREEN_GROUP = "evergreen_rod_import";\n')
    lines.append("function evergreenSource(url) {")
    lines.append("  return {")
    lines.append('    "group": EVERGREEN_GROUP,')
    lines.append('    "url": url')
    lines.append("  };")
    lines.append("}\n")
    lines.append(f"const techGlossary = {render_value(tech_glossary, 0)};\n")
    lines.append(f"const termsByBrand = {render_value(terms_by_brand, 0)};\n")
    lines.append("module.exports = {")
    lines.append("  meta,")
    lines.append("  techGlossary,")
    lines.append("  termsByBrand")
    lines.append("};")
    return "\n".join(lines) + "\n"


def main():
    current = load_current_module()
    tech_glossary = current["techGlossary"]
    terms_by_brand = current["termsByBrand"]
    terms = read_evergreen_terms()
    evidence = build_official_evidence(terms)

    for term in terms:
        official_text = evidence[term]["text"]
        urls = evidence[term]["sources"]
        if term in tech_glossary:
            entry = tech_glossary[term]
            old_text = entry.get("text", "")
            if "\nEVERGREEN: " in old_text:
                old_text = old_text.split("\nEVERGREEN: ", 1)[0]
            if official_text and official_text not in old_text:
                entry["text"] = f"{old_text}\nEVERGREEN: {official_text}" if old_text else official_text
            else:
                entry["text"] = old_text
            if not entry.get("text_simple"):
                entry["text_simple"] = simple_text(term)
            elif term in SIMPLE_OVERRIDES or term not in current["termsByBrand"].get(EVERGREEN_BRAND, []):
                # Keep explanations generic and brand-neutral for shared technology names.
                entry["text_simple"] = simple_text(term)
            entry["sources"] = merge_sources(entry.get("sources"), urls)
            entry["groups"] = merge_groups(entry.get("groups"))
        else:
            tech_glossary[term] = {
                "text": official_text,
                "text_simple": simple_text(term),
                "sources": [source_obj(url) for url in urls[:5]],
                "groups": [EVERGREEN_GROUP],
            }

    existing_brand_terms = terms_by_brand.get(EVERGREEN_BRAND, [])
    merged_brand_terms = list(existing_brand_terms)
    for term in terms:
        if term not in merged_brand_terms:
            merged_brand_terms.append(term)
    terms_by_brand[EVERGREEN_BRAND] = merged_brand_terms

    meta = current.get("meta", {})
    meta["generatedAt"] = datetime.now(timezone(timedelta(hours=8))).replace(microsecond=0).isoformat()
    meta["totalTerms"] = len(tech_glossary)
    meta["explainedTerms"] = sum(1 for item in tech_glossary.values() if norm(item.get("text")))

    GLOSSARY_PATH.write_text(render_js(meta, tech_glossary, terms_by_brand), encoding="utf-8")

    blank = [term for term in terms if not evidence[term]["text"]]
    print(json.dumps({
        "evergreen_terms": len(terms),
        "glossary_total": meta["totalTerms"],
        "glossary_explained": meta["explainedTerms"],
        "evergreen_blank_official_text": len(blank),
        "blank_terms_sample": blank[:30],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
